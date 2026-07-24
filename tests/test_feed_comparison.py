import io
import tempfile
import unittest
from unittest.mock import Mock, patch
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app import (
    download_comparison_feed,
    detect_supplier_feed_format,
    normalize_supplier_model_field,
    prepare_supplier_feed_item,
    read_supplier_feed_rows,
    supplier_feed_basic_auth_for_url,
    write_feed_comparison_workbook,
)


class FeedComparisonTest(unittest.TestCase):
    def test_supplier_model_field_is_exact_and_case_insensitive(self):
        content = b"""<?xml version="1.0" encoding="UTF-8"?>
        <catalog>
          <offer id="1">
            <name>Electric oven ABC 100</name>
            <vendorCode>ABC-100</vendorCode>
            <price>49990</price>
            <vendor>Example</vendor>
            <url>https://supplier.example/products/abc-100</url>
          </offer>
          <offer id="2">
            <name>Wrong field</name>
            <vendorCodeExtra>DO-NOT-USE</vendorCodeExtra>
          </offer>
        </catalog>
        """

        rows = read_supplier_feed_rows(content, "vendorcode")

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["source_model"], "ABC-100")
        self.assertEqual(rows[0]["name"], "Electric oven ABC 100")
        self.assertEqual(rows[0]["price"], "49990")

    def test_xml_tag_notation_is_normalized_to_field_name(self):
        self.assertEqual(normalize_supplier_model_field("<model>"), "model")
        self.assertEqual(normalize_supplier_model_field("</vendorCode>"), "vendorCode")
        self.assertEqual(normalize_supplier_model_field("offer/model"), "model")
        self.assertEqual(normalize_supplier_model_field("@sku"), "@sku")
        self.assertEqual(normalize_supplier_model_field("param:Модель"), "param:Модель")

    def test_supplier_param_and_attribute_fields_are_supported(self):
        content = b"""<catalog>
          <offer sku="SKU-10"><name>First</name></offer>
          <offer><name>Second</name><param name="Model">MODEL-20</param></offer>
        </catalog>"""

        attribute_rows = read_supplier_feed_rows(content, "@sku")
        param_rows = read_supplier_feed_rows(content, "param:Model")

        self.assertEqual([row["source_model"] for row in attribute_rows], ["SKU-10"])
        self.assertEqual([row["source_model"] for row in param_rows], ["MODEL-20"])

    def test_cp1251_semicolon_csv_is_detected_and_read_by_exact_model_column(self):
        content = (
            "Код товара;Категория;Фирма;Модель;Цена;Наличие на складе\r\n"
            "00079578;;Kitchen Aid;5KES8556ESX;199990;1\r\n"
            "00041425;;Kitchen Aid;5KCM1209EER;22990;3\r\n"
        ).encode("cp1251")

        rows = read_supplier_feed_rows(content, "Модель")

        self.assertEqual(detect_supplier_feed_format(content), "csv")
        self.assertEqual([row["source_model"] for row in rows], ["5KES8556ESX", "5KCM1209EER"])
        self.assertEqual(rows[0]["brand"], "Kitchen Aid")
        self.assertEqual(rows[0]["price"], "199990")
        self.assertEqual(rows[0]["name"], "5KES8556ESX")

    def test_utf8_comma_csv_is_supported_without_changing_xml_behavior(self):
        content = (
            "name,model,price,brand,url\n"
            "Electric oven ABC 100,ABC-100,49990,Example,https://supplier.example/abc-100\n"
        ).encode("utf-8")

        rows = read_supplier_feed_rows(content, "model")

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["source_model"], "ABC-100")
        self.assertEqual(rows[0]["name"], "Electric oven ABC 100")
        self.assertEqual(rows[0]["url"], "https://supplier.example/abc-100")

    def test_xlsx_feed_is_detected_and_uses_existing_model_candidate_logic(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([])
        sheet.append([])
        sheet.append(["code", "category", "brand", "name", "price"])
        sheet.append(
            [
                "00-00162357",
                "Автомобильные холодильники",
                "LIBHOF",
                "Автомобильный холодильник Libhof K-26 12В/24В",
                22200,
            ]
        )
        output = io.BytesIO()
        workbook.save(output)
        workbook.close()
        content = output.getvalue()

        rows = read_supplier_feed_rows(content, "name")
        prepared = prepare_supplier_feed_item(rows[0])

        self.assertEqual(detect_supplier_feed_format(content), "xlsx")
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["price"], "22200")
        self.assertEqual(rows[0]["brand"], "LIBHOF")
        self.assertIsNotNone(prepared)
        self.assertIn("K-26", prepared["model_candidates"])

    def test_csv_model_column_name_is_exact(self):
        content = "Модель товара;Цена\nABC-100;100\n".encode("cp1251")

        with self.assertRaisesRegex(ValueError, 'не найдена колонка модели "Модель"'):
            read_supplier_feed_rows(content, "Модель")

    @patch.dict(
        "os.environ",
        {
            "SUPPLIER_FEED_BASIC_AUTH_RULES_JSON": (
                '[{"host":"secure.example","username":"feed-user","password":"feed-password"}]'
            )
        },
        clear=False,
    )
    def test_supplier_basic_auth_is_selected_by_host(self):
        self.assertEqual(
            supplier_feed_basic_auth_for_url("https://secure.example/export/feed.csv"),
            ("feed-user", "feed-password"),
        )
        self.assertIsNone(supplier_feed_basic_auth_for_url("https://public.example/feed.xml"))

    @patch.dict(
        "os.environ",
        {
            "SUPPLIER_FEED_BASIC_AUTH_RULES_JSON": (
                '[{"host":"secure.example","username":"feed-user","password":"feed-password"}]'
            )
        },
        clear=False,
    )
    @patch("app.make_feed_session")
    def test_comparison_feed_download_uses_basic_auth_only_for_matching_feed(self, make_session):
        response = Mock(status_code=200, content=b"model;price\nABC-100;100\n", headers={})
        response.raise_for_status.return_value = None
        session = Mock()
        session.get.return_value = response
        make_session.return_value = session

        content = download_comparison_feed("https://secure.example/export/feed.csv")

        self.assertEqual(content, response.content)
        session.get.assert_called_once_with(
            "https://secure.example/export/feed.csv",
            timeout=60,
            auth=("feed-user", "feed-password"),
        )

    def test_workbook_has_one_sheet_per_supplier(self):
        supplier_results = [
            {
                "name": "Supplier A",
                "rows": [
                    {
                        "row": 1,
                        "name": "Product A",
                        "price": "100",
                        "brand": "Brand",
                        "model_candidates": "ABC-100",
                        "selected_model": "ABC-100",
                        "missing_on": "Нет на МегаКухня",
                    }
                ],
            },
            {"name": "Supplier B", "rows": []},
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "result.xlsx"
            write_feed_comparison_workbook(path, supplier_results)
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                self.assertEqual(workbook.sheetnames, ["Supplier A", "Supplier B"])
                self.assertEqual(workbook["Supplier A"]["G2"].value, "Нет на МегаКухня")
                self.assertEqual(workbook["Supplier B"].max_row, 1)
            finally:
                workbook.close()

    def test_supplier_exclusions_and_replace_rules_are_applied_before_comparison(self):
        item = {
            "source_model": "Brand ABC-100",
            "name": "Brand Electric oven ABC-100",
            "brand": "Brand",
        }

        prepared = prepare_supplier_feed_item(item, [], "Brand |")
        excluded = prepare_supplier_feed_item(item, ["Electric oven"], "Brand |")
        unchanged = prepare_supplier_feed_item(item)

        self.assertIsNotNone(prepared)
        self.assertEqual(prepared["name"], "Electric oven ABC-100")
        self.assertIn("ABC-100", prepared["model_candidates"])
        self.assertIsNone(excluded)
        self.assertIsNotNone(unchanged)
        self.assertEqual(unchanged["name"], item["name"])


if __name__ == "__main__":
    unittest.main()
