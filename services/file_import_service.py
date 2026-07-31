"""Extracted application service module."""

from services.core_service import (
    BeautifulSoup,
    Brand,
    Dict,
    ET,
    EXPORT_DIR,
    FILE_IMPORT_ACTIVE_STATUSES,
    FILE_IMPORT_ALLOWED_SUFFIXES,
    FILE_IMPORT_DIR,
    FILE_IMPORT_RESULT_FIELDS,
    FileImport,
    Iterable,
    List,
    MSK_TZ,
    Optional,
    Path,
    SessionLocal,
    Set,
    VISUAL_MODEL_TRANSLATION,
    csv,
    datetime,
    file_import_lock,
    file_import_stop_event,
    file_import_worker_thread,
    g,
    html_lib,
    io,
    news_lock,
    news_settings,
    normalize_file_import_exclusions,
    normalize_file_import_rules_text,
    normalize_model_key,
    output_text,
    re,
    safe_filename,
    select,
    session_scope,
    threading,
    time,
)
from services.scraping_service import clean_text, prepare_rule_model


class FileImportStopped(Exception):
    pass


def make_file_import_state(status: str = "idle") -> Dict[str, object]:
    return {
        "status": status,
        "stage": "",
        "percent": 0,
        "current_row": 0,
        "total_rows": 0,
        "processed_rows": 0,
        "excluded_rows": 0,
        "found_rows": 0,
        "missing_rows": 0,
        "model_not_found_rows": 0,
        "error": "",
        "started_at": "",
        "finished_at": "",
        "elapsed_seconds": 0,
        "result_filename": "",
    }


def normalize_file_import_state(value: object) -> Dict[str, object]:
    state = {**make_file_import_state(), **(value if isinstance(value, dict) else {})}
    state["status"] = str(state.get("status") or "idle")
    if state["status"] == "stopping":
        state["status"] = "stopped"
        state["stage"] = "Остановлено"
    for field in (
        "percent",
        "current_row",
        "total_rows",
        "processed_rows",
        "excluded_rows",
        "found_rows",
        "missing_rows",
        "model_not_found_rows",
        "elapsed_seconds",
    ):
        try:
            state[field] = int(float(state.get(field) or 0))
        except (TypeError, ValueError):
            state[field] = 0
    state["percent"] = max(0, min(100, state["percent"]))
    return state


def is_file_import_active_state(state: object) -> bool:
    return str((state if isinstance(state, dict) else {}).get("status") or "") in FILE_IMPORT_ACTIVE_STATUSES


def file_import_path_for_row(row: FileImport) -> Optional[Path]:
    file_meta = row.file if isinstance(row.file, dict) else {}
    filename = str(file_meta.get("stored_filename") or "").strip()
    if not filename:
        return None
    base_dir = FILE_IMPORT_DIR.resolve()
    path = (FILE_IMPORT_DIR / filename).resolve()
    if base_dir not in path.parents or not path.exists() or not path.is_file():
        return None
    return path


def update_file_import_state(db_session, **kwargs: object) -> Dict[str, object]:
    row = get_file_import_row(db_session)
    state = normalize_file_import_state(getattr(row, "state", {}) or {})
    state.update(kwargs)
    row.state = normalize_file_import_state(state)
    db_session.flush()
    return dict(row.state)


def stop_file_import_if_requested(stop_event: Optional[threading.Event]) -> None:
    if stop_event and stop_event.is_set():
        raise FileImportStopped()


def clear_file_import_storage() -> None:
    FILE_IMPORT_DIR.mkdir(parents=True, exist_ok=True)
    for path in FILE_IMPORT_DIR.iterdir():
        if path.is_file():
            try:
                path.unlink()
            except OSError:
                continue


def stored_file_import_files() -> List[Path]:
    FILE_IMPORT_DIR.mkdir(parents=True, exist_ok=True)
    return sorted(
        [
            path
            for path in FILE_IMPORT_DIR.iterdir()
            if path.is_file() and path.suffix.lower() in FILE_IMPORT_ALLOWED_SUFFIXES
        ],
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    )


def get_file_import_row(db_session=None) -> FileImport:
    db = db_session or g.db
    row = db.get(FileImport, 1)
    if row is None:
        row = FileImport(
            id=1,
            exclusions=[],
            model_field="",
            price_field="",
            replace_rules="",
            file={},
            state=make_file_import_state(),
        )
        db.add(row)
        db.flush()
    else:
        normalized_state = normalize_file_import_state(getattr(row, "state", {}) or {})
        if row.state != normalized_state:
            row.state = normalized_state
            db.flush()
        normalized_exclusions = normalize_file_import_exclusions(row.exclusions)
        if row.exclusions != normalized_exclusions:
            row.exclusions = normalized_exclusions
            db.flush()
        normalized_model_field = clean_text(str(getattr(row, "model_field", "") or ""))
        if row.model_field != normalized_model_field:
            row.model_field = normalized_model_field
            db.flush()
        normalized_price_field = clean_text(str(getattr(row, "price_field", "") or ""))
        if getattr(row, "price_field", "") != normalized_price_field:
            row.price_field = normalized_price_field
            db.flush()
        normalized_replace_rules = normalize_file_import_rules_text(getattr(row, "replace_rules", ""))
        if row.replace_rules != normalized_replace_rules:
            row.replace_rules = normalized_replace_rules
            db.flush()
    if not isinstance(row.file, dict) or not row.file.get("stored_filename"):
        stored_files = stored_file_import_files()
        if stored_files:
            path = stored_files[0]
            row.file = {
                "original_filename": path.name.split("_", 3)[-1] if "_" in path.name else path.name,
                "stored_filename": path.name,
                "uploaded_at": datetime.fromtimestamp(path.stat().st_mtime, MSK_TZ).isoformat(timespec="seconds"),
            }
            db.flush()
    return row


def current_file_import_path() -> Optional[Path]:
    row = get_file_import_row()
    return file_import_path_for_row(row)


def resolve_file_import_export_path(value: str) -> Optional[Path]:
    filename = Path(str(value or "")).name
    if not filename:
        return None
    path = (EXPORT_DIR / filename).resolve()
    if EXPORT_DIR.resolve() not in path.parents or not path.exists() or not path.is_file():
        return None
    return path


def remove_file_import_export(row: FileImport) -> None:
    candidates = [str(getattr(row, "export_path", "") or "")]
    file_meta = row.file if isinstance(row.file, dict) else {}
    candidates.append(str(file_meta.get("result_filename") or ""))
    for candidate in candidates:
        path = resolve_file_import_export_path(candidate)
        if path:
            try:
                path.unlink()
            except OSError:
                pass
    original_filename = str(file_meta.get("original_filename") or file_meta.get("filename") or "").strip()
    if original_filename:
        prefix = f"Новинки_{safe_filename(Path(original_filename).stem or 'file')}_"
        for pattern in (f"{prefix}*.csv", f"{prefix}*.xlsx"):
            for path in EXPORT_DIR.glob(pattern):
                if path.is_file():
                    try:
                        path.unlink()
                    except OSError:
                        pass


def public_file_import_state() -> Dict[str, object]:
    row = get_file_import_row()
    path = file_import_path_for_row(row)
    file_meta = row.file if isinstance(row.file, dict) else {}
    exclusions = normalize_file_import_exclusions(row.exclusions)
    exclusions_text = "\n".join(exclusions)
    model_field = clean_text(str(row.model_field or ""))
    price_field = clean_text(str(getattr(row, "price_field", "") or ""))
    replace_rules = normalize_file_import_rules_text(row.replace_rules)
    state = normalize_file_import_state(getattr(row, "state", {}) or {})
    active = is_file_import_active_state(state)
    result_filename = Path(str(row.export_path or file_meta.get("result_filename") or state.get("result_filename") or "")).name
    result_ready = bool(result_filename and not active and resolve_file_import_export_path(result_filename))
    if not path:
        return {
            "file": None,
            "exclusions": exclusions_text,
            "model_field": model_field,
            "price_field": price_field,
            "replace_rules": replace_rules,
            "result_filename": result_filename,
            "result_ready": result_ready,
            "state": state,
        }
    stat = path.stat()
    return {
        "exclusions": exclusions_text,
        "model_field": model_field,
        "price_field": price_field,
        "replace_rules": replace_rules,
        "result_filename": result_filename,
        "result_ready": result_ready,
        "state": state,
        "file": {
            "filename": output_text(str(file_meta.get("original_filename") or path.name)),
            "stored_filename": path.name,
            "size": stat.st_size,
            "uploaded_at": str(file_meta.get("uploaded_at") or datetime.fromtimestamp(stat.st_mtime, MSK_TZ).isoformat(timespec="seconds")),
        }
    }


def public_file_import_progress() -> Dict[str, object]:
    payload = public_file_import_state()
    return {
        "state": payload["state"],
        "result_filename": payload["result_filename"],
        "result_ready": payload["result_ready"],
    }


def public_file_import_settings() -> Dict[str, object]:
    payload = public_file_import_state()
    return {
        "model_field": payload["model_field"],
        "price_field": payload["price_field"],
        "exclusions": payload["exclusions"],
        "replace_rules": payload["replace_rules"],
    }


def decode_file_import_csv(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1251", "windows-1251"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def normalize_file_import_header(value: object) -> str:
    return clean_text(str(value or "")).casefold()


def file_import_column_index(headers: List[object], column_name: str, column_label: str = "модели") -> int:
    expected = normalize_file_import_header(column_name)
    if not expected:
        raise ValueError(f"Укажите название столбца {column_label}")
    for index, header in enumerate(headers):
        if normalize_file_import_header(header) == expected:
            return index
    raise ValueError(f"Столбец {column_label} не найден: {column_name}")


def file_import_optional_brand_index(headers: List[object]) -> Optional[int]:
    brand_names = {"brand", "бренд", "manufacturer", "производитель", "vendor", "марка"}
    for index, header in enumerate(headers):
        if normalize_file_import_header(header) in brand_names:
            return index
    return None


def file_import_optional_name_index(headers: List[object]) -> Optional[int]:
    name_headers = (
        "наименование товара",
        "название товара",
        "наименование",
        "название",
        "product name",
        "product_name",
        "item name",
        "item_name",
        "name",
    )
    normalized_headers = [normalize_file_import_header(header) for header in headers]
    for expected in name_headers:
        try:
            return normalized_headers.index(expected)
        except ValueError:
            continue
    return None


def file_import_cell_text(value: object) -> str:
    if value is None:
        return ""
    text = clean_text(str(value))
    return "" if text.casefold() in {"none", "null", "nan"} else text


def read_file_import_rows(path: Path, model_field: str, price_field: str = "") -> List[Dict[str, object]]:
    if path.suffix.lower() == ".csv":
        text = decode_file_import_csv(path.read_bytes())
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,|\t,")
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = ";"
        rows = list(csv.reader(io.StringIO(text), dialect))
    elif path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        workbook.close()
    elif path.suffix.lower() == ".xls":
        try:
            import xlrd
        except ImportError as exc:
            raise ValueError("Для обработки XLS установите зависимость xlrd") from exc

        workbook = xlrd.open_workbook(str(path))
        sheet = workbook.sheet_by_index(0)
        rows = [sheet.row_values(row_index) for row_index in range(sheet.nrows)]
    else:
        raise ValueError("Можно обработать только CSV, XLS или XLSX")

    header_index = next((index for index, row in enumerate(rows) if any(clean_text(str(cell or "")) for cell in row)), None)
    if header_index is None:
        return []
    headers = rows[header_index]
    model_index = file_import_column_index(headers, model_field, "модели")
    price_index = file_import_column_index(headers, price_field, "цены") if clean_text(str(price_field or "")) else None
    brand_index = file_import_optional_brand_index(headers)
    name_index = file_import_optional_name_index(headers)
    result: List[Dict[str, object]] = []
    for row_number, row in enumerate(rows[header_index + 1:], start=header_index + 2):
        source_model = file_import_cell_text(row[model_index] if model_index < len(row) else None)
        if not source_model:
            continue
        source_name = file_import_cell_text(
            row[name_index] if name_index is not None and name_index < len(row) else None
        )
        price = file_import_cell_text(
            row[price_index] if price_index is not None and price_index < len(row) else None
        )
        brand = file_import_cell_text(
            row[brand_index] if brand_index is not None and brand_index < len(row) else None
        )
        result.append(
            {
                "row_number": row_number,
                "source_model": source_model,
                "name": source_name or source_model,
                "price": price,
                "brand": brand,
            }
        )
    return result


def strip_file_import_model_special_chars(value: str) -> str:
    text = re.sub(r"[^0-9A-Za-zА-Яа-яЁё\s\-./\\|@+]+", " ", str(value or ""))
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"\s*([\-./\\|])\s*", r"\1", text)
    return clean_text(text)


def prepare_file_import_model(value: str, replace_rules: str) -> str:
    prepared = prepare_rule_model(
        str(value or ""),
        {"model_replace_rules": normalize_file_import_rules_text(replace_rules)},
    )
    return strip_file_import_model_special_chars(prepared)


def technical_clean_model_text(value: object, remove_brackets: bool = True) -> str:
    text = BeautifulSoup(str(value or ""), "html.parser").get_text(" ", strip=True)
    text = html_lib.unescape(text)
    text = text.replace("\xa0", " ").replace("\u2009", " ")
    text = re.sub(r"[–—−]", "-", text)
    text = re.sub(r"\s*([/_.+\-])\s*", r"\1", text)
    if remove_brackets:
        text = re.sub(r"\([^)]*\)|\[[^\]]*\]|\{[^}]*\}", " ", text)
    return clean_text(text)


def strip_file_import_non_model_phrases(value: str) -> str:
    text = clean_text(value)
    text = re.sub(r"(?<![A-Za-z0-9])side\s*[- ]\s*by\s*[- ]\s*side(?![A-Za-z0-9])", " ", text, flags=re.IGNORECASE)
    return clean_text(text)


def normalize_compare_key(value: object) -> str:
    text = technical_clean_model_text(value, remove_brackets=False)
    text = re.sub(r"[\"'`«»]", "", text)
    text = re.sub(r"\s+", " ", text).strip(" .,/\\_-+")
    return normalize_model_key(text)


def compact_compare_key(value: object) -> str:
    return re.sub(r"[\s./_+\-]+", "", normalize_compare_key(value))


def visual_compare_key(value: object) -> str:
    return compact_compare_key(str(value or "").translate(VISUAL_MODEL_TRANSLATION))


def compare_keys_for_value(value: object) -> Dict[str, str]:
    original = normalize_compare_key(value)
    compact = compact_compare_key(value)
    visual = visual_compare_key(value)
    keys = {
        "original": original,
        "normalized": original,
        "compact": compact,
        "visual": visual,
    }
    return {kind: key for kind, key in keys.items() if key}


def file_import_exclusion_matches(value: str, brand: str, exclusions: Iterable[str]) -> bool:
    haystack = f"{value} {brand}".casefold()
    visual_haystack = f"{value} {brand}".translate(VISUAL_MODEL_TRANSLATION).casefold()
    for exclusion in exclusions:
        pattern = clean_text(str(exclusion or ""))
        if not pattern:
            continue
        if pattern.casefold() in haystack or pattern.translate(VISUAL_MODEL_TRANSLATION).casefold() in visual_haystack:
            return True
    return False


def known_file_import_brands() -> List[str]:
    try:
        with session_scope() as session:
            rows = session.execute(select(Brand.name)).scalars().all()
    except Exception:
        rows = []
    brands = [clean_text(str(row or "")) for row in rows if clean_text(str(row or ""))]
    return sorted(set(brands), key=len, reverse=True)


def brand_match_pattern(brand: str) -> str:
    parts = [re.escape(part) for part in re.split(r"\s*&\s*|\s+", clean_text(brand)) if part]
    if not parts:
        return ""
    return r"\s*&\s*".join(parts) if "&" in brand else r"\s+".join(parts)


def find_brand_in_name(name: str, explicit_brand: str = "") -> str:
    if explicit_brand:
        return explicit_brand
    for brand in known_file_import_brands():
        pattern = brand_match_pattern(brand)
        if pattern and re.search(rf"(?<![A-Za-zА-Яа-яЁё0-9]){pattern}(?![A-Za-zА-Яа-яЁё0-9])", name, flags=re.IGNORECASE):
            return brand
    return ""


def tail_after_brand(name: str, brand: str) -> str:
    text = technical_clean_model_text(name)
    if not brand:
        return text
    pattern = brand_match_pattern(brand)
    match = re.search(pattern, text, flags=re.IGNORECASE) if pattern else None
    if match:
        before_brand = clean_text(text[:match.start()])
        if before_brand and any(model_signal_token(token) for token in before_brand.split()):
            return before_brand
        return clean_text(text[match.end():])
    visual_text = text.translate(VISUAL_MODEL_TRANSLATION)
    visual_brand = brand.translate(VISUAL_MODEL_TRANSLATION)
    visual_pattern = brand_match_pattern(visual_brand)
    visual_match = re.search(visual_pattern, visual_text, flags=re.IGNORECASE) if visual_pattern else None
    if visual_match:
        before_brand = clean_text(text[:visual_match.start()])
        if before_brand and any(model_signal_token(token) for token in before_brand.split()):
            return before_brand
        return clean_text(text[visual_match.end():])
    return text


def model_signal_token(token: str) -> bool:
    value = clean_text(token)
    if not value:
        return False
    if re.fullmatch(r"\d+(?:[.,]\d+)?(?:ВТ|BT|В|V|B|Л|МЛ|ML|КГ|KG|Г|G|СМ|CM|ММ|MM)(?:/\d+(?:[.,]\d+)?(?:ВТ|BT|В|V|B|Л|МЛ|ML|КГ|KG|Г|G|СМ|CM|ММ|MM))*", value.upper()):
        return False
    has_digit = bool(re.search(r"\d", value))
    has_latin = bool(re.search(r"[A-Za-z]", value))
    has_cyrillic = bool(re.search(r"[А-Яа-яЁё]", value))
    return has_latin or (has_digit and has_cyrillic) or bool(re.search(r"[A-Za-zА-Яа-яЁё]\d|\d[A-Za-zА-Яа-яЁё]", value))


def visual_model_suffix_token(token: str) -> bool:
    value = clean_text(token)
    return bool(value and re.fullmatch(r"[A-Za-zА-Яа-яЁё]{1,3}", value) and value.translate(VISUAL_MODEL_TRANSLATION) != value)


def candidate_until_russian_description(value: str) -> str:
    tokens = value.split()
    kept: List[str] = []
    seen_signal = False
    for token in tokens:
        if seen_signal and re.fullmatch(r"\d+(?:[.,]\d+)?(?:ВТ|W|BT|В|V|B|Л|L|МЛ|ML|КГ|KG|Г|G|СМ|CM|ММ|MM)(?:/\d+(?:[.,]\d+)?(?:ВТ|W|BT|В|V|B|Л|L|МЛ|ML|КГ|KG|Г|G|СМ|CM|ММ|MM))*", token.upper()):
            break
        if model_signal_token(token):
            kept.append(token)
            seen_signal = True
            continue
        if seen_signal and visual_model_suffix_token(token):
            kept.append(token)
            continue
        if seen_signal and re.fullmatch(r"[А-Яа-яЁё][А-Яа-яЁё/\-]*", token):
            break
        kept.append(token)
    return clean_text(" ".join(kept))


def first_model_block(value: str) -> str:
    tokens = re.findall(r"[A-Za-zА-Яа-яЁё0-9]+(?:[./_+\-][A-Za-zА-Яа-яЁё0-9]+)*", value)
    best: List[str] = []
    current: List[str] = []
    for token in tokens:
        if model_signal_token(token):
            current.append(token)
            continue
        if current:
            if len(" ".join(current)) > len(" ".join(best)):
                best = list(current)
            current = []
    if current and len(" ".join(current)) > len(" ".join(best)):
        best = current
    return clean_text(" ".join(best))


def normalize_candidate_display(value: str) -> str:
    text = technical_clean_model_text(value)
    text = re.sub(r"\s+", " ", text).strip(" .,/\\_-+")
    tokens = text.split()
    has_digit = any(char.isdigit() for char in text)
    normalized_tokens = [
        token.translate(VISUAL_MODEL_TRANSLATION).upper()
        if any(char.isdigit() for char in token) or (has_digit and re.fullmatch(r"[A-Za-zА-Яа-яЁё]{1,3}", token))
        else token
        for token in tokens
    ]
    return " ".join(normalized_tokens)


def add_model_candidate(candidates: List[str], value: str) -> None:
    candidate = normalize_candidate_display(value)
    if not candidate:
        return
    if re.match(r"^[А-Яа-яЁё]", candidate):
        return
    if re.search(r"[А-Яа-яЁё]", candidate):
        return
    if re.fullmatch(r"\d+(?:[.,]\d+)?(?:ВТ|W|BT|В|V|B|Л|L|МЛ|ML|КГ|KG|Г|G|СМ|CM|ММ|MM)(?:/\d+(?:[.,]\d+)?(?:ВТ|W|BT|В|V|B|Л|L|МЛ|ML|КГ|KG|Г|G|СМ|CM|ММ|MM))*", candidate.upper()):
        return
    key = normalize_compare_key(candidate)
    if len(key) < 2:
        return
    if key not in {normalize_compare_key(item) for item in candidates}:
        candidates.append(candidate)


def code_model_tokens(value: str) -> List[str]:
    result: List[str] = []
    for token in re.findall(r"[A-Za-zА-Яа-яЁё0-9]+(?:[./_+\-][A-Za-zА-Яа-яЁё0-9]+)*", value):
        if not model_signal_token(token):
            continue
        normalized = normalize_candidate_display(token)
        if not normalized or re.match(r"^[А-Яа-яЁё]", normalized):
            continue
        if not any(char.isdigit() for char in normalized):
            continue
        if normalized not in result:
            result.append(normalized)
    return result


def generate_model_candidates(name: str, brand: str = "") -> List[str]:
    cleaned = strip_file_import_non_model_phrases(technical_clean_model_text(name))
    detected_brand = find_brand_in_name(cleaned, brand)
    tail = tail_after_brand(cleaned, detected_brand)
    main_part = re.split(r"[,;|]", tail, maxsplit=1)[0]
    before_russian_description = candidate_until_russian_description(main_part)
    strong_block = first_model_block(main_part)

    candidates: List[str] = []
    for value in (before_russian_description, strong_block, main_part, tail):
        add_model_candidate(candidates, value)
    for token in code_model_tokens(main_part):
        add_model_candidate(candidates, token)

    base = before_russian_description or main_part or strong_block
    tokens = base.split()
    for start in range(1, min(4, len(tokens))):
        value = " ".join(tokens[start:])
        if any(char.isdigit() for char in value):
            add_model_candidate(candidates, value)
    for size in range(min(4, len(tokens)), 0, -1):
        value = " ".join(tokens[:size])
        if any(char.isdigit() for char in value):
            add_model_candidate(candidates, value)
    return candidates[:8]


def feed_index_add(index: Dict[str, Dict[str, object]], key: str, item: Dict[str, object]) -> None:
    if key and key not in index:
        index[key] = item


def index_feed_value(index: Dict[str, Dict[str, object]], value: str, item: Dict[str, object]) -> None:
    for key in compare_keys_for_value(value).values():
        feed_index_add(index, key, item)


def build_feed_index_from_xml(content: bytes, feed: Dict[str, object], stop_event: Optional[threading.Event] = None) -> Dict[str, Dict[str, object]]:
    index: Dict[str, Dict[str, object]] = {}
    for node_index, (_event, node) in enumerate(ET.iterparse(io.BytesIO(content), events=("end",)), start=1):
        if node_index % 500 == 0:
            stop_file_import_if_requested(stop_event)
        children = list(node)
        if not children:
            continue
        values: Dict[str, str] = {}
        for child in children:
            key = str(child.tag).split("}")[-1].lower()
            values[key] = clean_text(child.text or "")
        explicit_values = [
            values.get(key, "")
            for key in ("vendorcode", "vendor_code", "model", "sku", "article", "articul")
            if values.get(key)
        ]
        name = values.get("name") or values.get("title") or ""
        item = {
            "source": str(feed.get("source") or ""),
            "source_label": str(feed.get("source_label") or feed.get("url") or "Фид"),
            "feed_name": name,
            "feed_url": values.get("url") or "",
            "raw": values,
        }
        for value in explicit_values:
            index_feed_value(index, value, {**item, "matched_feed_key": normalize_compare_key(value)})
        if not explicit_values and name:
            for candidate in generate_model_candidates(name):
                index_feed_value(index, candidate, {**item, "matched_feed_key": normalize_compare_key(candidate)})
        node.clear()
    return index


def build_file_import_feed_indexes(stop_event: Optional[threading.Event] = None) -> List[Dict[str, object]]:
    from runtime.news_tasks import download_feed_files, feed_snapshot_path
    from services.log_service import save_logs
    from services.news_service import save_news_settings
    downloaded_feeds = download_feed_files(stop_event=stop_event)
    feed_indexes: List[Dict[str, object]] = []
    for feed in downloaded_feeds:
        stop_file_import_if_requested(stop_event)
        path = feed_snapshot_path(feed)
        try:
            if path is None:
                raise FileNotFoundError("Не задан путь к snapshot фида")
            index = build_feed_index_from_xml(path.read_bytes(), feed, stop_event=stop_event)
            feed_indexes.append({**feed, "index": index, "codes_count": len(index)})
        except FileImportStopped:
            raise
        except Exception as exc:
            feed_indexes.append({**feed, "index": {}, "codes_count": 0, "error": str(exc)})
    with news_lock:
        news_settings["feed_storage"] = [
            {key: value for key, value in feed.items() if key != "index"}
            for feed in feed_indexes
        ]
        save_news_settings()
    save_logs()
    return feed_indexes


def match_candidates_against_feed_indexes(
    candidates: List[str],
    feed_indexes: List[Dict[str, object]],
    stop_event: Optional[threading.Event] = None,
) -> Optional[Dict[str, object]]:
    for candidate in candidates:
        stop_file_import_if_requested(stop_event)
        keys = compare_keys_for_value(candidate)
        for reason, key in keys.items():
            for feed in feed_indexes:
                stop_file_import_if_requested(stop_event)
                index = feed.get("index", {})
                if isinstance(index, dict) and key in index:
                    match = dict(index[key])
                    match.update(
                        {
                            "selected_model": candidate,
                            "selected_reason": f"matched:{reason}",
                            "compare_key": key,
                            "matched_feed_key": str(match.get("matched_feed_key") or key),
                        }
                    )
                    return match
    return None


def missing_feed_labels(
    candidates: List[str],
    feed_indexes: List[Dict[str, object]],
    stop_event: Optional[threading.Event] = None,
) -> List[str]:
    labels: List[str] = []
    keys = set()
    for candidate in candidates:
        stop_file_import_if_requested(stop_event)
        keys.update(compare_keys_for_value(candidate).values())
    for feed in feed_indexes:
        stop_file_import_if_requested(stop_event)
        index = feed.get("index", {})
        if not isinstance(index, dict) or not any(key in index for key in keys):
            labels.append(str(feed.get("source_label") or feed.get("url") or "Фид"))
    return labels


def file_import_result_filename(original_filename: str) -> str:
    stem = safe_filename(Path(original_filename).stem or "file")
    return f"Новинки_{stem}_{datetime.now(MSK_TZ).strftime('%d-%m-%Y_%H-%M-%S')}.xlsx"


def excel_sheet_title(raw_title: str, used_titles: Set[str]) -> str:
    title = re.sub(r"[\[\]:*?/\\]", " ", clean_text(raw_title or "Результат"))
    title = re.sub(r"\s+", " ", title).strip() or "Результат"
    title = title[:31].rstrip() or "Результат"
    base = title
    counter = 2
    while title in used_titles:
        suffix = f" {counter}"
        title = f"{base[:31 - len(suffix)].rstrip()}{suffix}"
        counter += 1
    used_titles.add(title)
    return title


def write_file_import_workbook(result_path: Path, rows: List[Dict[str, object]], feed_indexes: List[Dict[str, object]]) -> None:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    feed_labels = [
        str(feed.get("source_label") or feed.get("url") or "Фид").strip() or "Фид"
        for feed in feed_indexes
    ]
    all_feed_labels = set(feed_labels)
    grouped_rows: Dict[str, List[Dict[str, object]]] = {}
    for row in rows:
        missing_labels = []
        for label in row.get("_missing_labels", []):
            label_text = str(label).strip()
            if label_text and label_text not in missing_labels:
                missing_labels.append(label_text)
        if not missing_labels:
            continue
        if all_feed_labels and set(missing_labels) == all_feed_labels:
            title = f"Не найдено на {', '.join(feed_labels)}"
            grouped_rows.setdefault(title, []).append(row)
            continue
        for label in missing_labels:
            title = f"Не найдено на {label}"
            grouped_rows.setdefault(title, []).append(row)

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    used_titles: Set[str] = set()
    groups = list(grouped_rows.items()) or [("Результат", [])]
    for raw_title, group_rows in groups:
        sheet = workbook.create_sheet(excel_sheet_title(raw_title, used_titles))
        sheet.append(FILE_IMPORT_RESULT_FIELDS)
        for row in group_rows:
            sheet.append([row.get(field, "") for field in FILE_IMPORT_RESULT_FIELDS])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_index, field in enumerate(FILE_IMPORT_RESULT_FIELDS, start=1):
            values = [str(field)]
            values.extend(str(row.get(field, "")) for row in group_rows[:200])
            width = min(max(len(value) for value in values) + 2, 60)
            sheet.column_dimensions[get_column_letter(column_index)].width = width
    workbook.save(result_path)


def compare_file_import_with_feeds(db_session=None, stop_event: Optional[threading.Event] = None) -> Dict[str, object]:
    db = db_session or g.db
    row = get_file_import_row(db)
    path = file_import_path_for_row(row)
    if not path:
        raise ValueError("Файл не загружен")
    model_field = clean_text(str(row.model_field or ""))
    if not model_field:
        raise ValueError("Укажите название столбца модели")

    exclusions = normalize_file_import_exclusions(row.exclusions)
    price_field = clean_text(str(getattr(row, "price_field", "") or ""))
    replace_rules = normalize_file_import_rules_text(row.replace_rules)
    started_at = time.time()
    update_file_import_state(
        db,
        status="running",
        stage="Читаю файл",
        percent=2,
        error="",
        started_at=datetime.now(MSK_TZ).isoformat(timespec="seconds"),
        finished_at="",
        elapsed_seconds=0,
        result_filename="",
    )
    db.commit()
    stop_file_import_if_requested(stop_event)
    source_rows = read_file_import_rows(path, model_field, price_field)
    total_rows = len(source_rows)
    update_file_import_state(db, stage="Загружаю фиды", percent=8, total_rows=total_rows)
    db.commit()
    stop_file_import_if_requested(stop_event)
    feed_indexes = build_file_import_feed_indexes(stop_event=stop_event)
    stop_file_import_if_requested(stop_event)
    update_file_import_state(db, stage="Сравниваю строки", percent=10, total_rows=total_rows)
    db.commit()

    result_rows: List[Dict[str, object]] = []
    processed = excluded = found = missing = empty_model = 0
    last_progress_time = 0.0
    last_progress_index = 0
    for index, item in enumerate(source_rows, start=1):
        if stop_event and stop_event.is_set():
            state = update_file_import_state(
                db,
                status="stopped",
                stage="Остановлено",
                percent=max(0, min(99, int((index - 1) / max(total_rows, 1) * 90) + 10)),
                current_row=index - 1,
                processed_rows=processed,
                excluded_rows=excluded,
                found_rows=found,
                missing_rows=missing,
                model_not_found_rows=empty_model,
                finished_at=datetime.now(MSK_TZ).isoformat(timespec="seconds"),
                elapsed_seconds=int(time.time() - started_at),
            )
            db.commit()
            return {
                "stopped": True,
                "total_rows": total_rows,
                "processed_rows": processed,
                "excluded_rows": excluded,
                "model_not_found_rows": empty_model,
                "found_rows": found,
                "missing_rows": missing,
                "state": state,
            }
        name = str(item.get("name") or "")
        source_model = str(item.get("source_model") or name)
        price = str(item.get("price") or "")
        brand = str(item.get("brand") or "")
        if file_import_exclusion_matches(f"{source_model} {name}", brand, exclusions):
            excluded += 1
        else:
            processed += 1
            prepared_model = prepare_file_import_model(source_model, replace_rules)
            candidates = generate_model_candidates(prepared_model, brand)
            if not candidates:
                empty_model += 1
            else:
                match = match_candidates_against_feed_indexes(candidates, feed_indexes, stop_event=stop_event)
                missing_labels = missing_feed_labels(candidates, feed_indexes, stop_event=stop_event)
                if missing_labels:
                    missing += 1
                    selected_model = str((match or {}).get("selected_model") or candidates[0])
                    result_rows.append(
                        {
                            "row": item.get("row_number"),
                            "name": name,
                            "price": price,
                            "brand": brand,
                            "model_candidates": " | ".join(candidates),
                            "selected_model": selected_model,
                            "missing_on": ", ".join(missing_labels),
                            "_missing_labels": missing_labels,
                        }
                    )
                else:
                    found += 1
        now = time.time()
        if index == total_rows or index - last_progress_index >= 10 or now - last_progress_time >= 1.0:
            last_progress_time = now
            last_progress_index = index
            update_file_import_state(
                db,
                stage="Сравниваю строки",
                percent=10 + int(index / max(total_rows, 1) * 85),
                current_row=index,
                processed_rows=processed,
                excluded_rows=excluded,
                found_rows=found,
                missing_rows=missing,
                model_not_found_rows=empty_model,
                elapsed_seconds=int(now - started_at),
            )
            db.commit()

    file_meta = dict(row.file) if isinstance(row.file, dict) else {}
    original_filename = str(file_meta.get("original_filename") or file_meta.get("filename") or path.name)
    remove_file_import_export(row)
    result_path = EXPORT_DIR / file_import_result_filename(original_filename)
    update_file_import_state(db, stage="Записываю XLSX", percent=98)
    db.commit()
    stop_file_import_if_requested(stop_event)
    write_file_import_workbook(result_path, result_rows, feed_indexes)

    file_meta["result_filename"] = result_path.name
    file_meta["result_created_at"] = datetime.now(MSK_TZ).isoformat(timespec="seconds")
    row.export_path = result_path.name
    row.file = file_meta
    state = update_file_import_state(
        db,
        status="completed",
        stage="Готово",
        percent=100,
        current_row=total_rows,
        total_rows=total_rows,
        processed_rows=processed,
        excluded_rows=excluded,
        found_rows=found,
        missing_rows=missing,
        model_not_found_rows=empty_model,
        finished_at=datetime.now(MSK_TZ).isoformat(timespec="seconds"),
        elapsed_seconds=int(time.time() - started_at),
        result_filename=result_path.name,
        error="",
    )
    db.commit()
    return {
        "total_rows": len(source_rows),
        "processed_rows": processed,
        "excluded_rows": excluded,
        "model_not_found_rows": empty_model,
        "found_rows": found,
        "missing_rows": missing,
        "result_filename": result_path.name,
        "result_url": "/api/file-import/download",
        "state": state,
    }


def run_file_import_compare(stop_event: threading.Event) -> None:
    global file_import_worker_thread
    db = SessionLocal()
    try:
        compare_file_import_with_feeds(db, stop_event=stop_event)
    except FileImportStopped:
        db.rollback()
        try:
            update_file_import_state(
                db,
                status="stopped",
                stage="Остановлено",
                finished_at=datetime.now(MSK_TZ).isoformat(timespec="seconds"),
            )
            db.commit()
        except Exception:
            db.rollback()
    except Exception as exc:
        db.rollback()
        try:
            update_file_import_state(
                db,
                status="error",
                stage="Ошибка",
                error=str(exc),
                finished_at=datetime.now(MSK_TZ).isoformat(timespec="seconds"),
            )
            db.commit()
        except Exception:
            db.rollback()
    finally:
        db.close()
        with file_import_lock:
            if threading.current_thread() is file_import_worker_thread:
                file_import_worker_thread = None
                file_import_stop_event.clear()


def recover_interrupted_file_import_scan() -> None:
    with session_scope() as db_session:
        row = get_file_import_row(db_session)
        state = normalize_file_import_state(getattr(row, "state", {}) or {})
        if not is_file_import_active_state(state):
            return
        row.state = {
            **state,
            "status": "error",
            "stage": "Прервано",
            "error": "Сравнение файла было прервано перезапуском сервера. Запустите его снова.",
            "finished_at": datetime.now(MSK_TZ).isoformat(timespec="seconds"),
        }


def start_file_import_compare() -> Dict[str, object]:
    global file_import_worker_thread
    with file_import_lock:
        if file_import_worker_thread and file_import_worker_thread.is_alive():
            return public_file_import_state()
        row = get_file_import_row()
        state = normalize_file_import_state(getattr(row, "state", {}) or {})
        if is_file_import_active_state(state):
            return public_file_import_state()
        path = file_import_path_for_row(row)
        if not path:
            raise ValueError("Файл не загружен")
        if not clean_text(str(row.model_field or "")):
            raise ValueError("Укажите название столбца модели")

        remove_file_import_export(row)
        file_meta = dict(row.file) if isinstance(row.file, dict) else {}
        file_meta.pop("result_filename", None)
        file_meta.pop("result_created_at", None)
        row.file = file_meta
        row.export_path = ""
        row.state = {
            **make_file_import_state("queued"),
            "stage": "В очереди",
            "started_at": datetime.now(MSK_TZ).isoformat(timespec="seconds"),
        }
        g.db.commit()

        file_import_stop_event.clear()
        file_import_worker_thread = threading.Thread(
            target=run_file_import_compare,
            args=(file_import_stop_event,),
            daemon=True,
        )
        file_import_worker_thread.start()
        return public_file_import_state()


def file_import_worker_alive() -> bool:
    return bool(file_import_worker_thread and file_import_worker_thread.is_alive())


__all__ = ['FileImportStopped', 'make_file_import_state', 'normalize_file_import_state', 'is_file_import_active_state', 'file_import_path_for_row', 'update_file_import_state', 'stop_file_import_if_requested', 'clear_file_import_storage', 'stored_file_import_files', 'get_file_import_row', 'current_file_import_path', 'resolve_file_import_export_path', 'remove_file_import_export', 'public_file_import_state', 'public_file_import_progress', 'public_file_import_settings', 'decode_file_import_csv', 'normalize_file_import_header', 'file_import_column_index', 'file_import_optional_brand_index', 'file_import_optional_name_index', 'file_import_cell_text', 'read_file_import_rows', 'strip_file_import_model_special_chars', 'prepare_file_import_model', 'technical_clean_model_text', 'strip_file_import_non_model_phrases', 'normalize_compare_key', 'compact_compare_key', 'visual_compare_key', 'compare_keys_for_value', 'file_import_exclusion_matches', 'known_file_import_brands', 'brand_match_pattern', 'find_brand_in_name', 'tail_after_brand', 'model_signal_token', 'visual_model_suffix_token', 'candidate_until_russian_description', 'first_model_block', 'normalize_candidate_display', 'add_model_candidate', 'code_model_tokens', 'generate_model_candidates', 'feed_index_add', 'index_feed_value', 'build_feed_index_from_xml', 'build_file_import_feed_indexes', 'match_candidates_against_feed_indexes', 'missing_feed_labels', 'file_import_result_filename', 'excel_sheet_title', 'write_file_import_workbook', 'compare_file_import_with_feeds', 'run_file_import_compare', 'recover_interrupted_file_import_scan', 'start_file_import_compare', 'file_import_worker_alive']
