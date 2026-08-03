"""Extracted application service module."""

import csv
import io
import json
import re
import threading
import time
import xml.etree.ElementTree as ET
import zipfile
from config import EXPORT_DIR, FEED_COMPARISON_ACTIVE_STATUSES, FEED_COMPARISON_PROGRESS_COMMIT_INTERVAL_SECONDS, FEED_COMPARISON_RESULT_FIELDS, MSK_TZ, env_str
from database.session import SessionLocal, session_scope
from datetime import datetime
from flask import g
from models import FeedComparison, OwnSite, SupplierFeed
from pathlib import Path
from runtime.state import feed_comparison_lock, feed_comparison_stop_event, feed_comparison_worker_thread, news_lock, news_settings
from services.normalization import normalize_feed_url, normalize_file_import_exclusions, normalize_file_import_rules_text
from services.domain_revisions import bump_domain_revision, domain_revision
from sqlalchemy import select
from typing import Dict, Iterable, List, Optional, Set
from urllib.parse import urlparse
from services.scraping import clean_text, prepare_rule_model


def make_feed_comparison_state(status: str = "idle") -> Dict[str, object]:
    return {
        "status": status,
        "stage": "",
        "percent": 0,
        "current_supplier": "",
        "current_row": 0,
        "total_rows": 0,
        "processed_rows": 0,
        "excluded_rows": 0,
        "missing_rows": 0,
        "suppliers_done": 0,
        "suppliers_total": 0,
        "error": "",
        "started_at": "",
        "finished_at": "",
        "elapsed_seconds": 0,
        "result_filename": "",
    }


def normalize_feed_comparison_state(value: object) -> Dict[str, object]:
    state = {**make_feed_comparison_state(), **(value if isinstance(value, dict) else {})}
    state["status"] = clean_text(str(state.get("status") or "idle"))
    for field in (
        "percent",
        "current_row",
        "total_rows",
        "processed_rows",
        "excluded_rows",
        "missing_rows",
        "suppliers_done",
        "suppliers_total",
        "elapsed_seconds",
    ):
        try:
            state[field] = int(float(state.get(field) or 0))
        except (TypeError, ValueError):
            state[field] = 0
    state["percent"] = max(0, min(100, state["percent"]))
    return state


def is_feed_comparison_active_state(state: object) -> bool:
    return str((state if isinstance(state, dict) else {}).get("status") or "") in FEED_COMPARISON_ACTIVE_STATUSES


def get_feed_comparison_row(db_session=None) -> FeedComparison:
    db = db_session or g.db
    row = db.get(FeedComparison, 1)
    if row is None:
        row = FeedComparison(id=1, state=make_feed_comparison_state(), export_path="")
        db.add(row)
        db.flush()
    normalized_state = normalize_feed_comparison_state(row.state)
    if row.state != normalized_state:
        row.state = normalized_state
        db.flush()
    return row


def resolve_feed_comparison_export_path(value: str) -> Optional[Path]:
    filename = Path(str(value or "")).name
    if not filename:
        return None
    path = (EXPORT_DIR / filename).resolve()
    if EXPORT_DIR.resolve() not in path.parents or not path.exists() or not path.is_file():
        return None
    return path


def remove_feed_comparison_export(row: FeedComparison) -> None:
    path = resolve_feed_comparison_export_path(str(row.export_path or ""))
    if path:
        try:
            path.unlink()
        except OSError:
            pass
    for old_path in EXPORT_DIR.glob("Сравнение_фидов_*.xlsx"):
        if old_path.is_file():
            try:
                old_path.unlink()
            except OSError:
                pass


def update_feed_comparison_state(db_session, **kwargs: object) -> Dict[str, object]:
    row = get_feed_comparison_row(db_session)
    state = normalize_feed_comparison_state(row.state)
    state.update(kwargs)
    row.state = normalize_feed_comparison_state(state)
    db_session.flush()
    return dict(row.state)


def public_own_site(row: OwnSite) -> Dict[str, object]:
    return {
        "id": row.id,
        "name": clean_text(str(row.name or "")),
        "feed_url": str(row.feed_url or ""),
        "feed_generate_url": str(row.feed_generate_url or ""),
    }


def public_supplier_feed(row: SupplierFeed) -> Dict[str, object]:
    exclusions = normalize_file_import_exclusions(getattr(row, "exclusions", []))
    return {
        "id": row.id,
        "name": clean_text(str(row.name or "")),
        "feed_url": str(row.feed_url or ""),
        "model_field": normalize_supplier_model_field(row.model_field),
        "name_field": normalize_supplier_model_field(getattr(row, "name_field", "")),
        "price_field": normalize_supplier_model_field(getattr(row, "price_field", "")),
        "brand_field": normalize_supplier_model_field(getattr(row, "brand_field", "")),
        "url_field": normalize_supplier_model_field(getattr(row, "url_field", "")),
        "exclusions": "\n".join(exclusions),
        "replace_rules": normalize_file_import_rules_text(getattr(row, "replace_rules", "")),
    }


def public_feed_comparison_state(db_session=None) -> Dict[str, object]:
    db = db_session or g.db
    comparison = get_feed_comparison_row(db)
    state = normalize_feed_comparison_state(comparison.state)
    result_path = resolve_feed_comparison_export_path(str(comparison.export_path or ""))
    own_sites = db.scalars(select(OwnSite).order_by(OwnSite.id)).all()
    suppliers = db.scalars(select(SupplierFeed).order_by(SupplierFeed.id)).all()
    return {
        "revision": domain_revision("feed_comparison"),
        "own_sites": [public_own_site(row) for row in own_sites],
        "suppliers": [public_supplier_feed(row) for row in suppliers],
        "state": state,
        "result_ready": bool(result_path and not is_feed_comparison_active_state(state)),
        "result_filename": result_path.name if result_path else "",
    }


def public_feed_comparison_progress(db_session=None) -> Dict[str, object]:
    db = db_session or g.db
    comparison = get_feed_comparison_row(db)
    state = normalize_feed_comparison_state(comparison.state)
    result_path = resolve_feed_comparison_export_path(str(comparison.export_path or ""))
    return {
        "revision": domain_revision("feed_comparison"),
        "state": state,
        "result_ready": bool(
            result_path and not is_feed_comparison_active_state(state)
        ),
        "result_filename": result_path.name if result_path else "",
    }


def sync_own_sites_runtime(db_session) -> None:
    sites = db_session.scalars(select(OwnSite).order_by(OwnSite.id)).all()
    payload = [public_own_site(site) for site in sites]
    with news_lock:
        news_settings["own_sites"] = [
            {
                "id": site.get("id"),
                "name": str(site.get("name") or ""),
                "feed_url": str(site.get("feed_url") or ""),
                "feed_generate_url": str(site.get("feed_generate_url") or ""),
            }
            for site in payload
        ]
        news_settings["feed_urls"] = [str(site["feed_url"]) for site in payload]
        news_settings["feed_generate_urls"] = [
            str(site["feed_generate_url"]) for site in payload if site.get("feed_generate_url")
        ]
        if payload:
            news_settings["feed_url"] = str(payload[0]["feed_url"])
            news_settings["feed_generate_url"] = str(payload[0].get("feed_generate_url") or "")
        else:
            news_settings["feed_url"] = ""
            news_settings["feed_generate_url"] = ""
    bump_domain_revision("feed_comparison")
    bump_domain_revision("settings")


def normalize_supplier_model_field(value: object) -> str:
    field = clean_text(str(value or ""))
    if not field:
        return ""
    if field.startswith("@") or field.casefold().startswith("param:"):
        return field
    tag_match = re.fullmatch(r"<\s*/?\s*([A-Za-z_][\w.:-]*)\s*(?:[^>]*)?>", field)
    if tag_match:
        return tag_match.group(1).split(":")[-1]
    field = field.strip().lstrip("./")
    if "/" in field:
        field = field.rsplit("/", 1)[-1]
    return field.split(":")[-1].strip()


def validate_feed_comparison_site_payload(payload: object, supplier: bool = False) -> Dict[str, object]:
    data = payload if isinstance(payload, dict) else {}
    name = clean_text(str(data.get("name") or ""))[:255]
    feed_url = normalize_feed_url(str(data.get("feed_url") or ""))
    if not name:
        raise ValueError("Укажите название фида")
    if not feed_url:
        raise ValueError("Укажите корректную ссылку на фид")
    result = {"name": name, "feed_url": feed_url}
    if supplier:
        model_field = normalize_supplier_model_field(data.get("model_field"))[:255]
        if not model_field:
            raise ValueError("Укажите название поля модели в фиде поставщика")
        result["model_field"] = model_field
        for field_name in ("name_field", "price_field", "brand_field", "url_field"):
            result[field_name] = normalize_supplier_model_field(data.get(field_name))[:255]
        result["exclusions"] = normalize_file_import_exclusions(data.get("exclusions"))
        result["replace_rules"] = normalize_file_import_rules_text(data.get("replace_rules"))
    else:
        raw_generate_url = clean_text(str(data.get("feed_generate_url") or ""))
        generate_url = normalize_feed_url(raw_generate_url) if raw_generate_url else ""
        if raw_generate_url and not generate_url:
            raise ValueError("Укажите корректную ссылку генерации фида")
        result["feed_generate_url"] = generate_url
    return result


def supplier_feed_basic_auth_rules() -> List[Dict[str, str]]:
    raw_value = env_str("SUPPLIER_FEED_BASIC_AUTH_RULES_JSON")
    if not raw_value:
        return []
    try:
        configured_rules = json.loads(raw_value)
    except json.JSONDecodeError as exc:
        raise ValueError("Некорректный JSON в SUPPLIER_FEED_BASIC_AUTH_RULES_JSON") from exc
    if isinstance(configured_rules, dict):
        configured_rules = [configured_rules]
    if not isinstance(configured_rules, list):
        raise ValueError("SUPPLIER_FEED_BASIC_AUTH_RULES_JSON должен содержать список правил")

    rules: List[Dict[str, str]] = []
    for item in configured_rules:
        if not isinstance(item, dict):
            continue
        host = clean_text(str(item.get("host") or "")).casefold()
        url_prefix = clean_text(str(item.get("url_prefix") or ""))
        username = str(item.get("username") or "")
        password = str(item.get("password") or "")
        if (host or url_prefix) and username and password:
            rules.append(
                {
                    "host": host,
                    "url_prefix": url_prefix,
                    "username": username,
                    "password": password,
                }
            )
    return rules


def supplier_feed_basic_auth_for_url(url: str) -> Optional[tuple[str, str]]:
    parsed = urlparse(url)
    request_host = (parsed.hostname or "").casefold()
    candidates: List[tuple[int, Dict[str, str]]] = []
    for rule in supplier_feed_basic_auth_rules():
        prefix = rule["url_prefix"]
        host = rule["host"]
        if prefix and url.startswith(prefix):
            candidates.append((10_000 + len(prefix), rule))
        elif host and request_host == host:
            candidates.append((len(host), rule))
    if not candidates:
        return None
    _score, selected = max(candidates, key=lambda item: item[0])
    return selected["username"], selected["password"]


def download_comparison_feed(url: str) -> bytes:
    from runtime.news_tasks import make_feed_session
    from services.file_validation import read_limited_response
    session = make_feed_session()
    request_kwargs: Dict[str, object] = {"timeout": 60}
    basic_auth = supplier_feed_basic_auth_for_url(url)
    if basic_auth:
        request_kwargs["auth"] = basic_auth
    response = session.get(url, stream=True, **request_kwargs)
    if response.status_code == 401:
        if basic_auth:
            raise ValueError("Фид отклонил сохранённые данные авторизации")
        if "basic" in str(response.headers.get("WWW-Authenticate") or "").casefold():
            raise ValueError("Фид требует HTTP Basic Authentication, но данные авторизации не настроены")
    response.raise_for_status()
    content = read_limited_response(response)
    if not content:
        raise ValueError("Фид вернул пустой ответ")
    return content


def xml_local_name(tag: object) -> str:
    return str(tag or "").split("}")[-1]


def supplier_xml_field_value(node: ET.Element, field_name: str) -> str:
    configured = normalize_supplier_model_field(field_name)
    if not configured:
        return ""
    if configured.startswith("@"):
        expected = configured[1:].casefold()
        for key, value in node.attrib.items():
            if xml_local_name(key).casefold() == expected:
                return clean_text(str(value or ""))
        return ""
    if configured.casefold().startswith("param:"):
        expected = configured.split(":", 1)[1].strip().casefold()
        for child in list(node):
            if xml_local_name(child.tag).casefold() != "param":
                continue
            param_name = clean_text(str(child.attrib.get("name") or "")).casefold()
            if param_name == expected:
                return clean_text(" ".join(child.itertext()))
        return ""
    expected = configured.casefold()
    for child in list(node):
        if xml_local_name(child.tag).casefold() == expected:
            return clean_text(" ".join(child.itertext()))
    return ""


def supplier_xml_available_fields(content: bytes, limit: int = 20) -> List[str]:
    fields: List[str] = []
    try:
        for _event, node in ET.iterparse(io.BytesIO(content), events=("end",)):
            if list(node):
                continue
            field = xml_local_name(node.tag)
            if field and field not in fields:
                fields.append(field)
                if len(fields) >= limit:
                    break
    except ET.ParseError:
        return []
    return fields


def detect_supplier_feed_format(content: bytes) -> str:
    sample = content[:4096]
    if sample.startswith(b"PK\x03\x04"):
        try:
            with zipfile.ZipFile(io.BytesIO(content)) as archive:
                if "xl/workbook.xml" in archive.namelist():
                    return "xlsx"
        except zipfile.BadZipFile:
            pass
    if sample.startswith((b"\xff\xfe", b"\xfe\xff")):
        try:
            return "xml" if sample.decode("utf-16").lstrip().startswith("<") else "csv"
        except UnicodeDecodeError:
            return "csv"
    sample = sample.removeprefix(b"\xef\xbb\xbf").lstrip()
    return "xml" if sample.startswith(b"<") else "csv"


def supplier_csv_rows(content: bytes) -> tuple[Iterable[List[str]], List[str], int]:
    from services.file_import_service import decode_file_import_csv, file_import_cell_text
    text = decode_file_import_csv(content)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,|\t,")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";"
    rows = csv.reader(io.StringIO(text), dialect)
    for header_index, row in enumerate(rows):
        if not any(file_import_cell_text(cell) for cell in row):
            continue
        headers = [file_import_cell_text(cell) for cell in row]
        return rows, headers, header_index + 2
    return iter(()), [], 1


def supplier_csv_optional_column_index(headers: List[object], field_name: str) -> Optional[int]:
    from services.file_import_service import normalize_file_import_header
    configured_field = normalize_supplier_model_field(field_name)
    if not configured_field:
        return None
    normalized_headers = [normalize_file_import_header(header) for header in headers]
    expected = normalize_file_import_header(configured_field)
    if expected in normalized_headers:
        return normalized_headers.index(expected)
    return None


def supplier_csv_available_fields(content: bytes, limit: int = 20) -> List[str]:
    try:
        _rows, headers, _first_row_number = supplier_csv_rows(content)
    except (csv.Error, UnicodeError):
        return []
    return [header for header in headers if header][:limit]


def supplier_xlsx_rows(content: bytes) -> tuple[object, Iterable[Iterable[object]], List[str], int]:
    from services.file_import_service import file_import_cell_text
    from services.file_validation import validate_xlsx_archive
    validate_xlsx_archive(content)
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Некорректный XLSX-фид поставщика: {exc}") from exc
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    for header_index, row in enumerate(rows):
        if not any(file_import_cell_text(cell) for cell in row):
            continue
        headers = [file_import_cell_text(cell) for cell in row]
        return workbook, rows, headers, header_index + 2
    return workbook, iter(()), [], 1


def supplier_xlsx_available_fields(content: bytes, limit: int = 20) -> List[str]:
    workbook = None
    try:
        workbook, _rows, headers, _first_row_number = supplier_xlsx_rows(content)
        return [header for header in headers if header][:limit]
    except ValueError:
        return []
    finally:
        if workbook is not None:
            workbook.close()


def read_supplier_xml_feed_rows(
    content: bytes,
    model_field: str,
    stop_event: Optional[threading.Event] = None,
    *,
    name_field: str = "",
    price_field: str = "",
    brand_field: str = "",
    url_field: str = "",
) -> List[Dict[str, object]]:
    from services.file_import_service import normalize_compare_key, stop_file_import_if_requested
    rows: List[Dict[str, object]] = []
    seen: Set[str] = set()
    try:
        iterator = ET.iterparse(io.BytesIO(content), events=("end",))
        for node_index, (_event, node) in enumerate(iterator, start=1):
            if node_index % 250 == 0:
                stop_file_import_if_requested(stop_event)
            source_model = supplier_xml_field_value(node, model_field)
            if not source_model:
                continue
            source_url = supplier_xml_field_value(node, url_field)
            dedupe_key = f"{normalize_compare_key(source_model)}|{source_url.casefold()}"
            if dedupe_key in seen:
                node.clear()
                continue
            seen.add(dedupe_key)
            name = supplier_xml_field_value(node, name_field) or source_model
            rows.append(
                {
                    "row_number": len(rows) + 1,
                    "source_model": source_model,
                    "name": name,
                    "price": supplier_xml_field_value(node, price_field),
                    "brand": supplier_xml_field_value(node, brand_field),
                    "url": source_url,
                }
            )
            node.clear()
    except ET.ParseError as exc:
        raise ValueError(f"Некорректный XML фида поставщика: {exc}") from exc
    return rows


def read_supplier_tabular_feed_rows(
    source_rows: Iterable[Iterable[object]],
    headers: List[str],
    first_row_number: int,
    model_field: str,
    feed_format: str,
    stop_event: Optional[threading.Event] = None,
    *,
    name_field: str = "",
    price_field: str = "",
    brand_field: str = "",
    url_field: str = "",
) -> List[Dict[str, object]]:
    from services.file_import_service import file_import_cell_text, file_import_column_index, normalize_compare_key, stop_file_import_if_requested
    if not headers:
        return []

    configured_field = normalize_supplier_model_field(model_field)
    try:
        model_index = file_import_column_index(headers, configured_field, "модели")
    except ValueError as exc:
        available_fields = ", ".join(header for header in headers if header)
        available_hint = f" Доступные колонки: {available_fields}." if available_fields else ""
        raise ValueError(
            f'В {feed_format}-фиде не найдена колонка модели "{configured_field}".{available_hint}'
        ) from exc

    name_index = supplier_csv_optional_column_index(headers, name_field)
    price_index = supplier_csv_optional_column_index(headers, price_field)
    brand_index = supplier_csv_optional_column_index(headers, brand_field)
    url_index = supplier_csv_optional_column_index(headers, url_field)

    rows: List[Dict[str, object]] = []
    seen: Set[str] = set()
    for row_number, source_row in enumerate(source_rows, start=first_row_number):
        if row_number % 250 == 0:
            stop_file_import_if_requested(stop_event)
        row = list(source_row)
        source_model = file_import_cell_text(row[model_index] if model_index < len(row) else None)
        if not source_model:
            continue
        source_url = file_import_cell_text(
            row[url_index] if url_index is not None and url_index < len(row) else None
        )
        dedupe_key = f"{normalize_compare_key(source_model)}|{source_url.casefold()}"
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        name = file_import_cell_text(
            row[name_index] if name_index is not None and name_index < len(row) else None
        )
        rows.append(
            {
                "row_number": row_number,
                "source_model": source_model,
                "name": name or source_model,
                "price": file_import_cell_text(
                    row[price_index] if price_index is not None and price_index < len(row) else None
                ),
                "brand": file_import_cell_text(
                    row[brand_index] if brand_index is not None and brand_index < len(row) else None
                ),
                "url": source_url,
            }
        )
    return rows


def read_supplier_csv_feed_rows(
    content: bytes,
    model_field: str,
    stop_event: Optional[threading.Event] = None,
    *,
    name_field: str = "",
    price_field: str = "",
    brand_field: str = "",
    url_field: str = "",
) -> List[Dict[str, object]]:
    try:
        source_rows, headers, first_row_number = supplier_csv_rows(content)
    except (csv.Error, UnicodeError) as exc:
        raise ValueError(f"Некорректный CSV фида поставщика: {exc}") from exc
    return read_supplier_tabular_feed_rows(
        source_rows,
        headers,
        first_row_number,
        model_field,
        "CSV",
        stop_event=stop_event,
        name_field=name_field,
        price_field=price_field,
        brand_field=brand_field,
        url_field=url_field,
    )


def read_supplier_xlsx_feed_rows(
    content: bytes,
    model_field: str,
    stop_event: Optional[threading.Event] = None,
    *,
    name_field: str = "",
    price_field: str = "",
    brand_field: str = "",
    url_field: str = "",
) -> List[Dict[str, object]]:
    workbook = None
    try:
        workbook, source_rows, headers, first_row_number = supplier_xlsx_rows(content)
        return read_supplier_tabular_feed_rows(
            source_rows,
            headers,
            first_row_number,
            model_field,
            "XLSX",
            stop_event=stop_event,
            name_field=name_field,
            price_field=price_field,
            brand_field=brand_field,
            url_field=url_field,
        )
    finally:
        if workbook is not None:
            workbook.close()


def read_supplier_feed_rows(
    content: bytes,
    model_field: str,
    stop_event: Optional[threading.Event] = None,
    *,
    name_field: str = "",
    price_field: str = "",
    brand_field: str = "",
    url_field: str = "",
) -> List[Dict[str, object]]:
    feed_format = detect_supplier_feed_format(content)
    field_kwargs = {
        "name_field": name_field,
        "price_field": price_field,
        "brand_field": brand_field,
        "url_field": url_field,
    }
    if feed_format == "xml":
        return read_supplier_xml_feed_rows(
            content, model_field, stop_event=stop_event, **field_kwargs
        )
    if feed_format == "xlsx":
        return read_supplier_xlsx_feed_rows(
            content, model_field, stop_event=stop_event, **field_kwargs
        )
    return read_supplier_csv_feed_rows(
        content, model_field, stop_event=stop_event, **field_kwargs
    )


def prepare_supplier_feed_item(
    item: Dict[str, object],
    exclusions: object = None,
    replace_rules: object = "",
) -> Optional[Dict[str, object]]:
    from services.file_import_service import file_import_exclusion_matches, generate_model_candidates, prepare_file_import_model
    raw_name = str(item.get("name") or item.get("source_model") or "")
    source_model = str(item.get("source_model") or raw_name)
    brand = str(item.get("brand") or "")
    if file_import_exclusion_matches(f"{source_model} {raw_name}", brand, normalize_file_import_exclusions(exclusions)):
        return None
    rules = normalize_file_import_rules_text(replace_rules)
    prepared_name = prepare_rule_model(raw_name, {"model_replace_rules": rules})
    prepared_model = prepare_file_import_model(source_model, rules)
    prepared_name_model = prepare_file_import_model(raw_name, rules)
    candidates: List[str] = []
    for source_value in (prepared_model, prepared_name_model):
        for candidate in generate_model_candidates(source_value, str(item.get("brand") or "")):
            if candidate and candidate not in candidates:
                candidates.append(candidate)
    return {
        **item,
        "name": prepared_name or raw_name,
        "model_candidates": candidates,
    }


def build_feed_comparison_own_indexes(
    own_sites: List[OwnSite],
    stop_event: Optional[threading.Event] = None,
) -> List[Dict[str, object]]:
    from runtime.news_tasks import feed_source_key, feed_source_label, trigger_feed_generation
    from services.file_import_service import build_feed_index_from_xml, stop_file_import_if_requested
    indexes: List[Dict[str, object]] = []
    for index, site in enumerate(own_sites, start=1):
        stop_file_import_if_requested(stop_event)
        if site.feed_generate_url:
            trigger_feed_generation(site.feed_generate_url)
        stop_file_import_if_requested(stop_event)
        content = download_comparison_feed(site.feed_url)
        feed = {
            "source": feed_source_key(site.feed_url),
            "source_label": site.name or feed_source_label(site.feed_url),
            "url": site.feed_url,
        }
        feed_index = build_feed_index_from_xml(content, feed, stop_event=stop_event)
        indexes.append({**feed, "index": feed_index, "codes_count": len(feed_index), "position": index})
    return indexes


def feed_comparison_result_filename() -> str:
    return f"Сравнение_фидов_{datetime.now(MSK_TZ).strftime('%d-%m-%Y_%H-%M-%S')}.xlsx"


def write_feed_comparison_workbook(
    result_path: Path,
    supplier_results: List[Dict[str, object]],
) -> None:
    from services.file_import_service import excel_sheet_title
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    workbook.remove(workbook.active)
    used_titles: Set[str] = set()
    for supplier_result in supplier_results:
        supplier_name = str(supplier_result.get("name") or "Поставщик")
        rows = supplier_result.get("rows") if isinstance(supplier_result.get("rows"), list) else []
        sheet = workbook.create_sheet(excel_sheet_title(supplier_name, used_titles))
        sheet.append(FEED_COMPARISON_RESULT_FIELDS)
        for row in rows:
            sheet.append([row.get(field, "") for field in FEED_COMPARISON_RESULT_FIELDS])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_index, field in enumerate(FEED_COMPARISON_RESULT_FIELDS, start=1):
            values = [str(field)]
            values.extend(str(row.get(field, "")) for row in rows[:200])
            width = min(max(len(value) for value in values) + 2, 60)
            sheet.column_dimensions[get_column_letter(column_index)].width = width
    if not workbook.sheetnames:
        sheet = workbook.create_sheet("Результат")
        sheet.append(FEED_COMPARISON_RESULT_FIELDS)
    workbook.save(result_path)


def compare_supplier_feeds(
    db_session=None,
    stop_event: Optional[threading.Event] = None,
) -> Dict[str, object]:
    from services.file_import_service import missing_feed_labels, stop_file_import_if_requested
    db = db_session or g.db
    own_sites = db.scalars(select(OwnSite).order_by(OwnSite.id)).all()
    suppliers = db.scalars(select(SupplierFeed).order_by(SupplierFeed.id)).all()
    if not own_sites:
        raise ValueError("Добавьте хотя бы один фид вашего сайта")
    if not suppliers:
        raise ValueError("Добавьте хотя бы один фид поставщика")

    started_at = time.time()
    update_feed_comparison_state(
        db,
        status="running",
        stage="Генерирую и загружаю фиды сайтов",
        percent=2,
        suppliers_total=len(suppliers),
        started_at=datetime.now(MSK_TZ).isoformat(timespec="seconds"),
        finished_at="",
        error="",
        result_filename="",
    )
    db.commit()
    own_indexes = build_feed_comparison_own_indexes(own_sites, stop_event=stop_event)
    stop_file_import_if_requested(stop_event)
    update_feed_comparison_state(db, stage="Загружаю фиды поставщиков", percent=15)
    db.commit()

    total_rows = 0
    supplier_results: List[Dict[str, object]] = []
    processed_rows = 0
    excluded_rows = 0
    missing_rows = 0
    progress_persisted_at = time.monotonic()

    for supplier_index, supplier in enumerate(suppliers, start=1):
        stop_file_import_if_requested(stop_event)
        update_feed_comparison_state(
            db,
            stage="Загружаю фид поставщика",
            current_supplier=supplier.name,
            suppliers_done=supplier_index - 1,
            percent=15 + int((supplier_index - 1) / max(len(suppliers), 1) * 75),
        )
        db.commit()
        content = download_comparison_feed(supplier.feed_url)
        stop_file_import_if_requested(stop_event)
        update_feed_comparison_state(
            db,
            stage="Читаю фид поставщика",
            current_supplier=supplier.name,
        )
        db.commit()
        source_rows = read_supplier_feed_rows(
            content,
            supplier.model_field,
            stop_event=stop_event,
            name_field=supplier.name_field,
            price_field=supplier.price_field,
            brand_field=supplier.brand_field,
            url_field=supplier.url_field,
        )
        if not source_rows:
            configured_field = normalize_supplier_model_field(supplier.model_field)
            feed_format = detect_supplier_feed_format(content)
            if feed_format == "xml":
                available_fields = supplier_xml_available_fields(content)
            elif feed_format == "xlsx":
                available_fields = supplier_xlsx_available_fields(content)
            else:
                available_fields = supplier_csv_available_fields(content)
            available_hint = (
                f" Найденные поля: {', '.join(available_fields)}."
                if available_fields
                else ""
            )
            raise ValueError(
                f'В фиде поставщика "{supplier.name}" не найдено товаров по полю '
                f'"{configured_field}". Проверьте имя поля модели в '
                f'{feed_format.upper()}.{available_hint}'
            )
        del content
        total_rows += len(source_rows)
        result_rows: List[Dict[str, object]] = []
        for row_index, item in enumerate(source_rows, start=1):
            stop_file_import_if_requested(stop_event)
            prepared_item = prepare_supplier_feed_item(
                item,
                exclusions=supplier.exclusions,
                replace_rules=supplier.replace_rules,
            )
            if prepared_item is None:
                excluded_rows += 1
            else:
                candidates = list(prepared_item.get("model_candidates") or [])
                if candidates:
                    missing_labels = missing_feed_labels(candidates, own_indexes, stop_event=stop_event)
                    if missing_labels:
                        missing_rows += 1
                        result_rows.append(
                            {
                                "row": item.get("row_number"),
                                "name": prepared_item.get("name") or "",
                                "price": item.get("price") or "",
                                "brand": item.get("brand") or "",
                                "model_candidates": " | ".join(candidates),
                                "selected_model": candidates[0],
                                "missing_on": ", ".join(f"Нет на {label}" for label in missing_labels),
                            }
                        )
            processed_rows += 1
            progress_checked_at = time.monotonic()
            if (
                row_index == len(source_rows)
                or progress_checked_at - progress_persisted_at
                >= FEED_COMPARISON_PROGRESS_COMMIT_INTERVAL_SECONDS
            ):
                supplier_progress = row_index / max(len(source_rows), 1)
                overall_progress = (supplier_index - 1 + supplier_progress) / max(len(suppliers), 1)
                update_feed_comparison_state(
                    db,
                    stage="Сравниваю модели",
                    current_supplier=supplier.name,
                    current_row=row_index,
                    total_rows=total_rows,
                    processed_rows=processed_rows,
                    excluded_rows=excluded_rows,
                    missing_rows=missing_rows,
                    percent=15 + int(overall_progress * 75),
                    elapsed_seconds=int(time.time() - started_at),
                )
                db.commit()
                progress_persisted_at = progress_checked_at
        supplier_results.append({"name": supplier.name, "rows": result_rows})
        del source_rows
        update_feed_comparison_state(db, suppliers_done=supplier_index)
        db.commit()

    comparison = get_feed_comparison_row(db)
    remove_feed_comparison_export(comparison)
    result_path = EXPORT_DIR / feed_comparison_result_filename()
    update_feed_comparison_state(db, stage="Записываю XLSX", percent=96)
    db.commit()
    stop_file_import_if_requested(stop_event)
    write_feed_comparison_workbook(result_path, supplier_results)
    comparison.export_path = result_path.name
    state = update_feed_comparison_state(
        db,
        status="completed",
        stage="Готово",
        percent=100,
        current_supplier="",
        current_row=total_rows,
        total_rows=total_rows,
        processed_rows=processed_rows,
        excluded_rows=excluded_rows,
        missing_rows=missing_rows,
        suppliers_done=len(suppliers),
        suppliers_total=len(suppliers),
        finished_at=datetime.now(MSK_TZ).isoformat(timespec="seconds"),
        elapsed_seconds=int(time.time() - started_at),
        result_filename=result_path.name,
        error="",
    )
    db.commit()
    return {"state": state, "result_filename": result_path.name}


def run_feed_comparison(stop_event: threading.Event) -> None:
    from services.file_import_service import FileImportStopped
    global feed_comparison_worker_thread
    db = SessionLocal()
    try:
        compare_supplier_feeds(db, stop_event=stop_event)
    except FileImportStopped:
        db.rollback()
        try:
            update_feed_comparison_state(
                db,
                status="stopped",
                stage="Остановлено",
                finished_at=datetime.now(MSK_TZ).isoformat(timespec="seconds"),
            )
            db.commit()
        except Exception:
            db.rollback()
    except Exception as exc:
        db.rollback()
        try:
            update_feed_comparison_state(
                db,
                status="error",
                stage="Ошибка",
                error=str(exc),
                finished_at=datetime.now(MSK_TZ).isoformat(timespec="seconds"),
            )
            db.commit()
        except Exception:
            db.rollback()
    finally:
        db.close()
        with feed_comparison_lock:
            if threading.current_thread() is feed_comparison_worker_thread:
                feed_comparison_worker_thread = None
                feed_comparison_stop_event.clear()


def recover_interrupted_feed_comparison() -> None:
    with session_scope() as db_session:
        row = get_feed_comparison_row(db_session)
        state = normalize_feed_comparison_state(row.state)
        if not is_feed_comparison_active_state(state):
            return
        row.state = {
            **state,
            "status": "error",
            "stage": "Прервано",
            "error": "Сравнение фидов было прервано перезапуском сервера. Запустите его снова.",
            "finished_at": datetime.now(MSK_TZ).isoformat(timespec="seconds"),
        }


def start_feed_comparison() -> Dict[str, object]:
    global feed_comparison_worker_thread
    with feed_comparison_lock:
        if feed_comparison_worker_thread and feed_comparison_worker_thread.is_alive():
            return public_feed_comparison_progress()
        comparison = get_feed_comparison_row()
        if is_feed_comparison_active_state(comparison.state):
            return public_feed_comparison_progress()
        if not g.db.scalar(select(OwnSite.id).limit(1)):
            raise ValueError("Добавьте хотя бы один фид вашего сайта")
        if not g.db.scalar(select(SupplierFeed.id).limit(1)):
            raise ValueError("Добавьте хотя бы один фид поставщика")
        remove_feed_comparison_export(comparison)
        comparison.export_path = ""
        comparison.state = {
            **make_feed_comparison_state("queued"),
            "stage": "В очереди",
            "started_at": datetime.now(MSK_TZ).isoformat(timespec="seconds"),
        }
        g.db.commit()
        feed_comparison_stop_event.clear()
        feed_comparison_worker_thread = threading.Thread(
            target=run_feed_comparison,
            args=(feed_comparison_stop_event,),
            daemon=True,
        )
        feed_comparison_worker_thread.start()
        return public_feed_comparison_progress()


def feed_comparison_worker_alive() -> bool:
    return bool(feed_comparison_worker_thread and feed_comparison_worker_thread.is_alive())
