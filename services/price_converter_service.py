"""Supplier price-list conversion into the service's normalized CSV format."""

from __future__ import annotations

import calendar
import csv
import io
import math
import os
import re
import time
from datetime import date, datetime
from pathlib import Path
from typing import Dict, Iterable, Iterator, List, Optional, Sequence, Tuple

from flask import g

from config import (
    EXPORT_DIR,
    MSK_TZ,
    PRICE_CONVERTER_ALLOWED_SUFFIXES,
    PRICE_CONVERTER_DIR,
)
from database.session import session_scope
from models import PriceConverter
from services.domain_revisions import bump_domain_revision, domain_revision
from services.normalization import output_text, safe_filename
from services.scraping import clean_text


def make_price_converter_state(status: str = "idle") -> Dict[str, object]:
    return {
        "status": status,
        "stage": "",
        "rows_written": 0,
        "matched_sheets": 0,
        "skipped_sheets": 0,
        "error": "",
        "started_at": "",
        "finished_at": "",
        "elapsed_seconds": 0,
        "result_filename": "",
    }


def normalize_price_converter_state(value: object) -> Dict[str, object]:
    state = {
        **make_price_converter_state(),
        **(value if isinstance(value, dict) else {}),
    }
    state["status"] = str(state.get("status") or "idle")
    for field in ("rows_written", "matched_sheets", "skipped_sheets", "elapsed_seconds"):
        try:
            state[field] = max(0, int(float(state.get(field) or 0)))
        except (TypeError, ValueError):
            state[field] = 0
    for field in ("stage", "error", "started_at", "finished_at", "result_filename"):
        state[field] = str(state.get(field) or "")
    return state


def normalize_sheet_number(value: object) -> Optional[int]:
    if value is None or str(value).strip() == "":
        return None
    text = str(value).strip()
    if not re.fullmatch(r"\d+", text):
        raise ValueError("Номер листа должен быть целым положительным числом")
    number = int(text)
    if number < 1:
        raise ValueError("Номер листа должен быть не меньше 1")
    return number


def normalize_promo_date(value: object) -> Optional[date]:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value).strip())
    except ValueError as exc:
        raise ValueError("Дата промо должна быть указана в формате ГГГГ-ММ-ДД") from exc


def normalize_promo_settings(
    promo_field: object,
    promo_date: object,
) -> Tuple[str, Optional[date]]:
    normalized_field = clean_text(str(promo_field or ""))[:255]
    normalized_date = normalize_promo_date(promo_date)
    if bool(normalized_field) != bool(normalized_date):
        raise ValueError(
            "Название столбца промо и дата промо должны быть заполнены вместе или оставлены пустыми"
        )
    return normalized_field, normalized_date


def price_converter_file_path(row: PriceConverter) -> Optional[Path]:
    metadata = row.file if isinstance(row.file, dict) else {}
    filename = str(metadata.get("stored_filename") or "").strip()
    if not filename:
        return None
    base_dir = PRICE_CONVERTER_DIR.resolve()
    path = (PRICE_CONVERTER_DIR / filename).resolve()
    if base_dir not in path.parents or not path.exists() or not path.is_file():
        return None
    return path


def clear_price_converter_storage() -> None:
    PRICE_CONVERTER_DIR.mkdir(parents=True, exist_ok=True)
    for path in PRICE_CONVERTER_DIR.iterdir():
        if path.is_file():
            try:
                path.unlink()
            except OSError:
                continue


def stored_price_converter_files() -> List[Path]:
    PRICE_CONVERTER_DIR.mkdir(parents=True, exist_ok=True)
    return sorted(
        [
            path
            for path in PRICE_CONVERTER_DIR.iterdir()
            if path.is_file() and path.suffix.lower() in PRICE_CONVERTER_ALLOWED_SUFFIXES
        ],
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    )


def get_price_converter_row(db_session=None) -> PriceConverter:
    db = db_session or g.db
    row = db.get(PriceConverter, 1)
    if row is None:
        row = PriceConverter(
            id=1,
            model_field="",
            price_field="",
            promo_field="",
            promo_date=None,
            sheet_number=None,
            export_path="",
            file={},
            state=make_price_converter_state(),
        )
        db.add(row)
        db.flush()
    else:
        normalized_state = normalize_price_converter_state(row.state)
        if row.state != normalized_state:
            row.state = normalized_state
        normalized_model = clean_text(str(row.model_field or ""))[:255]
        normalized_price = clean_text(str(row.price_field or ""))[:255]
        normalized_promo = clean_text(str(row.promo_field or ""))[:255]
        normalized_promo_date = normalize_promo_date(row.promo_date)
        normalized_sheet = normalize_sheet_number(row.sheet_number)
        if row.model_field != normalized_model:
            row.model_field = normalized_model
        if row.price_field != normalized_price:
            row.price_field = normalized_price
        if row.promo_field != normalized_promo:
            row.promo_field = normalized_promo
        if row.promo_date != normalized_promo_date:
            row.promo_date = normalized_promo_date
        if row.sheet_number != normalized_sheet:
            row.sheet_number = normalized_sheet
        db.flush()

    if not isinstance(row.file, dict) or not row.file.get("stored_filename"):
        stored_files = stored_price_converter_files()
        if stored_files:
            path = stored_files[0]
            row.file = {
                "original_filename": path.name.split("_", 3)[-1] if "_" in path.name else path.name,
                "stored_filename": path.name,
                "uploaded_at": datetime.fromtimestamp(path.stat().st_mtime, MSK_TZ).isoformat(timespec="seconds"),
            }
            db.flush()
    return row


def resolve_price_converter_export_path(value: object) -> Optional[Path]:
    filename = Path(str(value or "")).name
    if not filename:
        return None
    path = (EXPORT_DIR / filename).resolve()
    if EXPORT_DIR.resolve() not in path.parents or not path.exists() or not path.is_file():
        return None
    return path


def remove_price_converter_export(row: PriceConverter) -> None:
    state = row.state if isinstance(row.state, dict) else {}
    candidates = {
        str(row.export_path or ""),
        str(state.get("result_filename") or ""),
    }
    for candidate in candidates:
        path = resolve_price_converter_export_path(candidate)
        if path:
            try:
                path.unlink()
            except OSError:
                pass

    metadata = row.file if isinstance(row.file, dict) else {}
    original_filename = str(metadata.get("original_filename") or "").strip()
    if original_filename:
        prefix = f"Прайс_{safe_filename(Path(original_filename).stem or 'file')}_"
        for path in EXPORT_DIR.glob(f"{prefix}*.csv"):
            if path.is_file():
                try:
                    path.unlink()
                except OSError:
                    pass


def price_converter_settings(row: PriceConverter) -> Dict[str, object]:
    return {
        "model_field": str(row.model_field or ""),
        "price_field": str(row.price_field or ""),
        "promo_field": str(row.promo_field or ""),
        "promo_date": row.promo_date.isoformat() if row.promo_date else "",
        "sheet_number": row.sheet_number,
    }


def _price_converter_runtime(row: PriceConverter) -> Dict[str, object]:
    path = price_converter_file_path(row)
    metadata = row.file if isinstance(row.file, dict) else {}
    state = normalize_price_converter_state(row.state)
    result_filename = Path(
        str(row.export_path or state.get("result_filename") or "")
    ).name
    result_path = resolve_price_converter_export_path(result_filename)
    payload: Dict[str, object] = {
        "revision": domain_revision("price_converter"),
        "file": None,
        "state": state,
        "result_filename": result_filename,
        "result_ready": bool(result_path and state.get("status") != "running"),
    }
    if path:
        stat = path.stat()
        payload["file"] = {
            "filename": output_text(metadata.get("original_filename") or path.name),
            "stored_filename": path.name,
            "size": stat.st_size,
            "uploaded_at": str(
                metadata.get("uploaded_at")
                or datetime.fromtimestamp(stat.st_mtime, MSK_TZ).isoformat(timespec="seconds")
            ),
        }
    return payload


def public_price_converter_state(db_session=None) -> Dict[str, object]:
    row = get_price_converter_row(db_session)
    return {**_price_converter_runtime(row), **price_converter_settings(row)}


def public_price_converter_runtime(db_session=None) -> Dict[str, object]:
    db = db_session or g.db
    row = db.get(PriceConverter, 1)
    if row is None:
        return {
            "revision": domain_revision("price_converter"),
            "file": None,
            "state": make_price_converter_state(),
            "result_filename": "",
            "result_ready": False,
        }
    return _price_converter_runtime(row)


def public_price_converter_settings(db_session=None) -> Dict[str, object]:
    row = get_price_converter_row(db_session)
    return {
        "revision": domain_revision("price_converter"),
        **price_converter_settings(row),
    }


def decode_price_csv(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1251", "windows-1251"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def normalize_header(value: object) -> str:
    return clean_text(cell_text(value)).casefold()


def header_aliases(value: object) -> set[str]:
    aliases = {normalize_header(value)}
    if isinstance(value, (datetime, date)):
        aliases.update(
            {
                normalize_header(value.isoformat()),
                normalize_header(value.strftime("%d.%m.%Y")),
                normalize_header(value.strftime("%d.%m")),
                normalize_header(value.strftime("%Y-%m-%d")),
                normalize_header(calendar.month_abbr[value.month]),
                normalize_header(calendar.month_name[value.month]),
            }
        )
    return {alias for alias in aliases if alias}


def cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if not math.isfinite(value):
            return ""
        if value.is_integer():
            return str(int(value))
        return format(value, ".15g")
    return clean_text(str(value))


def normalize_price(value: object) -> str:
    return re.sub(r"\D+", "", cell_text(value))


def extract_price_converter_model(
    value: object,
    known_brands: Sequence[str],
) -> str:
    source_model = cell_text(value)
    if not source_model:
        return ""
    from services.file_import_service import generate_model_candidates

    candidates = generate_model_candidates(
        source_model,
        known_brands=known_brands,
    )
    return candidates[0] if candidates else source_model


def _header_indexes(
    row: Sequence[object],
    model_field: str,
    price_field: str,
    promo_field: str = "",
) -> Optional[Tuple[int, int, Optional[int]]]:
    expected_model = normalize_header(model_field)
    expected_price = normalize_header(price_field)
    expected_promo = normalize_header(promo_field)
    model_index = None
    price_index = None
    promo_index = None
    for index, value in enumerate(row):
        aliases = header_aliases(value)
        if model_index is None and expected_model in aliases:
            model_index = index
        if price_index is None and expected_price in aliases:
            price_index = index
        if expected_promo and promo_index is None and expected_promo in aliases:
            promo_index = index
    if (
        model_index is None
        or price_index is None
        or (expected_promo and promo_index is None)
    ):
        return None
    return model_index, price_index, promo_index


def _write_sheet_rows(
    rows: Iterator[Sequence[object]],
    writer,
    model_field: str,
    price_field: str,
    promo_field: str = "",
    promo_start_date: Optional[date] = None,
    promo_end_date: Optional[date] = None,
    known_brands: Sequence[str] = (),
) -> Optional[int]:
    indexes = None
    for row in rows:
        indexes = _header_indexes(row, model_field, price_field, promo_field)
        if indexes is not None:
            break
    if indexes is None:
        return None

    model_index, price_index, promo_index = indexes
    written = 0
    for row in rows:
        model = extract_price_converter_model(
            row[model_index] if model_index < len(row) else None,
            known_brands,
        )
        if not model:
            continue
        price = normalize_price(row[price_index] if price_index < len(row) else None)
        if not price:
            continue
        output_row = [model, price]
        if promo_index is not None and promo_start_date and promo_end_date:
            promo_price = normalize_price(
                row[promo_index] if promo_index < len(row) else None
            )
            special = (
                f"1,1,{promo_price},{promo_start_date.isoformat()},{promo_end_date.isoformat()}"
                if promo_price
                else ""
            )
            output_row.append(special)
        writer.writerow(output_row)
        written += 1
    return written


def _required_headers_text(
    model_field: str,
    price_field: str,
    promo_field: str = "",
) -> str:
    fields = [model_field, price_field]
    if promo_field:
        fields.append(promo_field)
    return " и ".join(f"«{field}»" for field in fields)


def _selected_indexes(total_sheets: int, sheet_number: Optional[int]) -> List[int]:
    if sheet_number is None:
        return list(range(total_sheets))
    if sheet_number > total_sheets:
        raise ValueError(
            f"Лист №{sheet_number} не найден. В файле листов: {total_sheets}"
        )
    return [sheet_number - 1]


def _iter_xls_sheet_rows(sheet) -> Iterator[Sequence[object]]:
    for row_index in range(sheet.nrows):
        yield sheet.row_values(row_index)


def _convert_sheet_sources(
    sources: Iterable[Tuple[int, str, Iterator[Sequence[object]]]],
    writer,
    model_field: str,
    price_field: str,
    explicit_sheet: bool,
    promo_field: str = "",
    promo_start_date: Optional[date] = None,
    promo_end_date: Optional[date] = None,
    known_brands: Sequence[str] = (),
) -> Dict[str, int]:
    matched = 0
    skipped = 0
    written = 0
    for sheet_index, sheet_name, rows in sources:
        sheet_written = _write_sheet_rows(
            rows,
            writer,
            model_field,
            price_field,
            promo_field,
            promo_start_date,
            promo_end_date,
            known_brands,
        )
        if sheet_written is None:
            skipped += 1
            if explicit_sheet:
                raise ValueError(
                    f"На листе №{sheet_index} «{sheet_name}» не найдена строка, "
                    "содержащая одновременно столбцы "
                    f"{_required_headers_text(model_field, price_field, promo_field)}"
                )
            continue
        matched += 1
        written += sheet_written

    if not matched:
        raise ValueError(
            f"Не найдена строка, содержащая одновременно столбцы "
            f"{_required_headers_text(model_field, price_field, promo_field)}"
        )
    return {
        "rows_written": written,
        "matched_sheets": matched,
        "skipped_sheets": skipped,
    }


def convert_price_source(
    source_path: Path,
    output_path: Path,
    model_field: str,
    price_field: str,
    sheet_number: Optional[int] = None,
    promo_field: str = "",
    promo_date: object = None,
    *,
    conversion_date: Optional[date] = None,
) -> Dict[str, int]:
    model_field = clean_text(str(model_field or ""))
    price_field = clean_text(str(price_field or ""))
    if not model_field:
        raise ValueError("Укажите название столбца модели")
    if not price_field:
        raise ValueError("Укажите название столбца цены")
    if normalize_header(model_field) == normalize_header(price_field):
        raise ValueError("Названия столбцов модели и цены должны отличаться")
    promo_field, normalized_promo_date = normalize_promo_settings(
        promo_field,
        promo_date,
    )
    configured_fields = [normalize_header(model_field), normalize_header(price_field)]
    if promo_field:
        configured_fields.append(normalize_header(promo_field))
    if len(set(configured_fields)) != len(configured_fields):
        raise ValueError("Названия столбцов модели, цены и промо должны отличаться")
    sheet_number = normalize_sheet_number(sheet_number)
    promo_start_date = (
        normalize_promo_date(conversion_date) or datetime.now(MSK_TZ).date()
    )
    from services.file_import_service import known_file_import_brands

    known_brands = known_file_import_brands()
    suffix = source_path.suffix.lower()
    if suffix not in PRICE_CONVERTER_ALLOWED_SUFFIXES:
        raise ValueError("Можно обработать только CSV, XLS или XLSX")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary = output_path.with_suffix(output_path.suffix + ".tmp")
    try:
        with temporary.open("w", encoding="utf-8-sig", newline="") as output:
            writer = csv.writer(output, delimiter=";", lineterminator="\n")
            headers = ["_MODEL_", "_PRICE_"]
            if promo_field:
                headers.append("_SPECIAL_")
            writer.writerow(headers)

            if suffix == ".csv":
                if sheet_number not in (None, 1):
                    raise ValueError("В CSV доступен только лист №1")
                text = decode_price_csv(source_path.read_bytes())
                sample = text[:4096]
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=";,|\t,")
                except csv.Error:
                    dialect = csv.excel
                    dialect.delimiter = ";"
                sources = [(1, "CSV", iter(csv.reader(io.StringIO(text), dialect)))]
                result = _convert_sheet_sources(
                    sources,
                    writer,
                    model_field,
                    price_field,
                    explicit_sheet=sheet_number is not None,
                    promo_field=promo_field,
                    promo_start_date=promo_start_date,
                    promo_end_date=normalized_promo_date,
                    known_brands=known_brands,
                )
            elif suffix == ".xlsx":
                from openpyxl import load_workbook

                workbook = load_workbook(source_path, read_only=True, data_only=True)
                try:
                    indexes = _selected_indexes(len(workbook.worksheets), sheet_number)
                    sources = (
                        (
                            index + 1,
                            workbook.worksheets[index].title,
                            iter(workbook.worksheets[index].iter_rows(values_only=True)),
                        )
                        for index in indexes
                    )
                    result = _convert_sheet_sources(
                        sources,
                        writer,
                        model_field,
                        price_field,
                        explicit_sheet=sheet_number is not None,
                        promo_field=promo_field,
                        promo_start_date=promo_start_date,
                        promo_end_date=normalized_promo_date,
                        known_brands=known_brands,
                    )
                finally:
                    workbook.close()
            else:
                try:
                    import xlrd
                except ImportError as exc:
                    raise ValueError("Для обработки XLS установите зависимость xlrd") from exc

                workbook = xlrd.open_workbook(str(source_path), on_demand=True)
                try:
                    indexes = _selected_indexes(workbook.nsheets, sheet_number)
                    sheets = [workbook.sheet_by_index(index) for index in indexes]
                    sources = (
                        (index + 1, sheet.name, _iter_xls_sheet_rows(sheet))
                        for index, sheet in zip(indexes, sheets)
                    )
                    result = _convert_sheet_sources(
                        sources,
                        writer,
                        model_field,
                        price_field,
                        explicit_sheet=sheet_number is not None,
                        promo_field=promo_field,
                        promo_start_date=promo_start_date,
                        promo_end_date=normalized_promo_date,
                        known_brands=known_brands,
                    )
                finally:
                    workbook.release_resources()
        os.replace(temporary, output_path)
        return result
    except Exception:
        temporary.unlink(missing_ok=True)
        raise


def price_converter_result_filename(original_filename: str) -> str:
    stem = safe_filename(Path(original_filename).stem or "file")
    timestamp = datetime.now(MSK_TZ).strftime("%Y%m%d_%H%M%S")
    return f"Прайс_{stem}_{timestamp}.csv"


def run_price_conversion(db_session=None) -> Dict[str, object]:
    db = db_session or g.db
    row = get_price_converter_row(db)
    source_path = price_converter_file_path(row)
    if not source_path:
        raise ValueError("Сначала прикрепите исходный файл")

    model_field = clean_text(str(row.model_field or ""))
    price_field = clean_text(str(row.price_field or ""))
    promo_field, promo_date = normalize_promo_settings(
        row.promo_field,
        row.promo_date,
    )
    sheet_number = normalize_sheet_number(row.sheet_number)
    if not model_field:
        raise ValueError("Укажите название столбца модели")
    if not price_field:
        raise ValueError("Укажите название столбца цены")

    started = time.time()
    started_at = datetime.now(MSK_TZ).isoformat(timespec="seconds")
    remove_price_converter_export(row)
    row.export_path = ""
    row.state = {
        **make_price_converter_state("running"),
        "stage": "Конвертирую файл",
        "started_at": started_at,
    }
    db.commit()

    metadata = row.file if isinstance(row.file, dict) else {}
    original_filename = str(metadata.get("original_filename") or source_path.name)
    result_filename = price_converter_result_filename(original_filename)
    result_path = EXPORT_DIR / result_filename
    try:
        result = convert_price_source(
            source_path,
            result_path,
            model_field,
            price_field,
            sheet_number,
            promo_field,
            promo_date,
        )
    except Exception as error:
        result_path.unlink(missing_ok=True)
        row.export_path = ""
        row.state = {
            **make_price_converter_state("error"),
            "stage": "Ошибка",
            "error": str(error),
            "started_at": started_at,
            "finished_at": datetime.now(MSK_TZ).isoformat(timespec="seconds"),
            "elapsed_seconds": int(time.time() - started),
        }
        db.commit()
        bump_domain_revision("price_converter")
        raise

    elapsed = int(time.time() - started)
    row.export_path = result_filename
    row.state = {
        **make_price_converter_state("completed"),
        **result,
        "stage": "Готово",
        "started_at": started_at,
        "finished_at": datetime.now(MSK_TZ).isoformat(timespec="seconds"),
        "elapsed_seconds": elapsed,
        "result_filename": result_filename,
    }
    db.commit()
    bump_domain_revision("price_converter")
    return public_price_converter_runtime(db)


def recover_interrupted_price_conversion() -> None:
    with session_scope() as db:
        row = db.get(PriceConverter, 1)
        if row is None:
            return
        state = normalize_price_converter_state(row.state)
        if state.get("status") != "running":
            return
        row.state = {
            **state,
            "status": "error",
            "stage": "Прервано",
            "error": "Конвертация была прервана перезапуском сервера. Запустите её снова.",
            "finished_at": datetime.now(MSK_TZ).isoformat(timespec="seconds"),
        }
