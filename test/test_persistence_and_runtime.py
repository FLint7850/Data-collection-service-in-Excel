import io
import threading
import unittest
import zipfile
from contextlib import contextmanager
from copy import deepcopy
from datetime import date, datetime, timedelta
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

from sqlalchemy import create_engine, inspect, text
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from database.repositories.projects import delete_project, update_project
from config import MSK_TZ
from models import Base, Brand, Donor, FileImport, PriceConverter, Project
from runtime.state import news_lock, news_settings
from services.file_validation import validate_xlsx_archive


class IsolatedDatabaseTestCase(unittest.TestCase):
    def setUp(self) -> None:
        self.engine = create_engine(
            "sqlite://",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(self.engine)
        self.Session = sessionmaker(bind=self.engine, expire_on_commit=False)

    def tearDown(self) -> None:
        self.engine.dispose()


class TargetedPersistenceTests(IsolatedDatabaseTestCase):
    def test_project_update_and_delete_do_not_touch_siblings(self) -> None:
        with self.Session.begin() as session:
            first = Project(legacy_id="first", name="First", start_urls=[])
            second = Project(legacy_id="second", name="Second", start_urls=[])
            session.add_all([first, second])
            session.flush()
            first_id, second_id = first.id, second.id

            update_project(first_id, {"name": "Updated", "unknown": "ignored"}, session)
            self.assertTrue(delete_project(first_id, session))

        with self.Session() as session:
            self.assertIsNone(session.get(Project, first_id))
            self.assertEqual(session.get(Project, second_id).name, "Second")

    def test_brand_deletion_removes_all_nested_donors(self) -> None:
        import services.news as news_service

        with self.Session.begin() as session:
            brand = Brand(
                name="Deleted brand",
                search_name="deleted brand",
                group_name="Маржа",
                state={"status": "idle"},
            )
            session.add(brand)
            session.flush()
            donors = [
                Donor(
                    legacy_id=f"deleted-{index}",
                    brand_id=brand.id,
                    site_url=f"https://example.test/{index}",
                    start_urls=[],
                )
                for index in range(2)
            ]
            session.add_all(donors)
            session.flush()
            brand_id = brand.id
            donor_ids = [donor.id for donor in donors]
            brand.primary_donor_id = donor_ids[0]

        @contextmanager
        def isolated_session_scope():
            session = self.Session()
            try:
                yield session
                session.commit()
            except Exception:
                session.rollback()
                raise
            finally:
                session.close()

        with patch.object(news_service, "session_scope", isolated_session_scope):
            primary_by_brand = news_service.delete_news_records(
                donor_ids,
                remove_brand=True,
                brand_id=brand_id,
            )

        self.assertEqual(primary_by_brand, {brand_id: None})
        with self.Session() as session:
            self.assertIsNone(session.get(Brand, brand_id))
            self.assertTrue(all(session.get(Donor, donor_id) is None for donor_id in donor_ids))

    def test_primary_donor_deletion_reassigns_brand_schedule_target(self) -> None:
        import services.news as news_service

        with self.Session.begin() as session:
            brand = Brand(
                name="Active brand",
                search_name="active brand",
                group_name="Маржа",
                state={"status": "idle"},
            )
            session.add(brand)
            session.flush()
            first = Donor(
                legacy_id="primary-first",
                brand_id=brand.id,
                site_url="https://first.example.test",
                start_urls=[],
            )
            second = Donor(
                legacy_id="primary-second",
                brand_id=brand.id,
                site_url="https://second.example.test",
                start_urls=[],
            )
            session.add_all([first, second])
            session.flush()
            brand_id = brand.id
            first_id = first.id
            second_id = second.id
            brand.primary_donor_id = first_id

        @contextmanager
        def isolated_session_scope():
            session = self.Session()
            try:
                yield session
                session.commit()
            except Exception:
                session.rollback()
                raise
            finally:
                session.close()

        with patch.object(news_service, "session_scope", isolated_session_scope):
            primary_by_brand = news_service.delete_news_records([first_id])

        self.assertEqual(primary_by_brand, {brand_id: second_id})
        with self.Session() as session:
            self.assertIsNone(session.get(Donor, first_id))
            self.assertEqual(session.get(Brand, brand_id).primary_donor_id, second_id)


class SupplierFeedConfigurationTests(unittest.TestCase):
    def test_supplier_payload_keeps_configured_optional_fields(self) -> None:
        from services.feeds import validate_feed_comparison_site_payload

        payload = validate_feed_comparison_site_payload(
            {
                "name": "Поставщик",
                "feed_url": "https://example.test/feed.xml",
                "model_field": "<model>",
                "name_field": "<product_name>",
                "price_field": "price",
                "brand_field": "param:Бренд",
                "url_field": " ",
            },
            supplier=True,
        )

        self.assertEqual(payload["model_field"], "model")
        self.assertEqual(payload["name_field"], "product_name")
        self.assertEqual(payload["price_field"], "price")
        self.assertEqual(payload["brand_field"], "param:Бренд")
        self.assertEqual(payload["url_field"], "")

    def test_csv_reads_only_explicitly_configured_optional_columns(self) -> None:
        from services.feeds import read_supplier_feed_rows

        content = (
            "SKU;Название поставщика;Стоимость;Марка поставщика;Карточка\n"
            "A-1;Чайник;12990;MAUNFELD;https://example.test/a-1\n"
        ).encode("utf-8")

        configured = read_supplier_feed_rows(
            content,
            "SKU",
            name_field="Название поставщика",
            price_field="Стоимость",
            brand_field="Марка поставщика",
            url_field="Карточка",
        )
        without_optional_fields = read_supplier_feed_rows(content, "SKU")

        self.assertEqual(
            configured[0],
            {
                "row_number": 2,
                "source_model": "A-1",
                "name": "Чайник",
                "price": "12990",
                "brand": "MAUNFELD",
                "url": "https://example.test/a-1",
            },
        )
        self.assertEqual(without_optional_fields[0]["name"], "A-1")
        self.assertEqual(without_optional_fields[0]["price"], "")
        self.assertEqual(without_optional_fields[0]["brand"], "")
        self.assertEqual(without_optional_fields[0]["url"], "")

    def test_xml_reads_only_explicitly_configured_optional_fields(self) -> None:
        from services.feeds import read_supplier_feed_rows

        content = b"""<?xml version="1.0" encoding="UTF-8"?>
        <offers><offer>
          <sku>A-1</sku><display>Chaynik</display><cost>12990</cost>
          <maker>MAUNFELD</maker><href>https://example.test/a-1</href>
        </offer></offers>"""

        rows = read_supplier_feed_rows(
            content,
            "sku",
            name_field="display",
            price_field="cost",
            brand_field="maker",
            url_field="href",
        )

        self.assertEqual(rows[0]["name"], "Chaynik")
        self.assertEqual(rows[0]["price"], "12990")
        self.assertEqual(rows[0]["brand"], "MAUNFELD")
        self.assertEqual(rows[0]["url"], "https://example.test/a-1")

    def test_runtime_migration_adds_supplier_field_columns(self) -> None:
        from database.session import migrate_supplier_feeds_table

        engine = create_engine("sqlite://")
        try:
            with engine.begin() as connection:
                connection.execute(
                    text(
                        "CREATE TABLE supplier_feeds ("
                        "id INTEGER PRIMARY KEY, model_field VARCHAR(255) NOT NULL, "
                        "exclusions JSON NOT NULL DEFAULT '[]', "
                        "replace_rules TEXT NOT NULL DEFAULT ''"
                        ")"
                    )
                )
                migrate_supplier_feeds_table(connection)
            columns = {column["name"] for column in inspect(engine).get_columns("supplier_feeds")}
        finally:
            engine.dispose()

        self.assertTrue(
            {"name_field", "price_field", "brand_field", "url_field"}.issubset(columns)
        )


class AttributeAssistantMigrationTests(unittest.TestCase):
    def test_runtime_migration_upgrades_legacy_attribute_tables_without_data_loss(self) -> None:
        from database.session import migrate_attribute_assistant_tables
        from models import AttributeBatch, AttributeMappingRule
        from sqlalchemy import select

        engine = create_engine("sqlite://")
        try:
            with engine.begin() as connection:
                connection.execute(
                    text(
                        "CREATE TABLE attribute_batches ("
                        "id INTEGER PRIMARY KEY, template_id INTEGER, "
                        "source_filename VARCHAR(500) NOT NULL, "
                        "stored_filename VARCHAR(500) NOT NULL, "
                        "export_filename VARCHAR(500) NOT NULL, "
                        "status VARCHAR(32) NOT NULL, "
                        "products_count INTEGER NOT NULL, "
                        "attributes_count INTEGER NOT NULL, "
                        "summary JSON NOT NULL, "
                        "created_at DATETIME NOT NULL, updated_at DATETIME NOT NULL, "
                        "report_filename VARCHAR(500) NOT NULL DEFAULT '', "
                        "input_mode VARCHAR(32) NOT NULL DEFAULT 'csv', "
                        "processing_mode VARCHAR(32) NOT NULL DEFAULT 'suggest', "
                        "source_urls JSON NOT NULL DEFAULT '[]'"
                        ")"
                    )
                )
                connection.execute(
                    text(
                        "INSERT INTO attribute_batches "
                        "(id, template_id, source_filename, stored_filename, export_filename, "
                        "status, products_count, attributes_count, summary, created_at, updated_at) "
                        "VALUES (2, 1, 'products.csv', 'stored.csv', 'result.csv', "
                        "'ready', 14, 1013, '{}', CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)"
                    )
                )
                connection.execute(
                    text(
                        "CREATE TABLE attribute_mapping_rules ("
                        "id INTEGER PRIMARY KEY, donor_id INTEGER NOT NULL, template_id INTEGER, "
                        "template_field_id INTEGER NOT NULL, donor_attribute_name VARCHAR(255) NOT NULL, "
                        "normalized_donor_name VARCHAR(255) NOT NULL, confidence INTEGER NOT NULL, "
                        "is_active BOOLEAN NOT NULL, created_at DATETIME NOT NULL, updated_at DATETIME NOT NULL"
                        ")"
                    )
                )
                connection.execute(
                    text(
                        "INSERT INTO attribute_mapping_rules "
                        "(id, donor_id, template_id, template_field_id, donor_attribute_name, "
                        "normalized_donor_name, confidence, is_active, created_at, updated_at) "
                        "VALUES (3, 4, 1, 5, 'Spin speed', 'spin speed', 100, 1, "
                        "CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)"
                    )
                )
                migrate_attribute_assistant_tables(connection)
                migrate_attribute_assistant_tables(connection)

            batch_columns = {column["name"] for column in inspect(engine).get_columns("attribute_batches")}
            mapping_columns = {
                column["name"] for column in inspect(engine).get_columns("attribute_mapping_rules")
            }
            Session = sessionmaker(bind=engine, expire_on_commit=False)
            with Session() as session:
                batch = session.scalar(select(AttributeBatch))
                rule = session.scalar(select(AttributeMappingRule))

            self.assertTrue(
                {
                    "name",
                    "original_path",
                    "export_path",
                    "stored_filename",
                    "products_count",
                    "source_urls",
                }.issubset(batch_columns)
            )
            self.assertIn("normalized_donor_attribute", mapping_columns)
            self.assertNotIn("normalized_donor_name", mapping_columns)
            self.assertEqual(batch.name, "products.csv")
            self.assertTrue(batch.original_path.endswith("stored.csv"))
            self.assertTrue(batch.export_path.endswith("result.csv"))
            self.assertEqual(batch.products_count, 14)
            self.assertEqual(rule.normalized_donor_attribute, "spin speed")
        finally:
            engine.dispose()


class PriceConverterTests(unittest.TestCase):
    def test_price_cleanup_keeps_only_digits_without_float_dot_zero(self) -> None:
        from services.price_converter_service import normalize_price

        self.assertEqual(normalize_price("49 490"), "49490")
        self.assertEqual(normalize_price("49 490Р"), "49490")
        self.assertEqual(normalize_price("49590руб"), "49590")
        self.assertEqual(normalize_price(49490.0), "49490")
        self.assertEqual(normalize_price(11674.16), "1167416")

    def test_xlsx_finds_headers_on_each_matching_sheet_and_skips_service_sheets(self) -> None:
        from openpyxl import Workbook
        from services.price_converter_service import convert_price_source

        with TemporaryDirectory() as directory:
            source = Path(directory) / "source.xlsx"
            output = Path(directory) / "result.csv"
            workbook = Workbook()
            first = workbook.active
            first.title = "Основной"
            first.append(["Прайс поставщика"])
            first.append([])
            first.append(["Модель", "Цена"])
            first.append(["Винные шкафы MEYVEL (Италия)", None])
            first.append(["A-1", "49 490 руб."])
            service = workbook.create_sheet("Пиктограммы")
            service.append(["Описание", "Картинка"])
            second = workbook.create_sheet("Дополнительный")
            second.append(["Название", "Цена", "Модель"])
            second.append(["Товар", 59990.0, "B-2"])
            workbook.save(source)
            workbook.close()

            result = convert_price_source(source, output, "Модель", "Цена")
            rows = output.read_text(encoding="utf-8-sig").splitlines()

        self.assertEqual(result, {"rows_written": 2, "matched_sheets": 2, "skipped_sheets": 1})
        self.assertEqual(rows, ["_MODEL_;_PRICE_", "A-1;49490", "B-2;59990"])

    def test_optional_promo_adds_special_column_and_cleans_promo_price(self) -> None:
        from openpyxl import Workbook
        from services.price_converter_service import convert_price_source

        with TemporaryDirectory() as directory:
            source = Path(directory) / "promo.xlsx"
            output = Path(directory) / "result.csv"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Прайс поставщика"])
            sheet.append(["Модель", "Цена", "Промо цена"])
            sheet.append(["A-1", "9 990 руб.", "7 999 руб."])
            sheet.append(["B-2", 11990, "9 999 ₽"])
            sheet.append(["C-3", 12990, None])
            workbook.save(source)
            workbook.close()

            result = convert_price_source(
                source,
                output,
                "Модель",
                "Цена",
                promo_field="Промо цена",
                promo_date="2026-09-01",
                conversion_date=date(2026, 8, 5),
            )
            rows = output.read_text(encoding="utf-8-sig").splitlines()

        self.assertEqual(result["rows_written"], 3)
        self.assertEqual(
            rows,
            [
                "_MODEL_;_PRICE_;_SPECIAL_",
                "A-1;9990;1,1,7999,2026-08-05,2026-09-01",
                "B-2;11990;1,1,9999,2026-08-05,2026-09-01",
                "C-3;12990;",
            ],
        )

    def test_model_candidates_are_applied_without_removing_duplicate_rows(self) -> None:
        from openpyxl import Workbook
        from services.price_converter_service import convert_price_source

        with TemporaryDirectory() as directory:
            source = Path(directory) / "models.xlsx"
            output = Path(directory) / "result.csv"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Модель", "Цена"])
            sheet.append(["CS6T-23038\nчерное стекло", 100])
            sheet.append(["CS6T-23038 платина", 200])
            sheet.append(["OCM64BSH Компактный духовой шкаф ASKO", 300])
            sheet.append(["12345", 400])
            workbook.save(source)
            workbook.close()

            result = convert_price_source(source, output, "Модель", "Цена")
            rows = output.read_text(encoding="utf-8-sig").splitlines()

        self.assertEqual(result["rows_written"], 4)
        self.assertEqual(
            rows,
            [
                "_MODEL_;_PRICE_",
                "CS6T-23038;100",
                "CS6T-23038;200",
                "OCM64BSH;300",
                "12345;400",
            ],
        )

    def test_promo_field_and_date_must_be_configured_together(self) -> None:
        from services.price_converter_service import normalize_promo_settings

        with self.assertRaisesRegex(ValueError, "заполнены вместе"):
            normalize_promo_settings("Промо", None)
        with self.assertRaisesRegex(ValueError, "заполнены вместе"):
            normalize_promo_settings("", "2026-09-01")
        self.assertEqual(normalize_promo_settings("", ""), ("", None))

    def test_specific_sheet_requires_both_headers_on_that_sheet(self) -> None:
        from openpyxl import Workbook
        from services.price_converter_service import convert_price_source

        with TemporaryDirectory() as directory:
            source = Path(directory) / "source.xlsx"
            output = Path(directory) / "result.csv"
            workbook = Workbook()
            workbook.active.append(["Модель", "Описание"])
            second = workbook.create_sheet("Данные")
            second.append(["Модель", "Цена"])
            second.append(["B-2", 100])
            workbook.save(source)
            workbook.close()

            with self.assertRaisesRegex(ValueError, "листе №1"):
                convert_price_source(source, output, "Модель", "Цена", 1)
            result = convert_price_source(source, output, "Модель", "Цена", 2)

        self.assertEqual(result["rows_written"], 1)

    def test_date_header_can_be_selected_by_displayed_date(self) -> None:
        from openpyxl import Workbook
        from services.price_converter_service import convert_price_source

        with TemporaryDirectory() as directory:
            source = Path(directory) / "dated.xlsx"
            output = Path(directory) / "result.csv"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Model", datetime(2026, 8, 3)])
            sheet.append(["DW-1", 74990])
            workbook.save(source)
            workbook.close()

            result = convert_price_source(source, output, "Model", "03.08.2026")

        self.assertEqual(result["rows_written"], 1)

    def test_csv_supports_leading_rows_and_rejects_nonexistent_sheet(self) -> None:
        from services.price_converter_service import convert_price_source

        with TemporaryDirectory() as directory:
            source = Path(directory) / "source.csv"
            output = Path(directory) / "result.csv"
            source.write_text(
                "Прайс поставщика\nКод;Стоимость\nA-1;49 490 руб.\n",
                encoding="utf-8-sig",
            )

            result = convert_price_source(source, output, "Код", "Стоимость")
            with self.assertRaisesRegex(ValueError, "только лист №1"):
                convert_price_source(source, output, "Код", "Стоимость", 2)

        self.assertEqual(result["rows_written"], 1)

    def test_source_and_export_cleanup_remove_only_converter_files(self) -> None:
        import services.price_converter_service as converter

        with TemporaryDirectory() as directory:
            root = Path(directory)
            storage = root / "storage"
            exports = root / "exports"
            storage.mkdir()
            exports.mkdir()
            source = storage / "source.xlsx"
            result = exports / "result.csv"
            unrelated = exports / "unrelated.csv"
            source.write_bytes(b"source")
            result.write_bytes(b"result")
            unrelated.write_bytes(b"keep")
            row = PriceConverter(
                id=1,
                export_path=result.name,
                file={"original_filename": "source.xlsx", "stored_filename": source.name},
                state={"result_filename": result.name},
            )

            with (
                patch.object(converter, "PRICE_CONVERTER_DIR", storage),
                patch.object(converter, "EXPORT_DIR", exports),
            ):
                converter.remove_price_converter_export(row)
                converter.clear_price_converter_storage()

            self.assertFalse(source.exists())
            self.assertFalse(result.exists())
            self.assertTrue(unrelated.exists())

    def test_runtime_migration_completes_price_converter_table(self) -> None:
        from database.session import migrate_price_converter_table

        engine = create_engine("sqlite://")
        try:
            with engine.begin() as connection:
                connection.execute(text("CREATE TABLE price_converter (id INTEGER PRIMARY KEY)"))
                migrate_price_converter_table(connection)
            columns = {column["name"] for column in inspect(engine).get_columns("price_converter")}
        finally:
            engine.dispose()

        self.assertTrue(
            {
                "model_field",
                "price_field",
                "promo_field",
                "promo_date",
                "sheet_number",
                "export_path",
                "file",
                "state",
            }.issubset(columns)
        )

    def test_price_converter_model_is_present_in_fresh_schema(self) -> None:
        engine = create_engine("sqlite://")
        try:
            Base.metadata.create_all(engine)
            with sessionmaker(bind=engine)() as session:
                session.add(PriceConverter(id=1))
                session.commit()
                self.assertIsNotNone(session.get(PriceConverter, 1))
        finally:
            engine.dispose()


class CompactPayloadTests(IsolatedDatabaseTestCase):
    def test_file_import_progress_skips_large_form_settings(self) -> None:
        import services.file_import_service as file_import_service

        with self.Session.begin() as session:
            session.add(
                FileImport(
                    id=1,
                    exclusions=["value"] * 100,
                    model_field="model",
                    price_field="price",
                    replace_rules="rule",
                    file={},
                    state={"status": "idle", "percent": 0},
                )
            )

        with (
            self.Session() as session,
            patch.object(
                file_import_service,
                "normalize_file_import_exclusions",
                side_effect=AssertionError("Compact polling touched form settings"),
            ),
        ):
            payload = file_import_service.public_file_import_progress(session)

        self.assertEqual(payload["state"]["status"], "idle")
        self.assertNotIn("exclusions", payload)

    def test_price_converter_runtime_skips_form_settings(self) -> None:
        import services.price_converter_service as converter

        with self.Session.begin() as session:
            session.add(
                PriceConverter(
                    id=1,
                    model_field="model",
                    price_field="price",
                    promo_field="promo",
                    promo_date=date(2026, 9, 1),
                    sheet_number=2,
                    file={},
                    state={"status": "idle"},
                )
            )

        with (
            self.Session() as session,
            patch.object(
                converter,
                "price_converter_settings",
                side_effect=AssertionError("Compact polling touched form settings"),
            ),
        ):
            payload = converter.public_price_converter_runtime(session)

        self.assertEqual(payload["state"]["status"], "idle")
        self.assertNotIn("model_field", payload)
        self.assertNotIn("promo_field", payload)


class LogStorageTests(IsolatedDatabaseTestCase):
    def test_sqlite_logs_support_initial_page_and_delta(self) -> None:
        import services.log_service as log_service

        @contextmanager
        def isolated_session_scope():
            session = self.Session()
            try:
                yield session
                session.commit()
            except Exception:
                session.rollback()
                raise
            finally:
                session.close()

        with patch.object(log_service, "session_scope", isolated_session_scope):
            first_id = log_service.append_log(
                {
                    "time": "2026-08-04T05:30:00+00:00",
                    "level": "info",
                    "message": "Первое",
                }
            )
            initial = log_service.query_logs(limit=10)
            second_id = log_service.append_log({"level": "error", "message": "Второе"})
            delta = log_service.query_logs(after_id=first_id, limit=10)
            log_service.clear_logs()
            after_clear = log_service.query_logs(after_id=second_id, limit=10)

        self.assertEqual(initial["logs_total"], 1)
        self.assertEqual(initial["logs"][0]["message"], "Первое")
        self.assertEqual(initial["logs"][0]["time"], "2026-08-04T08:30:00+03:00")
        self.assertEqual([item["id"] for item in delta["logs"]], [second_id])
        self.assertEqual(delta["logs_counts"], {"error": 1, "info": 1})
        self.assertFalse(after_clear["delta"])
        self.assertEqual(after_clear["logs"], [])


class SchedulerQueryTests(IsolatedDatabaseTestCase):
    def test_scheduler_selects_only_due_enabled_brands(self) -> None:
        import runtime.news_tasks as news_tasks

        now = datetime.now(MSK_TZ).replace(second=15, microsecond=0)
        with self.Session.begin() as session:
            brands = [
                    Brand(
                        name="Due",
                        search_name="due",
                        group_name="Маржа",
                        state={"status": "idle"},
                        enabled=True,
                        schedule_type="daily",
                        scan_time=now.strftime("%H:%M"),
                    ),
                    Brand(
                        name="Disabled",
                        search_name="disabled",
                        group_name="Маржа",
                        state={"status": "idle"},
                        enabled=False,
                        schedule_type="daily",
                        scan_time=now.strftime("%H:%M"),
                    ),
                    Brand(
                        name="Future",
                        search_name="future",
                        group_name="Маржа",
                        state={"status": "idle"},
                        enabled=True,
                        schedule_type="once",
                        next_run_at=(now + timedelta(hours=1)).replace(tzinfo=None),
                    ),
                    Brand(
                        name="Orphan",
                        search_name="orphan",
                        group_name="Маржа",
                        state={"status": "idle"},
                        enabled=True,
                        schedule_type="daily",
                        scan_time=now.strftime("%H:%M"),
                    ),
                ]
            session.add_all(brands)
            session.flush()
            session.add_all(
                Donor(
                    legacy_id=f"scheduler-{brand.id}",
                    brand_id=brand.id,
                    site_url=f"https://example.test/{brand.id}",
                    start_urls=[],
                )
                for brand in brands[:3]
            )

        @contextmanager
        def isolated_session_scope():
            session = self.Session()
            try:
                yield session
                session.commit()
            finally:
                session.close()

        with patch.object(news_tasks, "session_scope", isolated_session_scope):
            candidates = news_tasks.scheduled_brand_candidates(now)

        self.assertEqual([brand.name for brand in candidates], ["Due"])


class RuntimeSafetyTests(unittest.TestCase):
    def test_independent_news_scans_start_without_global_limit(self) -> None:
        import runtime.news_tasks as news_tasks
        import services.news as news_service

        started = {"first": threading.Event(), "second": threading.Event()}
        release = threading.Event()
        resume_flags = []

        def run_scan(monitor_id: str, _manual: bool, resume: bool = False) -> None:
            resume_flags.append((monitor_id, resume))
            started[monitor_id].set()
            release.wait(2)

        def get_monitor(monitor_id: str):
            return {"id": monitor_id, "state": {"status": "queued"}}

        with (
            patch.object(news_service, "get_news_monitor", side_effect=get_monitor),
            patch.object(news_tasks, "scan_news_monitor", side_effect=run_scan),
        ):
            self.assertTrue(news_tasks.start_news_scan("first", manual=True))
            self.assertTrue(news_tasks.start_news_scan("second", manual=True))
            self.assertTrue(started["first"].wait(1))
            self.assertTrue(started["second"].wait(1))
            self.assertFalse(news_tasks.start_news_scan("first", manual=True))
            with news_lock:
                threads = list(news_tasks.news_scan_threads.values())
            release.set()
            for thread in threads:
                thread.join(2)
            started["first"].clear()
            self.assertTrue(news_tasks.start_news_scan("first", manual=True, resume=True))
            self.assertTrue(started["first"].wait(1))
            with news_lock:
                resumed_thread = news_tasks.news_scan_threads.get("first")
            if resumed_thread:
                resumed_thread.join(2)

        self.assertIn(("first", True), resume_flags)

    def test_public_smtp_configuration_never_contains_password(self) -> None:
        from services.news import public_news_configuration

        with news_lock:
            original = deepcopy(news_settings)
            news_settings.clear()
            news_settings.update(
                {
                    "monitors": [],
                    "own_sites": [],
                    "smtp": {
                        "host": "smtp.example.test",
                        "password": "server-secret",
                        "username": "sender@example.test",
                    },
                }
            )
        try:
            payload = public_news_configuration()
        finally:
            with news_lock:
                news_settings.clear()
                news_settings.update(original)

        self.assertNotIn("password", payload["smtp"])
        self.assertTrue(payload["smtp"]["password_set"])

    def test_public_own_sites_keep_stable_database_ids(self) -> None:
        from services.news import own_sites_from_settings

        sites = own_sites_from_settings(
            {
                "own_sites": [
                    {
                        "id": 17,
                        "name": "Основной сайт",
                        "feed_url": "https://example.test/feed.xml",
                        "feed_generate_url": "",
                    }
                ]
            }
        )

        self.assertEqual(sites[0]["id"], 17)

    def test_explicit_empty_own_sites_do_not_restore_the_default_feed(self) -> None:
        from services.news import own_sites_from_settings

        self.assertEqual(own_sites_from_settings({"own_sites": []}), [])

    def test_suspicious_xlsx_compression_is_rejected(self) -> None:
        archive = io.BytesIO()
        with zipfile.ZipFile(archive, "w", compression=zipfile.ZIP_DEFLATED) as output:
            output.writestr("xl/sharedStrings.xml", b"0" * (1024 * 1024))
        archive.seek(0)

        with self.assertRaisesRegex(ValueError, "коэффициент сжатия"):
            validate_xlsx_archive(archive)


if __name__ == "__main__":
    unittest.main()
