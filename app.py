import json
import os
import csv
import hashlib
import html as html_lib
import io
import faulthandler
import re
import shutil
import smtplib
import ssl
import base64
import subprocess
import sys
import threading
import time
import traceback
import uuid
from collections import deque
from concurrent.futures import FIRST_COMPLETED, ThreadPoolExecutor, as_completed, wait
from dataclasses import dataclass
from datetime import datetime, time as datetime_time, timedelta, timezone
from email.message import EmailMessage
from fnmatch import fnmatch
from functools import lru_cache
from pathlib import Path
from queue import Empty, Queue
from typing import Any, Callable, Deque, Dict, Hashable, Iterable, List, Optional, Set
from urllib.parse import parse_qsl, urlencode, urldefrag, urljoin, urlparse, urlunparse
import xml.etree.ElementTree as ET

import requests
from bs4 import BeautifulSoup
from flask import Flask, Response, g, jsonify as flask_jsonify, redirect, render_template, request, send_file, session, url_for
from sqlalchemy import delete, select
from sqlalchemy.exc import OperationalError
from werkzeug.security import check_password_hash, generate_password_hash

from db import SessionLocal, init_db, session_scope
from models import AppSetting, Brand, ConnectionMethod, Donor, FileImport, OwnSite, Project, User

BASE_DIR = Path(__file__).resolve().parent


def load_local_env() -> None:
    env_path = BASE_DIR / ".env"
    if not env_path.exists():
        return
    try:
        for raw_line in env_path.read_text(encoding="utf-8").splitlines():
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key:
                os.environ[key] = value
    except OSError:
        return


load_local_env()
os.environ["PYTHON_DOTENV_DISABLED"] = "1"


def env_str(name: str, default: str = "") -> str:
    return str(os.environ.get(name, default) or default).strip()


def env_int(name: str, default: int, minimum: Optional[int] = None, maximum: Optional[int] = None) -> int:
    try:
        value = int(os.environ.get(name, str(default)) or default)
    except (TypeError, ValueError):
        value = default
    if minimum is not None:
        value = max(minimum, value)
    if maximum is not None:
        value = min(maximum, value)
    return value


def env_float(name: str, default: float, minimum: Optional[float] = None, maximum: Optional[float] = None) -> float:
    try:
        value = float(os.environ.get(name, str(default)) or default)
    except (TypeError, ValueError):
        value = default
    if minimum is not None:
        value = max(minimum, value)
    if maximum is not None:
        value = min(maximum, value)
    return value


def env_list(name: str, default: Iterable[str]) -> List[str]:
    raw_value = os.environ.get(name)
    if raw_value is None:
        return [str(item).strip() for item in default if str(item).strip()]
    return [item.strip() for item in raw_value.replace("\n", ",").split(",") if item.strip()]


def env_path(name: str, default: str) -> Path:
    value = env_str(name, default)
    path = Path(value)
    return path if path.is_absolute() else BASE_DIR / path


@lru_cache(maxsize=2)
def botasaurus_browser_executable(prefer_headless_shell: bool = True) -> Optional[str]:
    """Return a Playwright browser executable, preferring headless shell for headless work."""
    env_executable = os.environ.get("PLAYWRIGHT_BROWSER_EXECUTABLE", "").strip()
    if env_executable and Path(env_executable).is_file():
        return env_executable

    try:
        from playwright.sync_api import sync_playwright

        playwright = sync_playwright().start()
        try:
            executable = Path(playwright.chromium.executable_path)
        finally:
            playwright.stop()
        if prefer_headless_shell:
            browser_root = executable
            for parent in executable.parents:
                if parent.name.startswith("chromium-"):
                    browser_root = parent.parent
                    break
            shell_names = ("headless_shell.exe", "headless_shell")
            for shell_dir in sorted(browser_root.glob("chromium_headless_shell-*"), reverse=True):
                for shell_name in shell_names:
                    matches = list(shell_dir.rglob(shell_name))
                    if matches:
                        return str(matches[0])
        return str(executable) if executable.is_file() else None
    except Exception:
        return None


LOG_DIR = env_path("LOG_DIR", "logs")
FEED_DIR = env_path("FEED_DIR", "feeds")
LOGS_FILE = env_path("LOGS_FILE", str(LOG_DIR / "logs.json"))
UNIFIED_LOG_FILE = env_path("UNIFIED_LOG_FILE", str(LOG_DIR / "app.log"))
EXPORT_DIR = env_path("EXPORT_DIR", "exports")
FILE_IMPORT_DIR = env_path("FILE_IMPORT_DIR", "storage/file-import")
PROJECT_PROFILE_DIR = BASE_DIR / "profiles" / "projects"
DEFAULT_START_URL = env_str("DEFAULT_START_URL", "")
DEFAULT_FEED_URL = env_str("DEFAULT_FEED_URL", "https://mega-kuhnya.ru/price/last_modified.xml")
DEFAULT_FEED_GENERATE_URL = env_str(
    "DEFAULT_FEED_GENERATE_URL",
    "https://mega-kuhnya.ru/index.php?route=extension/feed/unixml/new_product",
)
MSK_TZ = timezone(timedelta(hours=env_int("APP_TIMEZONE_OFFSET_HOURS", 3)))
DEFAULT_EXCLUSIONS = env_list(
    "DEFAULT_EXCLUSIONS",
    [
        "/catalog/rasprodazha/",
        "/catalog/utsenka/",
        "/about/",
        "/contacts/",
    ],
)

REQUEST_TIMEOUT = env_int("REQUEST_TIMEOUT", 20, minimum=1)
CONNECTION_METHOD_TIMEOUT_SECONDS = env_int("CONNECTION_METHOD_TIMEOUT_SECONDS", 60, minimum=1)
REQUEST_DELAY_SECONDS = env_float("REQUEST_DELAY_SECONDS", 0.05, minimum=0.0)
MAX_RETRIES = env_int("MAX_RETRIES", 3, minimum=1)
FEED_WORKER_COUNT = env_int("FEED_WORKER_COUNT", 6, minimum=1, maximum=12)
FEED_SNAPSHOT_CACHE_SECONDS = env_int("FEED_SNAPSHOT_CACHE_SECONDS", 120, minimum=0, maximum=3600)
FEED_SNAPSHOT_RETAIN = env_int("FEED_SNAPSHOT_RETAIN", 3, minimum=2, maximum=10)
NEWS_SCAN_STALL_TIMEOUT = env_int("NEWS_SCAN_STALL_TIMEOUT", 180, minimum=1)
SCHEDULE_DUE_GRACE_SECONDS = env_int("SCHEDULE_DUE_GRACE_SECONDS", 90, minimum=0)
CONNECTION_METHOD_CACHE_SECONDS = env_int("CONNECTION_METHOD_CACHE_SECONDS", 30, minimum=1)
PRICE_RE = re.compile(r"\d[\d\s\u2009\xa0]{1,}(?:\u20bd|\u0440\u0443\u0431\.?)", re.IGNORECASE)
BLOCKED_PAGE_MARKERS = tuple(
    env_list(
        "BLOCKED_PAGE_MARKERS",
        [
            "cloudflare",
            "captcha",
            "access denied",
            "http 403",
            "__qrator",
            "qauth.js",
            "qrator",
            "доступ ограничен",
            "проверяем ваш браузер",
            "enable javascript",
        ],
    )
)

app = Flask(__name__)
app.secret_key = env_str("FLASK_SECRET_KEY", "change-this-secret-key")
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
)


@app.errorhandler(Exception)
def log_unhandled_exception(error: Exception):
    LOG_DIR.mkdir(exist_ok=True)
    with (LOG_DIR / "flask-error.log").open("a", encoding="utf-8") as error_file:
        error_file.write(f"\n[{datetime.now().isoformat(timespec='seconds')}] {request.method} {request.path}\n")
        error_file.write("".join(traceback.format_exception(type(error), error, error.__traceback__)))
    raise error


@app.before_request
def open_request_db_session() -> None:
    g.db = SessionLocal()


def ensure_default_user() -> None:
    with session_scope() as db_session:
        existing_user = db_session.scalar(select(User.id).limit(1))
        if existing_user:
            return
        username = env_str("AUTH_DEFAULT_USERNAME", "admin")
        password = env_str("AUTH_DEFAULT_PASSWORD", "admin")
        db_session.add(
            User(
                username=username,
                password_hash=generate_password_hash(password),
                is_active=True,
            )
        )


def is_public_endpoint() -> bool:
    endpoint = request.endpoint or ""
    return endpoint in {"healthcheck", "login", "static"} or request.path.startswith("/static/")


@app.before_request
def require_login() -> Optional[Response]:
    if is_public_endpoint():
        return None
    if session.get("user_id"):
        return None
    return redirect(url_for("login"))


@app.teardown_request
def close_request_db_session(error: Optional[BaseException] = None) -> None:
    db = g.pop("db", None)
    if db is None:
        return
    if error is None:
        db.commit()
    else:
        db.rollback()
    db.close()

state_lock = threading.RLock()
projects_lock = threading.RLock()
news_lock = threading.RLock()
file_import_lock = threading.RLock()
active_stop_event = threading.Event()
active_finish_event = threading.Event()
file_import_stop_event = threading.Event()
active_run_id = 0
worker_thread: Optional[threading.Thread] = None
file_import_worker_thread: Optional[threading.Thread] = None
active_crawler = None

projects: Dict[str, Dict[str, object]] = {}
news_settings: Dict[str, object] = {}
news_scheduler_thread: Optional[threading.Thread] = None
LOG_AUTO_CLEANUP = False
VISIBLE_BROWSER_LOCK = threading.Lock()
STANDALONE_BROWSER_SEMAPHORE = threading.BoundedSemaphore(1)
FEED_STORAGE_LOCK = threading.RLock()
feed_snapshot_cache: Dict[str, object] = {"signature": (), "created_at": 0.0, "feeds": []}
UNIFIED_LOG_LOCK = threading.Lock()
connection_method_cache_lock = threading.Lock()
connection_method_cache: Dict[str, object] = {"loaded_at": 0.0, "methods": []}
news_stop_events: Dict[str, threading.Event] = {}
news_stop_modes: Dict[str, str] = {}
news_scan_threads: Dict[str, threading.Thread] = {}
news_state_persisted_at: Dict[str, float] = {}
NEWS_TRANSITION_TIMEOUT_SECONDS = 180
PROGRESS_STREAM_INTERVAL_SECONDS = env_float("PROGRESS_STREAM_INTERVAL_SECONDS", 2.0, minimum=0.5, maximum=30.0)


@dataclass(frozen=True)
class ScanTask:
    key: Hashable
    run: Callable[[], None]
    on_start: Optional[Callable[[threading.Thread], None]] = None
    on_finish: Optional[Callable[[threading.Thread], None]] = None


class ScanDispatcher:
    """FIFO launcher shared by full project and news scans."""

    def __init__(self) -> None:
        self._queue: Deque[ScanTask] = deque()
        self._queued: Set[Hashable] = set()
        self._active: Set[Hashable] = set()
        self._lock = threading.RLock()

    def enqueue(
        self,
        key: Hashable,
        run: Callable[[], None],
        *,
        on_start: Optional[Callable[[threading.Thread], None]] = None,
        on_finish: Optional[Callable[[threading.Thread], None]] = None,
    ) -> bool:
        with self._lock:
            if key in self._queued or key in self._active:
                return False
            self._queued.add(key)
            self._queue.append(ScanTask(key, run, on_start, on_finish))
        self._dispatch()
        return True

    def cancel(self, key: Hashable) -> bool:
        with self._lock:
            if key not in self._queued:
                return False
            self._queued.remove(key)
            return True

    def contains(self, key: Hashable) -> bool:
        with self._lock:
            return key in self._queued or key in self._active

    def _dispatch(self) -> None:
        while True:
            with self._lock:
                if not self._queue:
                    return
                task = self._queue.popleft()
                if task.key not in self._queued:
                    continue
                self._queued.remove(task.key)
                self._active.add(task.key)
            threading.Thread(target=self._run, args=(task,), daemon=True).start()

    def _run(self, task: ScanTask) -> None:
        thread = threading.current_thread()
        try:
            if task.on_start:
                task.on_start(thread)
            task.run()
        finally:
            try:
                if task.on_finish:
                    task.on_finish(thread)
            finally:
                with self._lock:
                    self._active.discard(task.key)
            self._dispatch()


scan_dispatcher = ScanDispatcher()
BLOCKED_BROWSER_RESOURCE_TYPES = {"image", "media", "font", "stylesheet"}
BLOCKED_BROWSER_URL_PARTS = (
    "google-analytics",
    "googletagmanager",
    "doubleclick",
    "adservice",
    "adsystem",
    "yandex.ru/metrika",
    "mc.yandex",
    "metrika",
    "analytics",
    "counter",
    "facebook.net",
    "vk.com/rtrg",
    "top-fwz1.mail.ru",
    "mail.ru/counter",
)
SESSION_BROWSER_METHODS = {
    "protected-site",
}
BOTASAURUS_HEADLESS_METHODS = {
    "botasaurus-browser",
    "botasaurus-browser-direct",
    "botasaurus-visible",
}
DEBUG_VISIBLE_METHODS = {"botasaurus-debug-visible"}
STATIC_BROWSER_RENDER_METHODS = {
    *SESSION_BROWSER_METHODS,
    *BOTASAURUS_HEADLESS_METHODS,
    *DEBUG_VISIBLE_METHODS,
    "crawl4ai",
    "playwright",
    "scrapegraphai",
}
FETCH_DEBUG_HTML = env_str("FETCH_DEBUG_HTML", "0").lower() in {"1", "true", "yes", "on"}
FETCH_DEBUG_HTML_DIR = LOG_DIR / "debug-html"

scan_state: Dict[str, object] = {
    "status": "idle",
    "percent": 0,
    "currenturl": "",
    "totalprocessed": 0,
    "processed_products": 0,
    "found_products": 0,
    "skipped": 0,
    "error": "",
    "download_ready": False,
    "download_url": "",
    "filename": "",
    "thread_count": 4,
}


def make_state(thread_count: int = 4) -> Dict[str, object]:
    return {
        "status": "idle",
        "percent": 0,
        "currenturl": "",
        "totalprocessed": 0,
        "processed_products": 0,
        "found_products": 0,
        "skipped": 0,
        "error": "",
        "download_ready": False,
        "download_url": "",
        "filename": "",
        "thread_count": thread_count,
        "started_at": "",
        "finished_at": "",
        "elapsed_seconds": 0,
        "eta_seconds": None,
        "paused_with_result": False,
    }


_storage_init_lock = threading.RLock()
_storage_initialized = False


def ensure_storage() -> None:
    """Создает рабочие файлы при первом запуске."""
    global _storage_initialized
    if _storage_initialized:
        return
    with _storage_init_lock:
        if _storage_initialized:
            return
        EXPORT_DIR.mkdir(exist_ok=True)
        LOG_DIR.mkdir(exist_ok=True)
        FEED_DIR.mkdir(exist_ok=True)
        FILE_IMPORT_DIR.mkdir(parents=True, exist_ok=True)
        PROJECT_PROFILE_DIR.mkdir(parents=True, exist_ok=True)
        init_db()
        ensure_default_user()
        recover_interrupted_file_import_scan()
        load_projects()
        load_news_settings()
        start_news_scheduler()
        _storage_initialized = True


def normalize_start_urls(value: object, allow_empty: bool = False) -> List[str]:
    if isinstance(value, str):
        raw_items = re.split(r"[\n,]+", value)
    elif isinstance(value, list):
        raw_items = [str(item) for item in value]
    else:
        raw_items = [] if allow_empty else [DEFAULT_START_URL]

    urls = []
    for item in raw_items:
        item = item.strip()
        if not item:
            continue
        normalized = normalize_url(item, item)
        if normalized and normalized not in urls:
            urls.append(normalized)
    return urls or ([] if allow_empty else [DEFAULT_START_URL])


def normalize_feed_url(raw_url: str) -> str:
    raw_url = str(raw_url or "").strip()
    if not raw_url:
        return ""
    if not raw_url.startswith(("http://", "https://")):
        raw_url = "https://" + raw_url
    absolute_url, _fragment = urldefrag(raw_url)
    parsed = urlparse(absolute_url)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        return ""
    path = re.sub(r"/{2,}", "/", parsed.path or "/")
    return urlunparse((parsed.scheme.lower(), parsed.netloc.lower(), path, "", parsed.query, ""))


def normalize_feed_urls(value: object, fallback: str) -> List[str]:
    if isinstance(value, str):
        raw_items = re.split(r"[\n,]+", value)
    elif isinstance(value, list):
        raw_items = [str(item) for item in value]
    else:
        raw_items = [fallback]

    urls = []
    for item in raw_items:
        normalized = normalize_feed_url(item)
        if normalized and normalized not in urls:
            urls.append(normalized)
    return urls or [fallback]


def normalize_patterns(value: object) -> List[str]:
    if isinstance(value, str):
        raw_items = re.split(r"[\n,]+", value)
    elif isinstance(value, list):
        raw_items = [str(item) for item in value]
    else:
        raw_items = []

    patterns = []
    for item in raw_items:
        item = item.strip()
        if item and item not in patterns:
            patterns.append(item)
    return patterns


def normalize_file_import_exclusions(value: object) -> List[str]:
    if isinstance(value, str):
        raw_items = value.splitlines()
    elif isinstance(value, list):
        raw_items = []
        for item in value:
            raw_items.extend(str(item or "").splitlines())
    else:
        raw_items = []

    exclusions = []
    for item in raw_items:
        item = str(item or "").strip()
        if item and item not in exclusions:
            exclusions.append(item)
    return exclusions


def file_import_exclusions_text(value: object) -> str:
    return "\n".join(normalize_file_import_exclusions(value))


def normalize_file_import_rules_text(value: object) -> str:
    return str(value or "").replace("\r\n", "\n").replace("\r", "\n").strip()


def normalize_emails(value: object) -> List[str]:
    if isinstance(value, str):
        raw_items = re.split(r"[\n,;]+", value)
    elif isinstance(value, list):
        raw_items = []
        for item in value:
            raw_items.extend(re.split(r"[\n,;]+", str(item)))
    else:
        raw_items = []

    emails = []
    for item in raw_items:
        item = item.strip()
        if item and "@" in item and item not in emails:
            emails.append(item)
    return emails


def normalize_selector_settings(value: object) -> Dict[str, object]:
    if not isinstance(value, dict):
        value = {}
    allowed = {"name_selector", "availability_selector"}
    settings: Dict[str, object] = {}
    for key in allowed:
        text = clean_text(str(value.get(key, "")))
        if text:
            settings[key] = text
    status_exclusions = normalize_patterns(value.get("availability_exclusions", []))
    if status_exclusions:
        settings["availability_exclusions"] = status_exclusions
    return settings


def make_project(name: str = "Проект 1", start_urls: Optional[List[str]] = None) -> Dict[str, object]:
    project_id = uuid.uuid4().hex[:10]
    return {
        "id": project_id,
        "name": name,
        "start_urls": start_urls or [DEFAULT_START_URL],
        "thread_count": 4,
        "exclusions": DEFAULT_EXCLUSIONS.copy(),
        "product_url_filters": [],
        "product_url_exclusions": [],
        "extraction_rules": {},
        "state": make_state(4),
        "logs": [],
        "auto_cleanup": False,
        "connection_method": normalize_connection_method(None),
        "auto_connection_fallback": True,
        "persist_profile": False,
        "worker_thread": None,
        "stop_event": threading.Event(),
        "finish_event": threading.Event(),
        "crawler": None,
        "run_id": 0,
    }


PROJECT_PROGRESS_FIELDS = (
    "totalprocessed",
    "processed_products",
    "found_products",
    "in_memory_products",
    "queue_size",
    "active_tasks",
    "skipped",
    "failed_pages",
)

NEWS_PROGRESS_FIELDS = (
    "processed",
    "found_products",
    "candidate_products",
    "compared_products",
    "in_memory_products",
    "queue_size",
    "active_tasks",
    "failed_pages",
    "availability_skipped",
)


def is_active_status(status: object) -> bool:
    return str(status or "") in {"running", "queued", "pausing", "stopping"}


def progress_int(value: object) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def has_positive_progress_value(state: Dict[str, object], fields: Iterable[str]) -> bool:
    return any(progress_int(state.get(field, 0)) > 0 for field in fields)


def same_progress_run(current: Dict[str, object], previous: Dict[str, object]) -> bool:
    current_started = str(current.get("started_at") or "")
    previous_started = str(previous.get("started_at") or "")
    return not current_started or not previous_started or current_started == previous_started


def is_empty_active_progress_state(state: Dict[str, object], fields: Iterable[str]) -> bool:
    return (
        is_active_status(state.get("status"))
        and not has_positive_progress_value(state, fields)
        and not str(state.get("currenturl") or "").strip()
    )


def merge_stable_progress_state(
    state: Dict[str, object],
    previous: Optional[Dict[str, object]],
    fields: Iterable[str],
) -> Dict[str, object]:
    if not previous or not same_progress_run(state, previous) or not is_active_status(state.get("status")):
        return state
    merged = dict(state)
    for field in fields:
        if progress_int(merged.get(field, 0)) == 0 and progress_int(previous.get(field, 0)) > 0:
            merged[field] = previous[field]
    if not is_empty_active_progress_state(state, fields):
        return merged
    for field in (
        "percent",
        "currenturl",
        "active_urls",
        "last_event",
        "last_warning",
        "elapsed_seconds",
        "eta_seconds",
        "stall_seconds",
    ):
        value = merged.get(field)
        if value in ("", None, 0, []) and previous.get(field):
            merged[field] = previous[field]
    return merged


def stable_project_state(project: Dict[str, object], state: Dict[str, object]) -> Dict[str, object]:
    if not is_active_status(state.get("status")):
        return state
    previous = project.get("_last_progress_state")
    return merge_stable_progress_state(
        state,
        previous if isinstance(previous, dict) else None,
        PROJECT_PROGRESS_FIELDS,
    )


def public_project(project: Dict[str, object], include_details: bool = True) -> Dict[str, object]:
    state = repair_mojibake(stable_project_state(project, dict(project["state"])))
    filename = str(state.get("filename") or "")
    if filename and (EXPORT_DIR / filename).exists():
        state["download_ready"] = True
    payload = {
        "id": project["id"],
        "name": repair_mojibake_text(project["name"]),
        "thread_count": project["thread_count"],
        "state": state,
        "connection_method": project.get("connection_method", "requests"),
    }
    if not include_details:
        return payload
    payload.update(
        {
            "start_urls": project["start_urls"],
            "exclusions": project["exclusions"],
            "product_url_filters": project.get("product_url_filters", []),
            "product_url_exclusions": project.get("product_url_exclusions", []),
            "extraction_rules": project.get("extraction_rules", {}),
            "auto_cleanup": project.get("auto_cleanup", False),
            "auto_connection_fallback": project.get("auto_connection_fallback", True),
            "persist_profile": bool(project.get("persist_profile", False)),
        }
    )
    return payload


def payload_signature(value: object) -> str:
    return hashlib.sha256(
        json.dumps(value, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")
    ).hexdigest()


def projects_progress_payload() -> List[Dict[str, object]]:
    with projects_lock:
        return [public_project(project, include_details=False) for project in projects.values()]


def news_progress_payload() -> Dict[str, object]:
    return public_news_settings(include_connection_methods=False, include_monitor_details=False)


def progress_payload(include_projects: bool, include_news: bool) -> Dict[str, object]:
    payload: Dict[str, object] = {
        "connection_methods": public_connection_methods(),
        "logs_signature": logs_signature(),
    }
    if include_projects:
        payload["projects"] = projects_progress_payload()
    if include_news:
        payload["news"] = news_progress_payload()
    return payload


def project_model_to_dict(row: Project) -> Dict[str, object]:
    thread_count = parse_thread_count(row.thread_count)
    project = {
        "id": str(row.id),
        "name": row.name,
        "start_urls": normalize_start_urls(row.start_urls or [DEFAULT_START_URL]),
        "thread_count": thread_count,
        "exclusions": normalize_patterns(row.exclusions or DEFAULT_EXCLUSIONS),
        "product_url_filters": normalize_patterns(row.product_url_filters or []),
        "product_url_exclusions": normalize_patterns(getattr(row, "product_url_exclusions", None) or []),
        "extraction_rules": normalize_extraction_rules(row.extraction_rules or {}),
        "state": {**make_state(thread_count), **(row.state or {})},
        "logs": [],
        "auto_cleanup": bool(row.auto_cleanup),
        "connection_method": normalize_connection_method(row.connection_method),
        "auto_connection_fallback": bool(row.auto_connection_fallback),
        "persist_profile": bool(getattr(row, "persist_profile", False)),
        "worker_thread": None,
        "stop_event": threading.Event(),
        "finish_event": threading.Event(),
        "crawler": None,
        "run_id": 0,
    }
    if project["state"].get("status") == "running":
        project["state"]["status"] = "error"
        project["state"]["error"] = "Сбор был прерван перезапуском сервера. Запустите его снова."
    return project


def upsert_project_model(session, project: Dict[str, object]) -> int:
    row = get_project_row(session, project.get("id"))
    if row is None:
        legacy_id = str(project.get("id") or "").strip()
        row = Project(legacy_id=legacy_id if legacy_id and parse_db_int(legacy_id) is None else "", name=str(project.get("name") or "Проект"))
        session.add(row)
    row.name = str(project.get("name") or "Проект")
    row.start_urls = normalize_start_urls(project.get("start_urls") or DEFAULT_START_URL)
    row.thread_count = parse_thread_count(project.get("thread_count", 4))
    row.exclusions = normalize_patterns(project.get("exclusions", DEFAULT_EXCLUSIONS))
    row.product_url_filters = normalize_patterns(project.get("product_url_filters", []))
    row.product_url_exclusions = normalize_patterns(project.get("product_url_exclusions", []))
    row.extraction_rules = normalize_extraction_rules(project.get("extraction_rules", {}))
    row.state = dict(project.get("state") or make_state(row.thread_count))
    row.auto_cleanup = bool(project.get("auto_cleanup", False))
    row.connection_method = normalize_connection_method(project.get("connection_method"))
    row.auto_connection_fallback = bool(project.get("auto_connection_fallback", True))
    row.persist_profile = bool(project.get("persist_profile", False))
    session.flush()
    return int(row.id)

def save_projects() -> None:
    with projects_lock:
        with session_scope() as session:
            current_ids = set()
            rekey: List[tuple[str, str]] = []
            for old_key, project in list(projects.items()):
                db_id = upsert_project_model(session, project)
                public_id = str(db_id)
                current_ids.add(db_id)
                if str(project.get("id")) != public_id:
                    project["id"] = public_id
                if old_key != public_id:
                    rekey.append((old_key, public_id))
            for old_key, new_key in rekey:
                projects[new_key] = projects.pop(old_key)
            if current_ids:
                session.execute(delete(Project).where(Project.id.not_in(current_ids)))


def write_logs_file(data: List[Dict[str, object]]) -> None:
    LOGS_FILE.parent.mkdir(exist_ok=True)
    LOGS_FILE.write_text(json.dumps(repair_mojibake(data), ensure_ascii=False, indent=2), encoding="utf-8")


def append_unified_log(item: Dict[str, object]) -> None:
    item = repair_mojibake(item)
    LOG_DIR.mkdir(exist_ok=True)
    timestamp = str(item.get("time") or datetime.now(MSK_TZ).isoformat(timespec="seconds"))
    level = str(item.get("level") or "info").upper()
    project_name = repair_mojibake_text(item.get("project_name") or item.get("project_id") or "system")
    message = repair_mojibake_text(item.get("message") or "")
    line = f"{timestamp} [{level}] {project_name}: {message}\n"
    try:
        with UNIFIED_LOG_LOCK:
            UNIFIED_LOG_FILE.open("a", encoding="utf-8").write(line)
    except OSError:
        print(line, end="", flush=True)


def fetch_debug_log(message: str, level: str = "info") -> None:
    append_unified_log(
        {
            "project_id": "fetch-debug",
            "project_name": "fetch-debug",
            "level": level,
            "message": message,
        }
    )


def html_title_for_debug(html: str) -> str:
    if not html:
        return ""
    match = re.search(r"<title[^>]*>(.*?)</title>", html, flags=re.IGNORECASE | re.DOTALL)
    if not match:
        return ""
    return clean_text(re.sub(r"<[^>]+>", " ", html_lib.unescape(match.group(1))))[:180]


def debug_html_filename(url: str, method: str) -> str:
    parsed = urlparse(url)
    host = re.sub(r"[^A-Za-z0-9_-]+", "_", (parsed.hostname or "unknown").lower()).strip("_") or "unknown"
    method_name = re.sub(r"[^A-Za-z0-9_-]+", "_", str(method or "method")).strip("_") or "method"
    digest = hashlib.sha1(url.encode("utf-8", "ignore")).hexdigest()[:10]
    return f"{host}_{method_name}_{digest}.html"


def save_fetch_debug_html(url: str, method: str, html: str) -> None:
    if not FETCH_DEBUG_HTML or not html:
        return
    try:
        FETCH_DEBUG_HTML_DIR.mkdir(parents=True, exist_ok=True)
        path = FETCH_DEBUG_HTML_DIR / debug_html_filename(url, method)
        path.write_text(html, encoding="utf-8", errors="replace")
    except OSError as error:
        fetch_debug_log(f"Не удалось сохранить debug HTML для {method} {url}: {error}", "warning")


def log_fetch_result(method: str, url: str, html: Optional[str], elapsed: float = 0.0, extra: str = "") -> None:
    html_text = html or ""
    title = html_title_for_debug(html_text)
    blocked = looks_blocked_or_empty(html_text) if html_text else True
    suffix = f"; {extra}" if extra else ""
    fetch_debug_log(
        f"method={method}; url={url}; html_len={len(html_text)}; blocked={blocked}; "
        f"elapsed={elapsed:.2f}s; title={title}{suffix}",
        "info" if html_text and not blocked else "warning",
    )
    if html_text:
        save_fetch_debug_html(url, method, html_text)


def log_fetch_exception(method: str, url: str, error: BaseException) -> None:
    fetch_debug_log(f"method={method}; url={url}; error={type(error).__name__}: {error}", "warning")


def read_logs_file() -> List[Dict[str, object]]:
    if not LOGS_FILE.exists():
        return []
    try:
        data = json.loads(LOGS_FILE.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return []
    if not isinstance(data, list):
        return []
    return [repair_mojibake(item) for item in data if isinstance(item, dict)]


def get_log_auto_cleanup() -> bool:
    with session_scope() as db_session:
        app_setting = db_session.get(AppSetting, 1)
        return bool(app_setting.auto_cleanup) if app_setting else False


def set_log_auto_cleanup(value: bool) -> bool:
    global LOG_AUTO_CLEANUP
    LOG_AUTO_CLEANUP = bool(value)
    with session_scope() as db_session:
        app_setting = db_session.get(AppSetting, 1)
        if app_setting is None:
            app_setting = AppSetting(id=1)
            db_session.add(app_setting)
        app_setting.auto_cleanup = LOG_AUTO_CLEANUP
    return LOG_AUTO_CLEANUP


UNIFIED_LOG_RE = re.compile(r"^(?P<time>\S+) \[(?P<level>[^\]]+)\] (?P<project_name>.*?): (?P<message>.*)$")
LOG_TAIL_LINES = 2000


def log_time_from_path(path: Path) -> str:
    try:
        return datetime.fromtimestamp(path.stat().st_mtime, MSK_TZ).isoformat(timespec="seconds")
    except OSError:
        return datetime.now(MSK_TZ).isoformat(timespec="seconds")


def read_tail_lines(path: Path, limit: int = LOG_TAIL_LINES) -> List[str]:
    try:
        lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
    except OSError:
        return []
    if limit > 0:
        return lines[-limit:]
    return lines


def read_unified_log_file() -> List[Dict[str, object]]:
    entries: List[Dict[str, object]] = []
    for line in read_tail_lines(UNIFIED_LOG_FILE):
        line = line.strip()
        if not line:
            continue
        match = UNIFIED_LOG_RE.match(line)
        if match:
            entries.append(
                repair_mojibake(
                    {
                        "time": match.group("time"),
                        "level": match.group("level").lower(),
                        "project_name": match.group("project_name"),
                        "message": match.group("message"),
                    }
                )
            )
            continue
    return entries


def read_plain_log_file(path: Path, project_name: str, level: str) -> List[Dict[str, object]]:
    entries: List[Dict[str, object]] = []
    timestamp = log_time_from_path(path)
    for line in read_tail_lines(path):
        line = line.strip()
        if not line:
            continue
        if "PermissionError" in line and "app.log" in line:
            continue
        entries.append(
            repair_mojibake(
                {
                    "time": timestamp,
                    "level": level,
                    "project_name": project_name,
                    "message": line,
                }
            )
        )
    return entries


def iter_server_log_files() -> Iterable[Path]:
    for directory in (LOG_DIR / "server-output", LOG_DIR / "server-error"):
        if not directory.exists():
            continue
        try:
            files = sorted(
                [path for path in directory.iterdir() if path.is_file()],
                key=lambda item: item.stat().st_mtime,
            )
        except OSError:
            continue
        yield from files


def combined_log_entries() -> List[Dict[str, object]]:
    entries: List[Dict[str, object]] = []
    entries.extend(read_logs_file())
    entries.extend(read_plain_log_file(LOG_DIR / "flask-error.log", "flask-error.log", "error"))
    for path in iter_server_log_files():
        level = "error" if path.parent.name == "server-error" else "info"
        entries.extend(read_plain_log_file(path, path.name, level))

    deduped: List[Dict[str, object]] = []
    seen: Set[tuple[str, str, str, str]] = set()
    for item in entries:
        normalized = repair_mojibake(item)
        key = (
            str(normalized.get("time") or ""),
            str(normalized.get("level") or ""),
            str(normalized.get("project_name") or normalized.get("project_id") or ""),
            str(normalized.get("message") or ""),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(normalized)
    return deduped


def is_recent_log_entry(item: Dict[str, object], cutoff: float) -> bool:
    try:
        return datetime.fromisoformat(str(item.get("time") or "")).timestamp() >= cutoff
    except (TypeError, ValueError):
        return True


def log_line_timestamp(line: str) -> Optional[float]:
    match = re.match(r"^(?:\[(?P<bracket>[^\]]+)\]|(?P<plain>\S+))", line.strip())
    if not match:
        return None
    raw_value = match.group("bracket") or match.group("plain")
    try:
        return datetime.fromisoformat(raw_value).timestamp()
    except ValueError:
        return None


def iter_runtime_log_files() -> Iterable[Path]:
    for path in (LOGS_FILE, UNIFIED_LOG_FILE, LOG_DIR / "flask-error.log"):
        if path.exists() and path.is_file():
            yield path
    yield from iter_server_log_files()


def clear_runtime_log_files() -> None:
    LOG_DIR.mkdir(exist_ok=True)
    write_logs_file([])
    for path in iter_runtime_log_files():
        if path == LOGS_FILE:
            continue
        try:
            if path.parent.name in {"server-output", "server-error"}:
                path.unlink()
            else:
                path.write_text("", encoding="utf-8")
        except OSError:
            try:
                path.write_text("", encoding="utf-8")
            except OSError:
                continue


def prune_text_log_file(path: Path, cutoff: float) -> None:
    if not path.exists() or not path.is_file():
        return
    try:
        lines = path.read_text(encoding="utf-8", errors="replace").splitlines(keepends=True)
    except OSError:
        return
    filtered_lines = []
    for line in lines:
        timestamp = log_line_timestamp(line)
        if timestamp is None or timestamp >= cutoff:
            filtered_lines.append(line)
    if len(filtered_lines) == len(lines):
        return
    try:
        path.write_text("".join(filtered_lines), encoding="utf-8")
    except OSError:
        return


def prune_old_log_files(cutoff: float) -> None:
    prune_text_log_file(UNIFIED_LOG_FILE, cutoff)
    prune_text_log_file(LOG_DIR / "flask-error.log", cutoff)
    for path in iter_server_log_files():
        try:
            if path.stat().st_mtime < cutoff:
                path.unlink()
        except OSError:
            continue


def save_logs() -> None:
    with projects_lock:
        data = []
        for project in projects.values():
            data.extend(project.get("logs", []))
    with news_lock:
        data.extend(news_settings.get("logs", []) if isinstance(news_settings.get("logs"), list) else [])
    write_logs_file(data)


def logs_signature() -> str:
    parts = []
    for path in iter_runtime_log_files():
        try:
            stat = path.stat()
        except OSError:
            continue
        try:
            relative_path = path.relative_to(BASE_DIR)
        except ValueError:
            relative_path = path
        parts.append(f"{relative_path}:{stat.st_mtime_ns}:{stat.st_size}")
    if not parts:
        return "empty"
    return hashlib.sha256("|".join(sorted(parts)).encode("utf-8")).hexdigest()


def load_logs() -> None:
    for item in read_logs_file():
        project_id = item.get("project_id")
        project = projects.get(project_id)
        if project:
            project.setdefault("logs", []).append(item)


def load_news_logs_from_file() -> List[Dict[str, object]]:
    return [
        item
        for item in read_logs_file()
        if str(item.get("project_id") or "").startswith("news")
    ]


def load_projects() -> None:
    with projects_lock:
        if projects:
            return
        with session_scope() as session:
            rows = session.scalars(select(Project).order_by(Project.created_at, Project.id)).all()

        if not rows:
            project = make_project("Проект 1", [DEFAULT_START_URL])
            projects[project["id"]] = project
            save_projects()
        else:
            for row in rows:
                projects[str(row.id)] = project_model_to_dict(row)

        if not projects:
            project = make_project("Проект 1", [DEFAULT_START_URL])
            projects[project["id"]] = project
            save_projects()
        load_logs()


def get_project(project_id: str) -> Optional[Dict[str, object]]:
    ensure_storage()
    with projects_lock:
        return projects.get(project_id)


def update_project_state(project: Dict[str, object], **kwargs: object) -> None:
    with projects_lock:
        state = dict(project.get("state", make_state(parse_thread_count(project.get("thread_count", 4)))))
        state.update(kwargs)
        if is_active_status(state.get("status")):
            previous = project.get("_last_progress_state")
            state = merge_stable_progress_state(
                state,
                previous if isinstance(previous, dict) else None,
                PROJECT_PROGRESS_FIELDS,
            )
            if has_positive_progress_value(state, PROJECT_PROGRESS_FIELDS) or str(state.get("currenturl") or "").strip():
                project["_last_progress_state"] = dict(state)
        else:
            project.pop("_last_progress_state", None)
        project["state"] = state


def reset_project_state(project: Dict[str, object], status: str = "idle") -> None:
    thread_count = parse_thread_count(project.get("thread_count", 4))
    state = make_state(thread_count)
    state["status"] = status
    project["state"] = state
    project.pop("_last_progress_state", None)


def project_worker_alive(project: Dict[str, object]) -> bool:
    worker = project.get("worker_thread")
    return isinstance(worker, threading.Thread) and worker.is_alive()


def reset_project_state_after_form_save(project: Dict[str, object]) -> None:
    if project_worker_alive(project):
        return
    status = str((project.get("state") or {}).get("status") or "idle")
    if status in {"running", "queued", "stopping"}:
        return
    project["crawler"] = None
    project["stop_mode"] = ""
    reset_project_state(project, "idle")


def add_project_log(project: Dict[str, object], message: str, level: str = "info") -> None:
    with projects_lock:
        logs = project.setdefault("logs", [])
        item = {
            "time": datetime.now().isoformat(timespec="seconds"),
            "project_id": project["id"],
            "project_name": repair_mojibake_text(project["name"]),
            "level": level,
            "message": repair_mojibake_text(message),
        }
        logs.append(item)
        append_unified_log(item)
        if project.get("auto_cleanup"):
            cutoff = time.time() - 7 * 24 * 60 * 60
            logs[:] = [
                item
                for item in logs
                if datetime.fromisoformat(item["time"]).timestamp() >= cutoff
            ]
        save_logs()


def normalize_model_key(value: str) -> str:
    return re.sub(r"\s+", " ", clean_text(value)).upper()


def repair_mojibake_text(value: object) -> object:
    if not isinstance(value, str) or not value:
        return value
    markers = (
        "\u0402",
        "\u0405",
        "\u0406",
        "\u040e",
        "\u0451",
        "\u0452",
        "\u0455",
        "\u045f",
        "\u20ac",
        "С",
        "П",
        "О",
        "М",
        "Н",
        "Д",
        "Г",
        "Ц",
        "Р",
        "С",
    )
    markers = markers + ("\u0420", "\u0421")
    text = value
    for _ in range(3):
        if not any(marker in text for marker in markers):
            break
        try:
            repaired = text.encode("cp1251").decode("utf-8")
        except UnicodeError:
            break
        if repaired == text:
            break
        text = repaired
    return text


def repair_mojibake(value: object) -> object:
    if isinstance(value, dict):
        return {key: repair_mojibake(item) for key, item in value.items()}
    if isinstance(value, list):
        return [repair_mojibake(item) for item in value]
    return repair_mojibake_text(value)


def jsonify(*args: object, **kwargs: object):
    repaired_args = tuple(repair_mojibake(item) for item in args)
    repaired_kwargs = repair_mojibake(kwargs) if kwargs else {}
    return flask_jsonify(*repaired_args, **repaired_kwargs)


def output_text(value: object) -> str:
    if value is None:
        return ""
    return str(repair_mojibake_text(str(value)) or "")


def default_smtp_settings() -> Dict[str, object]:
    return {
        "host": env_str("SMTP_HOST", "smtp.yandex.ru"),
        "port": env_int("SMTP_PORT", 465, minimum=1, maximum=65535),
        "security": env_str("SMTP_SECURITY", "ssl"),
        "username": env_str("SMTP_USERNAME", ""),
        "password": env_str("SMTP_PASSWORD", env_str("YANDEX_SMTP_PASSWORD", "")),
        "recipients": env_list("SMTP_RECIPIENTS", []),
    }


def merge_smtp_settings(base: Dict[str, object], stored: Dict[str, object]) -> Dict[str, object]:
    smtp = dict(base)
    for key in ("host", "security", "username", "password"):
        value = str(stored.get(key) or "").strip()
        if value:
            smtp[key] = value
    if stored.get("port"):
        try:
            smtp["port"] = int(stored.get("port") or smtp.get("port") or 465)
        except (TypeError, ValueError):
            pass
    recipients = normalize_emails(stored.get("recipients", []))
    if recipients:
        smtp["recipients"] = recipients
    smtp.pop("sender", None)
    return smtp


def default_news_settings() -> Dict[str, object]:
    return {
        "feed_url": DEFAULT_FEED_URL,
        "feed_generate_url": DEFAULT_FEED_GENERATE_URL,
        "feed_urls": [DEFAULT_FEED_URL],
        "feed_generate_urls": [DEFAULT_FEED_GENERATE_URL],
        "auto_cleanup": False,
        "smtp": default_smtp_settings(),
        "monitors": [],
        "logs": [],
        "feed_storage": [],
    }


def make_news_monitor(group: str, brand: str, urls: List[str], site_url: str = "") -> Dict[str, object]:
    monitor_id = uuid.uuid4().hex[:10]
    site_url = str(site_url or "").strip()
    return {
        "id": monitor_id,
        "group": group,
        "brand": brand,
        "created_at": datetime.now().isoformat(timespec="milliseconds"),
        "site_url": site_url,
        "start_urls": list(urls),
        "enabled": True,
        "schedule_type": "daily",
        "scan_time": "01:00",
        "weekday": 0,
        "next_run_at": "",
        "thread_count": 4,
        "connection_method": normalize_connection_method(None),
        "auto_connection_fallback": True,
        "exclusions": DEFAULT_EXCLUSIONS.copy(),
        "product_url_filters": [],
        "product_url_exclusions": [],
        "extraction_rules": {},
        "selector_settings": {},
        "seen_models": [],
        "known_new_products": {},
        "state": make_news_state(),
        "collapsed": True,
    }


def make_news_state(status: str = "idle") -> Dict[str, object]:
    return {
        "status": status,
        "stage": "",
        "percent": 0,
        "currenturl": "",
        "processed": 0,
        "found_products": 0,
        "candidate_products": 0,
        "compared_products": 0,
        "queue_size": 0,
        "active_tasks": 0,
        "active_urls": [],
        "in_memory_products": 0,
        "availability_skipped": 0,
        "failed_pages": 0,
        "stall_seconds": 0,
        "last_event": "",
        "last_warning": "",
        "new_count": 0,
        "missing_by_feed": [],
        "skipped": 0,
        "last_scan_at": "",
        "last_csv": "",
        "error": "",
        "started_at": "",
        "finished_at": "",
        "elapsed_seconds": 0,
        "next_run_at": "",
    }


def normalize_news_monitor(item: Dict[str, object]) -> Dict[str, object]:
    start_urls = normalize_start_urls(item.get("start_urls") or "", allow_empty=True)
    monitor = make_news_monitor(
        clean_text(str(item.get("group") or "Маржа")),
        clean_text(str(item.get("brand") or "Донор")),
        start_urls,
        str(item.get("site_url") or ""),
    )
    monitor["id"] = str(item.get("id") or monitor["id"])
    monitor["created_at"] = str(item.get("created_at") or monitor["created_at"])
    monitor["site_url"] = str(item.get("site_url") or "")
    monitor["enabled"] = bool(item.get("enabled", True))
    monitor["schedule_type"] = str(item.get("schedule_type") or "daily")
    monitor["scan_time"] = str(item.get("scan_time") or "01:00")[:5]
    monitor["weekday"] = max(0, min(int(item.get("weekday", 0) or 0), 6))
    monitor["next_run_at"] = str(item.get("next_run_at") or "")
    monitor["thread_count"] = parse_thread_count(item.get("thread_count", 4))
    monitor["connection_method"] = normalize_connection_method(item.get("connection_method"))
    monitor["auto_connection_fallback"] = bool(item.get("auto_connection_fallback", True))
    monitor["exclusions"] = normalize_patterns(item.get("exclusions", DEFAULT_EXCLUSIONS))
    monitor["product_url_filters"] = normalize_patterns(item.get("product_url_filters", []))
    monitor["product_url_exclusions"] = normalize_patterns(item.get("product_url_exclusions", []))
    monitor["extraction_rules"] = normalize_extraction_rules(item.get("extraction_rules", {}))
    monitor["selector_settings"] = normalize_selector_settings(item.get("selector_settings", {}))
    monitor["seen_models"] = [normalize_model_key(str(value)) for value in item.get("seen_models", []) if str(value).strip()]
    known = item.get("known_new_products", {})
    monitor["known_new_products"] = known if isinstance(known, dict) else {}
    state = item.get("state", {})
    monitor["state"] = {**make_news_state(), **state} if isinstance(state, dict) else make_news_state()
    monitor["state"].pop("last_feeds", None)
    monitor["brand_state"] = dict(monitor["state"])
    monitor["collapsed"] = bool(item.get("collapsed", True))
    return monitor


def split_news_monitor_by_site(item: Dict[str, object]) -> List[Dict[str, object]]:
    urls = normalize_start_urls(item.get("start_urls") or "", allow_empty=True)
    monitors = []
    if not urls:
        return [normalize_news_monitor({**item, "site_url": "", "start_urls": []})]
    for index, url in enumerate(urls):
        copy_item = dict(item)
        copy_item["id"] = str(item.get("id") or uuid.uuid4().hex[:10]) if index == 0 else uuid.uuid4().hex[:10]
        copy_item["start_urls"] = [url]
        monitors.append(normalize_news_monitor(copy_item))
    return monitors


def unique_news_brand_name(group: str, base_name: str = "Новый бренд") -> str:
    base_name = clean_text(base_name) or "Новый бренд"
    group_name = clean_text(group)
    names = {
        clean_text(str(item.get("brand") or ""))
        for item in news_settings.get("monitors", [])
        if isinstance(item, dict) and clean_text(str(item.get("group") or "")) == group_name
    }
    if base_name not in names:
        return base_name
    index = 2
    while True:
        candidate = f"{base_name} {index}"
        if candidate not in names:
            return candidate
        index += 1


def donor_connection_code(row: Donor) -> str:
    method_row = getattr(row, "connection_method_row", None)
    return normalize_connection_method(getattr(method_row, "code", None))


def donor_model_to_monitor(row: Donor) -> Dict[str, object]:
    brand = row.brand
    brand_state = repair_mojibake({**make_news_state(), **(brand.state or {})}) if brand else make_news_state()
    site_url = str(row.site_url or "").strip()
    start_urls = normalize_start_urls(getattr(row, "start_urls", None) or "", allow_empty=True)
    monitor = {
        "id": str(row.id),
        "group": brand.group_name if brand else "",
        "brand": brand.name if brand else "Донор",
        "brand_id": brand.id if brand else None,
        "brand_created_at": brand.created_at.isoformat(timespec="milliseconds") if brand and brand.created_at else "",
        "primary_donor_id": brand.primary_donor_id if brand else None,
        "brand_state": brand_state,
        "created_at": row.created_at.isoformat(timespec="milliseconds") if row.created_at else "",
        "site_url": site_url,
        "start_urls": start_urls,
        "enabled": bool(brand.enabled) if brand else True,
        "schedule_type": brand.schedule_type if brand else "daily",
        "scan_time": brand.scan_time if brand else "01:00",
        "weekday": max(0, min(int((brand.weekday if brand else 0) or 0), 6)),
        "next_run_at": datetime_to_input_value(brand.next_run_at if brand else None),
        "thread_count": parse_thread_count(row.thread_count),
        "connection_method": donor_connection_code(row),
        "connection_id": row.connection_id,
        "auto_connection_fallback": bool(row.auto_connection_fallback),
        "exclusions": normalize_patterns(row.exclusions or DEFAULT_EXCLUSIONS),
        "product_url_filters": normalize_patterns(row.product_url_filters or []),
        "product_url_exclusions": normalize_patterns(getattr(row, "product_url_exclusions", None) or []),
        "extraction_rules": normalize_extraction_rules(row.extraction_rules or {}),
        "selector_settings": normalize_selector_settings(row.selector_settings or {}),
        "seen_models": [normalize_model_key(str(value)) for value in (row.seen_models or []) if str(value).strip()],
        "known_new_products": row.known_new_products or {},
        "state": dict(brand_state),
    }
    return monitor


def recover_interrupted_news_scans(monitors: Iterable[Dict[str, object]]) -> None:
    """Mark only scans inherited from a previous process as interrupted.

    This is called once during application startup, never during scheduler reloads.
    """
    for monitor in monitors:
        state = monitor.get("state")
        if not isinstance(state, dict) or state.get("status") not in {"running", "queued", "pausing", "stopping"}:
            continue
        state.update(
            {
                "status": "error",
                "stage": "Прервано",
                "error": "Сканирование было прервано перезапуском сервера. Запустите его снова.",
                "currenturl": "",
                "queue_size": 0,
                "active_tasks": 0,
                "active_urls": [],
                "in_memory_products": 0,
            }
        )
        state["stage"] = "\u041f\u0440\u0435\u0440\u0432\u0430\u043d\u043e"
        state["error"] = (
            "\u0421\u043a\u0430\u043d\u0438\u0440\u043e\u0432\u0430\u043d\u0438\u0435 \u0431\u044b\u043b\u043e \u043f\u0440\u0435\u0440\u0432\u0430\u043d\u043e "
            "\u043f\u0435\u0440\u0435\u0437\u0430\u043f\u0443\u0441\u043a\u043e\u043c \u0441\u0435\u0440\u0432\u0435\u0440\u0430. \u0417\u0430\u043f\u0443\u0441\u0442\u0438\u0442\u0435 \u0435\u0433\u043e \u0441\u043d\u043e\u0432\u0430."
        )
        monitor["brand_state"] = dict(state)


def get_or_create_brand(session, monitor: Dict[str, object]) -> Brand:
    name = clean_text(str(monitor.get("brand") or "Донор"))
    group_name = clean_text(str(monitor.get("group") or "Маржа"))
    row = session.scalar(select(Brand).where(Brand.name == name, Brand.group_name == group_name))
    if row is None:
        row = Brand(
            name=name,
            group_name=group_name,
            state={**make_news_state(), **(monitor.get("brand_state") or monitor.get("state") or {})},
            enabled=bool(monitor.get("enabled", True)),
            schedule_type=str(monitor.get("schedule_type") or "daily"),
            scan_time=str(monitor.get("scan_time") or "01:00")[:5],
            weekday=max(0, min(int(monitor.get("weekday", 0) or 0), 6)),
            next_run_at=parse_datetime_value(monitor.get("next_run_at")),
        )
        session.add(row)
        session.flush()
    else:
        row.group_name = group_name
        row.state = {**make_news_state(), **(monitor.get("brand_state") or monitor.get("state") or row.state or {})}
        row.enabled = bool(monitor.get("enabled", row.enabled))
        schedule_type = str(monitor.get("schedule_type") or row.schedule_type or "daily")
        row.schedule_type = schedule_type if schedule_type in {"daily", "weekly", "once"} else "daily"
        row.scan_time = str(monitor.get("scan_time") or row.scan_time or "01:00")[:5]
        row.weekday = max(0, min(int(monitor.get("weekday", row.weekday) or 0), 6))
        row.next_run_at = parse_datetime_value(monitor.get("next_run_at"))
    return row


def upsert_donor_model(session, monitor: Dict[str, object]) -> int:
    normalized = normalize_news_monitor(monitor)
    brand = get_or_create_brand(session, normalized)
    row = get_donor_row(session, normalized.get("id"))
    if row is None:
        legacy_id = str(normalized.get("id") or "").strip()
        row = Donor(
            legacy_id=legacy_id if legacy_id and parse_db_int(legacy_id) is None else "",
            brand_id=brand.id,
        )
        session.add(row)
    row.brand_id = brand.id
    start_urls = normalize_start_urls(normalized.get("start_urls") or "", allow_empty=True)
    site_url = str(normalized.get("site_url") or "").strip()
    row.site_url = site_url
    row.start_urls = start_urls
    row.thread_count = parse_thread_count(normalized.get("thread_count", 4))
    row.connection_id = connection_method_id_for(session, normalized.get("connection_method"))
    row.auto_connection_fallback = bool(normalized.get("auto_connection_fallback", True))
    row.exclusions = normalize_patterns(normalized.get("exclusions", DEFAULT_EXCLUSIONS))
    row.product_url_filters = normalize_patterns(normalized.get("product_url_filters", []))
    row.product_url_exclusions = normalize_patterns(normalized.get("product_url_exclusions", []))
    row.extraction_rules = normalize_extraction_rules(normalized.get("extraction_rules", {}))
    row.selector_settings = normalize_selector_settings(normalized.get("selector_settings", {}))
    row.seen_models = [normalize_model_key(str(value)) for value in normalized.get("seen_models", []) if str(value).strip()]
    row.known_new_products = normalized.get("known_new_products", {}) if isinstance(normalized.get("known_new_products"), dict) else {}
    session.flush()
    if not brand.primary_donor_id or not any(donor.id == brand.primary_donor_id for donor in brand.donors):
        brand.primary_donor_id = row.id
    session.flush()
    monitor["brand_id"] = brand.id
    monitor["brand_created_at"] = brand.created_at.isoformat(timespec="milliseconds") if brand.created_at else ""
    monitor["primary_donor_id"] = brand.primary_donor_id
    return int(row.id)


def aggregate_brand_state(monitors: List[Dict[str, object]]) -> Dict[str, object]:
    states = [{**make_news_state(), **(monitor.get("state") or {})} for monitor in monitors if isinstance(monitor, dict)]
    if not states:
        return make_news_state()
    priority = ["running", "queued", "pausing", "stopping", "error", "partial", "completed"]
    selected = next((state for status in priority for state in states if state.get("status") == status), states[0])
    result = {**make_news_state(), **selected}
    last_scan_at = max((str(state.get("last_scan_at") or state.get("finished_at") or "") for state in states), default="")
    if last_scan_at:
        result["last_scan_at"] = last_scan_at
    return result



def sync_brand_runtime_fields(source_monitor: Dict[str, object]) -> None:
    group = clean_text(str(source_monitor.get("group") or ""))
    brand = clean_text(str(source_monitor.get("brand") or ""))
    fields = ("enabled", "schedule_type", "scan_time", "weekday", "next_run_at", "state", "brand_state")
    for item in news_settings.get("monitors", []):
        if (
            isinstance(item, dict)
            and clean_text(str(item.get("group") or "")) == group
            and clean_text(str(item.get("brand") or "")) == brand
        ):
            for field in fields:
                if field in source_monitor:
                    item[field] = dict(source_monitor[field]) if isinstance(source_monitor[field], dict) else source_monitor[field]


def ensure_brand_primary_flags(monitors: List[Dict[str, object]]) -> None:
    grouped: Dict[tuple[str, str], List[Dict[str, object]]] = {}
    for item in monitors:
        if not isinstance(item, dict):
            continue
        key = (clean_text(str(item.get("group") or "")), clean_text(str(item.get("brand") or "")))
        grouped.setdefault(key, []).append(item)
    for items in grouped.values():
        primary_id = str(items[0].get("primary_donor_id") or items[0].get("id") or "")
        if not any(str(item.get("id")) == primary_id for item in items):
            primary_id = str(items[0].get("id") or "")
        for item in items:
            item["primary_donor_id"] = primary_id

def own_sites_from_settings(settings: Dict[str, object]) -> List[Dict[str, str]]:
    if isinstance(settings.get("own_sites"), list):
        sites = []
        for index, item in enumerate(settings.get("own_sites", []), start=1):
            if not isinstance(item, dict):
                continue
            feed_url = normalize_feed_url(str(item.get("feed_url") or "").strip())
            if not feed_url:
                continue
            sites.append(
                {
                    "name": clean_text(str(item.get("name") or "")) or f"Фид {index}",
                    "feed_url": feed_url,
                    "feed_generate_url": normalize_feed_url(str(item.get("feed_generate_url") or "").strip()),
                }
            )
        if sites:
            return sites
    feed_urls = normalize_feed_urls(settings.get("feed_urls") or settings.get("feed_url") or DEFAULT_FEED_URL, DEFAULT_FEED_URL)
    generate_urls = normalize_feed_urls(settings.get("feed_generate_urls") or settings.get("feed_generate_url") or DEFAULT_FEED_GENERATE_URL, DEFAULT_FEED_GENERATE_URL)
    sites = []
    for index, feed_url in enumerate(feed_urls):
        generate_url = generate_urls[index] if index < len(generate_urls) else (generate_urls[0] if generate_urls else "")
        sites.append({"name": feed_source_label(feed_url), "feed_url": feed_url, "feed_generate_url": generate_url})
    return sites


def save_news_settings() -> None:
    with news_lock:
        with session_scope() as session:
            smtp = dict(news_settings.get("smtp", {}))
            smtp.pop("sender", None)
            app_setting = session.get(AppSetting, 1)
            if app_setting is None:
                app_setting = AppSetting(id=1)
                session.add(app_setting)
            app_setting.auto_cleanup = bool(news_settings.get("auto_cleanup", False))
            app_setting.smtp = smtp
            app_setting.feed_storage = list(news_settings.get("feed_storage", [])) if isinstance(news_settings.get("feed_storage"), list) else []

            current_donor_ids = set()
            for monitor in news_settings.get("monitors", []):
                if not isinstance(monitor, dict):
                    continue
                db_id = upsert_donor_model(session, monitor)
                current_donor_ids.add(db_id)
                if str(monitor.get("id")) != str(db_id):
                    monitor["id"] = str(db_id)
            grouped_monitors: Dict[tuple[str, str], List[Dict[str, object]]] = {}
            for monitor in news_settings.get("monitors", []):
                if not isinstance(monitor, dict):
                    continue
                group_name = clean_text(str(monitor.get("group") or "Маржа"))
                brand_name = clean_text(str(monitor.get("brand") or "Донор"))
                grouped_monitors.setdefault((brand_name, group_name), []).append(monitor)
            for (brand_name, group_name), brand_monitors in grouped_monitors.items():
                brand_row = session.scalar(select(Brand).where(Brand.name == brand_name, Brand.group_name == group_name))
                if brand_row:
                    brand_row.state = aggregate_brand_state(brand_monitors)
            if current_donor_ids:
                session.execute(delete(Donor).where(Donor.id.not_in(current_donor_ids)))
            else:
                session.execute(delete(Donor))
            session.flush()
            for brand_row in session.scalars(select(Brand)).all():
                donor_ids = [donor.id for donor in brand_row.donors]
                if donor_ids and brand_row.primary_donor_id not in donor_ids:
                    brand_row.primary_donor_id = donor_ids[0]
            session.execute(delete(Brand).where(~Brand.donors.any()))

            current_feed_urls = set()
            for site in own_sites_from_settings(news_settings):
                current_feed_urls.add(site["feed_url"])
                row = session.scalar(select(OwnSite).where(OwnSite.feed_url == site["feed_url"]))
                if row is None:
                    row = OwnSite(name=site["name"], feed_url=site["feed_url"], feed_generate_url=site["feed_generate_url"])
                    session.add(row)
                else:
                    row.name = site["name"]
                    row.feed_generate_url = site["feed_generate_url"]
            if current_feed_urls:
                session.execute(delete(OwnSite).where(OwnSite.feed_url.not_in(current_feed_urls)))


def load_news_settings() -> None:
    global LOG_AUTO_CLEANUP
    with news_lock:
        if news_settings:
            return
        settings = default_news_settings()
        with session_scope() as session:
            donor_rows = session.scalars(
                select(Donor)
                .join(Brand, Donor.brand_id == Brand.id)
                .order_by(Brand.group_name, Brand.name, Donor.id)
            ).all()
            app_setting = session.get(AppSetting, 1)
            if app_setting:
                settings["auto_cleanup"] = bool(app_setting.auto_cleanup)
                LOG_AUTO_CLEANUP = bool(app_setting.auto_cleanup)
                if isinstance(app_setting.smtp, dict):
                    settings["smtp"] = merge_smtp_settings(dict(settings["smtp"]), app_setting.smtp)
                if isinstance(app_setting.feed_storage, list):
                    settings["feed_storage"] = app_setting.feed_storage
            own_sites = session.scalars(select(OwnSite).order_by(OwnSite.id)).all()
            if own_sites:
                settings["own_sites"] = [
                    {
                        "name": site.name or feed_source_label(site.feed_url),
                        "feed_url": site.feed_url,
                        "feed_generate_url": site.feed_generate_url,
                    }
                    for site in own_sites
                ]
                feed_urls = [site.feed_url for site in own_sites]
                generate_urls = [site.feed_generate_url for site in own_sites]
                settings["feed_urls"] = feed_urls
                settings["feed_generate_urls"] = generate_urls
                settings["feed_url"] = feed_urls[0]
                settings["feed_generate_url"] = generate_urls[0] if generate_urls else DEFAULT_FEED_GENERATE_URL
            settings["monitors"] = [donor_model_to_monitor(row) for row in donor_rows]
            recover_interrupted_news_scans(settings["monitors"])
            ensure_brand_primary_flags(settings["monitors"])
            settings["logs"] = load_news_logs_from_file()
        news_settings.update(settings)
        save_news_settings()


def reload_news_monitors_from_db() -> None:
    load_news_settings()
    cleanup_stale_news_transitions()
    with news_lock:
        active_by_id = {
            str(monitor.get("id")): monitor
            for monitor in news_settings.get("monitors", [])
            if isinstance(monitor, dict)
            and monitor.get("state", {}).get("status") in {"running", "queued", "pausing", "stopping"}
            and not is_stale_news_transition(monitor)
        }
    with session_scope() as session:
        donor_rows = session.scalars(
            select(Donor)
            .join(Brand, Donor.brand_id == Brand.id)
            .order_by(Brand.group_name, Brand.name, Donor.id)
        ).all()
        monitors = [donor_model_to_monitor(row) for row in donor_rows]
        ensure_brand_primary_flags(monitors)
    with news_lock:
        news_settings["monitors"] = [
            active_by_id.get(str(monitor.get("id")), monitor)
            for monitor in monitors
        ]


def add_news_log(monitor: Optional[Dict[str, object]], message: str, level: str = "info") -> None:
    with news_lock:
        logs = news_settings.setdefault("logs", [])
        logs.append(
            {
                "time": datetime.now().isoformat(timespec="seconds"),
                "project_id": f"news:{monitor.get('id')}" if monitor else "news",
                "project_name": repair_mojibake_text(f"Новинки: {monitor.get('brand')}") if monitor else "Новинки",
                "level": level,
                "message": repair_mojibake_text(message),
            }
        )
        append_unified_log(logs[-1])
        if news_settings.get("auto_cleanup"):
            cutoff = time.time() - 7 * 24 * 60 * 60
            logs[:] = [
                item
                for item in logs
                if datetime.fromisoformat(item["time"]).timestamp() >= cutoff
            ]
    save_logs()


def get_news_monitor(monitor_id: str) -> Optional[Dict[str, object]]:
    ensure_storage()
    with news_lock:
        for monitor in news_settings.get("monitors", []):
            if str(monitor.get("id")) == str(monitor_id):
                return monitor
    return None


def resolve_export_file(filename: str) -> Optional[Path]:
    if not filename:
        return None
    candidates = [filename]
    repaired = repair_mojibake_text(filename)
    if isinstance(repaired, str) and repaired and repaired not in candidates:
        candidates.append(repaired)
    for candidate in candidates:
        path = (EXPORT_DIR / candidate).resolve()
        if EXPORT_DIR.resolve() in path.parents and path.exists():
            return path
    return None


def news_csv_prefix(monitor: Dict[str, object]) -> str:
    brand = clean_text(str(monitor.get("brand") or "")).strip()
    source = safe_filename(brand or "unknown_site")
    return f"Новинки_{source}_"


def news_csv_filename(monitor: Dict[str, object], created_at: Optional[datetime] = None) -> str:
    created_at = created_at or datetime.now(MSK_TZ)
    return f"{news_csv_prefix(monitor)}{created_at.strftime('%d-%m-%Y_%H-%M-%S')}.csv"


def delete_news_csv_for_monitor(monitor: Dict[str, object], keep_filename: str = "") -> None:
    keep_filename = str(keep_filename or "").strip()
    filenames = {
        keep_filename,
        str((monitor.get("state") or {}).get("last_csv") or ""),
    }
    state = monitor.get("state", {}) if isinstance(monitor.get("state"), dict) else {}
    state_data = state.get("data", {}) if isinstance(state.get("data"), dict) else {}
    filenames.add(str(state_data.get("csv") or ""))
    prefix = news_csv_prefix(monitor)
    try:
        for path in EXPORT_DIR.glob(f"{prefix}*.csv"):
            if path.is_file() and path.name not in filenames:
                path.unlink(missing_ok=True)
    except OSError:
        pass
    for filename in filenames:
        if not filename:
            continue
        if filename == keep_filename:
            continue
        path = resolve_export_file(filename)
        if path:
            try:
                path.unlink(missing_ok=True)
            except OSError:
                pass


def project_csv_prefix(project: Optional[Dict[str, object]]) -> str:
    source = safe_filename(str((project or {}).get("name") or "project"))
    return f"{source}_"


def project_csv_filename(project: Optional[Dict[str, object]], created_at: Optional[datetime] = None) -> str:
    created_at = created_at or datetime.now()
    return f"{project_csv_prefix(project)}{created_at.strftime('%d-%m-%Y_%H-%M-%S')}.csv"


def delete_project_csv_for_project(project: Dict[str, object], keep_filename: str = "") -> None:
    keep_filename = str(keep_filename or "").strip()
    state = project.get("state", {}) if isinstance(project.get("state"), dict) else {}
    filenames = {
        keep_filename,
        str(state.get("filename") or ""),
    }
    prefix = project_csv_prefix(project)
    try:
        for path in EXPORT_DIR.glob(f"{prefix}*.csv"):
            if path.is_file() and path.name not in filenames:
                path.unlink(missing_ok=True)
    except OSError:
        pass
    for filename in filenames:
        if not filename or filename == keep_filename:
            continue
        path = resolve_export_file(filename)
        if path:
            try:
                path.unlink(missing_ok=True)
            except OSError:
                pass



def runtime_news_monitor_by_id(monitor_id: object) -> Optional[Dict[str, object]]:
    monitor_id_text = str(monitor_id or "")
    if not monitor_id_text:
        return None
    for monitor in news_settings.get("monitors", []):
        if isinstance(monitor, dict) and str(monitor.get("id")) == monitor_id_text:
            return monitor
    return None


def runtime_news_monitors_for_brand(brand_id: object) -> List[Dict[str, object]]:
    brand_id_text = str(brand_id or "")
    if not brand_id_text:
        return []
    return [
        monitor
        for monitor in news_settings.get("monitors", [])
        if isinstance(monitor, dict) and str(monitor.get("brand_id")) == brand_id_text
    ]


def public_news_brand(brand: Brand) -> Dict[str, object]:
    with news_lock:
        runtime_monitors = runtime_news_monitors_for_brand(brand.id)
        runtime_by_id = {str(monitor.get("id")): monitor for monitor in runtime_monitors}
        runtime_brand_state = aggregate_brand_state(runtime_monitors) if runtime_monitors else None

    brand_state = repair_mojibake({**make_news_state(), **(runtime_brand_state or brand.state or {})})
    primary_donor_id = brand.primary_donor_id
    if primary_donor_id and not any(donor.id == primary_donor_id for donor in brand.donors):
        primary_donor_id = brand.donors[0].id if brand.donors else None

    donors = []
    for donor in sorted(brand.donors, key=lambda item: item.id):
        runtime_monitor = runtime_by_id.get(str(donor.id)) or {}
        donor_state = repair_mojibake({**brand_state, **(runtime_monitor.get("state") or {})})
        site_url = str(donor.site_url or "").strip()
        donors.append(
            {
                "id": str(donor.id),
                "legacy_id": donor.legacy_id or "",
                "brand_id": brand.id,
                "site_url": site_url,
                "start_urls": normalize_start_urls(donor.start_urls or [], allow_empty=True),
                "thread_count": parse_thread_count(donor.thread_count),
                "connection_id": donor.connection_id,
                "connection_method": donor_connection_code(donor),
                "auto_connection_fallback": bool(donor.auto_connection_fallback),
                "exclusions": normalize_patterns(donor.exclusions or DEFAULT_EXCLUSIONS),
                "product_url_filters": normalize_patterns(donor.product_url_filters or []),
                "product_url_exclusions": normalize_patterns(getattr(donor, "product_url_exclusions", None) or []),
                "extraction_rules": normalize_extraction_rules(donor.extraction_rules or {}),
                "selector_settings": normalize_selector_settings(donor.selector_settings or {}),
                "seen_models": [normalize_model_key(str(value)) for value in (donor.seen_models or []) if str(value).strip()],
                "created_at": donor.created_at.isoformat(timespec="milliseconds") if donor.created_at else "",
                "updated_at": donor.updated_at.isoformat(timespec="milliseconds") if donor.updated_at else "",
                "state": donor_state,
                "group": brand.group_name,
                "brand": brand.name,
                "brand_created_at": brand.created_at.isoformat(timespec="milliseconds") if brand.created_at else "",
                "primary_donor_id": primary_donor_id,
                "enabled": bool(brand.enabled),
                "schedule_type": brand.schedule_type or "daily",
                "scan_time": brand.scan_time or "01:00",
                "weekday": max(0, min(int(brand.weekday or 0), 6)),
                "next_run_at": datetime_to_input_value(brand.next_run_at),
            }
        )

    return {
        "id": brand.id,
        "brand_id": brand.id,
        "name": brand.name,
        "brand": brand.name,
        "group_name": brand.group_name,
        "group": brand.group_name,
        "state": brand_state,
        "enabled": bool(brand.enabled),
        "schedule_type": brand.schedule_type or "daily",
        "scan_time": brand.scan_time or "01:00",
        "weekday": max(0, min(int(brand.weekday or 0), 6)),
        "next_run_at": datetime_to_input_value(brand.next_run_at),
        "primary_donor_id": primary_donor_id,
        "created_at": brand.created_at.isoformat(timespec="milliseconds") if brand.created_at else "",
        "updated_at": brand.updated_at.isoformat(timespec="milliseconds") if brand.updated_at else "",
        "donors": donors,
    }
def public_news_monitor(monitor: Dict[str, object], include_details: bool = True) -> Dict[str, object]:
    public_monitor = repair_mojibake(dict(monitor))
    public_monitor.pop("known_new_products", None)
    public_monitor.pop("brand_state", None)
    state = dict(public_monitor.get("state") or make_news_state())
    original_state = monitor.get("state", {}) if isinstance(monitor.get("state"), dict) else {}
    original_data = original_state.get("data", {}) if isinstance(original_state.get("data"), dict) else {}
    public_data = state.get("data", {}) if isinstance(state.get("data"), dict) else {}
    filename = str(
        original_state.get("last_csv")
        or state.get("last_csv")
        or original_data.get("csv")
        or public_data.get("csv")
        or ""
    )
    if filename and not state.get("last_csv"):
        state["last_csv"] = str(repair_mojibake_text(filename) or filename)
    state["csv_ready"] = bool(resolve_export_file(filename))
    if state.get("last_csv"):
        state["last_csv"] = str(repair_mojibake_text(state["last_csv"]) or state["last_csv"])
    public_monitor["state"] = state
    if not include_details:
        summary_keys = {
            "id",
            "brand_id",
            "primary_donor_id",
            "group",
            "brand",
            "site_url",
            "start_urls",
            "enabled",
            "state",
            "created_at",
            "brand_created_at",
        }
        public_monitor = {key: value for key, value in public_monitor.items() if key in summary_keys}
    return public_monitor


def public_news_settings(include_connection_methods: bool = True, include_monitor_details: bool = True, include_monitors: bool = True) -> Dict[str, object]:
    cleanup_stale_news_transitions()
    with news_lock:
        smtp = dict(news_settings.get("smtp", {}))
        smtp.pop("sender", None)
        smtp["password_set"] = bool(news_settings.get("smtp", {}).get("password"))
        own_sites = own_sites_from_settings(news_settings)
        feed_urls = [site["feed_url"] for site in own_sites]
        feed_generate_urls = [site["feed_generate_url"] for site in own_sites]
        payload = {
            "feed_url": feed_urls[0] if feed_urls else DEFAULT_FEED_URL,
            "feed_generate_url": feed_generate_urls[0] if feed_generate_urls else DEFAULT_FEED_GENERATE_URL,
            "feed_urls": feed_urls,
            "feed_generate_urls": feed_generate_urls,
            "own_sites": own_sites,
            "auto_cleanup": bool(news_settings.get("auto_cleanup", False)),
            "smtp": smtp,
            "feed_storage": list(news_settings.get("feed_storage", [])) if isinstance(news_settings.get("feed_storage"), list) else [],
            "monitors": [public_news_monitor(monitor, include_details=include_monitor_details) for monitor in news_settings.get("monitors", [])] if include_monitors else [],
        }
        if include_connection_methods:
            payload["connection_methods"] = public_connection_methods()
        return payload

def reset_state(status: str = "idle", run_id: Optional[int] = None, thread_count: Optional[int] = None) -> None:
    with state_lock:
        if run_id is not None and run_id != active_run_id:
            return
        current_thread_count = thread_count or int(scan_state.get("thread_count", 4) or 4)
        scan_state.update(
            {
                "status": status,
                "percent": 0,
                "currenturl": "",
                "totalprocessed": 0,
                "processed_products": 0,
                "found_products": 0,
                "skipped": 0,
                "error": "",
                "download_ready": False,
                "download_url": "",
                "filename": "",
                "thread_count": current_thread_count,
            }
        )


def update_state(run_id: Optional[int] = None, **kwargs: object) -> None:
    with state_lock:
        if run_id is not None and run_id != active_run_id:
            return
        scan_state.update(kwargs)


def snapshot_state() -> Dict[str, object]:
    with state_lock:
        return dict(scan_state)


def parse_thread_count(value: object) -> int:
    try:
        return max(1, min(int(value or 4), 16))
    except (TypeError, ValueError):
        return 4


def project_runtime_thread_count(project: Dict[str, object]) -> int:
    method = normalize_connection_method(project.get("connection_method"))
    if bool(project.get("persist_profile", False)) or is_debug_visible_method(method):
        return 1
    return parse_thread_count(project.get("thread_count", 4))


def load_connection_methods(force_refresh: bool = False) -> List[Dict[str, object]]:
    """Возвращает способы подключения из БД в порядке их id."""
    now = time.time()
    with connection_method_cache_lock:
        cached_methods = list(connection_method_cache.get("methods") or [])
        loaded_at = float(connection_method_cache.get("loaded_at") or 0.0)
        if cached_methods and not force_refresh and now - loaded_at < CONNECTION_METHOD_CACHE_SECONDS:
            return cached_methods

    methods: List[Dict[str, object]] = []
    try:
        with session_scope() as session:
            rows = session.execute(select(ConnectionMethod).order_by(ConnectionMethod.id)).scalars().all()
        seen_codes: Set[str] = set()
        for row in rows:
            code = str(row.code or "").strip()
            if not code or code in seen_codes:
                continue
            seen_codes.add(code)
            methods.append({
                "id": int(row.id),
                "code": code,
                "name": str(row.name or row.code or "").strip(),
                "is_browser_render": bool(row.is_browser_render),
                "is_debug_visible": bool(row.is_debug_visible),
            })
    except Exception as error:
        append_unified_log({
            "project_id": "system",
            "project_name": "system",
            "level": "warning",
            "message": f"Не удалось прочитать способы подключения из БД: {error}",
        })

    if not methods:
        methods = [{
            "id": 0,
            "code": "requests",
            "name": "Requests",
            "is_browser_render": False,
            "is_debug_visible": False,
        }]

    with connection_method_cache_lock:
        connection_method_cache["methods"] = list(methods)
        connection_method_cache["loaded_at"] = now
    return methods


def get_connection_method_codes(force_refresh: bool = False) -> List[str]:
    return [str(method["code"]) for method in load_connection_methods(force_refresh)]


def public_connection_methods() -> List[Dict[str, object]]:
    return [
        {
            "id": method["id"],
            "code": method["code"],
            "name": method["name"],
        }
        for method in load_connection_methods()
    ]


def connection_method_has_flag(method: str, flag_name: str) -> bool:
    for row in load_connection_methods():
        if row["code"] == method:
            return bool(row.get(flag_name))
    return False


def is_browser_render_method(method: str) -> bool:
    method = str(method or "").strip()
    return method in STATIC_BROWSER_RENDER_METHODS or connection_method_has_flag(method, "is_browser_render")


def is_debug_visible_method(method: str) -> bool:
    method = str(method or "").strip()
    return method in DEBUG_VISIBLE_METHODS or connection_method_has_flag(method, "is_debug_visible")


def ordered_db_connection_methods(
    preferred: Optional[Iterable[str]] = None,
) -> List[str]:
    """Строит fallback-цепочку только из методов, которые есть в БД."""
    db_codes = get_connection_method_codes()
    ordered: List[str] = []

    if preferred:
        for method in preferred:
            if method in db_codes and method not in ordered:
                ordered.append(method)

    for method in db_codes:
        if method not in ordered:
            ordered.append(method)
    return ordered


def normalize_connection_method(value: object) -> str:
    method = str(value or "requests").strip()
    codes = get_connection_method_codes()
    if method in codes:
        return method
    return codes[0] if codes else "requests"


def parse_db_int(value: object) -> Optional[int]:
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def parse_datetime_value(value: object) -> Optional[datetime]:
    if isinstance(value, datetime):
        return value.replace(tzinfo=None) if value.tzinfo else value
    text = str(value or "").strip()
    if not text:
        return None
    try:
        parsed = datetime.fromisoformat(text)
    except ValueError:
        return None
    return parsed.replace(tzinfo=None) if parsed.tzinfo else parsed


def datetime_to_input_value(value: object) -> str:
    if not value:
        return ""
    if isinstance(value, datetime):
        parsed = value
    else:
        try:
            parsed = datetime.fromisoformat(str(value))
        except ValueError:
            return str(value or "")
    if parsed.tzinfo is None:
        return parsed.isoformat(timespec="minutes")
    return parsed.astimezone(MSK_TZ).replace(tzinfo=None).isoformat(timespec="minutes")


def get_project_row(session, public_id: object) -> Optional[Project]:
    db_id = parse_db_int(public_id)
    if db_id is not None:
        row = session.get(Project, db_id)
        if row is not None:
            return row
    legacy_id = str(public_id or "").strip()
    if not legacy_id:
        return None
    return session.scalar(select(Project).where(Project.legacy_id == legacy_id))


def get_donor_row(session, public_id: object) -> Optional[Donor]:
    db_id = parse_db_int(public_id)
    if db_id is not None:
        row = session.get(Donor, db_id)
        if row is not None:
            return row
    legacy_id = str(public_id or "").strip()
    if not legacy_id:
        return None
    return session.scalar(select(Donor).where(Donor.legacy_id == legacy_id))


def connection_method_id_for(session, code: object) -> Optional[int]:
    method = normalize_connection_method(code)
    row = session.scalar(select(ConnectionMethod).where(ConnectionMethod.code == method))
    return row.id if row else None


def normalize_extraction_rules(value: object) -> Dict[str, str]:
    if not isinstance(value, dict):
        value = {}
    single_line_fields = {
        "product_card_selector",
        "product_url_selector",
        "model_selector",
        "price_selector",
        "model_start_marker",
        "model_end_marker",
    }
    multiline_fields = {"model_replace_rules"}
    rules = {}
    for key in single_line_fields:
        text = str(value.get(key, "")).strip()
        if text:
            rules[key] = text
    for key in multiline_fields:
        text = str(value.get(key, "")).strip()
        if text:
            rules[key] = text
    return rules


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def normalize_url(raw_url: str, base_url: str) -> Optional[str]:
    """Приводит ссылку к каноническому виду внутри сайта."""
    if not raw_url:
        return None
    raw_url = raw_url.strip()
    if raw_url.startswith(("mailto:", "tel:", "javascript:")):
        return None

    absolute_url = urljoin(base_url, raw_url)
    absolute_url, _fragment = urldefrag(absolute_url)
    parsed = urlparse(absolute_url)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        return None

    path = re.sub(r"/{2,}", "/", parsed.path or "/")

# Важно: не удаляем завершающий слэш.
# Для Bitrix/каталогов URL вида /catalog/category/ и /catalog/category
# могут обрабатываться сайтом по-разному. Например ZUGEL требует слэш.

    # Сохраняем только пагинацию. Остальные параметры обычно создают дубликаты:
    # сортировки, UTM-метки, сравнение, фильтры с теми же товарами.
    pagination_params = []
    for key, value in parse_qsl(parsed.query, keep_blank_values=False):
        key_lower = key.lower()
        if key_lower == "page" or key_lower.startswith("pagen_"):
            pagination_params.append((key, value))
    query = urlencode(pagination_params)

    result = urlunparse((parsed.scheme.lower(), parsed.netloc.lower(), path, "", query, ""))
    return result


def same_site(url: str, root_netloc: str) -> bool:
    netloc = urlparse(url).netloc.lower()
    root = root_netloc.lower()
    return netloc == root or netloc.endswith("." + root)


def is_domain_url(url: str, domain: str) -> bool:
    return urlparse(url).netloc.lower().removeprefix("www.").endswith(domain.lower().removeprefix("www."))


def has_static_extension(url: str) -> bool:
    return bool(
        re.search(
            r"\.(?:jpg|jpeg|png|gif|svg|webp|pdf|doc|docx|xls|xlsx|zip|rar|mp4|avi|css|js)$",
            urlparse(url).path,
            flags=re.IGNORECASE,
        )
    )


def is_obvious_service_path(path: str) -> bool:
    first = next((part.lower() for part in path.split("/") if part), "")
    return first in {
        "articles",
        "reviews",
        "delivery-and-payment",
        "services",
        "credit",
        "guarantee",
        "contacts",
        "favorites",
        "compare",
        "cart",
        "login",
        "personal",
        "search",
        "upload",
        "bitrix",
        "ajax",
    }


def looks_blocked_or_empty(html: str) -> bool:
    """Определяет страницы блокировки или почти пустые HTML-оболочки."""
    lowered = html.lower()
    soup = BeautifulSoup(html, "html.parser")
    text = clean_text(soup.get_text(" ", strip=True))
    links_count = len(soup.select("a[href]"))
    if PRICE_RE.search(html):
        return False
    if PRICE_RE.search(text) and links_count > 10:
        return False
    if any(marker in lowered for marker in BLOCKED_PAGE_MARKERS):
        return len(text) < 1200 or links_count < 10
    return len(text) < 250 and links_count < 5


def should_follow_url(url: str, start_url: str, root_netloc: str) -> bool:
    """Ограничивает обход страницами сайта, полезными для поиска товаров."""
    if not same_site(url, root_netloc) or has_static_extension(url):
        return False

    path = urlparse(url).path or "/"
    start_path = urlparse(start_url).path or "/"
    normalized_start_path = start_path.rstrip("/") or "/"

    if normalized_start_path not in {"/", "/catalog"}:
        return path == normalized_start_path or path.startswith(normalized_start_path + "/")

    if path == "/" or path == "/catalog" or path.startswith("/catalog/"):
        return True

    return False


def should_follow_project_url(url: str, start_urls: List[str], root_netloc: str) -> bool:
    if has_static_extension(url):
        return False

    path = urlparse(url).path or "/"
    if is_obvious_service_path(path):
        return False
    allowed_domain = False
    for start_url in start_urls:
        start_netloc = urlparse(start_url).netloc
        if not same_site(url, start_netloc or root_netloc):
            continue
        allowed_domain = True
        start_path = (urlparse(start_url).path or "/").rstrip("/") or "/"
        if start_path in {"/", "/catalog"}:
            return True
        if path == start_path or path.startswith(start_path + "/"):
            return True

    return False


def is_catalog_url(url: str) -> bool:
    path = urlparse(url).path or "/"
    return path == "/catalog" or path.startswith("/catalog/")


def exclusion_matches(url: str, pattern: str) -> bool:
    """Проверяет URL по пользовательскому шаблону исключения."""
    pattern = pattern.strip()
    if not pattern:
        return False

    parsed = urlparse(url)
    full_url = url.rstrip("/") + "/"
    path = (parsed.path or "/").rstrip("/") + "/"
    normalized_pattern = pattern.rstrip("/") + "/"

    if "*" in pattern or "?" in pattern:
        return fnmatch(full_url, pattern) or fnmatch(path, pattern)

    if pattern.startswith(("http://", "https://")):
        return full_url.startswith(normalized_pattern)

    return path.startswith(normalized_pattern) or pattern in parsed.path


def product_url_matches_filters(url: str, filters: Iterable[str]) -> bool:
    patterns = [str(pattern).strip().lower() for pattern in filters if str(pattern).strip()]
    if not patterns:
        return True

    parsed = urlparse(url)
    haystack = f"{url} {parsed.path}".lower()
    return any(pattern in haystack or fnmatch(haystack, pattern) for pattern in patterns)


def product_url_matches_any(url: str, patterns: Iterable[str]) -> bool:
    normalized_patterns = [str(pattern).strip().lower() for pattern in patterns if str(pattern).strip()]
    if not normalized_patterns:
        return False

    parsed = urlparse(url)
    haystack = f"{url} {parsed.path}".lower()
    return any(pattern in haystack or fnmatch(haystack, pattern) for pattern in normalized_patterns)


def product_url_filter_patterns(
    product_url_filters: Optional[Iterable[str]],
    rules: Optional[Dict[str, str]] = None,
) -> List[str]:
    patterns = normalize_patterns(product_url_filters or [])
    rules = rules or {}
    selector_value = str(rules.get("product_url_selector", "") or "").strip()
    if selector_value and (selector_value.startswith(("http://", "https://", "/")) or "://" in selector_value):
        if selector_value not in patterns:
            patterns.append(selector_value)
    return patterns


def canonicalize_product_url_by_filters(url: str, filters: Iterable[str]) -> str:
    if not url:
        return ""
    parsed = urlparse(url)
    path = parsed.path or "/"
    product_anchor = re.search(r"/goods?_\d+", path, flags=re.IGNORECASE)
    if product_anchor and product_anchor.start() > 0:
        path = path[product_anchor.start():]
    return urlunparse((parsed.scheme, parsed.netloc, path, "", parsed.query, ""))


def is_product_url_for_filters(url: str, filters: Iterable[str]) -> bool:
    return bool([pattern for pattern in filters if str(pattern).strip()]) and product_url_matches_filters(url, filters)


def clean_text(value: str) -> str:
    return re.sub(r"\s+", " ", value.replace("\xa0", " ").replace("\u2009", " ")).strip()


def split_text_lines(value: str) -> List[str]:
    return [clean_text(line) for line in re.split(r"[\n\r]+", value) if clean_text(line)]


MODEL_BRAND_CACHE_SECONDS = 60
model_brand_cache_lock = threading.Lock()
model_brand_cache: Dict[str, object] = {"loaded_at": 0.0, "brands": set()}


def model_brand_names(force_refresh: bool = False) -> Set[str]:
    """Возвращает бренды из БД только для fallback-очистки модели из названия."""
    now = time.time()
    with model_brand_cache_lock:
        cached = set(model_brand_cache.get("brands") or set())
        loaded_at = float(model_brand_cache.get("loaded_at") or 0.0)
        if cached and not force_refresh and now - loaded_at < MODEL_BRAND_CACHE_SECONDS:
            return cached

    brands: Set[str] = set()
    try:
        with session_scope() as session:
            rows = session.execute(select(Brand.name)).scalars().all()
        for name in rows:
            text = clean_text(str(name or "")).upper()
            if text:
                brands.add(text)
    except Exception:
        brands = set()

    with model_brand_cache_lock:
        model_brand_cache["brands"] = set(brands)
        model_brand_cache["loaded_at"] = now
    return brands


def known_brand_regex() -> str:
    return "|".join(re.escape(brand.lower()) for brand in sorted(model_brand_names(), key=len, reverse=True))

def normalize_model(value: str, product_url: str = "") -> str:
    """Возвращает маркировку модели без полного товарного названия."""
    text = clean_text(value)

    if not text:
        return ""

    # Сохраняем регистр и разделители моделей, которые уже выглядят как готовая модель.
    mixed_case_model = text.replace("\\", "/")
    mixed_case_model = re.sub(r"[–—]", "-", mixed_case_model)
    mixed_case_model = re.sub(r"\s*([/_.+])\s*", r"\1", mixed_case_model)
    mixed_case_model = re.sub(r"\s+-\s+", " - ", mixed_case_model)
    mixed_case_model = re.sub(r"\s{2,}", " ", mixed_case_model).strip()
    mixed_case_model = mixed_case_model.rstrip(".")
    text = mixed_case_model

    if (
        re.fullmatch(
            r"[A-Za-z0-9./_@-]+(?:\s+-\s+|\s+[A-Za-z0-9./_@-]+){0,8}",
            mixed_case_model,
        )
        and any(char.isdigit() for char in mixed_case_model)
        and any(char.isalpha() for char in mixed_case_model)
        and any(char.islower() for char in mixed_case_model)
    ):
        return mixed_case_model

    # Частый случай: "Бренд ABC123" -> "ABC123".
    brands_regex = known_brand_regex()
    if brands_regex:
        brand_match = re.search(
            rf"\b(?:{brands_regex})\b\s+([A-Z0-9][A-Z0-9./_@\\-]{{2,}})",
            text,
            re.IGNORECASE,
        )
        if brand_match:
            return brand_match.group(1).strip(" .,/\\_-").replace("\\", "/").upper()

    latin_model_text = text.replace("\\", "/")
    latin_model_tokens = re.findall(r"[A-Za-z0-9./_@-]+", latin_model_text)
    if (
        latin_model_tokens
        and " ".join(latin_model_tokens).strip() == re.sub(r"\s+", " ", latin_model_text).strip()
        and 1 <= len(latin_model_tokens) <= 6
        and any(any(char.isdigit() for char in token) for token in latin_model_tokens)
        and any(any(char.isalpha() for char in token) for token in latin_model_tokens)
        and latin_model_tokens[0].upper() not in {"SERIE", "SERIES"}
        and latin_model_tokens[0].upper() not in model_brand_names()
    ):
        return " ".join(token.strip(" .,/\\_-").upper() for token in latin_model_tokens if token.strip(" .,/\\_-"))

    ascii_tokens = re.findall(r"[A-Za-z0-9]+(?:[./_@-][A-Za-z0-9]+)*", text)
    for start_index in range(max(0, len(ascii_tokens) - 6), len(ascii_tokens)):
        candidate_tokens = [token.strip(" .,/\\_-") for token in ascii_tokens[start_index:] if token.strip(" .,/\\_-")]
        if not (2 <= len(candidate_tokens) <= 6):
            continue
        if not any(any(char.isdigit() for char in token) for token in candidate_tokens):
            continue
        if not all(re.fullmatch(r"[A-Z0-9./_@-]+", token) for token in candidate_tokens):
            continue
        if candidate_tokens[0].upper() in model_brand_names() or candidate_tokens[0].upper() in {"SERIE", "SERIES"}:
            continue
        return " ".join(candidate_tokens).upper()

    ignored_tokens = {*model_brand_names()}
    code_tokens = []
    for token in re.findall(r"[A-Z\u0400-\u04FF0-9][A-Z\u0400-\u04FF0-9./_@\\-]{2,}", text, flags=re.IGNORECASE):
        cleaned = token.strip(" .,/\\_-")
        if cleaned.upper() in ignored_tokens:
            continue
        has_digit = any(char.isdigit() for char in cleaned)
        has_letter = any(char.isalpha() for char in cleaned)
        if has_digit and has_letter:
            code_tokens.append(cleaned)

    if code_tokens:
        return code_tokens[-1].replace("\\", "/").upper()

    return text


def first_text(soup: BeautifulSoup, selectors: Iterable[str]) -> str:
    for selector in selectors:
        node = soup.select_one(selector)
        if node:
            text = clean_text(node.get_text(" ", strip=True))
            if text:
                return text
    return ""


def normalize_price_value(value: object) -> str:
    text = clean_text(str(value or ""))
    if not text:
        return ""
    match = PRICE_RE.search(text)
    if match:
        return clean_text(match.group(0))
    if not re.fullmatch(r"[^\w\u0400-\u04FF]*\d[\d\s\u2009\xa0.,]*[^\w\u0400-\u04FF]*", text):
        return ""
    digits = re.sub(r"[^\d]", "", text)
    if digits and len(digits) >= 2:
        return f"{int(digits):,}".replace(",", " ") + " \u20bd"
    return ""


def apply_extract_regex(value: str, pattern: str) -> str:
    text = clean_text(value)
    if not pattern:
        return text
    try:
        match = re.search(pattern, text, flags=re.IGNORECASE)
    except re.error:
        return text
    if not match:
        return text
    return clean_text(match.group(1) if match.groups() else match.group(0))


def replacement_flags(flag_text: str) -> int:
    flags = 0
    flags_text = flag_text.lower()
    if "i" in flags_text:
        flags |= re.IGNORECASE
    if "m" in flags_text:
        flags |= re.MULTILINE
    if "s" in flags_text:
        flags |= re.DOTALL
    return flags


def wildcard_rule_to_regex(pattern: str) -> str:
    result = []
    index = 0
    while index < len(pattern):
        if pattern.startswith("{skip}", index):
            result.append(".*?")
            index += len("{skip}")
        elif pattern.startswith("{.}", index):
            result.append(".")
            index += len("{.}")
        else:
            result.append(re.escape(pattern[index]))
            index += 1
    return "".join(result)


def apply_replace_rules(value: str, rules_text: str) -> str:
    text = html_lib.unescape(value or "")
    if not rules_text:
        return text

    for raw_line in rules_text.splitlines():
        line = raw_line.strip()
        if not line or "|" not in line:
            continue
        pattern, replacement = line.split("|", 1)
        pattern = pattern.strip()
        replacement = replacement.strip()

        try:
            if pattern == "{br}":
                text = re.sub(r"\r\n|\r|\n", replacement, text)
                continue

            regex_match = re.fullmatch(r"\{reg\[#(.*)#([a-zA-Z]*)\]\}", pattern)
            if regex_match:
                regex_pattern, flags_text = regex_match.groups()
                text = re.sub(regex_pattern, replacement, text, flags=replacement_flags(flags_text))
                continue

            if "{skip}" in pattern or "{.}" in pattern:
                text = re.sub(wildcard_rule_to_regex(pattern), replacement, text, flags=re.DOTALL)
                continue

            text = text.replace(pattern, replacement)
        except re.error:
            continue

    return text


def strip_html_to_text(value: str) -> str:
    raw = html_lib.unescape(value or "")
    if "<" in raw and ">" in raw:
        return BeautifulSoup(raw, "html.parser").get_text(" ", strip=True)
    return raw


def extract_between_markers(source: str, start_marker: str, end_marker: str) -> str:
    if not start_marker and not end_marker:
        return ""

    text = source or ""
    start_index = 0
    if start_marker:
        found_start = text.find(start_marker)
        if found_start < 0:
            return ""
        start_index = found_start + len(start_marker)

    end_index = len(text)
    if end_marker:
        found_end = text.find(end_marker, start_index)
        if found_end < 0:
            return ""
        end_index = found_end

    return text[start_index:end_index]


def extract_model_by_markers(source: str, rules: Dict[str, str]) -> str:
    return extract_between_markers(
        source,
        str(rules.get("model_start_marker", "")),
        str(rules.get("model_end_marker", "")),
    )


def prepare_rule_model(value: str, rules: Dict[str, str]) -> str:
    text = apply_replace_rules(value, str(rules.get("model_replace_rules", "")))
    text = strip_html_to_text(text)
    return clean_text(text)


def has_explicit_model_rules(rules: Optional[Dict[str, str]]) -> bool:
    if not rules:
        return False
    return any(
        str(rules.get(key, "")).strip()
        for key in ("model_selector", "model_start_marker", "model_end_marker", "model_replace_rules")
    )


def has_configured_model_source(rules: Optional[Dict[str, str]]) -> bool:
    if not rules:
        return False
    return any(
        str(rules.get(key, "")).strip()
        for key in ("model_selector", "model_start_marker", "model_end_marker")
    )


def has_model_replace_rules(rules: Optional[Dict[str, str]]) -> bool:
    return bool(rules and str(rules.get("model_replace_rules", "")).strip())


def finalize_scraped_model(
    value: str,
    product_url: str,
    rules: Optional[Dict[str, str]] = None,
    preserve_configured_model: bool = False,
) -> str:
    prepared = prepare_rule_model(value, rules or {})
    if not prepared:
        return ""
    if preserve_configured_model or has_model_replace_rules(rules):
        return prepared
    return normalize_model(prepared, product_url)


def first_by_selector(root, selector: str) -> str:
    if not selector or not hasattr(root, "select"):
        return ""
    try:
        node = root.select_one(selector)
    except Exception:
        return ""
    if not node:
        return ""
    return clean_text(node.get("content") or node.get("value") or node.get_text(" ", strip=True))


def extract_prices(value: str) -> List[str]:
    return [clean_text(match.group(0)) for match in PRICE_RE.finditer(value or "")]


def extract_price_from_container(container, selector: str = "") -> str:
    if not container or not hasattr(container, "select"):
        return ""
    if not selector:
        return ""
    selected = first_by_selector(container, selector)
    selected_price = (extract_prices(selected) or [normalize_price_value(selected)])[0]
    return selected_price or ""


def extract_listing_products_by_rules(
    soup: BeautifulSoup,
    current_url: str,
    rules: Dict[str, str],
    seen_urls: Set[str],
    product_url_filters: Optional[Iterable[str]] = None,
) -> List[Dict[str, str]]:
    card_selector = rules.get("product_card_selector", "")
    if not card_selector:
        return []
    filters = list(product_url_filters or [])
    products: List[Dict[str, str]] = []
    try:
        cards = soup.select(card_selector)
    except Exception:
        return []
    url_selector = rules.get("product_url_selector", "")
    for card in cards:
        link_node = None
        if url_selector and not product_url_filter_patterns([], {"product_url_selector": url_selector}):
            try:
                link_node = card.select_one(url_selector)
            except Exception:
                link_node = None
        product_url = canonicalize_product_url_by_filters(normalize_url(link_node.get("href", "") if link_node else "", current_url), filters)
        if not product_url or product_url in seen_urls or not product_url_matches_filters(product_url, filters):
            continue

        model = extract_model_by_markers(str(card), rules)
        model_from_config = bool(model)
        if not model:
            model = first_by_selector(card, rules.get("model_selector", ""))
            model_from_config = bool(model)
        if has_configured_model_source(rules) and not model_from_config:
            continue
        model = finalize_scraped_model(model, product_url, rules, model_from_config)

        price = extract_price_from_container(card, rules.get("price_selector", ""))
        if not model or not price:
            continue

        seen_urls.add(product_url)
        products.append({"url": product_url, "model": model, "price": price})
    return products


def extract_listing_products(
    current_url: str,
    html: str,
    rules: Optional[Dict[str, str]] = None,
    product_url_filters: Optional[Iterable[str]] = None,
) -> List[Dict[str, str]]:
    """Собирает товары прямо со страницы категории/каталога."""
    soup = BeautifulSoup(html, "html.parser")
    products: List[Dict[str, str]] = []
    seen_urls: Set[str] = set()
    rules = normalize_extraction_rules(rules or {})
    filters = product_url_filter_patterns(product_url_filters or [], rules)
    configured_card_source = bool(str(rules.get("product_card_selector", "")).strip())

    if configured_card_source:
        products.extend(extract_listing_products_by_rules(soup, current_url, rules, seen_urls, filters))
        seen_urls.update(product["url"] for product in products if product.get("url"))

    return products


def should_accept_extracted_product(
    url: str,
    soup: BeautifulSoup,
    model: str,
    price: str,
    h1: str,
    page_text: str,
    model_from_labeled_value: bool,
    rules: Optional[Dict[str, str]] = None,
    assume_product: bool = False,
    allow_empty_price: bool = False,
) -> bool:
    """Решает, можно ли сохранить страницу как товар.

    Важно: здесь нет проверок вида `is_maunfeld_url()` или `is_kuppersberg_url()`.
    Новый бренд должен работать через URL-фильтры, селекторы, schema.org и общие признаки товара.
    """
    if not model:
        return False
    if not price and not allow_empty_price:
        return False
    if assume_product:
        return True
    return bool(
        rules
        and any(
            str(rules.get(key, "")).strip()
            for key in ("model_selector", "price_selector", "model_start_marker", "model_end_marker")
        )
    )


def extract_product_data_by_rules(
    url: str,
    html: str,
    soup: BeautifulSoup,
    rules: Dict[str, str],
    fallback_price: str = "",
    assume_product: bool = False,
    allow_empty_price: bool = False,
) -> Optional[Dict[str, str]]:
    if not rules:
        return None
    model = extract_model_by_markers(html, rules)
    if not model:
        model = first_by_selector(soup, rules.get("model_selector", ""))
    price = first_by_selector(soup, rules.get("price_selector", ""))
    model = finalize_scraped_model(model, url, rules, True)
    prices = extract_prices(price)
    if prices:
        price = prices[-1]
    else:
        price = normalize_price_value(price or fallback_price)
    page_text = clean_text(soup.get_text(" ", strip=True))
    h1 = first_text(soup, ["h1"])
    if should_accept_extracted_product(
        url=url,
        soup=soup,
        model=model,
        price=price,
        h1=h1,
        page_text=page_text,
        model_from_labeled_value=False,
        rules=rules,
        assume_product=assume_product,
        allow_empty_price=allow_empty_price,
    ):
        return {"url": url, "model": model, "price": price}
    return None


def extract_product_data(
    url: str,
    html: str,
    fallback_price: str = "",
    rules: Optional[Dict[str, str]] = None,
    assume_product: bool = False,
    allow_empty_price: bool = False,
) -> Optional[Dict[str, str]]:
    soup = BeautifulSoup(html, "html.parser")
    rules = rules or {}
    ruled_product = extract_product_data_by_rules(
        url,
        html,
        soup,
        rules,
        fallback_price,
        assume_product,
        allow_empty_price=allow_empty_price,
    )
    if ruled_product:
        return ruled_product
    return None

def fetch_with_botasaurus_request(url: str) -> Optional[str]:
    """Fallback через Botasaurus Request: браузероподобный HTTP-запрос с Google Referrer."""
    try:
        from botasaurus.request import Request
        from botasaurus.request import request as botasaurus_request
    except ImportError as error:
        log_fetch_exception("botasaurus-request:import", url, error)
        return None

    @botasaurus_request(max_retry=MAX_RETRIES, output=None, create_error_logs=False)
    def _fetch_html(request_client: Request, target_url: str):
        response = request_client.get(target_url)
        response.raise_for_status()
        return {
            "content_type": response.headers.get("content-type", ""),
            "text": response.text,
        }

    try:
        result = _fetch_html(url)
    except Exception as error:
        log_fetch_exception("botasaurus-request", url, error)
        return None

    if isinstance(result, list):
        result = result[0] if result else None
    if not isinstance(result, dict):
        return None

    content_type = result.get("content_type", "")
    html = result.get("text", "")
    if html and ("text/html" in content_type or "application/xhtml" in content_type or not content_type):
        return html
    return None


def fetch_with_botasaurus_browser(url: str, navigation: str = "direct") -> Optional[str]:
    """Fallback через Botasaurus Browser для страниц, которым нужен настоящий рендеринг."""
    try:
        from botasaurus.browser import Driver
        from botasaurus.browser import browser
    except ImportError:
        return None

    chrome_executable_path = botasaurus_browser_executable(prefer_headless_shell=False)

    @browser(
        headless=True,
        chrome_executable_path=chrome_executable_path,
        add_arguments=[
            "--headless=new",
            "--disable-gpu",
            "--disable-dev-shm-usage",
            "--disable-background-networking",
            "--disable-sync",
            "--blink-settings=imagesEnabled=false",
        ],
        window_size=[1280, 720],
        block_images_and_css=True,
        wait_for_complete_page_load=False,
        max_retry=1,
        output=None,
        close_on_crash=True,
        create_error_logs=False,
    )
    def _render_html(driver: Driver, target_url: str):
        if navigation == "direct" and hasattr(driver, "get"):
            driver.get(target_url)
        else:
            driver.google_get(target_url)
        time.sleep(2)
        for _ in range(4):
            try:
                driver.run_js("window.scrollTo(0, document.body.scrollHeight)")
            except Exception:
                break
            time.sleep(0.8)
        return driver.page_html

    try:
        started = time.time()
        with STANDALONE_BROWSER_SEMAPHORE:
            result = _render_html(url)
    except Exception as error:
        log_fetch_exception(f"botasaurus-browser:{navigation}", url, error)
        return None

    if isinstance(result, list):
        result = result[0] if result else None
    if result:
        log_fetch_result(f"botasaurus-browser:{navigation}", url, result, time.time() - started)
    return result if isinstance(result, str) and result.strip() else None


def fetch_with_botasaurus_visible_browser(url: str) -> Optional[str]:
    """Совместимый скрытый вариант старого botasaurus-visible для автопереключения."""
    return fetch_with_botasaurus_browser(url, "direct")


def fetch_with_botasaurus_debug_visible_browser(url: str) -> Optional[str]:
    """Ручной диагностический режим: открывает видимый браузер только при явном выборе."""
    try:
        from botasaurus.browser import Driver
        from botasaurus.browser import browser
    except ImportError as error:
        log_fetch_exception("botasaurus-debug-visible:import", url, error)
        return None

    chrome_executable_path = botasaurus_browser_executable(prefer_headless_shell=False)

    @browser(
        headless=False,
        chrome_executable_path=chrome_executable_path,
        profile="protected_sites_debug_visible",
        window_size=[1280, 720],
        add_arguments=["--window-position=40,40"],
        block_images_and_css=False,
        wait_for_complete_page_load=True,
        reuse_driver=False,
        output=None,
        close_on_crash=True,
        create_error_logs=False,
        max_retry=1,
    )
    def _render_html(driver: Driver, target_url: str):
        driver.get(target_url)
        time.sleep(8)
        for _ in range(3):
            try:
                driver.run_js("window.scrollTo(0, document.body.scrollHeight)")
            except Exception:
                break
            time.sleep(0.8)
        return driver.page_html

    try:
        with VISIBLE_BROWSER_LOCK:
            result = _render_html(url)
    except Exception as error:
        log_fetch_exception("botasaurus-debug-visible", url, error)
        return None

    if isinstance(result, list):
        result = result[0] if result else None
    return result if isinstance(result, str) and result.strip() else None


class BotasaurusBrowserSession:
    """One hidden full Chromium session owned by one project/news scan.

    thread_count still controls parallel open pages, but every page now uses a
    selector-driven fast path and closes as soon as the needed DOM appears.
    """

    def __init__(
        self,
        stop_signal: Optional[threading.Event] = None,
        max_pages: int = 1,
        profile_dir: Optional[Path] = None,
    ) -> None:
        self.stop_signal = stop_signal
        self.max_pages = max(1, parse_thread_count(max_pages))
        self.profile_dir = profile_dir
        self._state_lock = threading.Lock()
        self._ready = threading.Event()
        self._thread: Optional[threading.Thread] = None
        self._loop = None
        self._playwright = None
        self._browser = None
        self._context = None
        self._semaphore = None
        self._closed = False
        self._start_error: Optional[BaseException] = None

    def _ensure_started(self) -> bool:
        with self._state_lock:
            if self._closed:
                return False
            if self._thread is None or not self._thread.is_alive():
                self._start_error = None
                self._ready.clear()
                self._thread = threading.Thread(target=self._run_loop, name="project-browser-session", daemon=True)
                self._thread.start()
        if not self._ready.wait(timeout=30):
            return False
        return self._start_error is None and self._loop is not None

    def _run_loop(self) -> None:
        import asyncio

        self._loop = asyncio.new_event_loop()
        asyncio.set_event_loop(self._loop)
        try:
            self._loop.run_until_complete(self._start_browser())
        except BaseException as error:  # noqa: BLE001
            self._start_error = error
            fetch_debug_log(f"browser-session start failed: {type(error).__name__}: {error}", "warning")
            self._ready.set()
            return
        self._ready.set()
        try:
            self._loop.run_forever()
        finally:
            if not self._loop.is_closed():
                self._loop.run_until_complete(self._shutdown_async())
                self._loop.close()

    async def _start_browser(self) -> None:
        import asyncio
        from playwright.async_api import async_playwright

        self._playwright = await async_playwright().start()
        # Используем полноценный Chromium, а нагрузку снижаем ранней остановкой
        # страницы после появления нужных селекторов.
        executable_path = env_str("PLAYWRIGHT_BROWSER_EXECUTABLE") or botasaurus_browser_executable(prefer_headless_shell=False)
        launch_options = {
            "headless": True,
            "args": [
                "--disable-blink-features=AutomationControlled",
                "--disable-gpu",
                "--disable-dev-shm-usage",
                "--disable-background-networking",
                "--disable-sync",
                "--no-sandbox",
                "--blink-settings=imagesEnabled=false",
                "--disable-renderer-backgrounding",
                "--disable-background-timer-throttling",
            ],
        }
        if executable_path:
            launch_options["executable_path"] = executable_path
        context_options = dict(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
            ),
            locale="ru-RU",
            viewport={"width": 1366, "height": 900},
            extra_http_headers={
                "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
            },
        )
        if self.profile_dir is not None:
            self.profile_dir.mkdir(parents=True, exist_ok=True)
            self._context = await self._playwright.chromium.launch_persistent_context(
                str(self.profile_dir),
                **launch_options,
                **context_options,
            )
            self._browser = self._context.browser
        else:
            self._browser = await self._playwright.chromium.launch(**launch_options)
            self._context = await self._browser.new_context(**context_options)
        try:
            await self._context.add_init_script(
                """
                Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
                Object.defineProperty(navigator, 'languages', { get: () => ['ru-RU', 'ru', 'en-US', 'en'] });
                Object.defineProperty(navigator, 'plugins', { get: () => [1, 2, 3, 4, 5] });
                window.chrome = window.chrome || { runtime: {} };
                """
            )
        except Exception:
            pass
        self._semaphore = asyncio.Semaphore(self.max_pages)

    async def _close_browser(self) -> None:
        if self._context is not None:
            try:
                await self._context.close()
            except Exception:
                pass
        if self.profile_dir is None and self._browser is not None:
            try:
                await self._browser.close()
            except Exception:
                pass
        if self._playwright is not None:
            try:
                await self._playwright.stop()
            except Exception:
                pass
        self._context = None
        self._browser = None
        self._playwright = None
        self._semaphore = None

    async def _shutdown_async(self) -> None:
        import asyncio

        current_task = asyncio.current_task()
        tasks = [
            task
            for task in asyncio.all_tasks()
            if task is not current_task and not task.done()
        ]
        for task in tasks:
            task.cancel()
        if tasks:
            await asyncio.gather(*tasks, return_exceptions=True)
        await self._close_browser()

    @staticmethod
    def _profiles_for_method(method: str) -> List[Dict[str, object]]:
        method = str(method or "")
        if method == "protected-site":
            return [
                {"name": "protected_direct", "block_stylesheet": False, "selector_timeout": 9000, "referer": ""},
                {"name": "protected_google_referrer", "block_stylesheet": False, "selector_timeout": 9000, "referer": "https://www.google.com/"},
            ]
        if method == "botasaurus-browser-direct":
            referer = ""
        elif method == "botasaurus-browser":
            referer = "https://www.google.com/"
        else:
            referer = ""
        return [
            {"name": "fast_full_browser", "block_stylesheet": True, "selector_timeout": 2500, "referer": referer},
            {"name": "compatible_full_browser", "block_stylesheet": False, "selector_timeout": 5500, "referer": referer},
        ]

    @staticmethod
    def _selector_list(rules: Optional[Dict[str, str]], allow_empty_price: bool = False) -> List[str]:
        rules = rules or {}
        selectors: List[str] = []
        for key in ("product_card_selector", "product_url_selector", "model_selector"):
            value = str(rules.get(key) or "").strip()
            if value and value not in selectors:
                selectors.append(value)
        price_selector = str(rules.get("price_selector") or "").strip()
        if price_selector and (not allow_empty_price or not selectors):
            selectors.append(price_selector)
        return selectors

    @staticmethod
    async def _count_locator(page, selector: str) -> int:
        try:
            return await page.locator(selector).count()
        except Exception:
            return 0

    async def _count_ready_nodes(self, page, rules: Optional[Dict[str, str]], allow_empty_price: bool = False) -> int:
        total = 0
        for selector in self._selector_list(rules, allow_empty_price):
            total += await self._count_locator(page, selector)
        return total

    async def _wait_for_ready_selectors(self, page, rules: Optional[Dict[str, str]], allow_empty_price: bool, timeout_ms: int) -> bool:
        for selector in self._selector_list(rules, allow_empty_price):
            try:
                await page.wait_for_selector(selector, state="attached", timeout=timeout_ms)
                return True
            except Exception:
                continue
        return False

    def _html_usable_for_parsing(
        self,
        url: str,
        html: str,
        rules: Optional[Dict[str, str]],
        product_url_filters: Optional[Iterable[str]],
        allow_empty_price: bool,
    ) -> bool:
        if not html or looks_blocked_or_empty(html):
            return False
        try:
            if extract_listing_products(url, html, rules or {}, product_url_filters or []):
                return True
        except Exception:
            pass
        try:
            product = extract_product_data(
                url,
                html,
                "",
                rules or {},
                assume_product=is_product_url_for_filters(url, product_url_filters or []),
                allow_empty_price=allow_empty_price,
            )
            if product:
                return True
        except Exception:
            pass
        return False

    async def _fetch_once(
        self,
        url: str,
        method: str,
        profile: Dict[str, object],
        rules: Optional[Dict[str, str]],
        product_url_filters: Optional[Iterable[str]],
        allow_empty_price: bool,
    ) -> Optional[str]:
        if self._context is None:
            return None
        page = await self._context.new_page()
        profile_name = str(profile.get("name") or "browser")
        block_stylesheet = bool(profile.get("block_stylesheet"))
        selector_timeout = int(profile.get("selector_timeout") or 2500)
        started = time.time()
        scroll_count = 0
        try:
            referer = str(profile.get("referer") or "")
            if referer:
                await page.set_extra_http_headers({"Referer": referer})

            async def route_handler(route, request):
                if self._should_block_resource(request, block_stylesheet=block_stylesheet):
                    await route.abort()
                else:
                    await route.continue_()

            await page.route("**/*", route_handler)
            timeout_ms = REQUEST_TIMEOUT * 1000
            await page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
            selector_found = await self._wait_for_ready_selectors(page, rules, allow_empty_price, selector_timeout)

            last_count = await self._count_ready_nodes(page, rules, allow_empty_price)
            stable_rounds = 0
            # Скроллим только до стабилизации DOM, а не фиксированное количество раз.
            for _ in range(5):
                if isinstance(self.stop_signal, threading.Event) and self.stop_signal.is_set():
                    return None
                await page.mouse.wheel(0, 1400)
                scroll_count += 1
                await page.wait_for_timeout(350)
                current_count = await self._count_ready_nodes(page, rules, allow_empty_price)
                if current_count <= last_count:
                    stable_rounds += 1
                else:
                    stable_rounds = 0
                    last_count = current_count
                if selector_found and stable_rounds >= 1:
                    break
                if stable_rounds >= 2:
                    break

            try:
                await page.evaluate("window.stop()")
            except Exception:
                pass
            html = await page.content()
            usable = self._html_usable_for_parsing(url, html, rules, product_url_filters, allow_empty_price)
            log_fetch_result(
                f"{method}:{profile_name}",
                url,
                html,
                time.time() - started,
                extra=f"selector_found={selector_found}; scroll_count={scroll_count}; usable={usable}",
            )
            return html if isinstance(html, str) and html.strip() else None
        except Exception as error:
            log_fetch_exception(f"{method}:{profile_name}", url, error)
            return None
        finally:
            try:
                await page.close()
            except Exception:
                pass

    async def _fetch_async(
        self,
        url: str,
        method: str,
        rules: Optional[Dict[str, str]] = None,
        product_url_filters: Optional[Iterable[str]] = None,
        allow_empty_price: bool = False,
    ) -> Optional[str]:
        if self._context is None or self._semaphore is None:
            return None

        async with self._semaphore:
            if isinstance(self.stop_signal, threading.Event) and self.stop_signal.is_set():
                return None
            best_html = ""
            for profile in self._profiles_for_method(method):
                html = await self._fetch_once(url, method, profile, rules, product_url_filters, allow_empty_price)
                if html and not best_html:
                    best_html = html
                if html and self._html_usable_for_parsing(url, html, rules, product_url_filters, allow_empty_price):
                    return html
            return best_html or None

    @staticmethod
    def _should_block_resource(request, block_stylesheet: bool = True) -> bool:
        resource_type = getattr(request, "resource_type", "")
        if resource_type in {"image", "media", "font"}:
            return True
        if block_stylesheet and resource_type == "stylesheet":
            return True
        request_url = str(getattr(request, "url", "") or "").lower()
        return any(part in request_url for part in BLOCKED_BROWSER_URL_PARTS)

    def fetch(
        self,
        url: str,
        method: str,
        rules: Optional[Dict[str, str]] = None,
        product_url_filters: Optional[Iterable[str]] = None,
        allow_empty_price: bool = False,
    ) -> Optional[str]:
        if isinstance(self.stop_signal, threading.Event) and self.stop_signal.is_set():
            return None
        with self._state_lock:
            if self._closed:
                return None
        if not self._ensure_started():
            return None
        future = None
        try:
            import asyncio

            future = asyncio.run_coroutine_threadsafe(
                self._fetch_async(url, method, rules, product_url_filters, allow_empty_price),
                self._loop,
            )
            result = future.result(timeout=REQUEST_TIMEOUT + 35)
        except Exception as error:
            if future is not None:
                future.cancel()
                try:
                    future.result(timeout=2)
                except Exception:
                    pass
            log_fetch_exception(method, url, error)
            return None
        return result if isinstance(result, str) and result.strip() else None

    def close(self) -> None:
        import asyncio

        thread = None
        loop = None
        with self._state_lock:
            if self._closed:
                return
            self._closed = True
            thread = self._thread
            loop = self._loop
        if loop is not None:
            try:
                shutdown = asyncio.run_coroutine_threadsafe(self._shutdown_async(), loop)
                shutdown.result(timeout=10)
            except Exception:
                pass
            try:
                loop.call_soon_threadsafe(loop.stop)
            except Exception:
                pass
        if thread is not None and thread is not threading.current_thread():
            thread.join(timeout=10)


class BotasaurusDebugVisibleSession:
    """One visible Botasaurus browser with a persistent project profile."""

    def __init__(self, stop_signal: Optional[threading.Event] = None, profile: str = "protected_sites_debug_visible") -> None:
        self.stop_signal = stop_signal
        self.profile = profile or "protected_sites_debug_visible"
        self._lock = threading.Lock()
        self._closed = False
        self._renderer = self._create_renderer()

    def _create_renderer(self):
        try:
            from botasaurus.browser import Driver
            from botasaurus.browser import browser
        except ImportError:
            return None

        chrome_executable_path = botasaurus_browser_executable(prefer_headless_shell=False)

        @browser(
            headless=False,
            chrome_executable_path=chrome_executable_path,
            profile=self.profile,
            window_size=[1280, 720],
            add_arguments=["--window-position=40,40"],
            block_images_and_css=False,
            wait_for_complete_page_load=True,
            reuse_driver=True,
            output=None,
            close_on_crash=True,
            create_error_logs=False,
            max_retry=1,
        )
        def _render_html(driver: Driver, target_url: str):
            driver.get(target_url)
            time.sleep(8)
            for _ in range(3):
                try:
                    driver.run_js("window.scrollTo(0, document.body.scrollHeight)")
                except Exception:
                    break
                time.sleep(0.8)
            return driver.page_html

        return _render_html

    def fetch(self, url: str, method: str) -> Optional[str]:
        if isinstance(self.stop_signal, threading.Event) and self.stop_signal.is_set():
            return None
        if self._renderer is None:
            return None
        with self._lock:
            if self._closed:
                return None
            try:
                result = self._renderer(url)
            except Exception:
                return None
        if isinstance(result, list):
            result = result[0] if result else None
        return result if isinstance(result, str) and result.strip() else None

    def close(self) -> None:
        with self._lock:
            self._closed = True
            renderer = self._renderer
            if renderer is not None and hasattr(renderer, "close"):
                try:
                    renderer.close()
                except Exception:
                    pass


def fetch_with_crawl4ai(url: str) -> Optional[str]:
    try:
        import asyncio
        from crawl4ai import AsyncWebCrawler, BrowserConfig, CrawlerRunConfig
    except ImportError:
        return None

    async def _fetch() -> Optional[str]:
        browser_config = BrowserConfig(
            browser_type="chromium",
            headless=True,
            channel="chromium",
            text_mode=True,
            light_mode=True,
            avoid_ads=True,
            avoid_css=True,
            viewport_width=1366,
            viewport_height=900,
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
            ),
            extra_args=[
                "--disable-gpu",
                "--disable-dev-shm-usage",
                "--disable-background-networking",
                "--disable-sync",
                "--blink-settings=imagesEnabled=false",
            ],
            verbose=False,
        )
        run_config = CrawlerRunConfig(
            wait_until="domcontentloaded",
            page_timeout=REQUEST_TIMEOUT * 1000,
            wait_for_images=False,
            delay_before_return_html=0.2,
            exclude_all_images=True,
            excluded_tags=["img", "picture", "source", "video", "audio", "svg", "style"],
            exclude_domains=list(BLOCKED_BROWSER_URL_PARTS),
            log_console=False,
            capture_network_requests=False,
            max_retries=0,
            verbose=False,
        )
        async with AsyncWebCrawler(config=browser_config) as crawler:
            result = await asyncio.wait_for(
                crawler.arun(url=url, config=run_config),
                timeout=REQUEST_TIMEOUT + 10,
            )
            html = getattr(result, "html", "") or getattr(result, "cleaned_html", "")
            return html if isinstance(html, str) else None

    try:
        with STANDALONE_BROWSER_SEMAPHORE:
            return asyncio.run(_fetch())
    except Exception:
        return None


ENGINE_OUTPUT_MARKER = "__PARSER_ENGINE_HTML_BASE64__:"

SCRAPY_FETCH_SCRIPT = r"""
import base64
import sys

import scrapy
from scrapy.crawler import CrawlerProcess

url = sys.argv[1]
timeout = int(float(sys.argv[2]))
marker = sys.argv[3]


class SinglePageSpider(scrapy.Spider):
    name = "single_page_fetch"
    body = b""
    handle_httpstatus_all = True
    custom_settings = {
        "LOG_ENABLED": False,
        "ROBOTSTXT_OBEY": False,
        "DOWNLOAD_TIMEOUT": timeout,
        "RETRY_ENABLED": False,
        "COOKIES_ENABLED": True,
        "HTTPERROR_ALLOW_ALL": True,
        "USER_AGENT": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36",
        "DEFAULT_REQUEST_HEADERS": {
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
        },
        "TELNETCONSOLE_ENABLED": False,
        "WARN_ON_GENERATOR_RETURN_VALUE": False,
    }

    async def start(self):
        yield scrapy.Request(url, dont_filter=True)

    def parse(self, response):
        SinglePageSpider.body = bytes(response.body or b"")


process = CrawlerProcess(settings=SinglePageSpider.custom_settings)
process.crawl(SinglePageSpider)
process.start(stop_after_crawl=True)
print(marker + base64.b64encode(SinglePageSpider.body).decode("ascii"))
"""

CRAWLEE_FETCH_SCRIPT = r"""
import asyncio
import base64
import os
import shutil
import sys
import tempfile
import uuid
from datetime import timedelta

storage_dir = os.path.join(tempfile.gettempdir(), "parser-crawlee", uuid.uuid4().hex)
os.environ["CRAWLEE_STORAGE_DIR"] = storage_dir

from crawlee.crawlers._http import HttpCrawler

url = sys.argv[1]
timeout = int(float(sys.argv[2]))
marker = sys.argv[3]


async def main():
    result = {"body": b""}
    crawler = HttpCrawler(
        max_requests_per_crawl=1,
        max_request_retries=0,
        request_handler_timeout=timedelta(seconds=timeout),
        configure_logging=False,
        ignore_http_error_status_codes=list(range(300, 600)),
    )

    @crawler.router.default_handler
    async def handler(context):
        result["body"] = await context.http_response.read()

    await crawler.run([url])
    print(marker + base64.b64encode(result["body"]).decode("ascii"))


try:
    asyncio.run(main())
finally:
    shutil.rmtree(storage_dir, ignore_errors=True)
"""

PLAYWRIGHT_FETCH_SCRIPT = r"""
import base64
import sys

from playwright.sync_api import sync_playwright

url = sys.argv[1]
timeout = int(float(sys.argv[2])) * 1000
marker = sys.argv[3]
blocked_resource_types = {"image", "media", "font", "stylesheet"}
blocked_url_parts = (
    "google-analytics",
    "googletagmanager",
    "doubleclick",
    "adservice",
    "adsystem",
    "yandex.ru/metrika",
    "mc.yandex",
    "metrika",
    "analytics",
    "counter",
    "facebook.net",
    "vk.com/rtrg",
    "top-fwz1.mail.ru",
    "mail.ru/counter",
)


def should_block(request):
    if request.resource_type in blocked_resource_types:
        return True
    request_url = (request.url or "").lower()
    return any(part in request_url for part in blocked_url_parts)

with sync_playwright() as p:
    browser = p.chromium.launch(
        headless=True,
        args=[
            "--disable-blink-features=AutomationControlled",
            "--disable-dev-shm-usage",
            "--no-sandbox",
        ],
    )
    context = browser.new_context(
        user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36",
        locale="ru-RU",
        viewport={"width": 1366, "height": 900},
    )
    page = context.new_page()
    page.route("**/*", lambda route, request: route.abort() if should_block(request) else route.continue_())
    page.goto(url, wait_until="domcontentloaded", timeout=timeout)
    try:
        page.wait_for_load_state("networkidle", timeout=min(timeout, 15000))
    except Exception:
        pass
    for _ in range(3):
        try:
            page.mouse.wheel(0, 1600)
            page.wait_for_timeout(500)
        except Exception:
            break
    html = page.content()
    browser.close()
    print(marker + base64.b64encode(html.encode("utf-8", "replace")).decode("ascii"))
"""

SCRAPEGRAPHAI_FETCH_SCRIPT = r"""
import base64
import sys

try:
    import langchain_community.chat_models as community_chat_models
    from langchain_ollama import ChatOllama

    if not hasattr(community_chat_models, "ChatOllama"):
        community_chat_models.ChatOllama = ChatOllama
except Exception:
    pass

from scrapegraphai.nodes.fetch_node import FetchNode

url = sys.argv[1]
timeout = int(float(sys.argv[2]))
marker = sys.argv[3]

node = FetchNode(
    input="url",
    output=["doc"],
    node_config={
        "headless": True,
        "timeout": timeout,
        "use_soup": False,
        "cut": False,
        "loader_kwargs": {
            "timeout": timeout,
            "requires_js_support": True,
            "load_state": "networkidle",
            "args": [
                "--disable-blink-features=AutomationControlled",
                "--disable-dev-shm-usage",
                "--no-sandbox",
            ],
        },
    },
)
state = node.execute({"url": url}) or {}
documents = state.get("doc") or state.get("document") or []
html = ""
if isinstance(documents, list) and documents:
    html = getattr(documents[0], "page_content", "") or str(documents[0] or "")
elif isinstance(documents, str):
    html = documents
print(marker + base64.b64encode(str(html).encode("utf-8", "replace")).decode("ascii"))
"""


class PlaywrightHeadlessRenderer:
    """Small pool of isolated Chromium workers for fallback rendering."""

    def __init__(self) -> None:
        self.jobs: Queue = Queue()
        self.threads: List[threading.Thread] = []
        self.lock = threading.Lock()

    def ensure_started(self) -> None:
        with self.lock:
            self.threads = [thread for thread in self.threads if thread.is_alive()]
            while len(self.threads) < 1:
                worker_number = len(self.threads) + 1
                thread = threading.Thread(
                    target=self._worker,
                    name=f"playwright-headless-renderer-{worker_number}",
                    daemon=True,
                )
                self.threads.append(thread)
                thread.start()

    def fetch(self, url: str, timeout_seconds: int) -> Optional[str]:
        self.ensure_started()
        result_queue: Queue = Queue(maxsize=1)
        self.jobs.put((url, timeout_seconds, result_queue))
        try:
            status, value = result_queue.get(timeout=timeout_seconds + 35)
        except Empty as error:
            raise RuntimeError("Playwright: внутренний headless browser не ответил вовремя") from error
        if status == "error":
            raise RuntimeError(f"Playwright: {value}")
        return value if isinstance(value, str) and value.strip() else None

    def _worker(self) -> None:
        from playwright.sync_api import sync_playwright

        def should_block_resource(request) -> bool:
            resource_type = getattr(request, "resource_type", "")
            if resource_type in BLOCKED_BROWSER_RESOURCE_TYPES:
                return True
            request_url = str(getattr(request, "url", "") or "").lower()
            return any(part in request_url for part in BLOCKED_BROWSER_URL_PARTS)

        with sync_playwright() as playwright:
            executable_path = botasaurus_browser_executable(prefer_headless_shell=True)
            launch_options = {
                "headless": True,
                "args": [
                    "--disable-blink-features=AutomationControlled",
                    "--disable-gpu",
                    "--disable-dev-shm-usage",
                    "--disable-background-networking",
                    "--disable-sync",
                    "--no-sandbox",
                    "--blink-settings=imagesEnabled=false",
                ],
            }
            if executable_path:
                launch_options["executable_path"] = executable_path
            browser = playwright.chromium.launch(**launch_options)
            context = browser.new_context(
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
                ),
                locale="ru-RU",
                viewport={"width": 1366, "height": 900},
            )
            try:
                while True:
                    url, timeout_seconds, result_queue = self.jobs.get()
                    page = None
                    try:
                        page = context.new_page()
                        page.route(
                            "**/*",
                            lambda route, request: route.abort()
                            if should_block_resource(request)
                            else route.continue_(),
                        )
                        timeout_ms = int(float(timeout_seconds)) * 1000
                        page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
                        try:
                            page.wait_for_load_state("networkidle", timeout=min(timeout_ms, 10000))
                        except Exception:
                            pass
                        for _ in range(3):
                            page.mouse.wheel(0, 1600)
                            page.wait_for_timeout(350)
                        html = page.content()
                        result_queue.put(("ok", html), block=False)
                    except Exception as error:
                        result_queue.put(("error", error), block=False)
                    finally:
                        if page is not None:
                            try:
                                page.close()
                            except Exception:
                                pass
            finally:
                context.close()
                browser.close()


playwright_headless_renderer = PlaywrightHeadlessRenderer()


def fetch_with_python_engine(script: str, url: str, timeout_seconds: int, engine_name: str = "engine") -> Optional[str]:
    try:
        completed = subprocess.run(
            [sys.executable, "-c", script, url, str(timeout_seconds), ENGINE_OUTPUT_MARKER],
            cwd=str(BASE_DIR),
            capture_output=True,
            timeout=timeout_seconds + 10,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )
    except Exception as error:
        raise RuntimeError(f"{engine_name}: не удалось запустить процесс: {error}") from error

    stdout = completed.stdout.decode("utf-8", "replace")
    stderr = completed.stderr.decode("utf-8", "replace")

    if completed.returncode != 0:
        details = (stderr or stdout or "").strip()
        if len(details) > 1200:
            details = details[-1200:]
        raise RuntimeError(f"{engine_name}: процесс завершился с кодом {completed.returncode}: {details}")

    for line in reversed(stdout.splitlines()):
        if ENGINE_OUTPUT_MARKER not in line:
            continue

        payload = line.split(ENGINE_OUTPUT_MARKER, 1)[1].strip()

        if not payload:
            raise RuntimeError(f"{engine_name}: движок вернул пустой HTML")

        try:
            html_bytes = base64.b64decode(payload)
        except Exception as error:
            raise RuntimeError(f"{engine_name}: не удалось декодировать HTML: {error}") from error

        html = html_bytes.decode("utf-8", "replace").strip()
        return html or None

    details = (stderr or stdout or "").strip()
    if len(details) > 1200:
        details = details[-1200:]
    raise RuntimeError(f"{engine_name}: движок не вернул HTML. Вывод: {details}")


def fetch_with_scrapy(url: str) -> Optional[str]:
    try:
        import scrapy  # noqa: F401
    except ImportError:
        return None
    return fetch_with_python_engine(SCRAPY_FETCH_SCRIPT, url, REQUEST_TIMEOUT, "Scrapy")


def fetch_with_crawlee(url: str) -> Optional[str]:
    try:
        import crawlee  # noqa: F401
    except ImportError:
        return None
    return fetch_with_python_engine(CRAWLEE_FETCH_SCRIPT, url, REQUEST_TIMEOUT, "Crawlee")


def fetch_with_playwright(url: str) -> Optional[str]:
    try:
        import playwright  # noqa: F401
    except ImportError:
        return None
    with STANDALONE_BROWSER_SEMAPHORE:
        return playwright_headless_renderer.fetch(url, REQUEST_TIMEOUT)


def fetch_with_scrapegraphai(url: str) -> Optional[str]:
    try:
        import scrapegraphai  # noqa: F401
    except ImportError:
        return None
    with STANDALONE_BROWSER_SEMAPHORE:
        return fetch_with_python_engine(SCRAPEGRAPHAI_FETCH_SCRIPT, url, REQUEST_TIMEOUT, "ScrapeGraphAI")


class ProductSiteCrawler:
    def __init__(
        self,
        start_urls: List[str],
        run_id: int,
        stop_signal: threading.Event,
        finish_signal: threading.Event,
        thread_count: int,
        project: Optional[Dict[str, object]] = None,
        exclusions: Optional[List[str]] = None,
        product_url_filters: Optional[List[str]] = None,
        product_url_exclusions: Optional[List[str]] = None,
        extraction_rules: Optional[Dict[str, str]] = None,
        connection_method: str = "requests",
        auto_connection_fallback: bool = True,
        connection_method_state: Optional[Dict[str, object]] = None,
        allow_empty_price: bool = False,
        browser_session: Optional[BotasaurusBrowserSession] = None,
        owns_browser_session: bool = True,
        profile_dir: Optional[Path] = None,
    ):
        self.run_id = run_id
        self.stop_signal = stop_signal
        self.finish_signal = finish_signal
        self.thread_count = parse_thread_count(thread_count)
        self.start_urls = normalize_start_urls(start_urls)
        self.start_url = self.start_urls[0]
        self.root_netloc = urlparse(self.start_url).netloc
        self.project = project
        self.exclusions = exclusions if exclusions is not None else DEFAULT_EXCLUSIONS.copy()
        self.extraction_rules = normalize_extraction_rules(extraction_rules or {})
        self.product_url_filters = product_url_filter_patterns(product_url_filters or [], self.extraction_rules)
        self.product_url_exclusions = normalize_patterns(product_url_exclusions or [])
        self.connection_method = normalize_connection_method(connection_method)
        self.auto_connection_fallback = bool(auto_connection_fallback)
        self.allow_empty_price = bool(allow_empty_price)
        self.profile_dir = profile_dir
        if connection_method_state is None:
            connection_method_state = {
                "active_method": self.connection_method,
                "lock": threading.Lock(),
            }
        else:
            connection_method_state.setdefault("active_method", self.connection_method)
            connection_method_state.setdefault("lock", threading.Lock())
        self.connection_method_state = connection_method_state
        self.active_connection_method = str(connection_method_state.get("active_method") or self.connection_method)
        debug_profile = str(profile_dir) if profile_dir is not None else "protected_sites_debug_visible"
        self.browser_session = browser_session or BotasaurusBrowserSession(
            self.stop_signal,
            self.thread_count,
            profile_dir=self.profile_dir,
        )
        self.debug_visible_session = BotasaurusDebugVisibleSession(self.stop_signal, debug_profile)
        self.owns_browser_session = owns_browser_session
        self.browser_sessions_lock = threading.Lock()
        self.browser_sessions: List[object] = []
        self.thread_local = threading.local()
        self.headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0 Safari/537.36"
            ),
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
        }
        self.queue: Queue[str] = Queue()
        self.queued: Set[str] = set()
        self.in_progress: Set[str] = set()
        self.visited: Set[str] = set()
        self.skipped_urls: Set[str] = set()
        self.result_urls: Set[str] = set()
        self.pending_prices: Dict[str, str] = {}
        self.results: List[Dict[str, str]] = []
        self.failed_attempts: Dict[str, int] = {}
        self.permanent_failures: Set[str] = set()
        self.data_lock = threading.Lock()
        self.excel_finalized = False
        self.started_at = 0.0
        self.elapsed_before_resume = 0.0
        self.last_progress_at = time.time()
        self.last_progress_signature: tuple = ()
        self.fatal_error = ""

    def update_state(self, **kwargs: object) -> None:
        if self.project is not None:
            if self.run_id != int(self.project.get("run_id", self.run_id)):
                return
            update_project_state(self.project, **kwargs)
        else:
            update_state(self.run_id, **kwargs)

    def reset_state(self, status: str = "idle") -> None:
        if self.project is not None:
            if self.run_id != int(self.project.get("run_id", self.run_id)):
                return
            reset_project_state(self.project, status)
        else:
            reset_state(status, self.run_id, self.thread_count)

    def log(self, message: str, level: str = "info") -> None:
        if self.project is not None:
            if self.run_id != int(self.project.get("run_id", self.run_id)):
                return
            add_project_log(self.project, message, level)

    def get_session(self) -> requests.Session:
        session = getattr(self.thread_local, "session", None)
        if session is None:
            session = requests.Session()
            session.headers.update(self.headers)
            self.thread_local.session = session
        return session

    def browser_session_for_worker(self) -> BotasaurusBrowserSession:
        return self.browser_session

    def debug_visible_session_for_worker(self) -> BotasaurusDebugVisibleSession:
        return self.debug_visible_session

    def close_browser_sessions(self) -> None:
        sessions: List[object] = []
        with self.browser_sessions_lock:
            for session in self.browser_sessions:
                if session not in sessions:
                    sessions.append(session)
            self.browser_sessions.clear()
        if self.owns_browser_session:
            if self.browser_session not in sessions:
                sessions.append(self.browser_session)
            if self.debug_visible_session not in sessions:
                sessions.append(self.debug_visible_session)
        for session in sessions:
            session.close()

    def fetch_with_requests(self, url: str) -> Optional[str]:
        last_error = ""
        candidate_urls = [url]
        for attempt in range(1, MAX_RETRIES + 1):
            if self.stop_signal.is_set():
                return None
            for candidate_url in candidate_urls:
                try:
                    response = self.get_session().get(candidate_url, timeout=REQUEST_TIMEOUT)
                    response.raise_for_status()
                    content_type = response.headers.get("content-type", "")
                    if "text/html" not in content_type and "application/xhtml" not in content_type:
                        return None
                    if not looks_blocked_or_empty(response.text):
                        return response.text

                    last_error = "страница похожа на блокировку или пустой JS-шаблон"
                    break
                except requests.RequestException as exc:
                    last_error = str(exc)
                    if isinstance(exc, requests.HTTPError) and exc.response is not None and exc.response.status_code in {404, 410}:
                        with self.data_lock:
                            self.permanent_failures.add(url)
                        self.log(f"URL пропущен: страница вернула {exc.response.status_code}: {url}", "warning")
                        return None
                    if isinstance(exc, requests.HTTPError) and exc.response is not None and exc.response.status_code in {401, 403}:
                        continue
                    break
            if last_error == "страница похожа на блокировку или пустой JS-шаблон":
                break
            if self.stop_signal.is_set():
                return None
            if attempt < MAX_RETRIES:
                time.sleep(min(attempt, 3))
        if last_error:
            self.log(f"requests не смог загрузить {url}: {last_error}", "warning")
        return None

    def fetch_by_method(self, url: str, method: str) -> Optional[str]:
        target_url = url
        if method == "requests":
            return self.fetch_with_requests(target_url)
        if method == "botasaurus-request":
            return fetch_with_botasaurus_request(target_url)
        if method in {"botasaurus-browser", "botasaurus-browser-direct", "botasaurus-visible", "playwright"}:
            return self.browser_session_for_worker().fetch(
                target_url,
                method,
                self.extraction_rules,
                self.product_url_filters,
                self.allow_empty_price,
            )
        if method == "botasaurus-debug-visible":
            return self.debug_visible_session_for_worker().fetch(target_url, method)
        if method in SESSION_BROWSER_METHODS:
            return self.browser_session_for_worker().fetch(
                target_url,
                method,
                self.extraction_rules,
                self.product_url_filters,
                self.allow_empty_price,
            )
        if method == "crawl4ai":
            return fetch_with_crawl4ai(target_url)
        if method == "scrapegraphai":
            return fetch_with_scrapegraphai(target_url)
        if method == "scrapy":
            return fetch_with_scrapy(target_url)
        if method == "crawlee":
            return fetch_with_crawlee(target_url)
        return None

    def fetch_by_method_with_timeout(self, url: str, method: str) -> Optional[str]:
        if is_browser_render_method(method) or is_debug_visible_method(method):
            try:
                return self.fetch_by_method(url, method)
            except Exception as error:
                self.log(f"Метод подключения {method} завершился ошибкой для {url}: {error}", "warning")
                return None

        result_queue: Queue = Queue(maxsize=1)

        def run_method() -> None:
            try:
                result_queue.put(("ok", self.fetch_by_method(url, method)), block=False)
            except Exception as error:
                result_queue.put(("error", error), block=False)

        thread = threading.Thread(target=run_method, daemon=True)
        thread.start()
        try:
            status, value = result_queue.get(timeout=CONNECTION_METHOD_TIMEOUT_SECONDS)
        except Empty:
            self.log(
                f"Метод подключения {method} превысил таймаут {CONNECTION_METHOD_TIMEOUT_SECONDS} сек. для {url}",
                "warning",
            )
            return None
        if status == "error":
            self.log(f"Метод подключения {method} завершился ошибкой для {url}: {value}", "warning")
            return None
        return value if isinstance(value, str) else None

    def fallback_method_sequence(self) -> List[str]:
        """Возвращает fallback-методы из БД без схлопывания разных браузерных движков."""
        methods: List[str] = []
        for method in ordered_db_connection_methods():
            # Видимый debug-браузер не запускаем автоматически, только если выбран явно.
            if is_debug_visible_method(method) and method != self.connection_method:
                continue
            if method not in methods:
                methods.append(method)
        return methods

    def current_connection_method(self) -> str:
        lock = self.connection_method_state["lock"]
        with lock:
            current = normalize_connection_method(self.connection_method_state.get("active_method"))
            self.connection_method_state["active_method"] = current
            self.active_connection_method = current
            return current

    def set_active_connection_method(self, method: str) -> None:
        previous = self.active_connection_method
        lock = self.connection_method_state["lock"]
        with lock:
            current = normalize_connection_method(method)
            self.connection_method_state["active_method"] = current
            self.active_connection_method = current
        if current != previous and self.owns_browser_session:
            self.close_browser_sessions()
            self.browser_session = BotasaurusBrowserSession(
                self.stop_signal,
                self.thread_count,
                profile_dir=self.profile_dir,
            )
            self.debug_visible_session = BotasaurusDebugVisibleSession(
                self.stop_signal,
                str(self.profile_dir) if self.profile_dir is not None else "protected_sites_debug_visible",
            )

    def fetch_with_connection_method(self, url: str, method: str) -> Optional[str]:
        self.last_progress_at = time.time()
        self.log(f"Пробую метод подключения {method} для {url}", "info")
        started = time.time()
        html = self.fetch_by_method_with_timeout(url, method)
        self.last_progress_at = time.time()
        if html:
            log_fetch_result(method, url, html, time.time() - started)
        if html and not looks_blocked_or_empty(html):
            return html
        self.log(f"Метод подключения {method} не сработал для {url}", "warning")
        return None

    def fetch(self, url: str) -> Optional[str]:
        current_method = self.current_connection_method()
        last_method = current_method

        if self.stop_signal.is_set():
            return None

        html = self.fetch_with_connection_method(url, current_method)
        if html:
            return html

        with self.data_lock:
            if url in self.permanent_failures:
                return None

        if not self.auto_connection_fallback:
            self.update_state(
                error=(
                    f"Не удалось загрузить {url}. Последний метод: {last_method}. "
                    "Проверьте способ подключения или включите автопереключение."
                ),
            )
            self.log(f"Не удалось загрузить {url}. Последний метод: {last_method}", "error")
            return None

        for method in self.fallback_method_sequence():
            if self.stop_signal.is_set():
                return None
            if method == current_method:
                continue
            last_method = method
            html = self.fetch_with_connection_method(url, method)
            if html:
                if method != current_method:
                    self.set_active_connection_method(method)
                    self.log(f"Автопереключение подключения: {method} для {url}", "warning")
                return html
            with self.data_lock:
                if url in self.permanent_failures:
                    break

        self.update_state(
            error=(
                f"Не удалось загрузить {url}. Последний метод: {last_method}. "
                "Проверьте способ подключения или включите автопереключение."
            ),
        )
        self.log(f"Не удалось загрузить {url}. Последний метод: {last_method}", "error")
        return None

    def is_excluded(self, url: str) -> bool:
        patterns = self.exclusions

        matched = any(exclusion_matches(url, pattern) for pattern in patterns)
        if matched and url not in self.skipped_urls:
            with self.data_lock:
                self.skipped_urls.add(url)
                skipped_count = len(self.skipped_urls)
            self.update_state(skipped=skipped_count)
        return matched

    def is_product_allowed(self, url: str) -> bool:
        return product_url_matches_filters(url, self.product_url_filters) and not product_url_matches_any(url, self.product_url_exclusions)

    def is_filter_marked_product(self, url: str) -> bool:
        return bool(self.product_url_filters) and self.is_product_allowed(url)

    def is_product_url(self, url: str) -> bool:
        return is_product_url_for_filters(url, self.product_url_filters)

    def is_start_url_path(self, url: str) -> bool:
        parsed = urlparse(url)
        normalized_path = (parsed.path or "/").rstrip("/") or "/"
        for start_url in self.start_urls:
            start_parsed = urlparse(start_url)
            if not same_site(url, start_parsed.netloc or self.root_netloc):
                continue
            start_path = (start_parsed.path or "/").rstrip("/") or "/"
            if normalized_path == start_path:
                return True
        return False

    def is_current_product_page(self, url: str) -> bool:
        return self.is_product_url(url) and self.is_product_allowed(url) and not self.is_start_url_path(url)

    def remember_listing_price(self, product_url: str, price: str) -> None:
        product_url = canonicalize_product_url_by_filters(product_url, self.product_url_filters)
        with self.data_lock:
            if product_url and price:
                self.pending_prices[product_url] = price

    def get_listing_price(self, product_url: str) -> str:
        product_url = canonicalize_product_url_by_filters(product_url, self.product_url_filters)
        with self.data_lock:
            return self.pending_prices.get(product_url, "")

    def enqueue(self, url: Optional[str], force: bool = False) -> None:
        url = canonicalize_product_url_by_filters(url or "", self.product_url_filters)
        if not url or url in self.visited or url in self.queued or url in self.in_progress:
            return
        is_product = self.is_product_url(url)
        if is_product and not self.is_product_allowed(url):
            return
        with self.data_lock:
            if url in self.result_urls:
                return
        if not force:
            if is_product:
                if not same_site(url, self.root_netloc) or has_static_extension(url):
                    return
            elif not should_follow_project_url(url, self.start_urls, self.root_netloc):
                return
        if force and (not same_site(url, self.root_netloc) or has_static_extension(url)):
            return
        if self.is_excluded(url):
            return
        self.queue.put(url)
        self.queued.add(url)

    def requeue_pending(self, pending_urls: Iterable[str]) -> None:
        with self.data_lock:
            for url in pending_urls:
                self.in_progress.discard(url)
        for url in pending_urls:
            self.enqueue(url, force=True)

    def extract_links(self, html: str, current_url: str) -> None:
        soup = BeautifulSoup(html, "html.parser")
        for link in soup.select("a[href]"):
            normalized = normalize_url(link.get("href", ""), current_url)
            self.enqueue(normalized)

    def add_products(self, products: Iterable[Dict[str, str]]) -> int:
        added = 0
        with self.data_lock:
            for product in products:
                product_url = canonicalize_product_url_by_filters(product.get("url", ""), self.product_url_filters)
                model = product.get("model", "")
                price = product.get("price", "")
                if product_url and not self.is_product_allowed(product_url):
                    continue
                if not product_url or not model or (not price and not self.allow_empty_price) or product_url in self.result_urls:
                    continue
                product["url"] = product_url
                self.result_urls.add(product_url)
                self.results.append(product)
                added += 1
        return added

    def snapshot_counts(self) -> Dict[str, int]:
        with self.data_lock:
            return {
                "visited": len(self.visited),
                "results": len(self.results),
                "skipped": len(self.skipped_urls),
                "queued": self.queue.qsize(),
                "active": len(self.in_progress),
                "failed": len(self.failed_attempts),
            }

    def snapshot_results(self) -> List[Dict[str, str]]:
        with self.data_lock:
            return [dict(item) for item in self.results]

    def refresh_progress(self, current_url: str = "") -> None:
        remaining = self.queue.qsize()
        counts = self.snapshot_counts()
        processed = counts["visited"]
        total_known = processed + remaining
        percent = int((processed / total_known) * 100) if total_known else 0
        elapsed = self.elapsed_seconds()
        eta = None
        if processed > 0 and remaining > 0:
            eta = int((elapsed / processed) * remaining)
        self.update_state(
            percent=percent,
            currenturl=current_url,
            totalprocessed=processed,
            processed_products=counts["results"],
            found_products=counts["results"],
            in_memory_products=counts["results"],
            queue_size=remaining,
            active_tasks=counts["active"],
            active_urls=sorted(self.in_progress)[:8],
            failed_pages=counts["failed"],
            stall_seconds=max(0, int(time.time() - self.last_progress_at)),
            skipped=counts["skipped"],
            thread_count=self.thread_count,
            elapsed_seconds=int(elapsed),
            eta_seconds=eta,
        )

    def progress_signature(self, pending_urls: Iterable[str]) -> tuple:
        counts = self.snapshot_counts()
        return (
            counts["visited"],
            counts["results"],
            counts["skipped"],
            self.queue.qsize(),
            tuple(sorted(pending_urls)),
        )

    def note_progress_activity(self, pending_urls: Iterable[str]) -> None:
        signature = self.progress_signature(pending_urls)
        if signature != self.last_progress_signature:
            self.last_progress_signature = signature
            self.last_progress_at = time.time()

    def mark_stalled(self, pending_urls: Iterable[str]) -> None:
        active_urls = list(pending_urls)
        counts = self.snapshot_counts()
        message = (
            f"Сбор не двигается {NEWS_SCAN_STALL_TIMEOUT} секунд. "
            f"Активных задач: {len(active_urls)}; очередь: {self.queue.qsize()}; "
            f"товаров в памяти: {counts['results']}. "
            f"Активные URL: {', '.join(active_urls[:5])}"
        )
        self.fatal_error = message
        self.update_state(
            status="error",
            error=message,
            currenturl=active_urls[0] if active_urls else "",
            active_urls=active_urls[:8],
            active_tasks=len(active_urls),
            queue_size=self.queue.qsize(),
            in_memory_products=counts["results"],
            stall_seconds=NEWS_SCAN_STALL_TIMEOUT,
        )
        self.log(message, "error")
        self.stop_signal.set()

    def elapsed_seconds(self) -> float:
        if self.started_at:
            return self.elapsed_before_resume + max(0.0, time.time() - self.started_at)
        return self.elapsed_before_resume

    def process_page(self, url: str, html: str) -> None:
        current_is_product = self.is_current_product_page(url)
        listing_products: List[Dict[str, str]] = []
        listing_price = self.get_listing_price(url)

        if current_is_product:
            product = extract_product_data(
                url,
                html,
                listing_price,
                self.extraction_rules,
                assume_product=True,
                allow_empty_price=self.allow_empty_price,
            )
            if product:
                self.add_products([product])
                return

            current_is_product = False

        listing_products = extract_listing_products(url, html, self.extraction_rules, self.product_url_filters)

        if (
            self.current_connection_method() != "requests"
            and not current_is_product
            and not listing_products
            and (is_catalog_url(url) or not self.is_product_url(url))
        ):
            self.update_state(
                error=f"На странице каталога не извлечены товары. Рендерю через браузер: {url}",
            )
            self.log(f"Рендеринг каталога через браузер: {url}", "warning")
            rendered_html = self.browser_session.fetch(
                url,
                "protected-site",
                self.extraction_rules,
                self.product_url_filters,
                self.allow_empty_price,
            )
            if rendered_html and not looks_blocked_or_empty(rendered_html):
                html = rendered_html
                listing_products = extract_listing_products(url, html, self.extraction_rules, self.product_url_filters)
                self.update_state(error="")

        for product in listing_products:
            product_url = canonicalize_product_url_by_filters(product.get("url", ""), self.product_url_filters)
            if product_url and not self.is_product_allowed(product_url):
                continue
            product["url"] = product_url
            self.remember_listing_price(product_url, product.get("price", ""))
            self.enqueue(product_url, force=True)
        if not self.product_url_filters and not has_explicit_model_rules(self.extraction_rules):
            self.add_products(listing_products)

        should_extract_current_product = (
            not self.is_start_url_path(url)
            and (not listing_products or bool(self.get_listing_price(url)))
        )
        product = None if not should_extract_current_product else extract_product_data(
            url,
            html,
            self.get_listing_price(url),
            self.extraction_rules,
            assume_product=self.is_product_url(url),
            allow_empty_price=self.allow_empty_price,
        )
        if product:
            self.add_products([product])

        if not current_is_product:
            self.extract_links(html, url)

    def finish_with_excel(self, partial: bool = False) -> None:
        with self.data_lock:
            if self.excel_finalized:
                return
            self.excel_finalized = True

        rows = self.snapshot_results()
        counts = self.snapshot_counts()
        filename = create_export_file(rows, self.project)
        if self.project is not None:
            delete_project_csv_for_project(self.project, keep_filename=filename.name)
        final_error = ""
        if partial:
            final_error = "Сбор приостановлен. CSV сформирован по уже найденным товарам."
        elif not self.results:
            final_error = (
                "Сбор завершен, но товары не найдены. Проверьте стартовый URL и исключения; "
                "для защищенных страниц убедитесь, что Botasaurus установился через run.ps1."
            )

        self.update_state(
            status="partial" if partial else "completed",
            percent=100 if not partial else int((self.project or {}).get("state", {}).get("percent", 0) or 0),
            currenturl="",
            totalprocessed=counts["visited"],
            processed_products=counts["results"],
            found_products=counts["results"],
            skipped=counts["skipped"],
            download_ready=True,
            download_url="/download",
            filename=filename.name,
            error=final_error,
            thread_count=self.thread_count,
            elapsed_seconds=int(self.elapsed_seconds()),
            eta_seconds=None,
            finished_at=now_iso() if not partial else "",
            paused_with_result=partial,
        )
        self.log(f"CSV сформирован: {filename.name}. Товаров: {counts['results']}", "success")
        if self.project is not None:
            save_projects()

    def run(self, resume: bool = False) -> None:
        if not self.started_at:
            self.started_at = time.time()
        self.update_state(
            status="running",
            thread_count=self.thread_count,
            started_at=(self.project or {}).get("state", {}).get("started_at") or now_iso(),
            paused_with_result=False,
        )
        self.log("Сбор продолжен" if resume else "Сбор запущен", "info")
        if not resume:
            for start_url in self.start_urls:
                self.enqueue(start_url)

        executor = ThreadPoolExecutor(max_workers=self.thread_count)
        pending = {}
        pending_urls_to_requeue = []
        self.note_progress_activity([])

        try:
            while not self.stop_signal.is_set():
                while len(pending) < self.thread_count:
                    try:
                        url = self.queue.get_nowait()
                    except Empty:
                        break

                    self.queued.discard(url)
                    if url in self.visited or url in self.in_progress:
                        continue
                    with self.data_lock:
                        if url in self.result_urls and self.is_product_url(url):
                            continue
                    self.in_progress.add(url)
                    self.update_state(
                        currenturl=url,
                        active_urls=sorted(self.in_progress)[:8],
                        active_tasks=len(self.in_progress),
                        queue_size=self.queue.qsize(),
                    )
                    pending[executor.submit(self.fetch, url)] = url
                    self.note_progress_activity(pending.values())
                    time.sleep(REQUEST_DELAY_SECONDS)

                if not pending:
                    if self.queue.empty():
                        break
                    continue

                done, _pending = wait(pending.keys(), timeout=0.5, return_when=FIRST_COMPLETED)
                if not done:
                    self.refresh_progress()
                    self.note_progress_activity(pending.values())
                    if pending and time.time() - self.last_progress_at >= NEWS_SCAN_STALL_TIMEOUT:
                        self.mark_stalled(pending.values())
                    continue

                for future in done:
                    url = pending.pop(future)
                    self.note_progress_activity(pending.values())
                    with self.data_lock:
                        self.in_progress.discard(url)
                        self.visited.add(url)
                    html = None
                    try:
                        html = future.result()
                    except Exception as exc:  # noqa: BLE001 - ошибку показываем в интерфейсе.
                        self.update_state(error=f"Ошибка обработки {url}: {exc}")
                        self.log(f"Ошибка обработки {url}: {exc}", "error")

                    if html and not self.stop_signal.is_set():
                        self.process_page(url, html)
                    elif not self.stop_signal.is_set():
                        with self.data_lock:
                            permanent_failure = url in self.permanent_failures
                        if permanent_failure:
                            self.log(f"URL пропущен без повторов: {url}", "warning")
                            self.refresh_progress(url)
                            continue
                        retry_count = self.failed_attempts.get(url, 0) + 1
                        self.failed_attempts[url] = retry_count
                        if retry_count <= 2:
                            with self.data_lock:
                                self.visited.discard(url)
                            self.enqueue(url, force=True)
                            self.log(f"Повторная попытка загрузки {retry_count}/2: {url}", "warning")
                        else:
                            self.log(f"URL пропущен после повторных попыток загрузки: {url}", "error")

                    self.refresh_progress(url)
        finally:
            if self.stop_signal.is_set():
                pending_urls_to_requeue = list(pending.values())
                self.requeue_pending(pending_urls_to_requeue)
            executor.shutdown(wait=False, cancel_futures=True)
            self.close_browser_sessions()

        if self.stop_signal.is_set():
            self.elapsed_before_resume = self.elapsed_seconds()
            self.started_at = 0.0
            if self.fatal_error:
                self.update_state(
                    status="error",
                    currenturl="",
                    active_urls=[],
                    active_tasks=0,
                    queue_size=self.queue.qsize(),
                    elapsed_seconds=int(self.elapsed_before_resume),
                    eta_seconds=None,
                    error=self.fatal_error,
                )
                return
            stop_mode = str((self.project or {}).get("stop_mode") or "")
            if self.finish_signal.is_set():
                self.finish_with_excel(partial=True)
            elif stop_mode == "pause":
                self.update_state(
                    status="paused",
                    currenturl="",
                    elapsed_seconds=int(self.elapsed_before_resume),
                    eta_seconds=None,
                    error="Сбор на паузе",
                )
                if (self.project or {}).get("state", {}).get("status") == "paused":
                    self.log("Сбор поставлен на паузу", "warning")
            else:
                self.update_state(
                    status="idle",
                    currenturl="",
                    active_urls=[],
                    active_tasks=0,
                    queue_size=0,
                    elapsed_seconds=int(self.elapsed_before_resume),
                    eta_seconds=None,
                    error="",
                    paused_with_result=False,
                )
                if stop_mode == "stop":
                    self.log("Сбор остановлен", "warning")
            return

        self.elapsed_before_resume = self.elapsed_seconds()
        self.started_at = 0.0
        self.finish_with_excel()


def safe_filename(value: str) -> str:
    value = output_text(value)
    cleaned = re.sub(r"[^A-Za-z\u0400-\u04FF0-9_-]+", "_", value, flags=re.IGNORECASE).strip("_")
    return cleaned or "project"


FILE_IMPORT_ALLOWED_SUFFIXES = {".csv", ".xlsx", ".xls"}
FILE_IMPORT_ACTIVE_STATUSES = {"queued", "running"}


class FileImportStopped(Exception):
    pass


def make_file_import_state(status: str = "idle") -> Dict[str, object]:
    return {
        "status": status,
        "stage": "",
        "percent": 0,
        "current_row": 0,
        "total_rows": 0,
        "processed_rows": 0,
        "excluded_rows": 0,
        "found_rows": 0,
        "missing_rows": 0,
        "model_not_found_rows": 0,
        "error": "",
        "started_at": "",
        "finished_at": "",
        "elapsed_seconds": 0,
        "result_filename": "",
    }


def normalize_file_import_state(value: object) -> Dict[str, object]:
    state = {**make_file_import_state(), **(value if isinstance(value, dict) else {})}
    state["status"] = str(state.get("status") or "idle")
    if state["status"] == "stopping":
        state["status"] = "stopped"
        state["stage"] = "Остановлено"
    for field in (
        "percent",
        "current_row",
        "total_rows",
        "processed_rows",
        "excluded_rows",
        "found_rows",
        "missing_rows",
        "model_not_found_rows",
        "elapsed_seconds",
    ):
        try:
            state[field] = int(float(state.get(field) or 0))
        except (TypeError, ValueError):
            state[field] = 0
    state["percent"] = max(0, min(100, state["percent"]))
    return state


def is_file_import_active_state(state: object) -> bool:
    return str((state if isinstance(state, dict) else {}).get("status") or "") in FILE_IMPORT_ACTIVE_STATUSES


def file_import_path_for_row(row: FileImport) -> Optional[Path]:
    file_meta = row.file if isinstance(row.file, dict) else {}
    filename = str(file_meta.get("stored_filename") or "").strip()
    if not filename:
        return None
    base_dir = FILE_IMPORT_DIR.resolve()
    path = (FILE_IMPORT_DIR / filename).resolve()
    if base_dir not in path.parents or not path.exists() or not path.is_file():
        return None
    return path


def update_file_import_state(db_session, **kwargs: object) -> Dict[str, object]:
    row = get_file_import_row(db_session)
    state = normalize_file_import_state(getattr(row, "state", {}) or {})
    state.update(kwargs)
    row.state = normalize_file_import_state(state)
    db_session.flush()
    return dict(row.state)


def stop_file_import_if_requested(stop_event: Optional[threading.Event]) -> None:
    if stop_event and stop_event.is_set():
        raise FileImportStopped()


def clear_file_import_storage() -> None:
    FILE_IMPORT_DIR.mkdir(parents=True, exist_ok=True)
    for path in FILE_IMPORT_DIR.iterdir():
        if path.is_file():
            try:
                path.unlink()
            except OSError:
                continue


def stored_file_import_files() -> List[Path]:
    FILE_IMPORT_DIR.mkdir(parents=True, exist_ok=True)
    return sorted(
        [
            path
            for path in FILE_IMPORT_DIR.iterdir()
            if path.is_file() and path.suffix.lower() in FILE_IMPORT_ALLOWED_SUFFIXES
        ],
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    )


def get_file_import_row(db_session=None) -> FileImport:
    db = db_session or g.db
    row = db.get(FileImport, 1)
    if row is None:
        row = FileImport(
            id=1,
            exclusions=[],
            model_field="",
            replace_rules="",
            file={},
            state=make_file_import_state(),
        )
        db.add(row)
        db.flush()
    else:
        normalized_state = normalize_file_import_state(getattr(row, "state", {}) or {})
        if row.state != normalized_state:
            row.state = normalized_state
            db.flush()
        normalized_exclusions = normalize_file_import_exclusions(row.exclusions)
        if row.exclusions != normalized_exclusions:
            row.exclusions = normalized_exclusions
            db.flush()
        normalized_model_field = clean_text(str(getattr(row, "model_field", "") or ""))
        if row.model_field != normalized_model_field:
            row.model_field = normalized_model_field
            db.flush()
        normalized_replace_rules = normalize_file_import_rules_text(getattr(row, "replace_rules", ""))
        if row.replace_rules != normalized_replace_rules:
            row.replace_rules = normalized_replace_rules
            db.flush()
    if not isinstance(row.file, dict) or not row.file.get("stored_filename"):
        stored_files = stored_file_import_files()
        if stored_files:
            path = stored_files[0]
            row.file = {
                "original_filename": path.name.split("_", 3)[-1] if "_" in path.name else path.name,
                "stored_filename": path.name,
                "uploaded_at": datetime.fromtimestamp(path.stat().st_mtime, MSK_TZ).isoformat(timespec="seconds"),
            }
            db.flush()
    return row


def current_file_import_path() -> Optional[Path]:
    row = get_file_import_row()
    return file_import_path_for_row(row)


def resolve_file_import_export_path(value: str) -> Optional[Path]:
    filename = Path(str(value or "")).name
    if not filename:
        return None
    path = (EXPORT_DIR / filename).resolve()
    if EXPORT_DIR.resolve() not in path.parents or not path.exists() or not path.is_file():
        return None
    return path


def remove_file_import_export(row: FileImport) -> None:
    candidates = [str(getattr(row, "export_path", "") or "")]
    file_meta = row.file if isinstance(row.file, dict) else {}
    candidates.append(str(file_meta.get("result_filename") or ""))
    for candidate in candidates:
        path = resolve_file_import_export_path(candidate)
        if path:
            try:
                path.unlink()
            except OSError:
                pass
    original_filename = str(file_meta.get("original_filename") or file_meta.get("filename") or "").strip()
    if original_filename:
        prefix = f"Новинки_{safe_filename(Path(original_filename).stem or 'file')}_"
        for path in EXPORT_DIR.glob(f"{prefix}*.csv"):
            if path.is_file():
                try:
                    path.unlink()
                except OSError:
                    pass


def public_file_import_state() -> Dict[str, object]:
    row = get_file_import_row()
    path = file_import_path_for_row(row)
    file_meta = row.file if isinstance(row.file, dict) else {}
    exclusions = normalize_file_import_exclusions(row.exclusions)
    exclusions_text = "\n".join(exclusions)
    model_field = clean_text(str(row.model_field or ""))
    replace_rules = normalize_file_import_rules_text(row.replace_rules)
    state = normalize_file_import_state(getattr(row, "state", {}) or {})
    active = is_file_import_active_state(state)
    result_filename = Path(str(row.export_path or file_meta.get("result_filename") or state.get("result_filename") or "")).name
    result_ready = bool(result_filename and not active and resolve_file_import_export_path(result_filename))
    if not path:
        return {
            "file": None,
            "exclusions": exclusions_text,
            "exclusions_list": exclusions,
            "model_field": model_field,
            "replace_rules": replace_rules,
            "result_filename": result_filename,
            "result_ready": result_ready,
            "state": state,
        }
    stat = path.stat()
    return {
        "exclusions": exclusions_text,
        "exclusions_list": exclusions,
        "model_field": model_field,
        "replace_rules": replace_rules,
        "result_filename": result_filename,
        "result_ready": result_ready,
        "state": state,
        "file": {
            "filename": output_text(str(file_meta.get("original_filename") or path.name)),
            "stored_filename": path.name,
            "size": stat.st_size,
            "uploaded_at": str(file_meta.get("uploaded_at") or datetime.fromtimestamp(stat.st_mtime, MSK_TZ).isoformat(timespec="seconds")),
        }
    }


def decode_file_import_csv(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1251", "windows-1251"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def normalize_file_import_header(value: object) -> str:
    return clean_text(str(value or "")).casefold()


def file_import_column_index(headers: List[object], column_name: str) -> int:
    expected = normalize_file_import_header(column_name)
    if not expected:
        raise ValueError("Укажите название столбца модели")
    for index, header in enumerate(headers):
        if normalize_file_import_header(header) == expected:
            return index
    raise ValueError(f"Столбец модели не найден: {column_name}")


def file_import_optional_brand_index(headers: List[object]) -> Optional[int]:
    brand_names = {"brand", "бренд", "manufacturer", "производитель", "vendor", "марка"}
    for index, header in enumerate(headers):
        if normalize_file_import_header(header) in brand_names:
            return index
    return None


def read_file_import_rows(path: Path, model_field: str) -> List[Dict[str, object]]:
    if path.suffix.lower() == ".csv":
        text = decode_file_import_csv(path.read_bytes())
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,|\t,")
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = ";"
        rows = list(csv.reader(io.StringIO(text), dialect))
    elif path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        workbook.close()
    elif path.suffix.lower() == ".xls":
        try:
            import xlrd
        except ImportError as exc:
            raise ValueError("Для обработки XLS установите зависимость xlrd") from exc

        workbook = xlrd.open_workbook(str(path))
        sheet = workbook.sheet_by_index(0)
        rows = [sheet.row_values(row_index) for row_index in range(sheet.nrows)]
    else:
        raise ValueError("Можно обработать только CSV, XLS или XLSX")

    header_index = next((index for index, row in enumerate(rows) if any(clean_text(str(cell or "")) for cell in row)), None)
    if header_index is None:
        return []
    headers = rows[header_index]
    model_index = file_import_column_index(headers, model_field)
    brand_index = file_import_optional_brand_index(headers)
    result: List[Dict[str, object]] = []
    for row_number, row in enumerate(rows[header_index + 1:], start=header_index + 2):
        source = clean_text(str(row[model_index] if model_index < len(row) else ""))
        if not source:
            continue
        brand = clean_text(str(row[brand_index] if brand_index is not None and brand_index < len(row) else ""))
        result.append({"row_number": row_number, "name": source, "brand": brand})
    return result


def strip_file_import_model_special_chars(value: str) -> str:
    text = re.sub(r"[^0-9A-Za-zА-Яа-яЁё\s\-./\\|@+]+", " ", str(value or ""))
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"\s*([\-./\\|])\s*", r"\1", text)
    return clean_text(text)


def prepare_file_import_model(value: str, replace_rules: str) -> str:
    prepared = prepare_rule_model(
        str(value or ""),
        {"model_replace_rules": normalize_file_import_rules_text(replace_rules)},
    )
    return strip_file_import_model_special_chars(prepared)


VISUAL_MODEL_TRANSLATION = str.maketrans(
    {
        "А": "A",
        "В": "B",
        "Е": "E",
        "К": "K",
        "М": "M",
        "Н": "H",
        "О": "O",
        "Р": "P",
        "С": "C",
        "Т": "T",
        "Х": "X",
        "а": "A",
        "в": "B",
        "е": "E",
        "к": "K",
        "м": "M",
        "н": "H",
        "о": "O",
        "р": "P",
        "с": "C",
        "т": "T",
        "х": "X",
    }
)


def technical_clean_model_text(value: object, remove_brackets: bool = True) -> str:
    text = BeautifulSoup(str(value or ""), "html.parser").get_text(" ", strip=True)
    text = html_lib.unescape(text)
    text = text.replace("\xa0", " ").replace("\u2009", " ")
    text = re.sub(r"[–—−]", "-", text)
    text = re.sub(r"\s*([/_.+\-])\s*", r"\1", text)
    if remove_brackets:
        text = re.sub(r"\([^)]*\)|\[[^\]]*\]|\{[^}]*\}", " ", text)
    return clean_text(text)


def strip_file_import_non_model_phrases(value: str) -> str:
    text = clean_text(value)
    text = re.sub(r"(?<![A-Za-z0-9])side\s*[- ]\s*by\s*[- ]\s*side(?![A-Za-z0-9])", " ", text, flags=re.IGNORECASE)
    return clean_text(text)


def normalize_compare_key(value: object) -> str:
    text = technical_clean_model_text(value, remove_brackets=False)
    text = re.sub(r"[\"'`«»]", "", text)
    text = re.sub(r"\s+", " ", text).strip(" .,/\\_-+")
    return normalize_model_key(text)


def compact_compare_key(value: object) -> str:
    return re.sub(r"[\s./_+\-]+", "", normalize_compare_key(value))


def visual_compare_key(value: object) -> str:
    return compact_compare_key(str(value or "").translate(VISUAL_MODEL_TRANSLATION))


def compare_keys_for_value(value: object) -> Dict[str, str]:
    original = normalize_compare_key(value)
    compact = compact_compare_key(value)
    visual = visual_compare_key(value)
    keys = {
        "original": original,
        "normalized": original,
        "compact": compact,
        "visual": visual,
    }
    return {kind: key for kind, key in keys.items() if key}


def file_import_exclusion_matches(value: str, brand: str, exclusions: Iterable[str]) -> bool:
    haystack = f"{value} {brand}".casefold()
    visual_haystack = f"{value} {brand}".translate(VISUAL_MODEL_TRANSLATION).casefold()
    for exclusion in exclusions:
        pattern = clean_text(str(exclusion or ""))
        if not pattern:
            continue
        if pattern.casefold() in haystack or pattern.translate(VISUAL_MODEL_TRANSLATION).casefold() in visual_haystack:
            return True
    return False


def known_file_import_brands() -> List[str]:
    try:
        with session_scope() as session:
            rows = session.execute(select(Brand.name)).scalars().all()
    except Exception:
        rows = []
    brands = [clean_text(str(row or "")) for row in rows if clean_text(str(row or ""))]
    return sorted(set(brands), key=len, reverse=True)


def brand_match_pattern(brand: str) -> str:
    parts = [re.escape(part) for part in re.split(r"\s*&\s*|\s+", clean_text(brand)) if part]
    if not parts:
        return ""
    return r"\s*&\s*".join(parts) if "&" in brand else r"\s+".join(parts)


def find_brand_in_name(name: str, explicit_brand: str = "") -> str:
    if explicit_brand:
        return explicit_brand
    for brand in known_file_import_brands():
        pattern = brand_match_pattern(brand)
        if pattern and re.search(rf"(?<![A-Za-zА-Яа-яЁё0-9]){pattern}(?![A-Za-zА-Яа-яЁё0-9])", name, flags=re.IGNORECASE):
            return brand
    return ""


def tail_after_brand(name: str, brand: str) -> str:
    text = technical_clean_model_text(name)
    if not brand:
        return text
    pattern = brand_match_pattern(brand)
    match = re.search(pattern, text, flags=re.IGNORECASE) if pattern else None
    if match:
        before_brand = clean_text(text[:match.start()])
        if before_brand and any(model_signal_token(token) for token in before_brand.split()):
            return before_brand
        return clean_text(text[match.end():])
    visual_text = text.translate(VISUAL_MODEL_TRANSLATION)
    visual_brand = brand.translate(VISUAL_MODEL_TRANSLATION)
    visual_pattern = brand_match_pattern(visual_brand)
    visual_match = re.search(visual_pattern, visual_text, flags=re.IGNORECASE) if visual_pattern else None
    if visual_match:
        before_brand = clean_text(text[:visual_match.start()])
        if before_brand and any(model_signal_token(token) for token in before_brand.split()):
            return before_brand
        return clean_text(text[visual_match.end():])
    return text


def model_signal_token(token: str) -> bool:
    value = clean_text(token)
    if not value:
        return False
    if re.fullmatch(r"\d+(?:[.,]\d+)?(?:ВТ|BT|В|V|B|Л|МЛ|ML|КГ|KG|Г|G|СМ|CM|ММ|MM)(?:/\d+(?:[.,]\d+)?(?:ВТ|BT|В|V|B|Л|МЛ|ML|КГ|KG|Г|G|СМ|CM|ММ|MM))*", value.upper()):
        return False
    has_digit = bool(re.search(r"\d", value))
    has_latin = bool(re.search(r"[A-Za-z]", value))
    has_cyrillic = bool(re.search(r"[А-Яа-яЁё]", value))
    return has_latin or (has_digit and has_cyrillic) or bool(re.search(r"[A-Za-zА-Яа-яЁё]\d|\d[A-Za-zА-Яа-яЁё]", value))


def visual_model_suffix_token(token: str) -> bool:
    value = clean_text(token)
    return bool(value and re.fullmatch(r"[A-Za-zА-Яа-яЁё]{1,3}", value) and value.translate(VISUAL_MODEL_TRANSLATION) != value)


def candidate_until_russian_description(value: str) -> str:
    tokens = value.split()
    kept: List[str] = []
    seen_signal = False
    for token in tokens:
        if seen_signal and re.fullmatch(r"\d+(?:[.,]\d+)?(?:ВТ|W|BT|В|V|B|Л|L|МЛ|ML|КГ|KG|Г|G|СМ|CM|ММ|MM)(?:/\d+(?:[.,]\d+)?(?:ВТ|W|BT|В|V|B|Л|L|МЛ|ML|КГ|KG|Г|G|СМ|CM|ММ|MM))*", token.upper()):
            break
        if model_signal_token(token):
            kept.append(token)
            seen_signal = True
            continue
        if seen_signal and visual_model_suffix_token(token):
            kept.append(token)
            continue
        if seen_signal and re.fullmatch(r"[А-Яа-яЁё][А-Яа-яЁё/\-]*", token):
            break
        kept.append(token)
    return clean_text(" ".join(kept))


def first_model_block(value: str) -> str:
    tokens = re.findall(r"[A-Za-zА-Яа-яЁё0-9]+(?:[./_+\-][A-Za-zА-Яа-яЁё0-9]+)*", value)
    best: List[str] = []
    current: List[str] = []
    for token in tokens:
        if model_signal_token(token):
            current.append(token)
            continue
        if current:
            if len(" ".join(current)) > len(" ".join(best)):
                best = list(current)
            current = []
    if current and len(" ".join(current)) > len(" ".join(best)):
        best = current
    return clean_text(" ".join(best))


def normalize_candidate_display(value: str) -> str:
    text = technical_clean_model_text(value)
    text = re.sub(r"\s+", " ", text).strip(" .,/\\_-+")
    tokens = text.split()
    has_digit = any(char.isdigit() for char in text)
    normalized_tokens = [
        token.translate(VISUAL_MODEL_TRANSLATION).upper()
        if any(char.isdigit() for char in token) or (has_digit and re.fullmatch(r"[A-Za-zА-Яа-яЁё]{1,3}", token))
        else token
        for token in tokens
    ]
    return " ".join(normalized_tokens)


def add_model_candidate(candidates: List[str], value: str) -> None:
    candidate = normalize_candidate_display(value)
    if not candidate:
        return
    if re.match(r"^[А-Яа-яЁё]", candidate):
        return
    if re.search(r"[А-Яа-яЁё]", candidate):
        return
    if re.fullmatch(r"\d+(?:[.,]\d+)?(?:ВТ|W|BT|В|V|B|Л|L|МЛ|ML|КГ|KG|Г|G|СМ|CM|ММ|MM)(?:/\d+(?:[.,]\d+)?(?:ВТ|W|BT|В|V|B|Л|L|МЛ|ML|КГ|KG|Г|G|СМ|CM|ММ|MM))*", candidate.upper()):
        return
    key = normalize_compare_key(candidate)
    if len(key) < 2:
        return
    if key not in {normalize_compare_key(item) for item in candidates}:
        candidates.append(candidate)


def code_model_tokens(value: str) -> List[str]:
    result: List[str] = []
    for token in re.findall(r"[A-Za-zА-Яа-яЁё0-9]+(?:[./_+\-][A-Za-zА-Яа-яЁё0-9]+)*", value):
        if not model_signal_token(token):
            continue
        normalized = normalize_candidate_display(token)
        if not normalized or re.match(r"^[А-Яа-яЁё]", normalized):
            continue
        if not any(char.isdigit() for char in normalized):
            continue
        if normalized not in result:
            result.append(normalized)
    return result


def generate_model_candidates(name: str, brand: str = "") -> List[str]:
    cleaned = strip_file_import_non_model_phrases(technical_clean_model_text(name))
    detected_brand = find_brand_in_name(cleaned, brand)
    tail = tail_after_brand(cleaned, detected_brand)
    main_part = re.split(r"[,;|]", tail, maxsplit=1)[0]
    before_russian_description = candidate_until_russian_description(main_part)
    strong_block = first_model_block(main_part)

    candidates: List[str] = []
    for value in (before_russian_description, strong_block, main_part, tail):
        add_model_candidate(candidates, value)
    for token in code_model_tokens(main_part):
        add_model_candidate(candidates, token)

    base = before_russian_description or main_part or strong_block
    tokens = base.split()
    for start in range(1, min(4, len(tokens))):
        value = " ".join(tokens[start:])
        if any(char.isdigit() for char in value):
            add_model_candidate(candidates, value)
    for size in range(min(4, len(tokens)), 0, -1):
        value = " ".join(tokens[:size])
        if any(char.isdigit() for char in value):
            add_model_candidate(candidates, value)
    return candidates[:8]


def feed_index_add(index: Dict[str, Dict[str, object]], key: str, item: Dict[str, object]) -> None:
    if key and key not in index:
        index[key] = item


def index_feed_value(index: Dict[str, Dict[str, object]], value: str, item: Dict[str, object]) -> None:
    for key in compare_keys_for_value(value).values():
        feed_index_add(index, key, item)


def build_feed_index_from_xml(content: bytes, feed: Dict[str, object], stop_event: Optional[threading.Event] = None) -> Dict[str, Dict[str, object]]:
    index: Dict[str, Dict[str, object]] = {}
    for node_index, (_event, node) in enumerate(ET.iterparse(io.BytesIO(content), events=("end",)), start=1):
        if node_index % 500 == 0:
            stop_file_import_if_requested(stop_event)
        children = list(node)
        if not children:
            continue
        values: Dict[str, str] = {}
        for child in children:
            key = str(child.tag).split("}")[-1].lower()
            values[key] = clean_text(child.text or "")
        explicit_values = [
            values.get(key, "")
            for key in ("vendorcode", "vendor_code", "model", "sku", "article", "articul")
            if values.get(key)
        ]
        name = values.get("name") or values.get("title") or ""
        item = {
            "source": str(feed.get("source") or ""),
            "source_label": str(feed.get("source_label") or feed.get("url") or "Фид"),
            "feed_name": name,
            "feed_url": values.get("url") or "",
            "raw": values,
        }
        for value in explicit_values:
            index_feed_value(index, value, {**item, "matched_feed_key": normalize_compare_key(value)})
        if not explicit_values and name:
            for candidate in generate_model_candidates(name):
                index_feed_value(index, candidate, {**item, "matched_feed_key": normalize_compare_key(candidate)})
        node.clear()
    return index


def build_file_import_feed_indexes(stop_event: Optional[threading.Event] = None) -> List[Dict[str, object]]:
    downloaded_feeds = download_feed_files(stop_event=stop_event)
    feed_indexes: List[Dict[str, object]] = []
    for feed in downloaded_feeds:
        stop_file_import_if_requested(stop_event)
        path = feed_snapshot_path(feed)
        try:
            if path is None:
                raise FileNotFoundError("Не задан путь к snapshot фида")
            index = build_feed_index_from_xml(path.read_bytes(), feed, stop_event=stop_event)
            feed_indexes.append({**feed, "index": index, "codes_count": len(index)})
        except FileImportStopped:
            raise
        except Exception as exc:
            feed_indexes.append({**feed, "index": {}, "codes_count": 0, "error": str(exc)})
    with news_lock:
        news_settings["feed_storage"] = [
            {key: value for key, value in feed.items() if key != "index"}
            for feed in feed_indexes
        ]
        save_news_settings()
    save_logs()
    return feed_indexes


def match_candidates_against_feed_indexes(
    candidates: List[str],
    feed_indexes: List[Dict[str, object]],
    stop_event: Optional[threading.Event] = None,
) -> Optional[Dict[str, object]]:
    for candidate in candidates:
        stop_file_import_if_requested(stop_event)
        keys = compare_keys_for_value(candidate)
        for reason, key in keys.items():
            for feed in feed_indexes:
                stop_file_import_if_requested(stop_event)
                index = feed.get("index", {})
                if isinstance(index, dict) and key in index:
                    match = dict(index[key])
                    match.update(
                        {
                            "selected_model": candidate,
                            "selected_reason": f"matched:{reason}",
                            "compare_key": key,
                            "matched_feed_key": str(match.get("matched_feed_key") or key),
                        }
                    )
                    return match
    return None


def missing_feed_labels(
    candidates: List[str],
    feed_indexes: List[Dict[str, object]],
    stop_event: Optional[threading.Event] = None,
) -> List[str]:
    labels: List[str] = []
    keys = set()
    for candidate in candidates:
        stop_file_import_if_requested(stop_event)
        keys.update(compare_keys_for_value(candidate).values())
    for feed in feed_indexes:
        stop_file_import_if_requested(stop_event)
        index = feed.get("index", {})
        if not isinstance(index, dict) or not (keys & set(index.keys())):
            labels.append(str(feed.get("source_label") or feed.get("url") or "Фид"))
    return labels


def file_import_result_filename(original_filename: str) -> str:
    stem = safe_filename(Path(original_filename).stem or "file")
    return f"Новинки_{stem}_{datetime.now(MSK_TZ).strftime('%d-%m-%Y_%H-%M-%S')}.csv"


def compare_file_import_with_feeds(db_session=None, stop_event: Optional[threading.Event] = None) -> Dict[str, object]:
    db = db_session or g.db
    row = get_file_import_row(db)
    path = file_import_path_for_row(row)
    if not path:
        raise ValueError("Файл не загружен")
    model_field = clean_text(str(row.model_field or ""))
    if not model_field:
        raise ValueError("Укажите название столбца модели")

    exclusions = normalize_file_import_exclusions(row.exclusions)
    replace_rules = normalize_file_import_rules_text(row.replace_rules)
    started_at = time.time()
    update_file_import_state(
        db,
        status="running",
        stage="Читаю файл",
        percent=2,
        error="",
        started_at=datetime.now(MSK_TZ).isoformat(timespec="seconds"),
        finished_at="",
        elapsed_seconds=0,
        result_filename="",
    )
    db.commit()
    stop_file_import_if_requested(stop_event)
    source_rows = read_file_import_rows(path, model_field)
    total_rows = len(source_rows)
    update_file_import_state(db, stage="Загружаю фиды", percent=8, total_rows=total_rows)
    db.commit()
    stop_file_import_if_requested(stop_event)
    feed_indexes = build_file_import_feed_indexes(stop_event=stop_event)
    stop_file_import_if_requested(stop_event)
    update_file_import_state(db, stage="Сравниваю строки", percent=10, total_rows=total_rows)
    db.commit()

    result_rows: List[Dict[str, object]] = []
    processed = excluded = found = missing = empty_model = 0
    last_progress_time = 0.0
    last_progress_index = 0
    for index, item in enumerate(source_rows, start=1):
        if stop_event and stop_event.is_set():
            state = update_file_import_state(
                db,
                status="stopped",
                stage="Остановлено",
                percent=max(0, min(99, int((index - 1) / max(total_rows, 1) * 90) + 10)),
                current_row=index - 1,
                processed_rows=processed,
                excluded_rows=excluded,
                found_rows=found,
                missing_rows=missing,
                model_not_found_rows=empty_model,
                finished_at=datetime.now(MSK_TZ).isoformat(timespec="seconds"),
                elapsed_seconds=int(time.time() - started_at),
            )
            db.commit()
            return {
                "stopped": True,
                "total_rows": total_rows,
                "processed_rows": processed,
                "excluded_rows": excluded,
                "model_not_found_rows": empty_model,
                "found_rows": found,
                "missing_rows": missing,
                "state": state,
            }
        name = str(item.get("name") or "")
        brand = str(item.get("brand") or "")
        if file_import_exclusion_matches(name, brand, exclusions):
            excluded += 1
        else:
            processed += 1
            prepared_model = prepare_file_import_model(name, replace_rules)
            candidates = generate_model_candidates(prepared_model, brand)
            if not candidates:
                empty_model += 1
                result_rows.append(
                    {
                        "row": item.get("row_number"),
                        "name": name,
                        "brand": brand,
                        "model_candidates": "",
                        "selected_model": "",
                        "missing_on": "",
                    }
                )
            else:
                match = match_candidates_against_feed_indexes(candidates, feed_indexes, stop_event=stop_event)
                if match:
                    found += 1
                else:
                    missing += 1
                    selected_model = candidates[0]
                    result_rows.append(
                        {
                            "row": item.get("row_number"),
                            "name": name,
                            "brand": brand,
                            "model_candidates": " | ".join(candidates),
                            "selected_model": selected_model,
                            "missing_on": ", ".join(missing_feed_labels(candidates, feed_indexes, stop_event=stop_event)),
                        }
                    )
        now = time.time()
        if index == total_rows or index - last_progress_index >= 10 or now - last_progress_time >= 1.0:
            last_progress_time = now
            last_progress_index = index
            update_file_import_state(
                db,
                stage="Сравниваю строки",
                percent=10 + int(index / max(total_rows, 1) * 85),
                current_row=index,
                processed_rows=processed,
                excluded_rows=excluded,
                found_rows=found,
                missing_rows=missing,
                model_not_found_rows=empty_model,
                elapsed_seconds=int(now - started_at),
            )
            db.commit()

    file_meta = dict(row.file) if isinstance(row.file, dict) else {}
    original_filename = str(file_meta.get("original_filename") or file_meta.get("filename") or path.name)
    remove_file_import_export(row)
    result_path = EXPORT_DIR / file_import_result_filename(original_filename)
    update_file_import_state(db, stage="Записываю CSV", percent=98)
    db.commit()
    stop_file_import_if_requested(stop_event)
    with result_path.open("w", encoding="utf-8-sig", newline="") as csv_file:
        fieldnames = [
            "row",
            "name",
            "brand",
            "model_candidates",
            "selected_model",
            "missing_on",
        ]
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames, delimiter=";")
        writer.writeheader()
        writer.writerows(result_rows)

    file_meta["result_filename"] = result_path.name
    file_meta["result_created_at"] = datetime.now(MSK_TZ).isoformat(timespec="seconds")
    row.export_path = result_path.name
    row.file = file_meta
    state = update_file_import_state(
        db,
        status="completed",
        stage="Готово",
        percent=100,
        current_row=total_rows,
        total_rows=total_rows,
        processed_rows=processed,
        excluded_rows=excluded,
        found_rows=found,
        missing_rows=missing,
        model_not_found_rows=empty_model,
        finished_at=datetime.now(MSK_TZ).isoformat(timespec="seconds"),
        elapsed_seconds=int(time.time() - started_at),
        result_filename=result_path.name,
        error="",
    )
    db.commit()
    return {
        "total_rows": len(source_rows),
        "processed_rows": processed,
        "excluded_rows": excluded,
        "model_not_found_rows": empty_model,
        "found_rows": found,
        "missing_rows": missing,
        "result_filename": result_path.name,
        "result_url": "/api/file-import/download",
        "state": state,
    }


def run_file_import_compare(stop_event: threading.Event) -> None:
    global file_import_worker_thread
    db = SessionLocal()
    try:
        compare_file_import_with_feeds(db, stop_event=stop_event)
    except FileImportStopped:
        db.rollback()
        try:
            update_file_import_state(
                db,
                status="stopped",
                stage="Остановлено",
                finished_at=datetime.now(MSK_TZ).isoformat(timespec="seconds"),
            )
            db.commit()
        except Exception:
            db.rollback()
    except Exception as exc:
        db.rollback()
        try:
            update_file_import_state(
                db,
                status="error",
                stage="Ошибка",
                error=str(exc),
                finished_at=datetime.now(MSK_TZ).isoformat(timespec="seconds"),
            )
            db.commit()
        except Exception:
            db.rollback()
    finally:
        db.close()
        with file_import_lock:
            if threading.current_thread() is file_import_worker_thread:
                file_import_worker_thread = None
                file_import_stop_event.clear()


def recover_interrupted_file_import_scan() -> None:
    with session_scope() as db_session:
        row = get_file_import_row(db_session)
        state = normalize_file_import_state(getattr(row, "state", {}) or {})
        if not is_file_import_active_state(state):
            return
        row.state = {
            **state,
            "status": "error",
            "stage": "Прервано",
            "error": "Сравнение файла было прервано перезапуском сервера. Запустите его снова.",
            "finished_at": datetime.now(MSK_TZ).isoformat(timespec="seconds"),
        }


def start_file_import_compare() -> Dict[str, object]:
    global file_import_worker_thread
    with file_import_lock:
        if file_import_worker_thread and file_import_worker_thread.is_alive():
            return public_file_import_state()
        row = get_file_import_row()
        state = normalize_file_import_state(getattr(row, "state", {}) or {})
        if is_file_import_active_state(state):
            return public_file_import_state()
        path = file_import_path_for_row(row)
        if not path:
            raise ValueError("Файл не загружен")
        if not clean_text(str(row.model_field or "")):
            raise ValueError("Укажите название столбца модели")

        remove_file_import_export(row)
        file_meta = dict(row.file) if isinstance(row.file, dict) else {}
        file_meta.pop("result_filename", None)
        file_meta.pop("result_created_at", None)
        row.file = file_meta
        row.export_path = ""
        row.state = {
            **make_file_import_state("queued"),
            "stage": "В очереди",
            "started_at": datetime.now(MSK_TZ).isoformat(timespec="seconds"),
        }
        g.db.commit()

        file_import_stop_event.clear()
        file_import_worker_thread = threading.Thread(
            target=run_file_import_compare,
            args=(file_import_stop_event,),
            daemon=True,
        )
        file_import_worker_thread.start()
        return public_file_import_state()


def create_export_file(rows: List[Dict[str, str]], project: Optional[Dict[str, object]] = None) -> Path:
    if project:
        filename = project_csv_filename(project)
    else:
        filename = f"export_{datetime.now().strftime('%d-%m-%Y_%H-%M-%S')}.csv"
    path = EXPORT_DIR / filename

    with path.open("w", encoding="utf-8-sig", newline="") as csv_file:
        writer = csv.writer(csv_file, delimiter=";")
        writer.writerow(["URL товара", "_MODEL_", "_PRICE_"])
        for row in rows:
            writer.writerow([output_text(row.get("url", "")), output_text(row.get("model", "")), output_text(row.get("price", ""))])

    return path


class CollectOnlyCrawler(ProductSiteCrawler):
    def __init__(self, *args, progress_callback=None, **kwargs):
        kwargs.setdefault("allow_empty_price", True)
        super().__init__(*args, **kwargs)
        self.progress_callback = progress_callback

    def update_state(self, **kwargs: object) -> None:
        if self.progress_callback:
            self.progress_callback(kwargs)

    def log(self, message: str, level: str = "info") -> None:
        if self.progress_callback:
            self.progress_callback(
                {
                    "log_message": message,
                    "log_level": level,
                }
            )

    def finish_with_excel(self, partial: bool = False) -> None:
        with self.data_lock:
            self.excel_finalized = True


def text_by_selector(soup: BeautifulSoup, selector: str) -> str:
    if not selector:
        return ""
    try:
        return first_by_selector(soup, selector)
    except Exception:
        return ""


def extract_availability(soup: BeautifulSoup, selector: str = "") -> str:
    selected = text_by_selector(soup, selector)
    if selected:
        return selected
    page_text = clean_text(soup.get_text(" ", strip=True))
    patterns = [
        r"В наличии",
        r"Нет в наличии",
        r"Под заказ",
        r"Ожидается",
        r"Сообщить о поступлении",
        r"available",
        r"out of stock",
        r"in stock",
    ]
    for pattern in patterns:
        match = re.search(pattern, page_text, flags=re.IGNORECASE)
        if match:
            return clean_text(match.group(0))
    return ""


def availability_is_excluded(availability: str, rules: object) -> bool:
    """Проверяет статус наличия по построчным правилам исключения до сравнения с фидами."""
    status = clean_text(availability).lower()
    if not status:
        return False
    for rule in normalize_patterns(rules):
        normalized_rule = clean_text(rule).lower()
        if normalized_rule and normalized_rule in status:
            return True
    return False


def extract_product_name(soup: BeautifulSoup, selector: str = "") -> str:
    selected = text_by_selector(soup, selector)
    if selected:
        return selected
    meta = soup.select_one("meta[property='og:title'], meta[name='twitter:title']")
    if meta and meta.get("content"):
        return clean_text(meta.get("content", ""))
    return first_text(soup, ["h1", "[itemprop='name']", ".product-title", ".product__title"])


def feed_source_key(url: str) -> str:
    hostname = urlparse(url).hostname or "feed"
    return safe_filename(hostname.lower().removeprefix("www."))


def feed_source_label(url: str) -> str:
    hostname = (urlparse(url).hostname or "").lower().removeprefix("www.")
    if "mega-kuhnya" in hostname:
        return "Мега-кухня"
    if "vsya-tehnika" in hostname:
        return "Вся техника"
    return hostname or "Фид"


def source_feed_dir(source: str) -> Path:
    return FEED_DIR / safe_filename(source)


def feed_snapshots_dir() -> Path:
    return FEED_DIR / ".snapshots"


def feed_snapshot_path(feed: Dict[str, object]) -> Optional[Path]:
    snapshot_value = str(feed.get("snapshot") or "").strip()
    source_value = str(feed.get("source") or "").strip()
    filename = Path(str(feed.get("filename") or "")).name
    if not source_value or not filename:
        return None
    if not snapshot_value:
        # Backward compatibility for feed metadata written before snapshots.
        legacy_root = source_feed_dir(source_value).resolve()
        legacy_path = (legacy_root / filename).resolve()
        return legacy_path if legacy_root in legacy_path.parents else None
    snapshot = safe_filename(snapshot_value)
    source = safe_filename(source_value)
    path = (feed_snapshots_dir() / snapshot / source / filename).resolve()
    snapshot_root = (feed_snapshots_dir() / snapshot).resolve()
    return path if snapshot_root in path.parents else None


def local_feed_filename(kind: str, index: int, url: str) -> str:
    parsed = urlparse(url)
    raw_name = Path(parsed.path).name or kind
    stem = Path(raw_name).stem or kind
    suffix = Path(raw_name).suffix.lower()
    if suffix not in {".xml", ".yml"}:
        suffix = ".xml"
    return f"{index:02d}_{safe_filename(kind)}_{safe_filename(stem)}{suffix}"


def generation_file_url(url: str) -> str:
    parsed = urlparse(url)
    query = dict(parse_qsl(parsed.query, keep_blank_values=True))
    query["cron"] = "file"
    return urlunparse(parsed._replace(query=urlencode(query)))


def clear_source_feeds(source: str) -> Path:
    feed_dir = source_feed_dir(source)
    if feed_dir.exists():
        for path in feed_dir.glob("*"):
            if path.is_file():
                path.unlink(missing_ok=True)
    feed_dir.mkdir(parents=True, exist_ok=True)
    return feed_dir


def make_feed_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
            )
        }
    )
    return session


def trigger_feed_generation(generate_url: str) -> None:
    try:
        response = make_feed_session().get(generation_file_url(generate_url), timeout=60)
        response.raise_for_status()
    except Exception:
        pass


def download_feed_site(index: int, site: Dict[str, str], snapshot_dir: Path, snapshot: str) -> Optional[Dict[str, object]]:
    url = site["feed_url"]
    try:
        response = make_feed_session().get(url, timeout=60)
        response.raise_for_status()
        source = feed_source_key(url)
        feed_dir = snapshot_dir / source
        feed_dir.mkdir(parents=True, exist_ok=True)
        filename = local_feed_filename("feed", index, url)
        path = feed_dir / filename
        temporary_path = path.with_suffix(path.suffix + ".part")
        temporary_path.write_bytes(response.content)
        os.replace(temporary_path, path)
        return {
            "kind": "feed",
            "source": source,
            "source_label": site.get("name") or feed_source_label(url),
            "url": url,
            "filename": filename,
            "snapshot": snapshot,
            "size": path.stat().st_size,
            "downloaded_at": datetime.now(MSK_TZ).isoformat(timespec="seconds"),
        }
    except Exception:
        return None


def cleanup_feed_snapshots() -> None:
    snapshots_dir = feed_snapshots_dir()
    if not snapshots_dir.exists():
        return
    try:
        snapshots = sorted(
            [path for path in snapshots_dir.iterdir() if path.is_dir()],
            key=lambda path: path.stat().st_mtime,
            reverse=True,
        )
    except OSError:
        return
    for path in snapshots[FEED_SNAPSHOT_RETAIN:]:
        try:
            shutil.rmtree(path)
        except OSError:
            # Windows can keep a file open while it is being downloaded. Keep this
            # old immutable snapshot and retry cleanup after a future refresh.
            continue


def wait_feed_futures(futures, stop_event: Optional[threading.Event]) -> List[object]:
    pending = set(futures)
    results: List[object] = []
    while pending:
        stop_file_import_if_requested(stop_event)
        completed, pending = wait(pending, timeout=0.5, return_when=FIRST_COMPLETED)
        if not completed:
            continue
        for future in completed:
            results.append(future.result())
    return results


def download_feed_files(stop_event: Optional[threading.Event] = None) -> List[Dict[str, object]]:
    with news_lock:
        own_sites = own_sites_from_settings(news_settings)
        feed_urls = [site["feed_url"] for site in own_sites]
        generate_urls = [site["feed_generate_url"] for site in own_sites if site.get("feed_generate_url")]
    signature = tuple((site["feed_url"], site.get("feed_generate_url", "")) for site in own_sites)
    with FEED_STORAGE_LOCK:
        cache_feeds = feed_snapshot_cache.get("feeds")
        cache_age = time.time() - float(feed_snapshot_cache.get("created_at") or 0.0)
        if (
            signature == feed_snapshot_cache.get("signature")
            and cache_age <= FEED_SNAPSHOT_CACHE_SECONDS
            and isinstance(cache_feeds, list)
            and cache_feeds
            and all((path := feed_snapshot_path(feed)) and path.exists() for feed in cache_feeds if isinstance(feed, dict))
        ):
            return [dict(feed) for feed in cache_feeds if isinstance(feed, dict)]

        stop_file_import_if_requested(stop_event)
        if generate_urls:
            executor = ThreadPoolExecutor(max_workers=min(FEED_WORKER_COUNT, len(generate_urls)))
            try:
                wait_feed_futures([executor.submit(trigger_feed_generation, url) for url in generate_urls], stop_event)
            except FileImportStopped:
                executor.shutdown(wait=False, cancel_futures=True)
                raise
            else:
                executor.shutdown(wait=True)

        snapshot = f"{int(time.time() * 1000)}_{uuid.uuid4().hex[:8]}"
        snapshot_dir = feed_snapshots_dir() / snapshot
        snapshot_dir.mkdir(parents=True, exist_ok=False)
        downloaded: List[Dict[str, object]] = []
        if own_sites:
            executor = ThreadPoolExecutor(max_workers=min(FEED_WORKER_COUNT, len(own_sites)))
            try:
                futures = [
                    executor.submit(download_feed_site, index, site, snapshot_dir, snapshot)
                    for index, site in enumerate(own_sites, start=1)
                ]
                for feed in wait_feed_futures(futures, stop_event):
                    if feed:
                        downloaded.append(feed)
            except FileImportStopped:
                executor.shutdown(wait=False, cancel_futures=True)
                try:
                    shutil.rmtree(snapshot_dir)
                except OSError:
                    pass
                raise
            else:
                executor.shutdown(wait=True)
        if not downloaded:
            try:
                shutil.rmtree(snapshot_dir)
            except OSError:
                pass
            return []

        feed_snapshot_cache.update(
            {"signature": signature, "created_at": time.time(), "feeds": [dict(feed) for feed in downloaded]}
        )
        cleanup_feed_snapshots()
        return downloaded


def fetch_existing_vendor_codes() -> tuple[Set[str], List[Dict[str, object]]]:
    codes, feeds, _feed_code_sets = fetch_existing_vendor_code_sets()
    return codes, feeds


def fetch_existing_vendor_code_sets() -> tuple[Set[str], List[Dict[str, object]], List[Dict[str, object]]]:
    downloaded_feeds = download_feed_files()
    codes: Set[str] = set()
    feeds: List[Dict[str, object]] = []
    feed_code_sets: List[Dict[str, object]] = []
    for feed in downloaded_feeds:
        path = feed_snapshot_path(feed)
        try:
            if path is None:
                raise FileNotFoundError("Не задан путь к snapshot фида")
            feed_codes = parse_vendor_codes_from_xml(path.read_bytes())
            codes.update(feed_codes)
            feeds.append({**feed, "codes_count": len(feed_codes)})
            feed_code_sets.append({**feed, "codes_count": len(feed_codes), "codes": feed_codes})
        except Exception as exc:
            feeds.append({**feed, "codes_count": 0, "error": str(exc)})
            feed_code_sets.append({**feed, "codes_count": 0, "codes": set(), "error": str(exc)})
    with news_lock:
        news_settings["feed_storage"] = feeds
        save_news_settings()
    save_logs()
    return codes, feeds, feed_code_sets


def product_compare_keys(product: Dict[str, str]) -> Set[str]:
    keys = {
        normalize_model_key(str(product.get("model", ""))),
        normalize_model_key(str(product.get("vendor_code", ""))),
    }
    keys.discard("")
    return keys


def build_missing_summary(new_items: List[Dict[str, str]], feed_code_sets: List[Dict[str, object]]) -> List[Dict[str, object]]:
    summary: List[Dict[str, object]] = []
    for feed in feed_code_sets:
        feed_codes = feed.get("codes", set())
        if not isinstance(feed_codes, set):
            feed_codes = set(feed_codes) if isinstance(feed_codes, list) else set()
        count = 0
        for item in new_items:
            keys = product_compare_keys(item)
            if keys and not (keys & feed_codes):
                count += 1
        summary.append(
            {
                "source": str(feed.get("source") or ""),
                "source_label": str(feed.get("source_label") or feed.get("url") or "Фид"),
                "url": str(feed.get("url") or ""),
                "count": count,
                "codes_count": int(feed.get("codes_count") or 0),
                "error": str(feed.get("error") or ""),
            }
        )
    return summary


def parse_vendor_codes_from_xml(content: bytes) -> Set[str]:
    codes: Set[str] = set()
    try:
        for _event, node in ET.iterparse(io.BytesIO(content), events=("end",)):
            children = list(node)
            if children:
                values: Dict[str, str] = {}
                for child in children:
                    key = str(child.tag).split("}")[-1].lower()
                    if key in {"vendorcode", "model", "name", "title"}:
                        values[key] = clean_text(child.text or "")
                vendor_code = normalize_model_key(values.get("vendorcode", ""))
                model = values.get("model") or values.get("name") or values.get("title") or vendor_code
                model_key = normalize_model_key(model)
                if vendor_code:
                    codes.add(vendor_code)
                if model_key:
                    codes.add(model_key)
                node.clear()
    except ET.ParseError:
        raise
    return codes


def parse_feed_products_from_xml(content: bytes) -> List[Dict[str, object]]:
    products: List[Dict[str, object]] = []
    root = ET.fromstring(content)
    for node in root.iter():
        children = list(node)
        if not children:
            continue
        values: Dict[str, str] = {}
        for child in children:
            key = str(child.tag).split("}")[-1].lower()
            values[key] = clean_text(child.text or "")
        vendor_code = normalize_model_key(values.get("vendorcode", ""))
        model = values.get("model") or values.get("name") or values.get("title") or vendor_code
        model_key = normalize_model_key(model)
        if not vendor_code and not model_key:
            continue
        products.append(
            {
                "vendor_code": vendor_code,
                "model_key": model_key,
                "name": values.get("name") or values.get("model") or values.get("title") or "",
                "url": values.get("url") or "",
                "raw": values,
            }
        )
    return products


def validate_monitor_selectors(monitor: Dict[str, object]) -> None:
    selector_fields = []
    rules = monitor.get("extraction_rules", {}) if isinstance(monitor.get("extraction_rules"), dict) else {}
    selectors = monitor.get("selector_settings", {}) if isinstance(monitor.get("selector_settings"), dict) else {}
    for key in ("product_card_selector", "product_url_selector", "model_selector", "price_selector"):
        if rules.get(key):
            selector_fields.append((key, str(rules[key])))
    for key in ("name_selector", "availability_selector"):
        if selectors.get(key):
            selector_fields.append((key, str(selectors[key])))
    soup = BeautifulSoup("", "html.parser")
    for key, selector in selector_fields:
        try:
            soup.select(selector)
        except Exception as exc:
            raise ValueError(f"Ошибка CSS-селектора {key}: {selector}. {exc}") from exc


def update_news_monitor_state(monitor: Dict[str, object], persist: bool = True, **kwargs: object) -> None:
    with news_lock:
        previous_state = dict(monitor.get("state", make_news_state()))
        state = dict(previous_state)
        current_status = str(state.get("status") or "")
        terminal_statuses = {"completed", "error", "partial", "idle", "stopped"}
        if state.get("finished_at") and current_status in terminal_statuses and kwargs.get("status") != current_status:
            return
        state.update(kwargs)
        previous_progress = monitor.get("_last_progress_state")
        state = merge_stable_progress_state(
            state,
            previous_progress if isinstance(previous_progress, dict) else previous_state,
            NEWS_PROGRESS_FIELDS,
        )
        state = repair_mojibake(state)
        if state.get("started_at") and state.get("status") in {"running", "queued"}:
            try:
                started_at = datetime.fromisoformat(str(state.get("started_at")))
                if started_at.tzinfo is None:
                    started_at = started_at.replace(tzinfo=MSK_TZ)
                state["elapsed_seconds"] = int((datetime.now(MSK_TZ) - started_at).total_seconds())
            except ValueError:
                pass
        monitor["state"] = state
        monitor["brand_state"] = dict(state)
        if is_active_status(state.get("status")):
            if has_positive_progress_value(state, NEWS_PROGRESS_FIELDS) or str(state.get("currenturl") or "").strip():
                monitor["_last_progress_state"] = dict(state)
        else:
            monitor.pop("_last_progress_state", None)
        group = clean_text(str(monitor.get("group") or ""))
        brand = clean_text(str(monitor.get("brand") or ""))
        for item in news_settings.get("monitors", []):
            if (
                isinstance(item, dict)
                and clean_text(str(item.get("group") or "")) == group
                and clean_text(str(item.get("brand") or "")) == brand
            ):
                item["state"] = dict(state)
                item["brand_state"] = dict(state)
    if persist:
        persist_news_monitor_state(monitor)


def persist_news_monitor_state(monitor: Dict[str, object], force: bool = False) -> None:
    monitor_id = str(monitor.get("id") or "").strip()
    if not monitor_id:
        return
    now = time.time()
    if not force and now - news_state_persisted_at.get(monitor_id, 0) < 1:
        return
    news_state_persisted_at[monitor_id] = now
    state = repair_mojibake({**make_news_state(), **(monitor.get("state") or {})})
    for attempt in range(4):
        try:
            with session_scope() as session:
                donor = get_donor_row(session, monitor_id)
                if donor is None:
                    return
                donor.updated_at = datetime.utcnow()
                if donor.brand:
                    donor.brand.state = state
            return
        except OperationalError as exc:
            if "database is locked" not in str(exc).lower() or attempt == 3:
                print(f"Failed to persist news monitor state {monitor_id}: {exc}", flush=True)
                return
            time.sleep(0.25 * (attempt + 1))
        except Exception as exc:
            print(f"Failed to persist news monitor state {monitor_id}: {exc}", flush=True)
            return


def news_monitor_thread_alive(monitor_id: object) -> bool:
    thread = news_scan_threads.get(str(monitor_id))
    return isinstance(thread, threading.Thread) and thread.is_alive()


def enqueue_news_scan(monitor_id: str, manual: bool) -> bool:
    monitor_id = str(monitor_id)

    def run() -> None:
        monitor = get_news_monitor(monitor_id)
        state = monitor.get("state", {}) if monitor else {}
        if monitor and state.get("status") == "queued":
            scan_news_monitor(monitor_id, manual)

    def on_start(thread: threading.Thread) -> None:
        with news_lock:
            news_scan_threads[monitor_id] = thread

    def on_finish(thread: threading.Thread) -> None:
        with news_lock:
            if news_scan_threads.get(monitor_id) is thread:
                news_scan_threads.pop(monitor_id, None)

    return scan_dispatcher.enqueue(
        ("news", monitor_id), run, on_start=on_start, on_finish=on_finish
    )


def transition_requested_at(monitor: Dict[str, object]) -> Optional[datetime]:
    state = monitor.get("state", {}) if isinstance(monitor.get("state"), dict) else {}
    for key in ("stop_requested_at", "finished_at", "started_at"):
        parsed = parse_schedule_datetime(state.get(key))
        if parsed:
            return parsed
    return None


def is_stale_news_transition(monitor: Dict[str, object]) -> bool:
    state = monitor.get("state", {}) if isinstance(monitor.get("state"), dict) else {}
    status = str(state.get("status") or "")
    if status not in {"pausing", "stopping"}:
        return False
    requested_at = transition_requested_at(monitor)
    timed_out = bool(requested_at and (datetime.now(MSK_TZ) - requested_at).total_seconds() >= NEWS_TRANSITION_TIMEOUT_SECONDS)
    return timed_out or not news_monitor_thread_alive(monitor.get("id"))


def finalize_stale_news_transition(monitor: Dict[str, object]) -> bool:
    if not is_stale_news_transition(monitor):
        return False
    state = dict(monitor.get("state", {}) if isinstance(monitor.get("state"), dict) else {})
    was_pausing = state.get("status") == "pausing"
    state.update(
        {
            "status": "partial" if was_pausing else "stopped",
            "stage": "Приостановлено" if was_pausing else "Остановлено",
            "error": "",
            "finished_at": datetime.now(MSK_TZ).isoformat(timespec="seconds"),
            "currenturl": "",
            "queue_size": 0,
            "active_tasks": 0,
            "active_urls": [],
        }
    )
    with news_lock:
        monitor["state"] = state
        monitor["brand_state"] = dict(state)
    return True


def cleanup_stale_news_transitions() -> None:
    changed: List[Dict[str, object]] = []
    with news_lock:
        monitors = [item for item in news_settings.get("monitors", []) if isinstance(item, dict)]
    for monitor in monitors:
        if finalize_stale_news_transition(monitor):
            changed.append(monitor)
    for monitor in changed:
        threading.Thread(target=persist_news_monitor_state, args=(monitor, True), daemon=True).start()


def update_brand_scan_state(
    target_type: str,
    target_id: str,
    status: str,
    started_at: float,
    found_products: int = 0,
    new_count: int = 0,
    data: Optional[Dict[str, object]] = None,
) -> None:
    if target_type not in {"news", "donor"}:
        return
    data = data or {}
    existing_state: Dict[str, object] = {}
    with news_lock:
        monitor = next((item for item in news_settings.get("monitors", []) if str(item.get("id")) == str(target_id)), None)
        if monitor and isinstance(monitor.get("state"), dict):
            existing_state = dict(monitor.get("state") or {})
    finished_at = datetime.now(MSK_TZ).isoformat(timespec="seconds")
    state = {
        **make_news_state(),
        **existing_state,
        "status": status or "idle",
        "started_at": existing_state.get("started_at") or datetime.fromtimestamp(started_at, MSK_TZ).isoformat(timespec="seconds"),
        "finished_at": finished_at,
        "last_scan_at": existing_state.get("last_scan_at") or finished_at,
        "found_products": found_products,
        "new_count": new_count,
        "data": data,
    }
    if data.get("csv"):
        state["last_csv"] = str(data.get("csv") or "")
    if data.get("missing_by_feed"):
        state["missing_by_feed"] = data.get("missing_by_feed")
    if data.get("error"):
        state["error"] = str(data.get("error") or "")
    state = repair_mojibake(state)
    with session_scope() as session:
        donor = get_donor_row(session, target_id)
        if donor and donor.brand:
            donor.brand.state = state
    with news_lock:
        if monitor:
            group = clean_text(str(monitor.get("group") or ""))
            brand = clean_text(str(monitor.get("brand") or ""))
            for item in news_settings.get("monitors", []):
                if (
                    isinstance(item, dict)
                    and clean_text(str(item.get("group") or "")) == group
                    and clean_text(str(item.get("brand") or "")) == brand
                ):
                    item["brand_state"] = dict(state)
                    item["state"] = dict(state)


class NewsScanStopped(Exception):
    pass


def get_news_stop_event(monitor_id: str) -> threading.Event:
    with news_lock:
        event = news_stop_events.get(monitor_id)
        if not event:
            event = threading.Event()
            news_stop_events[monitor_id] = event
        return event


def request_news_stop(monitor_id: str, mode: str) -> threading.Event:
    event = get_news_stop_event(monitor_id)
    monitor: Optional[Dict[str, object]] = None
    with news_lock:
        news_stop_modes[monitor_id] = mode
    monitor = get_news_monitor(monitor_id)
    if monitor:
        update_news_monitor_state(
            monitor,
            persist=False,
            status="pausing" if mode == "pause" else "stopping",
            stage="Приостановка" if mode == "pause" else "Остановка",
            currenturl="",
            stop_requested_at=datetime.now(MSK_TZ).isoformat(timespec="seconds"),
        )
    event.set()
    if monitor:
        threading.Thread(target=persist_news_monitor_state, args=(monitor, True), daemon=True).start()
    return event


def collect_products_for_monitor(
    monitor: Dict[str, object],
    stop_signal: threading.Event,
    browser_session: BotasaurusBrowserSession,
) -> List[Dict[str, str]]:
    finish_signal = threading.Event()
    start_urls = normalize_start_urls(monitor.get("start_urls") or "", allow_empty=True)
    if not start_urls:
        raise RuntimeError("У донора не указаны стартовые URL для сканирования.")

    def progress_callback(payload: Dict[str, object]) -> None:
        log_message = str(payload.get("log_message") or "").strip()
        log_level = str(payload.get("log_level") or "info").strip() or "info"
        if log_message:
            add_news_log(monitor, log_message, log_level)
            if stop_signal.is_set():
                return
            event_state = {"last_event": log_message}
            if log_level == "warning":
                event_state["last_warning"] = log_message
            elif log_level == "error":
                event_state["error"] = log_message
            update_news_monitor_state(monitor, **event_state)
            if not any(key in payload for key in ("percent", "currenturl", "totalprocessed", "found_products")):
                return
        if stop_signal.is_set():
            return
        update_news_monitor_state(
            monitor,
            status="running",
            stage="Сканирование сайта-донора",
            percent=min(85, int(payload.get("percent", 0) or 0)),
            currenturl=str(payload.get("currenturl", "")),
            processed=int(payload.get("totalprocessed", 0) or 0),
            found_products=int(payload.get("found_products", 0) or 0),
            in_memory_products=int(payload.get("in_memory_products", payload.get("found_products", 0)) or 0),
            queue_size=int(payload.get("queue_size", 0) or 0),
            active_tasks=int(payload.get("active_tasks", 0) or 0),
            active_urls=list(payload.get("active_urls", []) or [])[:8],
            failed_pages=int(payload.get("failed_pages", 0) or 0),
            stall_seconds=int(payload.get("stall_seconds", 0) or 0),
            skipped=int(payload.get("skipped", 0) or 0),
            error=str(payload.get("error", "") or ""),
        )

    crawler = CollectOnlyCrawler(
        start_urls,
        int(time.time()),
        stop_signal,
        finish_signal,
        parse_thread_count(monitor.get("thread_count", 4)),
        project=None,
        exclusions=list(monitor.get("exclusions", DEFAULT_EXCLUSIONS)),
        product_url_filters=list(monitor.get("product_url_filters", [])),
        product_url_exclusions=list(monitor.get("product_url_exclusions", [])),
        extraction_rules=normalize_extraction_rules(monitor.get("extraction_rules", {})),
        connection_method=normalize_connection_method(monitor.get("connection_method")),
        auto_connection_fallback=bool(monitor.get("auto_connection_fallback", True)),
        allow_empty_price=True,
        browser_session=browser_session,
        owns_browser_session=False,
        progress_callback=progress_callback,
    )
    crawler.run()
    products = crawler.snapshot_results()
    if crawler.fatal_error:
        message = f"{crawler.fatal_error}. Промежуточно хранится в памяти товаров: {len(products)}."
        update_news_monitor_state(
            monitor,
            status="error",
            error=message,
            found_products=len(products),
            in_memory_products=len(products),
            currenturl="",
        )
        raise RuntimeError(message)
    return products


def enrich_news_product(
    product: Dict[str, str],
    monitor: Dict[str, object],
    stop_signal: threading.Event,
    browser_session: BotasaurusBrowserSession,
    connection_method_state: Optional[Dict[str, object]] = None,
) -> Dict[str, str]:
    url = product.get("url", "")
    selector_settings = monitor.get("selector_settings", {}) if isinstance(monitor.get("selector_settings"), dict) else {}
    extraction_rules = normalize_extraction_rules(monitor.get("extraction_rules", {}))
    details = {
        "date_found": datetime.now(MSK_TZ).strftime("%d.%m.%Y %H:%M:%S"),
        "group": str(monitor.get("group") or ""),
        "brand": str(monitor.get("brand") or ""),
        "name": product.get("model", ""),
        "model": product.get("model", ""),
        "price": product.get("price", ""),
        "availability": "",
        "url": url,
    }
    fetcher = CollectOnlyCrawler(
        [url],
        int(time.time()),
        stop_signal,
        threading.Event(),
        1,
        connection_method=normalize_connection_method(monitor.get("connection_method")),
        auto_connection_fallback=bool(monitor.get("auto_connection_fallback", True)),
        connection_method_state=connection_method_state,
        allow_empty_price=True,
        browser_session=browser_session,
        owns_browser_session=False,
    )
    html = fetcher.fetch(url) if url else ""
    if not html:
        return details
    soup = BeautifulSoup(html, "html.parser")
    name = extract_product_name(soup, str(selector_settings.get("name_selector", "")))
    product_data = extract_product_data(
        url,
        html,
        product.get("price", ""),
        extraction_rules,
        allow_empty_price=True,
    )
    if product_data:
        details["model"] = product_data.get("model", details["model"])
        details["price"] = product_data.get("price", details["price"])
    else:
        marker_model = extract_model_by_markers(html, extraction_rules)
        model_candidate = marker_model or details["model"] or name
        prepared_model = finalize_scraped_model(
            model_candidate,
            url,
            extraction_rules,
            preserve_configured_model=bool(marker_model),
        )
        if prepared_model:
            details["model"] = prepared_model
    details["name"] = name or details["name"]
    details["availability"] = extract_availability(soup, str(selector_settings.get("availability_selector", "")))
    return details


def create_news_csv(rows: List[Dict[str, str]], monitor: Dict[str, object], filename: str = "") -> Path:
    if not filename:
        filename = news_csv_filename(monitor)
    filename = output_text(filename)
    path = EXPORT_DIR / filename
    with path.open("w", encoding="utf-8-sig", newline="") as csv_file:
        writer = csv.writer(csv_file, delimiter=";")
        writer.writerow(
            [
                "Дата появления",
                "Группа",
                "Сайт/бренд",
                "Наименование",
                "Модель",
                "Цена",
                "Наличие",
                "Нет на сайтах",
                "URL товара",
            ]
        )
        for row in rows:
            writer.writerow(
                [
                    output_text(row.get("date_found", "")),
                    output_text(row.get("group", "")),
                    output_text(row.get("brand", "")),
                    output_text(row.get("name", "")),
                    output_text(row.get("model", "")),
                    output_text(row.get("price", "")),
                    output_text(row.get("availability", "")),
                    output_text(row.get("missing_on", "")),
                    output_text(row.get("url", "")),
                ]
            )
    return path


def build_email_message(
    sender_email: str,
    recipient: str,
    subject: str,
    body: str,
    csv_path: Optional[Path] = None,
) -> EmailMessage:
    message = EmailMessage()
    message["Subject"] = subject
    message["From"] = sender_email
    message["To"] = recipient
    message.set_content(body)
    if csv_path:
        message.add_attachment(
            csv_path.read_bytes(),
            maintype="text",
            subtype="csv",
            filename=str(repair_mojibake_text(csv_path.name) or csv_path.name),
        )
    return message


def send_messages_to_recipients(
    host: str,
    port: int,
    security_mode: str,
    username: str,
    password: str,
    sender_email: str,
    recipients: List[str],
    subject: str,
    body: str,
    csv_path: Optional[Path] = None,
) -> None:
    context = ssl.create_default_context()
    failures: List[str] = []
    if security_mode == "tls":
        with smtplib.SMTP(host, port, timeout=30) as server:
            server.starttls(context=context)
            server.login(username, password)
            for recipient in recipients:
                try:
                    server.send_message(
                        build_email_message(sender_email, recipient, subject, body, csv_path),
                        from_addr=sender_email,
                        to_addrs=[recipient],
                    )
                except Exception as exc:
                    failures.append(f"{recipient}: {exc}")
    else:
        with smtplib.SMTP_SSL(host, port, context=context, timeout=30) as server:
            server.login(username, password)
            for recipient in recipients:
                try:
                    server.send_message(
                        build_email_message(sender_email, recipient, subject, body, csv_path),
                        from_addr=sender_email,
                        to_addrs=[recipient],
                    )
                except Exception as exc:
                    failures.append(f"{recipient}: {exc}")
    if failures:
        raise RuntimeError("Не удалось отправить на: " + "; ".join(failures))


def feed_missing_labels(keys: Set[str], feed_code_sets: List[Dict[str, object]]) -> List[str]:
    if not keys:
        return []
    missing_feeds = []
    for feed in feed_code_sets:
        feed_codes = feed.get("codes", set())
        if not isinstance(feed_codes, set):
            feed_codes = set(feed_codes) if isinstance(feed_codes, list) else set()
        if not (keys & feed_codes):
            missing_feeds.append(str(feed.get("source_label") or feed.get("url") or "Фид"))
    return missing_feeds


def news_monitor_profile_storage_dir(monitor: Dict[str, object]) -> Path:
    monitor_id = safe_filename(str(monitor.get("id") or monitor.get("brand") or "monitor"))
    return PROJECT_PROFILE_DIR / "news" / monitor_id


def news_monitor_should_keep_browser_profile(monitor: Dict[str, object]) -> bool:
    method = normalize_connection_method(monitor.get("connection_method"))
    return method == "protected-site" or bool(monitor.get("auto_connection_fallback", True))


def enrich_news_candidates(
    products: List[Dict[str, str]],
    monitor: Dict[str, object],
    feed_code_sets: List[Dict[str, object]],
    stop_signal: threading.Event,
    browser_session: BotasaurusBrowserSession,
    progress_callback,
) -> List[Dict[str, str]]:
    candidates: List[tuple[int, Dict[str, str]]] = []
    resolved: List[Optional[Dict[str, str]]] = [None] * len(products)

    for index, product in enumerate(products):
        keys = product_compare_keys(product)
        if not keys:
            candidates.append((index, product))
            continue
        missing_feeds = feed_missing_labels(keys, feed_code_sets)
        if not missing_feeds:
            details = {
                "date_found": datetime.now(MSK_TZ).strftime("%d.%m.%Y %H:%M:%S"),
                "group": str(monitor.get("group") or ""),
                "brand": str(monitor.get("brand") or ""),
                "name": product.get("model", ""),
                "model": product.get("model", ""),
                "price": product.get("price", ""),
                "availability": "",
                        "url": product.get("url", ""),
            }
            resolved[index] = details
            progress_callback(index + 1, details.get("url", ""))
        else:
            candidates.append((index, product))

    if candidates and not stop_signal.is_set():
        max_workers = min(parse_thread_count(monitor.get("thread_count", 4)), len(candidates))
        connection_method_state = {
            "active_method": normalize_connection_method(monitor.get("connection_method")),
            "lock": threading.Lock(),
        }
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_to_index = {
                executor.submit(
                    enrich_news_product,
                    product,
                    monitor,
                    stop_signal,
                    browser_session,
                    connection_method_state,
                ): (index, product)
                for index, product in candidates
            }
            completed = len(products) - len(candidates)
            for future in as_completed(future_to_index):
                if stop_signal.is_set():
                    raise NewsScanStopped()
                index, product = future_to_index[future]
                details = future.result()
                details["model"] = details.get("model") or product.get("model", "")
                resolved[index] = details
                completed += 1
                progress_callback(completed, details.get("url", product.get("url", "")))

    return [item for item in resolved if item]


def send_news_email_legacy(
    monitor: Optional[Dict[str, object]],
    new_count: int,
    test: bool = False,
    error_holder: Optional[List[str]] = None,
    missing_summary: Optional[List[Dict[str, object]]] = None,
) -> bool:
    with news_lock:
        smtp_config = dict(news_settings.get("smtp", {}))
    recipients = normalize_emails(smtp_config.get("recipients", []))
    username = str(smtp_config.get("username") or "").strip()
    password = str(smtp_config.get("password") or "").strip()
    sender_emails = normalize_emails(username)
    if not username or not password or not recipients:
        error_message = "Email не отправлен: заполните email-логин, пароль приложения и получателей SMTP"
        error_message = str(repair_mojibake_text(error_message))
        if error_holder is not None:
            error_holder.append(error_message)
        add_news_log(monitor, error_message, "warning")
        return False
    if not sender_emails:
        error_message = "Email не отправлен: email-логин должен быть адресом почты"
        error_message = str(repair_mojibake_text(error_message))
        if error_holder is not None:
            error_holder.append(error_message)
        add_news_log(monitor, error_message, "warning")
        return False
    sender_email = sender_emails[0]

    if test:
        subject = "Тест email-уведомлений"
        body = "Тестовое письмо отправлено из мониторинга новинок. SMTP-настройки работают."
    else:
        brand = str((monitor or {}).get("brand") or "донор")
        site_url = str((monitor or {}).get("site_url") or "")
        subject = f"Уведомление о новинках на сайте {brand}"
        lines = [f"На {site_url or brand} найдено всего: {new_count}"]
        for item in missing_summary or []:
            count = int(item.get("count") or 0)
            label = str(item.get("source_label") or item.get("url") or "сайт")
            lines.append(f"На сайте {label} не было найдено {count} новинок.")
        body = "\n".join(lines)
    subject = str(repair_mojibake_text(subject))
    body = str(repair_mojibake_text(body))

    smtp_defaults = default_smtp_settings()
    host = str(smtp_config.get("host") or smtp_defaults["host"])
    port = int(smtp_config.get("port") or smtp_defaults["port"])
    security_mode = str(smtp_config.get("security") or smtp_defaults["security"]).lower()
    try:
        send_messages_to_recipients(host, port, security_mode, username, password, sender_email, recipients, subject, body)
    except Exception as exc:
        error_message = f"Ошибка отправки email: {exc}"
        error_message = str(repair_mojibake_text(error_message))
        if error_holder is not None:
            error_holder.append(error_message)
        add_news_log(monitor, error_message, "error")
        return False
    add_news_log(monitor, "Тестовое email-сообщение отправлено" if test else f"Email-уведомление отправлено. Новинок: {new_count}", "success")
    return True


def send_news_email(
    monitor: Optional[Dict[str, object]],
    new_count: int,
    test: bool = False,
    error_holder: Optional[List[str]] = None,
    missing_summary: Optional[List[Dict[str, object]]] = None,
) -> bool:
    with news_lock:
        smtp_config = dict(news_settings.get("smtp", {}))
    recipients = normalize_emails(smtp_config.get("recipients", []))
    username = str(smtp_config.get("username") or "").strip()
    password = str(smtp_config.get("password") or "").strip()
    sender_emails = normalize_emails(username)
    if not username or not password or not recipients:
        error_message = "Email не отправлен: заполните email-логин, пароль приложения и получателей SMTP"
        if error_holder is not None:
            error_holder.append(error_message)
        add_news_log(monitor, error_message, "warning")
        return False
    if not sender_emails:
        error_message = "Email не отправлен: email-логин должен быть адресом почты"
        if error_holder is not None:
            error_holder.append(error_message)
        add_news_log(monitor, error_message, "warning")
        return False

    sender_email = sender_emails[0]
    csv_path: Optional[Path] = None
    if test:
        subject = "Тест email-уведомления"
        body = "Тестовое письмо отправлено из мониторинга новинок. SMTP-настройки работают."
    else:
        brand = str(repair_mojibake_text((monitor or {}).get("brand") or "донор"))
        site_url = str((monitor or {}).get("site_url") or "")
        subject = f"Уведомление о новинках на сайте {brand}"
        lines = [f"На {site_url or brand} найдено всего: {new_count}"]
        for item in missing_summary or []:
            count = int(item.get("count") or 0)
            label = str(repair_mojibake_text(item.get("source_label") or item.get("url") or "сайт"))
            lines.append(f"На сайте {label} не было найдено {count} новинок.")
        body = "\n".join(lines)

        state = (monitor or {}).get("state", {}) if isinstance((monitor or {}).get("state"), dict) else {}
        state_data = state.get("data", {}) if isinstance(state.get("data"), dict) else {}
        csv_filename = str(state.get("last_csv") or state_data.get("csv") or "")
        csv_path = resolve_export_file(csv_filename)

    smtp_defaults = default_smtp_settings()
    host = str(smtp_config.get("host") or smtp_defaults["host"])
    port = int(smtp_config.get("port") or smtp_defaults["port"])
    security_mode = str(smtp_config.get("security") or smtp_defaults["security"]).lower()
    try:
        send_messages_to_recipients(host, port, security_mode, username, password, sender_email, recipients, subject, body, csv_path)
    except Exception as exc:
        error_message = f"Ошибка отправки email: {exc}"
        if error_holder is not None:
            error_holder.append(error_message)
        add_news_log(monitor, error_message, "error")
        return False

    add_news_log(
        monitor,
        "Тестовое email-сообщение отправлено" if test else f"Email-уведомление отправлено. Новинок: {new_count}",
        "success",
    )
    return True


def scan_news_monitor(monitor_id: str, manual: bool = False) -> None:
    monitor = get_news_monitor(monitor_id)
    if not monitor:
        return
    refresh_monitor_schedule_from_brand(monitor)
    started = time.time()
    stop_event = get_news_stop_event(monitor_id)
    stop_event.clear()
    with news_lock:
        news_stop_modes.pop(monitor_id, None)
    new_items: List[Dict[str, str]] = []
    products: List[Dict[str, str]] = []
    local_feeds: List[Dict[str, object]] = []
    feed_code_sets: List[Dict[str, object]] = []
    missing_summary: List[Dict[str, object]] = []
    availability_skipped = 0
    profile_dir = news_monitor_profile_storage_dir(monitor) if news_monitor_should_keep_browser_profile(monitor) else None
    browser_session = BotasaurusBrowserSession(
        stop_event,
        parse_thread_count(monitor.get("thread_count", 4)),
        profile_dir=profile_dir,
    )

    def check_stop_requested() -> None:
        if stop_event.is_set():
            raise NewsScanStopped()

    with news_lock:
        monitor["state"] = {
            **make_news_state("running"),
            "stage": "Подготовка",
            "started_at": datetime.now(MSK_TZ).isoformat(timespec="seconds"),
        }
        monitor["brand_state"] = dict(monitor["state"])
        persist_news_monitor_state(monitor, force=True)
    add_news_log(monitor, "Ручное сканирование новинок запущено" if manual else "Плановое сканирование новинок запущено", "info")
    add_news_log(monitor, "Начал сбор", "info")

    try:
        update_news_monitor_state(monitor, stage="Подготовка", percent=2)
        validate_monitor_selectors(monitor)
        add_news_log(
            monitor,
            f"Scan settings: URL={', '.join(monitor.get('start_urls', []))}; "
            f"method={monitor.get('connection_method')}; threads={monitor.get('thread_count')}",
            "info",
        )
        update_news_monitor_state(monitor, stage="Сканирование сайта-донора", percent=5)
        products = collect_products_for_monitor(monitor, stop_event, browser_session)
        check_stop_requested()
        add_news_log(monitor, "Сбор закончил", "info")
        add_news_log(monitor, f"Сканирование сайта завершено. Найдено товаров: {len(products)}", "info")
        update_news_monitor_state(monitor, stage="Генерация и загрузка фидов ваших сайтов", percent=84, currenturl="")
        add_news_log(monitor, "Скачивание фида", "info")
        all_existing_codes, local_feeds, feed_code_sets = fetch_existing_vendor_code_sets()
        check_stop_requested()
        add_news_log(monitor, "Фид скачался", "info")
        add_news_log(
            monitor,
            f"Фиды обновлены после сбора донора: {len(local_feeds)}. Моделей всего: {len(all_existing_codes)}",
            "info",
        )
        update_news_monitor_state(
            monitor,
            stage="Сравнение с фидами",
            percent=86,
            candidate_products=len(products),
            found_products=len(products),
            compared_products=0,
            currenturl="",
        )
        add_news_log(monitor, "Началось сравнение", "info")
        known = monitor.get("known_new_products", {}) if isinstance(monitor.get("known_new_products"), dict) else {}
        def update_compare_progress(index: int, current_url: str = "") -> None:
            check_stop_requested()
            update_news_monitor_state(
                monitor,
                stage="Сравнение с фидами",
                percent=86 + int((index / max(1, len(products))) * 12),
                compared_products=index,
                currenturl=current_url,
            )

        enriched_products = enrich_news_candidates(
            products,
            monitor,
            feed_code_sets,
            stop_event,
            browser_session,
            update_compare_progress,
        )
        availability_exclusions = normalize_patterns((monitor.get("selector_settings") or {}).get("availability_exclusions", []))
        for details, product in zip(enriched_products, products):
            check_stop_requested()
            details["model"] = details.get("model") or product.get("model", "")
            if availability_is_excluded(details.get("availability", ""), availability_exclusions):
                availability_skipped += 1
                continue
            detail_keys = product_compare_keys(details) | product_compare_keys(product)
            if not detail_keys:
                continue
            missing_feeds = feed_missing_labels(detail_keys, feed_code_sets)
            if not missing_feeds:
                continue
            model_key = sorted(detail_keys)[0]
            details["missing_on"] = ", ".join(missing_feeds)
            details["missing_on_count"] = len(missing_feeds)
            new_items.append(details)
            known[model_key] = details
        missing_summary = build_missing_summary(new_items, feed_code_sets)
        add_news_log(monitor, "Сравнение закончилось", "info")
        if availability_skipped:
            add_news_log(monitor, f"Исключено по статусу наличия: {availability_skipped}", "info")
        for item in missing_summary:
            add_news_log(
                monitor,
                f"Нет на {item.get('source_label')}: {int(item.get('count') or 0)}",
                "info",
            )

        update_news_monitor_state(monitor, stage="Формирование CSV", percent=99, currenturl="")
        csv_path = create_news_csv(new_items, monitor)
        delete_news_csv_for_monitor(monitor, keep_filename=csv_path.name)
        elapsed = int(time.time() - started)
        with news_lock:
            monitor["known_new_products"] = known
            monitor["state"] = {
                **monitor.get("state", {}),
                "status": "completed",
                "stage": "Завершено",
                "percent": 100,
                "processed": len(products),
                "found_products": len(products),
                "candidate_products": len(products),
                "compared_products": len(products),
                "in_memory_products": len(products),
                "availability_skipped": availability_skipped,
                "queue_size": 0,
                "active_tasks": 0,
                "active_urls": [],
                "new_count": len(new_items),
                "missing_by_feed": missing_summary,
                "last_scan_at": datetime.now(MSK_TZ).isoformat(timespec="seconds"),
                "last_csv": csv_path.name,
                "error": "",
                "finished_at": datetime.now(MSK_TZ).isoformat(timespec="seconds"),
                "elapsed_seconds": elapsed,
                "currenturl": "",
            }
            monitor["brand_state"] = dict(monitor["state"])
            if normalize_schedule_type(monitor.get("schedule_type")) != "once":
                monitor["next_run_at"] = update_brand_next_run_at(monitor.get("brand_id"))
            save_news_settings()
        add_news_log(monitor, f"Сканирование завершено. Найдено новинок: {len(new_items)}. CSV: {csv_path.name}", "success")
        update_brand_scan_state(
            "donor",
            monitor_id,
            "completed",
            started,
            found_products=len(products),
            new_count=len(new_items),
            data={
                "csv": csv_path.name,
                "feeds": local_feeds,
                "missing_by_feed": missing_summary,
                "availability_skipped": availability_skipped,
            },
        )
        if new_items:
            send_news_email(monitor, len(new_items), missing_summary=missing_summary)
    except NewsScanStopped:
        elapsed = int(time.time() - started)
        with news_lock:
            stop_mode = news_stop_modes.get(monitor_id, "stop")
        partial_csv = ""
        if stop_mode == "pause" and new_items:
            partial_path = create_news_csv(new_items, monitor)
            partial_csv = partial_path.name
            delete_news_csv_for_monitor(monitor, keep_filename=partial_csv)
        missing_summary = build_missing_summary(new_items, feed_code_sets) if feed_code_sets else []
        with news_lock:
            monitor["state"] = {
                **monitor.get("state", {}),
                "status": "partial" if stop_mode == "pause" else "idle",
                "stage": "Приостановлено" if stop_mode == "pause" else "Ожидание",
                "error": "",
                "finished_at": datetime.now(MSK_TZ).isoformat(timespec="seconds"),
                "elapsed_seconds": elapsed,
                "currenturl": "",
                "last_csv": partial_csv or monitor.get("state", {}).get("last_csv", ""),
                "new_count": len(new_items),
                "missing_by_feed": missing_summary,
                "processed": len(products),
                "found_products": len(products),
                "in_memory_products": len(products),
                "availability_skipped": availability_skipped,
                "queue_size": int(monitor.get("state", {}).get("queue_size", 0) or 0),
                "active_tasks": 0,
                "active_urls": [],
            }
            monitor["brand_state"] = dict(monitor["state"])
            save_news_settings()
        add_news_log(
            monitor,
            f"Сканирование новинок приостановлено. CSV: {partial_csv}" if stop_mode == "pause" else "Сканирование новинок остановлено",
            "warning",
        )
        update_brand_scan_state(
            "donor",
            monitor_id,
            "partial" if stop_mode == "pause" else "idle",
            started,
            found_products=len(products),
            new_count=len(new_items),
            data={"csv": partial_csv, "availability_skipped": availability_skipped},
        )
    except Exception as exc:
        elapsed = int(time.time() - started)
        with news_lock:
            monitor["state"] = {
                **monitor.get("state", {}),
                "status": "error",
                "stage": "Ошибка",
                "error": str(exc),
                "finished_at": datetime.now(MSK_TZ).isoformat(timespec="seconds"),
                "elapsed_seconds": elapsed,
                "found_products": len(products),
                "in_memory_products": len(products),
                "availability_skipped": availability_skipped,
                "active_tasks": 0,
                "active_urls": [],
            }
            monitor["brand_state"] = dict(monitor["state"])
            save_news_settings()
        add_news_log(monitor, f"Ошибка сканирования новинок: {exc}", "error")
        update_brand_scan_state(
            "donor",
            monitor_id,
            "error",
            started,
            found_products=len(products),
            new_count=len(new_items),
            data={"error": str(exc)},
        )
    finally:
        browser_session.close()
        with news_lock:
            news_stop_modes.pop(monitor_id, None)
            thread = news_scan_threads.get(monitor_id)
            if thread is threading.current_thread():
                news_scan_threads.pop(monitor_id, None)
        stop_event.clear()


def parse_scan_time(value: object) -> datetime_time:
    text = str(value or "01:00")
    try:
        hour, minute = [int(part) for part in text[:5].split(":", 1)]
        return datetime_time(max(0, min(hour, 23)), max(0, min(minute, 59)), tzinfo=MSK_TZ)
    except Exception:
        return datetime_time(1, 0, tzinfo=MSK_TZ)


def parse_schedule_datetime(value: object) -> Optional[datetime]:
    if isinstance(value, datetime):
        return value.astimezone(MSK_TZ) if value.tzinfo else value.replace(tzinfo=MSK_TZ)
    text = str(value or "").strip()
    if not text:
        return None
    try:
        parsed = datetime.fromisoformat(text)
    except ValueError:
        return None
    return parsed.astimezone(MSK_TZ) if parsed.tzinfo else parsed.replace(tzinfo=MSK_TZ)


def normalize_schedule_type(value: object) -> str:
    schedule_type = str(value or "daily")
    return schedule_type if schedule_type in {"daily", "weekly", "once"} else "daily"


def normalize_weekday(value: object) -> int:
    try:
        return max(0, min(int(value or 0), 6))
    except (TypeError, ValueError):
        return 0


def compute_schedule_run_at(
    schedule_type: object,
    scan_time: object,
    weekday: object = 0,
    once_at: object = None,
    now: Optional[datetime] = None,
) -> Optional[datetime]:
    now = now or datetime.now(MSK_TZ)
    schedule = normalize_schedule_type(schedule_type)
    if schedule == "once":
        return parse_schedule_datetime(once_at)
    run_time = parse_scan_time(scan_time)
    candidate = now.replace(hour=run_time.hour, minute=run_time.minute, second=0, microsecond=0)
    if schedule == "weekly":
        candidate += timedelta(days=normalize_weekday(weekday) - now.weekday())
    return candidate


def compute_next_schedule_at(
    schedule_type: object,
    scan_time: object,
    weekday: object = 0,
    once_at: object = None,
    now: Optional[datetime] = None,
) -> Optional[datetime]:
    now = now or datetime.now(MSK_TZ)
    schedule = normalize_schedule_type(schedule_type)
    if schedule == "once":
        return parse_schedule_datetime(once_at)
    candidate = compute_schedule_run_at(schedule, scan_time, weekday, now=now)
    if candidate is None:
        return None
    if schedule == "weekly":
        if candidate <= now:
            candidate += timedelta(days=7)
    elif candidate <= now:
        candidate += timedelta(days=1)
    return candidate


def compute_next_run_at(monitor: Dict[str, object]) -> str:
    candidate = compute_next_schedule_at(
        monitor.get("schedule_type"),
        monitor.get("scan_time"),
        monitor.get("weekday"),
        monitor.get("next_run_at"),
    )
    return candidate.isoformat(timespec="minutes") if candidate else ""


def brand_schedule_fields(brand: Brand) -> Dict[str, object]:
    return {
        "enabled": bool(brand.enabled),
        "schedule_type": normalize_schedule_type(brand.schedule_type),
        "scan_time": str(brand.scan_time or "01:00")[:5],
        "weekday": normalize_weekday(brand.weekday),
        "next_run_at": datetime_to_input_value(brand.next_run_at),
        "primary_donor_id": brand.primary_donor_id,
    }


def is_brand_due(brand: Brand, now: Optional[datetime] = None) -> bool:
    if not bool(brand.enabled):
        return False
    state = brand.state if isinstance(brand.state, dict) else {}
    if state.get("status") in {"running", "queued", "pausing", "stopping"}:
        return False
    now = now or datetime.now(MSK_TZ)
    schedule_type = normalize_schedule_type(brand.schedule_type)
    due_at = compute_schedule_run_at(schedule_type, brand.scan_time, brand.weekday, brand.next_run_at, now)
    if not due_at:
        return False
    if schedule_type in {"daily", "weekly"}:
        seconds_after_due = (now - due_at).total_seconds()
        if seconds_after_due < 0 or seconds_after_due >= SCHEDULE_DUE_GRACE_SECONDS:
            return False
    elif now < due_at:
        return False
    last_scan = str(state.get("last_scan_at") or "")
    if last_scan:
        last_scan_at = parse_schedule_datetime(last_scan)
        if last_scan_at and last_scan_at >= due_at:
            return False
    return True


def update_brand_next_run_at(brand_id: object) -> str:
    with session_scope() as session:
        brand = session.get(Brand, parse_db_int(brand_id))
        if not brand:
            return ""
        next_at = compute_next_schedule_at(brand.schedule_type, brand.scan_time, brand.weekday, brand.next_run_at)
        if normalize_schedule_type(brand.schedule_type) != "once":
            brand.next_run_at = next_at.replace(tzinfo=None) if next_at else None
        return datetime_to_input_value(brand.next_run_at)


def refresh_monitor_schedule_from_brand(monitor: Dict[str, object]) -> None:
    brand_id = parse_db_int(monitor.get("brand_id"))
    if not brand_id:
        return
    with session_scope() as session:
        brand = session.get(Brand, brand_id)
        if not brand:
            return
        monitor.update(brand_schedule_fields(brand))


def safe_next_path(value: object) -> str:
    text = str(value or "").strip()
    if not text.startswith("/") or text.startswith("//"):
        return url_for("index")
    parsed = urlparse(text)
    if parsed.scheme or parsed.netloc:
        return url_for("index")
    return urlunparse(("", "", parsed.path or "/", "", parsed.query, ""))


def start_news_scheduler() -> None:
    global news_scheduler_thread
    if isinstance(news_scheduler_thread, threading.Thread) and news_scheduler_thread.is_alive():
        return

    def scheduler_loop() -> None:
        while True:
            try:
                reload_news_monitors_from_db()
                due_ids: List[str] = []
                with news_lock:
                    monitor_by_id = {
                        str(monitor.get("id")): monitor
                        for monitor in news_settings.get("monitors", [])
                        if isinstance(monitor, dict)
                    }
                    with session_scope() as session:
                        brand_rows = session.scalars(select(Brand).order_by(Brand.id)).all()
                        due_brands = [brand for brand in brand_rows if is_brand_due(brand)]
                        due_brand_data = [
                            {
                                "brand_id": brand.id,
                                "brand_name": brand.name,
                                "primary_id": brand.primary_donor_id,
                                "schedule": brand_schedule_fields(brand),
                                "donor_ids": [donor.id for donor in brand.donors],
                            }
                            for brand in due_brands
                        ]
                    for brand_data in due_brand_data:
                        primary_id = str(brand_data.get("primary_id") or "")
                        selected = monitor_by_id.get(primary_id)
                        if selected is None:
                            fallback_id = next((str(donor_id) for donor_id in brand_data.get("donor_ids", []) if str(donor_id) in monitor_by_id), "")
                            selected = monitor_by_id.get(fallback_id)
                        if selected is None:
                            add_news_log(
                                None,
                                f"Плановый запуск пропущен: основной донор бренда {brand_data.get('brand_name')} не найден.",
                                "warning",
                            )
                            continue
                        selected.update(brand_data["schedule"])
                        selected["state"] = {**selected.get("state", {}), "status": "queued"}
                        selected["brand_state"] = dict(selected["state"])
                        sync_brand_runtime_fields(selected)
                        due_ids.append(str(selected.get("id")))
                    if due_ids:
                        save_news_settings()
                for monitor_id in due_ids:
                    enqueue_news_scan(monitor_id, manual=False)
            except Exception:
                pass
            time.sleep(30)

    news_scheduler_thread = threading.Thread(target=scheduler_loop, daemon=True)
    news_scheduler_thread.start()


@app.route("/login", methods=["GET", "POST"])
def login() -> str | Response:
    ensure_storage()
    if session.get("user_id"):
        return redirect(url_for("index"))
    error = ""
    if request.method == "POST":
        username = str(request.form.get("username") or "").strip()
        password = str(request.form.get("password") or "")
        user = g.db.scalar(select(User).where(User.username == username, User.is_active.is_(True)))
        if user and check_password_hash(user.password_hash, password):
            session.clear()
            session["user_id"] = int(user.id)
            session["username"] = user.username
            return redirect(url_for("index"))
        error = "Неверный логин или пароль"
    return render_template("login.html", error=error)


@app.get("/api/health")
def healthcheck():
    ensure_storage()
    return jsonify({"ok": True})


@app.post("/logout")
def logout() -> Response:
    session.clear()
    return redirect(url_for("login"))


def render_app_page(active_view: str = "projects", active_project_id: str = "", active_news_id: str = "") -> str:
    ensure_storage()
    static_files = [BASE_DIR / "static" / "js" / "app.js", BASE_DIR / "static" / "css" / "styles.css"]
    static_version = max((int(path.stat().st_mtime) for path in static_files if path.exists()), default=0)
    return render_template(
        "index.html",
        default_start_url=DEFAULT_START_URL,
        static_version=static_version,
        active_view=active_view,
        active_project_id=active_project_id,
        active_news_id=active_news_id,
        progress_interval_ms=int(PROGRESS_STREAM_INTERVAL_SECONDS * 1000),
    )


@app.route("/")
def index() -> Response:
    return redirect(url_for("projects_page"))


@app.get("/projects")
def projects_page() -> str:
    return render_app_page("projects")


@app.get("/projects/edit/<project_id>")
def project_edit_page(project_id: str) -> str:
    return render_app_page("projects", active_project_id=project_id)


@app.get("/news")
def news_page() -> str:
    return render_app_page("news")


@app.get("/news/edit/<brand_id>")
def news_edit_page(brand_id: str) -> str:
    return render_app_page("news", active_news_id=brand_id)


@app.get("/file-import")
def file_import_page() -> str:
    return render_app_page("import")


@app.get("/logs")
def logs_page() -> str:
    return render_app_page("logs")


@app.get("/settings")
def settings_page() -> str:
    return render_app_page("settings")


@app.get("/api/state")
def api_state():
    return jsonify(snapshot_state())


@app.get("/api/connection-methods")
def api_connection_methods():
    ensure_storage()
    return jsonify({"connection_methods": public_connection_methods()})


@app.get("/api/news")
def api_news():
    ensure_storage()
    summary = request.args.get("summary") == "1"
    include_monitors = request.args.get("monitors", "1") != "0"
    return jsonify(public_news_settings(include_monitor_details=not summary, include_monitors=include_monitors))


@app.patch("/api/news/settings")
def api_update_news_settings():
    ensure_storage()
    payload = request.get_json(silent=True) or {}
    with news_lock:
        if "own_sites" in payload and isinstance(payload.get("own_sites"), list):
            own_sites_payload = [item for item in payload.get("own_sites", []) if isinstance(item, dict)]
            own_sites = []
            for index, item in enumerate(own_sites_payload, start=1):
                feed_url = normalize_feed_url(str(item.get("feed_url") or "").strip())
                if not feed_url:
                    continue
                feed_generate_url = normalize_feed_url(str(item.get("feed_generate_url") or "").strip())
                own_sites.append(
                    {
                        "name": clean_text(str(item.get("name") or "")) or f"Фид {index}",
                        "feed_url": feed_url,
                        "feed_generate_url": feed_generate_url,
                    }
                )
            feed_urls = [
                item["feed_url"]
                for item in own_sites
            ]
            feed_generate_urls = [
                item["feed_generate_url"]
                for item in own_sites
                if item.get("feed_generate_url")
            ]
            feed_urls = feed_urls or [DEFAULT_FEED_URL]
            feed_generate_urls = feed_generate_urls or [DEFAULT_FEED_GENERATE_URL]
            news_settings["own_sites"] = own_sites or [{"name": feed_source_label(DEFAULT_FEED_URL), "feed_url": DEFAULT_FEED_URL, "feed_generate_url": DEFAULT_FEED_GENERATE_URL}]
            news_settings["feed_urls"] = feed_urls
            news_settings["feed_url"] = feed_urls[0] if feed_urls else DEFAULT_FEED_URL
            news_settings["feed_generate_urls"] = feed_generate_urls
            news_settings["feed_generate_url"] = feed_generate_urls[0] if feed_generate_urls else DEFAULT_FEED_GENERATE_URL
        if "feed_url" in payload:
            news_settings["feed_url"] = str(payload.get("feed_url") or DEFAULT_FEED_URL).strip()
        if "feed_generate_url" in payload:
            news_settings["feed_generate_url"] = str(payload.get("feed_generate_url") or DEFAULT_FEED_GENERATE_URL).strip()
        if "feed_urls" in payload:
            feed_urls = normalize_feed_urls(payload.get("feed_urls") or DEFAULT_FEED_URL, DEFAULT_FEED_URL)
            news_settings["feed_urls"] = feed_urls
            news_settings["feed_url"] = feed_urls[0] if feed_urls else DEFAULT_FEED_URL
        if "feed_generate_urls" in payload:
            feed_generate_urls = normalize_feed_urls(payload.get("feed_generate_urls") or DEFAULT_FEED_GENERATE_URL, DEFAULT_FEED_GENERATE_URL)
            news_settings["feed_generate_urls"] = feed_generate_urls
            news_settings["feed_generate_url"] = feed_generate_urls[0] if feed_generate_urls else DEFAULT_FEED_GENERATE_URL
        if "auto_cleanup" in payload:
            news_settings["auto_cleanup"] = bool(payload.get("auto_cleanup"))
        if "smtp" in payload and isinstance(payload.get("smtp"), dict):
            smtp_payload = payload["smtp"]
            smtp = dict(news_settings.get("smtp", {}))
            smtp.pop("sender", None)
            for key in ("host", "security", "username"):
                if key in smtp_payload:
                    smtp[key] = str(smtp_payload.get(key) or "").strip()
            if "port" in smtp_payload:
                try:
                    smtp["port"] = int(smtp_payload.get("port") or 465)
                except (TypeError, ValueError):
                    smtp["port"] = 465
            if "password" in smtp_payload and str(smtp_payload.get("password") or "").strip():
                smtp["password"] = str(smtp_payload.get("password")).strip()
            if "recipients" in smtp_payload:
                smtp["recipients"] = normalize_emails(smtp_payload.get("recipients"))
            news_settings["smtp"] = smtp
        save_news_settings()
    return jsonify(public_news_settings())


@app.post("/api/news/email/test")
def api_test_news_email():
    ensure_storage()
    errors: List[str] = []
    if not send_news_email(None, 0, test=True, error_holder=errors):
        return jsonify({"error": errors[-1] if errors else "Email не отправлен. Проверьте SMTP-настройки и логи мониторинга."}), 500
    return jsonify({"ok": True})


@app.get("/api/news/brands/<brand_id>")
def api_get_news_brand(brand_id: str):
    ensure_storage()
    brand_pk = parse_db_int(brand_id)
    if not brand_pk:
        return jsonify({"error": "Бренд не найден"}), 404
    with session_scope() as session:
        brand = session.get(Brand, brand_pk)
        if not brand:
            return jsonify({"error": "Бренд не найден"}), 404
        payload = public_news_brand(brand)
    return jsonify({"brand": payload})

@app.get("/api/news/monitors/<monitor_id>")
def api_get_news_monitor(monitor_id: str):
    monitor = get_news_monitor(monitor_id)
    if not monitor:
        return jsonify({"error": "Монитор не найден"}), 404
    group = clean_text(str(monitor.get("group") or ""))
    brand = clean_text(str(monitor.get("brand") or ""))
    with news_lock:
        brand_monitors = [
            public_news_monitor(item)
            for item in news_settings.get("monitors", [])
            if isinstance(item, dict)
            and clean_text(str(item.get("group") or "")) == group
            and clean_text(str(item.get("brand") or "")) == brand
        ]
    return jsonify({"monitor": public_news_monitor(monitor), "brand_monitors": brand_monitors})


@app.patch("/api/news/monitors/<monitor_id>")
def api_update_news_monitor(monitor_id: str):
    monitor = get_news_monitor(monitor_id)
    if not monitor:
        return jsonify({"error": "Монитор не найден"}), 404
    payload = request.get_json(silent=True) or {}
    with news_lock:
        old_group = clean_text(str(monitor.get("group") or ""))
        old_brand = clean_text(str(monitor.get("brand") or ""))
        if "brand" in payload:
            new_brand = clean_text(str(payload.get("brand") or monitor.get("brand") or ""))
            if new_brand:
                for item in news_settings.get("monitors", []):
                    if (
                        isinstance(item, dict)
                        and clean_text(str(item.get("group") or "")) == old_group
                        and clean_text(str(item.get("brand") or "")) == old_brand
                    ):
                        item["brand"] = new_brand
                monitor["brand"] = new_brand
        if "start_urls" in payload:
            start_urls = normalize_start_urls(payload.get("start_urls"), allow_empty=True)
            if start_urls:
                monitor["start_urls"] = start_urls
        if "site_url" in payload:
            monitor["site_url"] = str(payload.get("site_url") or "").strip()
        if "enabled" in payload:
            monitor["enabled"] = bool(payload.get("enabled"))
        if "schedule_type" in payload:
            schedule_type = str(payload.get("schedule_type") or "daily")
            monitor["schedule_type"] = schedule_type if schedule_type in {"daily", "weekly", "once"} else "daily"
        if "scan_time" in payload:
            monitor["scan_time"] = str(payload.get("scan_time") or "01:00")[:5]
        if "weekday" in payload:
            try:
                monitor["weekday"] = max(0, min(int(payload.get("weekday") or 0), 6))
            except (TypeError, ValueError):
                monitor["weekday"] = 0
        if "next_run_at" in payload:
            monitor["next_run_at"] = str(payload.get("next_run_at") or "")
        if "thread_count" in payload:
            monitor["thread_count"] = parse_thread_count(payload.get("thread_count"))
        if "connection_method" in payload:
            monitor["connection_method"] = normalize_connection_method(payload.get("connection_method"))
        if "auto_connection_fallback" in payload:
            monitor["auto_connection_fallback"] = bool(payload.get("auto_connection_fallback"))
        if "exclusions" in payload:
            exclusions = normalize_patterns(payload.get("exclusions"))
            monitor["exclusions"] = exclusions
        if "product_url_filters" in payload:
            monitor["product_url_filters"] = normalize_patterns(payload.get("product_url_filters"))
        if "product_url_exclusions" in payload:
            monitor["product_url_exclusions"] = normalize_patterns(payload.get("product_url_exclusions"))
        if "extraction_rules" in payload:
            monitor["extraction_rules"] = normalize_extraction_rules(payload.get("extraction_rules"))
        if "selector_settings" in payload:
            monitor["selector_settings"] = normalize_selector_settings(payload.get("selector_settings"))
        if "primary_donor_id" in payload:
            primary_donor_id = str(payload.get("primary_donor_id") or "").strip()
            primary_donor_pk = parse_db_int(primary_donor_id)
            if primary_donor_pk:
                with session_scope() as session:
                    brand = session.get(Brand, parse_db_int(monitor.get("brand_id")))
                    if brand and any(donor.id == primary_donor_pk for donor in brand.donors):
                        brand.primary_donor_id = primary_donor_pk
        sync_brand_runtime_fields(monitor)
        save_news_settings()
    response_monitor = public_news_monitor(monitor)
    if "primary_donor_id" in payload:
        response_monitor["primary_donor_id"] = parse_db_int(payload.get("primary_donor_id"))
    return jsonify({"monitor": response_monitor})


@app.post("/api/news/monitors/<monitor_id>/scan")
def api_scan_news_monitor(monitor_id: str):
    monitor = get_news_monitor(monitor_id)
    if not monitor:
        return jsonify({"error": "Монитор не найден"}), 404
    if not normalize_start_urls(monitor.get("start_urls") or "", allow_empty=True):
        return jsonify({"error": "У выбранного донора не указаны стартовые URL. Заполните поле \"Стартовые URL\" и сохраните настройки."}), 400
    if monitor.get("state", {}).get("status") in {"running", "queued"}:
        return jsonify({"error": "Сканирование уже выполняется"}), 409
    with news_lock:
        monitor["state"] = {
            **make_news_state("queued"),
            "stage": "В очереди запуска",
            "started_at": datetime.now(MSK_TZ).isoformat(timespec="seconds"),
            "last_csv": str(monitor.get("state", {}).get("last_csv") or ""),
        }
        monitor["brand_state"] = dict(monitor["state"])
        sync_brand_runtime_fields(monitor)
        persist_news_monitor_state(monitor, force=True)
        response_monitor = public_news_monitor(monitor)
    if not enqueue_news_scan(monitor_id, manual=True):
        return jsonify({"error": "Сканирование уже выполняется или ожидает очереди"}), 409
    return jsonify({"monitor": response_monitor})


@app.post("/api/news/monitors/<monitor_id>/stop")
def api_stop_news_monitor(monitor_id: str):
    monitor = get_news_monitor(monitor_id)
    if not monitor:
        return jsonify({"error": "Монитор не найден"}), 404
    request_news_stop(monitor_id, "stop")
    with news_lock:
        response_monitor = public_news_monitor(monitor)
    threading.Thread(
        target=add_news_log,
        args=(monitor, "Запрошена остановка сканирования новинок", "warning"),
        daemon=True,
    ).start()
    return jsonify({"monitor": response_monitor})


@app.post("/api/news/monitors/<monitor_id>/pause")
def api_pause_news_monitor(monitor_id: str):
    monitor = get_news_monitor(monitor_id)
    if not monitor:
        return jsonify({"error": "Монитор не найден"}), 404
    request_news_stop(monitor_id, "pause")
    with news_lock:
        response_monitor = public_news_monitor(monitor)
    threading.Thread(
        target=add_news_log,
        args=(monitor, "Запрошена приостановка сканирования новинок с сохранением результата", "warning"),
        daemon=True,
    ).start()
    return jsonify({"monitor": response_monitor})


@app.post("/api/news/monitors/<monitor_id>/resume")
def api_resume_news_monitor(monitor_id: str):
    monitor = get_news_monitor(monitor_id)
    if not monitor:
        return jsonify({"error": "Монитор не найден"}), 404
    if monitor.get("state", {}).get("status") in {"running", "queued", "pausing", "stopping"}:
        return jsonify({"error": "Сканирование уже выполняется"}), 409
    with news_lock:
        monitor["state"] = {**monitor.get("state", {}), "status": "queued", "stage": "Продолжение"}
        monitor["brand_state"] = dict(monitor["state"])
        persist_news_monitor_state(monitor, force=True)
        response_monitor = public_news_monitor(monitor)
    if not enqueue_news_scan(monitor_id, manual=True):
        return jsonify({"error": "Сканирование уже выполняется или ожидает очереди"}), 409
    add_news_log(monitor, "Продолжение сканирования новинок поставлено в очередь", "info")
    return jsonify({"monitor": response_monitor})

@app.post("/api/news/monitors/<monitor_id>/reset-visual")
def api_reset_news_monitor_visual(monitor_id: str):
    monitor = get_news_monitor(monitor_id)
    if not monitor:
        return jsonify({"error": "Монитор не найден"}), 404
    group = clean_text(str(monitor.get("group") or ""))
    brand = clean_text(str(monitor.get("brand") or ""))
    active_statuses = {"running", "queued", "pausing", "stopping"}
    with news_lock:
        brand_monitors = [
            item
            for item in news_settings.get("monitors", [])
            if isinstance(item, dict)
            and clean_text(str(item.get("group") or "")) == group
            and clean_text(str(item.get("brand") or "")) == brand
        ]
        if any(str((item.get("state") or {}).get("status") or "") in active_statuses for item in brand_monitors):
            return jsonify({"error": "Нельзя сбрасывать статус, пока сканирование выполняется."}), 409
        previous_state = monitor.get("state", {}) if isinstance(monitor.get("state"), dict) else {}
        reset_state = make_news_state("idle")
        last_csv = str(previous_state.get("last_csv") or "")
        if last_csv:
            reset_state["last_csv"] = last_csv
        for item in brand_monitors:
            item["state"] = dict(reset_state)
            item["brand_state"] = dict(reset_state)
            item.pop("_last_progress_state", None)
        persist_news_monitor_state(monitor, force=True)
    state_patch = {
        "status": "idle",
        "stage": "",
        "percent": 0,
        "currenturl": "",
        "processed": 0,
        "found_products": 0,
        "candidate_products": 0,
        "compared_products": 0,
        "queue_size": 0,
        "active_tasks": 0,
        "active_urls": [],
        "in_memory_products": 0,
        "availability_skipped": 0,
        "failed_pages": 0,
        "stall_seconds": 0,
        "last_event": "",
        "last_warning": "",
        "new_count": 0,
        "missing_by_feed": [],
        "skipped": 0,
        "last_scan_at": "",
        "error": "",
        "started_at": "",
        "finished_at": "",
        "elapsed_seconds": 0,
    }
    return jsonify({"monitor": {"id": str(monitor_id), "state": state_patch}})

@app.post("/api/news/monitors")
def api_create_news_monitor():
    ensure_storage()
    payload = request.get_json(silent=True) or {}
    urls = normalize_start_urls(payload.get("start_urls") or "", allow_empty=True)
    site_url = str(payload.get("site_url") or "").strip()
    group = clean_text(str(payload.get("group") or "Маржа"))
    brand = clean_text(str(payload.get("brand") or "Новый донор"))
    if payload.get("create_new_brand"):
        with news_lock:
            brand = unique_news_brand_name(group, brand if brand and brand != "Новый донор" else "Новый бренд")
    monitor = make_news_monitor(
        group,
        brand,
        urls,
        site_url,
    )
    with news_lock:
        news_settings.setdefault("monitors", []).append(monitor)
        save_news_settings()
    add_news_log(monitor, "Монитор новинок создан", "success")
    return jsonify({"monitor": public_news_monitor(monitor)})


@app.delete("/api/news/monitors/<monitor_id>")
def api_delete_news_monitor(monitor_id: str):
    monitor = get_news_monitor(monitor_id)
    if not monitor:
        return jsonify({"error": "Монитор не найден"}), 404
    if request.args.get("mode") != "brand":
        group = clean_text(str(monitor.get("group") or ""))
        brand = clean_text(str(monitor.get("brand") or ""))
        with news_lock:
            brand_monitors = [
                item
                for item in news_settings.get("monitors", [])
                if isinstance(item, dict)
                and clean_text(str(item.get("group") or "")) == group
                and clean_text(str(item.get("brand") or "")) == brand
            ]
        if len(brand_monitors) < 2:
            return jsonify({"error": "Нельзя удалить единственного донора бренда"}), 409
    request_news_stop(monitor_id, "stop")
    delete_news_csv_for_monitor(monitor)
    with news_lock:
        monitors = news_settings.get("monitors", [])
        news_settings["monitors"] = [
            item
            for item in monitors
            if isinstance(item, dict) and str(item.get("id")) != monitor_id
        ]
        news_stop_events.pop(monitor_id, None)
        save_news_settings()
    add_news_log(monitor, "Монитор новинок удален", "warning")
    return jsonify({"ok": True, "monitors": [public_news_monitor(item) for item in news_settings.get("monitors", []) if isinstance(item, dict)]})


@app.get("/api/news/monitors/<monitor_id>/download")
def api_download_news_csv(monitor_id: str):
    monitor = get_news_monitor(monitor_id)
    if not monitor:
        return jsonify({"error": "Монитор не найден"}), 404
    state = monitor.get("state", {}) if isinstance(monitor.get("state"), dict) else {}
    state_data = state.get("data", {}) if isinstance(state.get("data"), dict) else {}
    filename = str(state.get("last_csv") or state_data.get("csv") or "")
    path = resolve_export_file(filename)
    if not path:
        return jsonify({"error": "CSV еще не готов"}), 404
    download_name = output_text(path.name)
    return send_file(path, as_attachment=True, download_name=download_name)


@app.get("/api/news/feeds/<source>/<path:filename>")
def api_download_news_feed(source: str, filename: str):
    ensure_storage()
    allowed_names = {
        str(feed.get("filename") or "")
        for feed in news_settings.get("feed_storage", [])
        if isinstance(feed, dict)
        and str(feed.get("source") or "") == source
    }
    if filename not in allowed_names:
        return jsonify({"error": "Фид не найден"}), 404
    feed = next(
        (
            item for item in news_settings.get("feed_storage", [])
            if isinstance(item, dict)
            and str(item.get("source") or "") == source
            and str(item.get("filename") or "") == filename
        ),
        None,
    )
    path = feed_snapshot_path(feed) if isinstance(feed, dict) else None
    if path is None or not path.exists():
        return jsonify({"error": "Фид не найден"}), 404
    return send_file(path, as_attachment=True, download_name=output_text(filename))


@app.get("/api/file-import")
def api_file_import_state():
    ensure_storage()
    return jsonify(public_file_import_state())


@app.patch("/api/file-import")
def api_update_file_import():
    ensure_storage()
    payload = request.get_json(silent=True) or {}
    row = get_file_import_row()
    if "exclusions" in payload:
        row.exclusions = normalize_file_import_exclusions(payload.get("exclusions"))
    if "model_field" in payload:
        row.model_field = clean_text(str(payload.get("model_field") or ""))[:255]
    if "replace_rules" in payload:
        row.replace_rules = normalize_file_import_rules_text(payload.get("replace_rules"))
    if "file" in payload:
        file_payload = payload.get("file")
        if not file_payload:
            row.file = {}
        elif isinstance(file_payload, dict):
            stored_filename = str(file_payload.get("stored_filename") or "").strip()
            base_dir = FILE_IMPORT_DIR.resolve()
            path = (FILE_IMPORT_DIR / stored_filename).resolve()
            if stored_filename and base_dir in path.parents and path.exists() and path.is_file():
                row.file = {
                    "original_filename": output_text(str(file_payload.get("filename") or file_payload.get("original_filename") or path.name)),
                    "stored_filename": path.name,
                    "uploaded_at": str(file_payload.get("uploaded_at") or datetime.fromtimestamp(path.stat().st_mtime, MSK_TZ).isoformat(timespec="seconds")),
                }
    return jsonify(public_file_import_state())


@app.post("/api/file-import")
def api_upload_file_import():
    ensure_storage()
    if is_file_import_active_state(public_file_import_state().get("state")):
        return jsonify({"error": "Дождитесь окончания сравнения файла"}), 409
    uploads = request.files.getlist("file")
    if len(uploads) > 1:
        return jsonify({"error": "Можно загрузить только один файл"}), 400
    upload = uploads[0] if uploads else None
    if not upload or not upload.filename:
        return jsonify({"error": "Файл не выбран"}), 400
    original_filename = output_text(upload.filename)
    suffix = Path(original_filename).suffix.lower()
    if suffix not in FILE_IMPORT_ALLOWED_SUFFIXES:
        return jsonify({"error": "Можно загрузить только CSV, XLS или XLSX"}), 400

    row = get_file_import_row()
    remove_file_import_export(row)
    clear_file_import_storage()
    stored_filename = f"{datetime.now(MSK_TZ).strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}_{safe_filename(Path(original_filename).stem)}{suffix}"
    target = (FILE_IMPORT_DIR / stored_filename).resolve()
    if FILE_IMPORT_DIR.resolve() not in target.parents:
        return jsonify({"error": "Некорректное имя файла"}), 400
    upload.save(target)
    row.file = {
        "original_filename": original_filename,
        "stored_filename": stored_filename,
        "uploaded_at": datetime.now(MSK_TZ).isoformat(timespec="seconds"),
    }
    row.export_path = ""
    row.state = make_file_import_state()
    return jsonify(public_file_import_state())


@app.delete("/api/file-import")
def api_delete_file_import():
    ensure_storage()
    if is_file_import_active_state(public_file_import_state().get("state")):
        return jsonify({"error": "Дождитесь окончания сравнения файла"}), 409
    row = get_file_import_row()
    remove_file_import_export(row)
    clear_file_import_storage()
    row.export_path = ""
    row.file = {}
    row.state = make_file_import_state()
    return jsonify(public_file_import_state())


@app.post("/api/file-import/compare")
def api_compare_file_import():
    ensure_storage()
    try:
        state = start_file_import_compare()
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        return jsonify({"error": f"Ошибка запуска сравнения: {exc}"}), 500
    return jsonify(state)


@app.post("/api/file-import/stop")
def api_stop_file_import():
    ensure_storage()
    with file_import_lock:
        active_thread = bool(file_import_worker_thread and file_import_worker_thread.is_alive())
        if active_thread:
            file_import_stop_event.set()
    row = get_file_import_row()
    state = normalize_file_import_state(getattr(row, "state", {}) or {})
    if is_file_import_active_state(state):
        row.state = {
            **state,
            "status": "stopped",
            "stage": "Остановлено",
            "finished_at": datetime.now(MSK_TZ).isoformat(timespec="seconds"),
        }
    return jsonify(public_file_import_state())


@app.get("/api/file-import/download")
def api_download_file_import_result():
    ensure_storage()
    row = get_file_import_row()
    if is_file_import_active_state(getattr(row, "state", {}) or {}):
        return jsonify({"error": "CSV будет доступен после окончания сравнения"}), 409
    file_meta = row.file if isinstance(row.file, dict) else {}
    filename = str(row.export_path or file_meta.get("result_filename") or "")
    path = resolve_file_import_export_path(filename)
    if not path:
        return jsonify({"error": "CSV еще не готов"}), 404
    return send_file(path, as_attachment=True, download_name=output_text(path.name))


@app.get("/api/projects")
def api_projects():
    ensure_storage()
    summary = request.args.get("summary") == "1"
    with projects_lock:
        return jsonify(
            {
                "projects": [public_project(project, include_details=not summary) for project in projects.values()],
                "connection_methods": public_connection_methods(),
            }
        )


@app.get("/api/projects/<project_id>")
def api_get_project(project_id: str):
    project = get_project(project_id)
    if not project:
        return jsonify({"error": "Проект не найден"}), 404
    return jsonify({"project": public_project(project)})


@app.post("/api/projects")
def api_create_project():
    ensure_storage()
    payload = request.get_json(silent=True) or {}
    name = str(payload.get("name") or f"Проект {len(projects) + 1}").strip()
    start_urls = normalize_start_urls(payload.get("start_urls") or DEFAULT_START_URL)
    project = make_project(name, start_urls)
    with projects_lock:
        projects[project["id"]] = project
        save_projects()
    add_project_log(project, "Проект создан", "success")
    return jsonify({"project": public_project(project)})


@app.patch("/api/projects/<project_id>")
def api_update_project(project_id: str):
    project = get_project(project_id)
    if not project:
        return jsonify({"error": "Проект не найден"}), 404

    payload = request.get_json(silent=True) or {}
    with projects_lock:
        if "name" in payload:
            project["name"] = str(payload.get("name") or project["name"]).strip() or project["name"]
        if "start_urls" in payload:
            project["start_urls"] = normalize_start_urls(payload.get("start_urls"))
        if "product_url_filters" in payload:
            project["product_url_filters"] = normalize_patterns(payload.get("product_url_filters"))
        if "product_url_exclusions" in payload:
            project["product_url_exclusions"] = normalize_patterns(payload.get("product_url_exclusions"))
        if "extraction_rules" in payload:
            project["extraction_rules"] = normalize_extraction_rules(payload.get("extraction_rules"))
        if "thread_count" in payload:
            thread_count = parse_thread_count(payload.get("thread_count"))
            project["thread_count"] = thread_count
            state = dict(project["state"])
            state["thread_count"] = thread_count
            project["state"] = state
        if "connection_method" in payload:
            project["connection_method"] = normalize_connection_method(payload.get("connection_method"))
        if "auto_connection_fallback" in payload:
            project["auto_connection_fallback"] = bool(payload.get("auto_connection_fallback"))
        if "persist_profile" in payload:
            project["persist_profile"] = bool(payload.get("persist_profile"))
        if "auto_cleanup" in payload:
            project["auto_cleanup"] = bool(payload.get("auto_cleanup"))
        reset_project_state_after_form_save(project)
        save_projects()
    return jsonify({"project": public_project(project)})


@app.delete("/api/projects/<project_id>")
def api_delete_project(project_id: str):
    project = get_project(project_id)
    if not project:
        return jsonify({"error": "Проект не найден"}), 404
    with projects_lock:
        if len(projects) <= 1:
            return jsonify({"error": "Нельзя удалить последний проект"}), 400
        stop_event = project.get("stop_event")
        if isinstance(stop_event, threading.Event):
            stop_event.set()
        projects.pop(project_id, None)
        save_projects()
    return jsonify({"ok": True})


@app.get("/api/projects/<project_id>/exclusions")
def api_project_exclusions(project_id: str):
    project = get_project(project_id)
    if not project:
        return jsonify({"error": "Проект не найден"}), 404
    return jsonify({"exclusions": project.get("exclusions", [])})


@app.post("/api/projects/<project_id>/exclusions")
def api_project_add_exclusion(project_id: str):
    project = get_project(project_id)
    if not project:
        return jsonify({"error": "Проект не найден"}), 404
    payload = request.get_json(silent=True) or {}
    pattern = str(payload.get("pattern", "")).strip()
    if not pattern:
        return jsonify({"error": "Пустое исключение"}), 400
    with projects_lock:
        exclusions = project.setdefault("exclusions", [])
        if pattern not in exclusions:
            exclusions.append(pattern)
            save_projects()
    return jsonify({"exclusions": project.get("exclusions", [])})


@app.delete("/api/projects/<project_id>/exclusions/<int:index>")
def api_project_delete_exclusion(project_id: str, index: int):
    project = get_project(project_id)
    if not project:
        return jsonify({"error": "Проект не найден"}), 404
    with projects_lock:
        exclusions = project.setdefault("exclusions", [])
        if index < 0 or index >= len(exclusions):
            return jsonify({"error": "Исключение не найдено"}), 404
        exclusions.pop(index)
        save_projects()
    return jsonify({"exclusions": project.get("exclusions", [])})


@app.get("/api/projects/<project_id>/product-url-filters")
def api_project_product_url_filters(project_id: str):
    project = get_project(project_id)
    if not project:
        return jsonify({"error": "Проект не найден"}), 404
    return jsonify({"product_url_filters": project.get("product_url_filters", [])})


@app.post("/api/projects/<project_id>/product-url-filters")
def api_project_add_product_url_filter(project_id: str):
    project = get_project(project_id)
    if not project:
        return jsonify({"error": "Проект не найден"}), 404
    payload = request.get_json(silent=True) or {}
    pattern = str(payload.get("pattern", "")).strip()
    if not pattern:
        return jsonify({"error": "Пустой фильтр ссылки"}), 400
    with projects_lock:
        filters = project.setdefault("product_url_filters", [])
        if pattern not in filters:
            filters.append(pattern)
            save_projects()
    return jsonify({"product_url_filters": project.get("product_url_filters", [])})


@app.delete("/api/projects/<project_id>/product-url-filters/<int:index>")
def api_project_delete_product_url_filter(project_id: str, index: int):
    project = get_project(project_id)
    if not project:
        return jsonify({"error": "Проект не найден"}), 404
    with projects_lock:
        filters = project.setdefault("product_url_filters", [])
        if index < 0 or index >= len(filters):
            return jsonify({"error": "Фильтр ссылки не найден"}), 404
        filters.pop(index)
        save_projects()
    return jsonify({"product_url_filters": project.get("product_url_filters", [])})


@app.get("/api/projects/<project_id>/product-url-exclusions")
def api_project_product_url_exclusions(project_id: str):
    project = get_project(project_id)
    if not project:
        return jsonify({"error": "Проект не найден"}), 404
    return jsonify({"product_url_exclusions": project.get("product_url_exclusions", [])})


@app.post("/api/projects/<project_id>/product-url-exclusions")
def api_project_add_product_url_exclusion(project_id: str):
    project = get_project(project_id)
    if not project:
        return jsonify({"error": "Проект не найден"}), 404
    payload = request.get_json(silent=True) or {}
    pattern = str(payload.get("pattern", "")).strip()
    if not pattern:
        return jsonify({"error": "Пустое исключение товарной ссылки"}), 400
    with projects_lock:
        exclusions = project.setdefault("product_url_exclusions", [])
        if pattern not in exclusions:
            exclusions.append(pattern)
            save_projects()
    return jsonify({"product_url_exclusions": project.get("product_url_exclusions", [])})


@app.delete("/api/projects/<project_id>/product-url-exclusions/<int:index>")
def api_project_delete_product_url_exclusion(project_id: str, index: int):
    project = get_project(project_id)
    if not project:
        return jsonify({"error": "Проект не найден"}), 404
    with projects_lock:
        exclusions = project.setdefault("product_url_exclusions", [])
        if index < 0 or index >= len(exclusions):
            return jsonify({"error": "Исключение товарной ссылки не найдено"}), 404
        exclusions.pop(index)
        save_projects()
    return jsonify({"product_url_exclusions": project.get("product_url_exclusions", [])})


def close_project_browser_session(project: Dict[str, object]) -> None:
    crawler = project.get("crawler")
    if isinstance(crawler, ProductSiteCrawler):
        crawler.close_browser_sessions()


def project_profile_storage_dir(project: Dict[str, object]) -> Path:
    project_id = safe_filename(str(project.get("id") or "project"))
    return PROJECT_PROFILE_DIR / project_id


def project_should_keep_browser_profile(project: Dict[str, object]) -> bool:
    if bool(project.get("persist_profile", False)):
        return True
    method = normalize_connection_method(project.get("connection_method"))
    return method == "protected-site" or bool(project.get("auto_connection_fallback", True))


def project_browser_profile_dir(project: Dict[str, object]) -> Optional[Path]:
    if not project_should_keep_browser_profile(project):
        return None
    return project_profile_storage_dir(project)


def cleanup_project_profile_if_disabled(project: Dict[str, object]) -> None:
    if project_should_keep_browser_profile(project):
        return
    profile_dir = project_profile_storage_dir(project)
    try:
        root = PROJECT_PROFILE_DIR.resolve()
        target = profile_dir.resolve()
        if target == root or root not in target.parents or not target.exists():
            return
        shutil.rmtree(target, ignore_errors=True)
    except Exception:
        pass


def start_project(project: Dict[str, object], resume: bool = False) -> Dict[str, object]:
    task_key = ("project", str(project["id"]))
    if scan_dispatcher.contains(task_key):
        raise RuntimeError("Сбор уже выполняется или ожидает очереди")
    worker = project.get("worker_thread")
    state = project.get("state", {})
    if isinstance(worker, threading.Thread) and worker.is_alive():
        if state.get("status") == "running":
            raise RuntimeError("Сбор уже выполняется")
        worker.join(timeout=2)
        if worker.is_alive():
            raise RuntimeError("Предыдущий поток еще завершается. Повторите через несколько секунд.")

    project["stop_event"] = threading.Event()
    project["finish_event"] = threading.Event()
    project["stop_mode"] = ""
    project["run_id"] = int(project.get("run_id", 0)) + 1

    crawler = project.get("crawler") if resume else None
    profile_dir = project_browser_profile_dir(project)
    runtime_thread_count = project_runtime_thread_count(project)
    if crawler:
        crawler.close_browser_sessions()
        cleanup_project_profile_if_disabled(project)
        crawler.run_id = int(project["run_id"])
        crawler.stop_signal = project["stop_event"]
        crawler.finish_signal = project["finish_event"]
        crawler.thread_count = runtime_thread_count
        crawler.exclusions = list(project.get("exclusions", DEFAULT_EXCLUSIONS))
        crawler.extraction_rules = normalize_extraction_rules(project.get("extraction_rules", {}))
        crawler.product_url_filters = product_url_filter_patterns(project.get("product_url_filters", []), crawler.extraction_rules)
        crawler.product_url_exclusions = normalize_patterns(project.get("product_url_exclusions", []))
        crawler.connection_method = normalize_connection_method(project.get("connection_method"))
        crawler.auto_connection_fallback = bool(project.get("auto_connection_fallback", True))
        crawler.profile_dir = profile_dir
        crawler.browser_session = BotasaurusBrowserSession(
            project["stop_event"],
            crawler.thread_count,
            profile_dir=crawler.profile_dir,
        )
        crawler.debug_visible_session = BotasaurusDebugVisibleSession(
            project["stop_event"],
            str(crawler.profile_dir) if crawler.profile_dir is not None else "protected_sites_debug_visible",
        )
        crawler.active_connection_method = crawler.connection_method
        crawler.connection_method_state["active_method"] = crawler.connection_method
        crawler.excel_finalized = False
    else:
        cleanup_project_profile_if_disabled(project)
        reset_project_state(project, "queued")
        crawler = ProductSiteCrawler(
            list(project.get("start_urls", [DEFAULT_START_URL])),
            int(project["run_id"]),
            project["stop_event"],
            project["finish_event"],
            runtime_thread_count,
            project=project,
            exclusions=list(project.get("exclusions", DEFAULT_EXCLUSIONS)),
            product_url_filters=list(project.get("product_url_filters", [])),
            product_url_exclusions=list(project.get("product_url_exclusions", [])),
            extraction_rules=normalize_extraction_rules(project.get("extraction_rules", {})),
            connection_method=project.get("connection_method", "requests"),
            auto_connection_fallback=bool(project.get("auto_connection_fallback", True)),
            profile_dir=profile_dir,
        )
        project["crawler"] = crawler

    def target() -> None:
        try:
            if resume:
                update_project_state(project, status="running", error="")
            else:
                reset_project_state(project, "running")
            crawler.run(resume=resume)
        except Exception as exc:  # noqa: BLE001
            update_project_state(project, status="error", error=str(exc), currenturl="", download_ready=False)
            add_project_log(project, f"Критическая ошибка: {exc}", "error")

    def on_start(thread: threading.Thread) -> None:
        project["worker_thread"] = thread

    def on_finish(thread: threading.Thread) -> None:
        if project.get("worker_thread") is thread:
            project["worker_thread"] = None

    if not scan_dispatcher.enqueue(task_key, target, on_start=on_start, on_finish=on_finish):
        raise RuntimeError("Сбор уже выполняется или ожидает очереди")
    add_project_log(project, "Продолжение поставлено в очередь" if resume else "Сбор поставлен в очередь запуска", "info")
    return project["state"]


@app.post("/api/projects/<project_id>/start")
def api_project_start(project_id: str):
    project = get_project(project_id)
    if not project:
        return jsonify({"error": "Проект не найден"}), 404

    payload = request.get_json(silent=True) or {}
    with projects_lock:
        if "start_urls" in payload:
            project["start_urls"] = normalize_start_urls(payload.get("start_urls"))
        if "product_url_filters" in payload:
            project["product_url_filters"] = normalize_patterns(payload.get("product_url_filters"))
        if "product_url_exclusions" in payload:
            project["product_url_exclusions"] = normalize_patterns(payload.get("product_url_exclusions"))
        if "extraction_rules" in payload:
            project["extraction_rules"] = normalize_extraction_rules(payload.get("extraction_rules"))
        if "thread_count" in payload:
            project["thread_count"] = parse_thread_count(payload.get("thread_count"))
        if "connection_method" in payload:
            project["connection_method"] = normalize_connection_method(payload.get("connection_method"))
        if "auto_connection_fallback" in payload:
            project["auto_connection_fallback"] = bool(payload.get("auto_connection_fallback"))
        if "persist_profile" in payload:
            project["persist_profile"] = bool(payload.get("persist_profile"))
        save_projects()

    try:
        state = start_project(project)
    except RuntimeError as exc:
        return jsonify({"error": str(exc)}), 409
    return jsonify(state)


@app.post("/api/projects/<project_id>/pause")
def api_project_pause(project_id: str):
    project = get_project(project_id)
    if not project:
        return jsonify({"error": "Проект не найден"}), 404
    status = project.get("state", {}).get("status")
    if status not in {"running", "paused"}:
        return jsonify({"error": "Сбор не выполняется"}), 409
    finish_event = project.get("finish_event")
    stop_event = project.get("stop_event")
    project["stop_mode"] = "pause"
    if isinstance(finish_event, threading.Event):
        finish_event.set()
    if status == "running" and isinstance(stop_event, threading.Event):
        stop_event.set()
    crawler = project.get("crawler")
    if crawler:
        close_project_browser_session(project)
        crawler.finish_with_excel(partial=True)
    add_project_log(project, "Сбор приостановлен с формированием CSV", "warning")
    return jsonify(project["state"])


@app.post("/api/projects/<project_id>/soft-pause")
def api_project_soft_pause(project_id: str):
    project = get_project(project_id)
    if not project:
        return jsonify({"error": "Проект не найден"}), 404
    if project.get("state", {}).get("status") != "running":
        return jsonify({"error": "Сбор не выполняется"}), 409
    stop_event = project.get("stop_event")
    project["stop_mode"] = "pause"
    if isinstance(stop_event, threading.Event):
        stop_event.set()
    close_project_browser_session(project)
    update_project_state(project, error="Ставлю сбор на паузу...", currenturl="")
    add_project_log(project, "Запрошена обычная пауза", "warning")
    return jsonify(project["state"])


@app.post("/api/projects/<project_id>/resume")
def api_project_resume(project_id: str):
    project = get_project(project_id)
    if not project:
        return jsonify({"error": "Проект не найден"}), 404
    status = project.get("state", {}).get("status")
    if status not in {"paused", "partial"}:
        return jsonify({"error": "Продолжить можно только после паузы"}), 409
    try:
        state = start_project(project, resume=True)
    except RuntimeError as exc:
        return jsonify({"error": str(exc)}), 409
    return jsonify(state)


@app.post("/api/projects/<project_id>/stop")
def api_project_stop(project_id: str):
    project = get_project(project_id)
    if not project:
        return jsonify({"error": "Проект не найден"}), 404
    scan_dispatcher.cancel(("project", str(project["id"])))
    stop_event = project.get("stop_event")
    if isinstance(stop_event, threading.Event):
        stop_event.set()
    close_project_browser_session(project)
    with projects_lock:
        project["stop_mode"] = "stop"
        project["run_id"] = int(project.get("run_id", 0)) + 1
        project["crawler"] = None
        state = dict(project.get("state") or make_state(parse_thread_count(project.get("thread_count", 4))))
        state.update(
            {
                "status": "idle",
                "currenturl": "",
                "active_urls": [],
                "active_tasks": 0,
                "queue_size": 0,
                "error": "",
                "eta_seconds": None,
                "finished_at": now_iso(),
                "paused_with_result": False,
            }
        )
        project["state"] = state
        save_projects()
    add_project_log(project, "Сбор остановлен", "warning")
    return jsonify(project["state"])


@app.post("/api/projects/<project_id>/restart")
def api_project_restart(project_id: str):
    project = get_project(project_id)
    if not project:
        return jsonify({"error": "Проект не найден"}), 404
    stop_event = project.get("stop_event")
    if isinstance(stop_event, threading.Event):
        stop_event.set()
    close_project_browser_session(project)
    worker = project.get("worker_thread")
    if isinstance(worker, threading.Thread) and worker.is_alive():
        with projects_lock:
            project["stop_mode"] = "stop"
            project["run_id"] = int(project.get("run_id", 0)) + 1
            project["crawler"] = None
        worker.join(timeout=3)
        if worker.is_alive():
            return jsonify({"error": "Предыдущий сбор еще завершается. Повторите перезапуск через несколько секунд."}), 409
    try:
        state = start_project(project)
    except RuntimeError as exc:
        return jsonify({"error": str(exc)}), 409
    return jsonify(state)


@app.get("/api/logs")
def api_logs():
    ensure_storage()
    global LOG_AUTO_CLEANUP
    auto_cleanup = get_log_auto_cleanup()
    LOG_AUTO_CLEANUP = auto_cleanup

    json_logs = read_logs_file()
    if auto_cleanup:
        cutoff = time.time() - 7 * 24 * 60 * 60
        filtered_logs = [
            item
            for item in json_logs
            if is_recent_log_entry(item, cutoff)
        ]
        if len(filtered_logs) != len(json_logs):
            json_logs = filtered_logs
            write_logs_file(json_logs)
        prune_old_log_files(cutoff)
        with projects_lock:
            for project in projects.values():
                logs = project.get("logs", [])
                project["logs"] = [
                    item
                    for item in logs
                    if is_recent_log_entry(item, cutoff)
                ]
        with news_lock:
            logs = news_settings.get("logs", [])
            news_settings["logs"] = [
                item
                for item in logs
                if is_recent_log_entry(item, cutoff)
            ]

    all_logs = combined_log_entries()
    all_logs.sort(key=lambda item: item.get("time", ""))
    try:
        limit = max(1, min(int(request.args.get("limit") or 200), 1000))
    except (TypeError, ValueError):
        limit = 200
    try:
        page = max(1, int(request.args.get("page") or 1))
    except (TypeError, ValueError):
        page = 1
    total = len(all_logs)
    end = max(0, total - (page - 1) * limit)
    start = max(0, end - limit)
    page_logs = all_logs[start:end]
    return jsonify(
        {
            "logs": page_logs,
            "logs_total": total,
            "logs_page": page,
            "logs_limit": limit,
            "auto_cleanup": auto_cleanup,
            "logs_signature": logs_signature(),
        }
    )

@app.delete("/api/logs")
def api_clear_logs():
    ensure_storage()
    with projects_lock:
        for project in projects.values():
            project["logs"] = []
    with news_lock:
        news_settings["logs"] = []
        save_news_settings()
    clear_runtime_log_files()
    return jsonify({"ok": True})


@app.post("/api/logs/settings")
def api_logs_settings():
    ensure_storage()
    payload = request.get_json(silent=True) or {}
    auto_cleanup = set_log_auto_cleanup(bool(payload.get("auto_cleanup")))
    with news_lock:
        news_settings["auto_cleanup"] = auto_cleanup
    return jsonify({"auto_cleanup": auto_cleanup})


@app.post("/start")
def start_scan():
    global active_crawler, active_finish_event, active_run_id, active_stop_event, worker_thread

    current_status = snapshot_state()["status"]
    if current_status == "running" and worker_thread and worker_thread.is_alive():
        return jsonify({"error": "Сбор уже выполняется"}), 409

    payload = request.get_json(silent=True) or {}
    start_url = str(payload.get("start_url") or DEFAULT_START_URL).strip()
    thread_count = parse_thread_count(payload.get("thread_count"))

    with state_lock:
        active_run_id += 1
        run_id = active_run_id
        active_stop_event = threading.Event()
        active_finish_event = threading.Event()
        stop_signal = active_stop_event
        finish_signal = active_finish_event

    reset_state("running", thread_count=thread_count)

    crawler = ProductSiteCrawler([start_url], run_id, stop_signal, finish_signal, thread_count)
    active_crawler = crawler

    def target() -> None:
        try:
            crawler.run()
        except Exception as exc:  # noqa: BLE001 - показываем ошибку пользователю в интерфейсе.
            update_state(run_id, status="error", error=str(exc), currenturl="", download_ready=False)

    worker_thread = threading.Thread(target=target, daemon=True)
    worker_thread.start()
    return jsonify(snapshot_state())


@app.post("/stop")
def stop_scan():
    global active_crawler, active_run_id

    active_stop_event.set()
    with state_lock:
        active_run_id += 1
        active_crawler = None
    reset_state("idle")
    return jsonify(snapshot_state())


@app.post("/pause")
def pause_scan_with_result():
    global active_crawler, active_run_id

    if snapshot_state()["status"] != "running":
        return jsonify({"error": "Сбор не выполняется"}), 409

    active_finish_event.set()
    active_stop_event.set()
    crawler = active_crawler
    if crawler:
        crawler.finish_with_excel(partial=True)
        with state_lock:
            active_run_id += 1
            active_crawler = None
        return jsonify(snapshot_state())

    update_state(
        active_run_id,
        error="Останавливаю сбор и формирую Excel по уже найденным товарам...",
        currenturl="",
    )
    return jsonify(snapshot_state())


@app.post("/restart")
def restart_scan():
    global active_crawler, active_run_id

    active_stop_event.set()
    with state_lock:
        active_run_id += 1
        active_crawler = None
    reset_state("idle")
    return start_scan()


@app.get("/progress")
def progress_stream():
    include_projects = request.args.get("projects", "1") == "1"
    include_news = request.args.get("news") == "1"
    if request.args.get("once") == "1":
        return jsonify(progress_payload(include_projects, include_news))

    def stream():
        last_projects_signature = ""
        last_news_signature = ""
        last_logs_signature = ""
        last_connection_methods_signature = ""
        while True:
            ensure_storage()
            payload: Dict[str, object] = {}

            if include_projects:
                projects_payload = projects_progress_payload()
                projects_signature = payload_signature(projects_payload)
                if projects_signature != last_projects_signature:
                    payload["projects"] = projects_payload
                    last_projects_signature = projects_signature

            if include_news:
                news_payload = news_progress_payload()
                news_signature = payload_signature(news_payload)
                if news_signature != last_news_signature:
                    payload["news"] = news_payload
                    last_news_signature = news_signature

            connection_methods = public_connection_methods()
            connection_methods_signature = payload_signature(connection_methods)
            if connection_methods_signature != last_connection_methods_signature:
                payload["connection_methods"] = connection_methods
                last_connection_methods_signature = connection_methods_signature

            current_logs_signature = logs_signature()
            if current_logs_signature != last_logs_signature:
                payload["logs_signature"] = current_logs_signature
                last_logs_signature = current_logs_signature

            if payload:
                data = json.dumps(payload, ensure_ascii=False)
                yield f"event: progress\ndata: {data}\n\n"
            else:
                yield ": keep-alive\n\n"
            time.sleep(PROGRESS_STREAM_INTERVAL_SECONDS)

    response = Response(stream(), mimetype="text/event-stream")
    response.headers["Cache-Control"] = "no-cache"
    response.headers["X-Accel-Buffering"] = "no"
    return response


@app.get("/download")
def download_excel():
    current_state = snapshot_state()
    filename = str(current_state.get("filename") or "")
    path = EXPORT_DIR / filename
    if not filename or not path.exists():
        return jsonify({"error": "Файл еще не готов"}), 404
    return send_file(path, as_attachment=True, download_name=output_text(filename))


@app.get("/api/projects/<project_id>/download")
def download_project_csv(project_id: str):
    project = get_project(project_id)
    if not project:
        return jsonify({"error": "Проект не найден"}), 404
    current_state = project.get("state", {})
    filename = str(current_state.get("filename") or "")
    path = EXPORT_DIR / filename
    if not filename or not path.exists():
        return jsonify({"error": "Файл еще не готов"}), 404
    return send_file(path, as_attachment=True, download_name=output_text(filename))


if __name__ == "__main__":
    ensure_storage()
    port = env_int("PORT", 5000, minimum=1, maximum=65535)
    if env_str("DEBUG_HANG_DUMP", "0") == "1":
        faulthandler.dump_traceback_later(10, repeat=True)
    from socketserver import ThreadingMixIn
    from wsgiref.simple_server import WSGIRequestHandler, WSGIServer, make_server

    class ThreadingWSGIServer(ThreadingMixIn, WSGIServer):
        daemon_threads = True

    with make_server("127.0.0.1", port, app, server_class=ThreadingWSGIServer, handler_class=WSGIRequestHandler) as server:
        print(f"Serving on http://127.0.0.1:{port}", flush=True)
        server.serve_forever()




