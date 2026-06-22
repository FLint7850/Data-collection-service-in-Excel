"""Spreadsheet import and comparison against feeds."""



from parser_app.runtime import *  # noqa: F401,F403



def safe_filename(value: str) -> str:
    value = output_text(value)
    cleaned = re.sub(r"[^A-Za-z\u0400-\u04FF0-9_-]+", "_", value, flags=re.IGNORECASE).strip("_")
    return cleaned or "project"

def clear_file_import_storage() -> None:
    FILE_IMPORT_DIR.mkdir(parents=True, exist_ok=True)
    for path in FILE_IMPORT_DIR.iterdir():
        if path.is_file():
            try:
                path.unlink()
            except OSError:
                continue

def stored_file_import_files() -> List[Path]:
    FILE_IMPORT_DIR.mkdir(parents=True, exist_ok=True)
    return sorted(
        [
            path
            for path in FILE_IMPORT_DIR.iterdir()
            if path.is_file() and path.suffix.lower() in FILE_IMPORT_ALLOWED_SUFFIXES
        ],
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    )

def get_file_import_row(db_session=None) -> FileImport:
    db = db_session or g.db
    row = db.get(FileImport, 1)
    if row is None:
        row = FileImport(
            id=1,
            exclusions=[],
            model_field="",
            replace_rules="",
            file={},
        )
        db.add(row)
        db.flush()
    else:
        normalized_exclusions = normalize_file_import_exclusions(row.exclusions)
        if row.exclusions != normalized_exclusions:
            row.exclusions = normalized_exclusions
            db.flush()
        normalized_model_field = clean_text(str(getattr(row, "model_field", "") or ""))
        if row.model_field != normalized_model_field:
            row.model_field = normalized_model_field
            db.flush()
        normalized_replace_rules = normalize_file_import_rules_text(getattr(row, "replace_rules", ""))
        if row.replace_rules != normalized_replace_rules:
            row.replace_rules = normalized_replace_rules
            db.flush()
    if not isinstance(row.file, dict) or not row.file.get("stored_filename"):
        stored_files = stored_file_import_files()
        if stored_files:
            path = stored_files[0]
            row.file = {
                "original_filename": path.name.split("_", 3)[-1] if "_" in path.name else path.name,
                "stored_filename": path.name,
                "uploaded_at": datetime.fromtimestamp(path.stat().st_mtime, MSK_TZ).isoformat(timespec="seconds"),
            }
            db.flush()
    return row

def current_file_import_path() -> Optional[Path]:
    row = get_file_import_row()
    file_meta = row.file if isinstance(row.file, dict) else {}
    filename = str(file_meta.get("stored_filename") or "").strip()
    if not filename:
        return None
    base_dir = FILE_IMPORT_DIR.resolve()
    path = (FILE_IMPORT_DIR / filename).resolve()
    if base_dir not in path.parents or not path.exists() or not path.is_file():
        return None
    return path

def resolve_file_import_export_path(value: str) -> Optional[Path]:
    filename = Path(str(value or "")).name
    if not filename:
        return None
    path = (EXPORT_DIR / filename).resolve()
    if EXPORT_DIR.resolve() not in path.parents or not path.exists() or not path.is_file():
        return None
    return path

def remove_file_import_export(row: FileImport) -> None:
    candidates = [str(getattr(row, "export_path", "") or "")]
    file_meta = row.file if isinstance(row.file, dict) else {}
    candidates.append(str(file_meta.get("result_filename") or ""))
    for candidate in candidates:
        path = resolve_file_import_export_path(candidate)
        if path:
            try:
                path.unlink()
            except OSError:
                pass

def public_file_import_state() -> Dict[str, object]:
    row = get_file_import_row()
    path = current_file_import_path()
    file_meta = row.file if isinstance(row.file, dict) else {}
    exclusions = normalize_file_import_exclusions(row.exclusions)
    exclusions_text = "\n".join(exclusions)
    model_field = clean_text(str(row.model_field or ""))
    replace_rules = normalize_file_import_rules_text(row.replace_rules)
    result_filename = Path(str(row.export_path or file_meta.get("result_filename") or "")).name
    if not path:
        return {
            "file": None,
            "exclusions": exclusions_text,
            "exclusions_list": exclusions,
            "model_field": model_field,
            "replace_rules": replace_rules,
            "result_filename": result_filename,
            "result_ready": bool(resolve_file_import_export_path(result_filename)),
        }
    stat = path.stat()
    return {
        "exclusions": exclusions_text,
        "exclusions_list": exclusions,
        "model_field": model_field,
        "replace_rules": replace_rules,
        "result_filename": result_filename,
        "result_ready": bool(resolve_file_import_export_path(result_filename)),
        "file": {
            "filename": output_text(str(file_meta.get("original_filename") or path.name)),
            "stored_filename": path.name,
            "size": stat.st_size,
            "uploaded_at": str(file_meta.get("uploaded_at") or datetime.fromtimestamp(stat.st_mtime, MSK_TZ).isoformat(timespec="seconds")),
        }
    }

def decode_file_import_csv(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1251", "windows-1251"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")

def normalize_file_import_header(value: object) -> str:
    return clean_text(str(value or "")).casefold()

def file_import_column_index(headers: List[object], column_name: str) -> int:
    expected = normalize_file_import_header(column_name)
    if not expected:
        raise ValueError("Укажите название столбца модели")
    for index, header in enumerate(headers):
        if normalize_file_import_header(header) == expected:
            return index
    raise ValueError(f"Столбец модели не найден: {column_name}")

def file_import_optional_brand_index(headers: List[object]) -> Optional[int]:
    brand_names = {"brand", "бренд", "manufacturer", "производитель", "vendor", "марка"}
    for index, header in enumerate(headers):
        if normalize_file_import_header(header) in brand_names:
            return index
    return None

def read_file_import_rows(path: Path, model_field: str) -> List[Dict[str, object]]:
    if path.suffix.lower() == ".csv":
        text = decode_file_import_csv(path.read_bytes())
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,|\t,")
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = ";"
        rows = list(csv.reader(io.StringIO(text), dialect))
    elif path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        workbook.close()
    else:
        raise ValueError("Можно обработать только CSV или XLSX")

    header_index = next((index for index, row in enumerate(rows) if any(clean_text(str(cell or "")) for cell in row)), None)
    if header_index is None:
        return []
    headers = rows[header_index]
    model_index = file_import_column_index(headers, model_field)
    brand_index = file_import_optional_brand_index(headers)
    result: List[Dict[str, object]] = []
    for row_number, row in enumerate(rows[header_index + 1:], start=header_index + 2):
        source = clean_text(str(row[model_index] if model_index < len(row) else ""))
        if not source:
            continue
        brand = clean_text(str(row[brand_index] if brand_index is not None and brand_index < len(row) else ""))
        result.append({"row_number": row_number, "name": source, "brand": brand})
    return result

def prepare_file_import_model(value: str, replace_rules: str) -> str:
    return prepare_rule_model(
        str(value or ""),
        {"model_replace_rules": normalize_file_import_rules_text(replace_rules)},
    )

def technical_clean_model_text(value: object, remove_brackets: bool = True) -> str:
    text = BeautifulSoup(str(value or ""), "html.parser").get_text(" ", strip=True)
    text = html_lib.unescape(text)
    text = text.replace("\xa0", " ").replace("\u2009", " ")
    text = re.sub(r"[–—−]", "-", text)
    text = re.sub(r"\s*([/_.+\-])\s*", r"\1", text)
    if remove_brackets:
        text = re.sub(r"\([^)]*\)|\[[^\]]*\]|\{[^}]*\}", " ", text)
    return clean_text(text)

def normalize_compare_key(value: object) -> str:
    text = technical_clean_model_text(value, remove_brackets=False)
    text = re.sub(r"[\"'`«»]", "", text)
    text = re.sub(r"\s+", " ", text).strip(" .,/\\_-+")
    return normalize_model_key(text)

def compact_compare_key(value: object) -> str:
    return re.sub(r"[\s./_+\-]+", "", normalize_compare_key(value))

def visual_compare_key(value: object) -> str:
    return compact_compare_key(str(value or "").translate(VISUAL_MODEL_TRANSLATION))

def compare_keys_for_value(value: object) -> Dict[str, str]:
    original = normalize_compare_key(value)
    compact = compact_compare_key(value)
    visual = visual_compare_key(value)
    keys = {
        "original": original,
        "normalized": original,
        "compact": compact,
        "visual": visual,
    }
    return {kind: key for kind, key in keys.items() if key}

def file_import_exclusion_matches(value: str, brand: str, exclusions: Iterable[str]) -> bool:
    haystack = f"{value} {brand}".casefold()
    visual_haystack = f"{value} {brand}".translate(VISUAL_MODEL_TRANSLATION).casefold()
    for exclusion in exclusions:
        pattern = clean_text(str(exclusion or ""))
        if not pattern:
            continue
        if pattern.casefold() in haystack or pattern.translate(VISUAL_MODEL_TRANSLATION).casefold() in visual_haystack:
            return True
    return False

def known_file_import_brands() -> List[str]:
    try:
        with session_scope() as session:
            rows = session.execute(select(Brand.name)).scalars().all()
    except Exception:
        rows = []
    brands = [clean_text(str(row or "")) for row in rows if clean_text(str(row or ""))]
    return sorted(set(brands), key=len, reverse=True)

def brand_match_pattern(brand: str) -> str:
    parts = [re.escape(part) for part in re.split(r"\s*&\s*|\s+", clean_text(brand)) if part]
    if not parts:
        return ""
    return r"\s*&\s*".join(parts) if "&" in brand else r"\s+".join(parts)

def find_brand_in_name(name: str, explicit_brand: str = "") -> str:
    if explicit_brand:
        return explicit_brand
    for brand in known_file_import_brands():
        pattern = brand_match_pattern(brand)
        if pattern and re.search(rf"(?<![A-Za-zА-Яа-яЁё0-9]){pattern}(?![A-Za-zА-Яа-яЁё0-9])", name, flags=re.IGNORECASE):
            return brand
    return ""

def tail_after_brand(name: str, brand: str) -> str:
    text = technical_clean_model_text(name)
    if not brand:
        return text
    pattern = brand_match_pattern(brand)
    match = re.search(pattern, text, flags=re.IGNORECASE) if pattern else None
    if match:
        return clean_text(text[match.end():])
    visual_text = text.translate(VISUAL_MODEL_TRANSLATION)
    visual_brand = brand.translate(VISUAL_MODEL_TRANSLATION)
    visual_pattern = brand_match_pattern(visual_brand)
    visual_match = re.search(visual_pattern, visual_text, flags=re.IGNORECASE) if visual_pattern else None
    if visual_match:
        return clean_text(text[visual_match.end():])
    return text

def is_measurement_token(value: str) -> bool:
    """Recognize standalone technical units without mistaking Latin model suffixes for litres."""
    return bool(
        re.fullmatch(
            r"\d+(?:[.,]\d+)?(?:ВТ|W|BT|В|V|B|Л|МЛ|КГ|KG|Г|G|СМ|CM|ММ|MM)"
            r"(?:/\d+(?:[.,]\d+)?(?:ВТ|W|BT|В|V|B|Л|МЛ|КГ|KG|Г|G|СМ|CM|ММ|MM))*",
            value.upper(),
        )
    )

def model_signal_token(token: str) -> bool:
    value = clean_text(token)
    if not value:
        return False
    if is_measurement_token(value):
        return False
    has_digit = bool(re.search(r"\d", value))
    has_latin = bool(re.search(r"[A-Za-z]", value))
    has_cyrillic = bool(re.search(r"[А-Яа-яЁё]", value))
    return has_latin or (has_digit and has_cyrillic) or bool(re.search(r"[A-Za-zА-Яа-яЁё]\d|\d[A-Za-zА-Яа-яЁё]", value))

def candidate_until_russian_description(value: str) -> str:
    tokens = value.split()
    kept: List[str] = []
    seen_signal = False
    for token in tokens:
        if seen_signal and is_measurement_token(token):
            break
        if model_signal_token(token):
            kept.append(token)
            seen_signal = True
            continue
        if seen_signal and re.fullmatch(r"[А-Яа-яЁё][А-Яа-яЁё/\-]*", token):
            break
        kept.append(token)
    return clean_text(" ".join(kept))

def first_model_block(value: str) -> str:
    tokens = re.findall(r"[A-Za-zА-Яа-яЁё0-9]+(?:[./_+\-][A-Za-zА-Яа-яЁё0-9]+)*", value)
    best: List[str] = []
    current: List[str] = []
    for token in tokens:
        if model_signal_token(token):
            current.append(token)
            continue
        if current:
            if len(" ".join(current)) > len(" ".join(best)):
                best = list(current)
            current = []
    if current and len(" ".join(current)) > len(" ".join(best)):
        best = current
    return clean_text(" ".join(best))

def normalize_candidate_display(value: str) -> str:
    text = technical_clean_model_text(value)
    text = re.sub(r"\s+", " ", text).strip(" .,/\\_-+")
    tokens = text.split()
    normalized_tokens = [
        token.translate(VISUAL_MODEL_TRANSLATION).upper() if any(char.isdigit() for char in token) else token
        for token in tokens
    ]
    return " ".join(normalized_tokens)

def add_model_candidate(candidates: List[str], value: str) -> None:
    candidate = normalize_candidate_display(value)
    if not candidate:
        return
    if re.match(r"^[А-Яа-яЁё]", candidate):
        return
    if re.search(r"[А-Яа-яЁё]", candidate):
        return
    if is_measurement_token(candidate):
        return
    key = normalize_compare_key(candidate)
    if len(key) < 2:
        return
    if key not in {normalize_compare_key(item) for item in candidates}:
        candidates.append(candidate)

def code_model_tokens(value: str) -> List[str]:
    result: List[str] = []
    for token in re.findall(r"[A-Za-zА-Яа-яЁё0-9]+(?:[./_+\-][A-Za-zА-Яа-яЁё0-9]+)*", value):
        if not model_signal_token(token):
            continue
        normalized = normalize_candidate_display(token)
        if not normalized or re.match(r"^[А-Яа-яЁё]", normalized):
            continue
        if not any(char.isdigit() for char in normalized):
            continue
        if normalized not in result:
            result.append(normalized)
    return result

def generate_model_candidates(name: str, brand: str = "") -> List[str]:
    cleaned = technical_clean_model_text(name)
    detected_brand = find_brand_in_name(cleaned, brand)
    tail = tail_after_brand(cleaned, detected_brand)
    main_part = re.split(r"[,;|]", tail, maxsplit=1)[0]
    before_russian_description = candidate_until_russian_description(main_part)
    strong_block = first_model_block(main_part)

    candidates: List[str] = []
    for value in (before_russian_description, strong_block, main_part, tail):
        add_model_candidate(candidates, value)
    for token in code_model_tokens(main_part):
        add_model_candidate(candidates, token)

    base = before_russian_description or main_part or strong_block
    tokens = base.split()
    for start in range(1, min(4, len(tokens))):
        add_model_candidate(candidates, " ".join(tokens[start:]))
    for size in range(min(4, len(tokens)), 0, -1):
        add_model_candidate(candidates, " ".join(tokens[:size]))
    return candidates[:8]

def feed_index_add(index: Dict[str, Dict[str, object]], key: str, item: Dict[str, object]) -> None:
    if key and key not in index:
        index[key] = item

def index_feed_value(index: Dict[str, Dict[str, object]], value: str, item: Dict[str, object]) -> None:
    for key in compare_keys_for_value(value).values():
        feed_index_add(index, key, item)

def build_feed_index_from_xml(content: bytes, feed: Dict[str, object]) -> Dict[str, Dict[str, object]]:
    index: Dict[str, Dict[str, object]] = {}
    for _event, node in ET.iterparse(io.BytesIO(content), events=("end",)):
        children = list(node)
        if not children:
            continue
        values: Dict[str, str] = {}
        for child in children:
            key = str(child.tag).split("}")[-1].lower()
            values[key] = clean_text(child.text or "")
        explicit_values = [
            values.get(key, "")
            for key in ("vendorcode", "vendor_code", "model", "sku", "article", "articul")
            if values.get(key)
        ]
        name = values.get("name") or values.get("title") or ""
        item = {
            "source": str(feed.get("source") or ""),
            "source_label": str(feed.get("source_label") or feed.get("url") or "Фид"),
            "feed_name": name,
            "feed_url": values.get("url") or "",
            "raw": values,
        }
        for value in explicit_values:
            index_feed_value(index, value, {**item, "matched_feed_key": normalize_compare_key(value)})
        if not explicit_values and name:
            for candidate in generate_model_candidates(name):
                index_feed_value(index, candidate, {**item, "matched_feed_key": normalize_compare_key(candidate)})
        node.clear()
    return index

def build_file_import_feed_indexes() -> List[Dict[str, object]]:
    downloaded_feeds = download_feed_files()
    feed_indexes: List[Dict[str, object]] = []
    for feed in downloaded_feeds:
        filename = str(feed.get("filename") or "")
        path = source_feed_dir(str(feed.get("source") or "")) / filename
        try:
            index = build_feed_index_from_xml(path.read_bytes(), feed)
            feed_indexes.append({**feed, "index": index, "codes_count": len(index)})
        except Exception as exc:
            feed_indexes.append({**feed, "index": {}, "codes_count": 0, "error": str(exc)})
    with news_lock:
        news_settings["feed_storage"] = [
            {key: value for key, value in feed.items() if key != "index"}
            for feed in feed_indexes
        ]
        save_news_settings()
    save_logs()
    return feed_indexes

def match_candidates_against_feed_indexes(candidates: List[str], feed_indexes: List[Dict[str, object]]) -> Optional[Dict[str, object]]:
    for candidate in candidates:
        keys = compare_keys_for_value(candidate)
        for reason, key in keys.items():
            for feed in feed_indexes:
                index = feed.get("index", {})
                if isinstance(index, dict) and key in index:
                    match = dict(index[key])
                    match.update(
                        {
                            "selected_model": candidate,
                            "selected_reason": f"matched:{reason}",
                            "compare_key": key,
                            "matched_feed_key": str(match.get("matched_feed_key") or key),
                        }
                    )
                    return match
    return None

def missing_feed_labels(candidates: List[str], feed_indexes: List[Dict[str, object]]) -> List[str]:
    labels: List[str] = []
    keys = set()
    for candidate in candidates:
        keys.update(compare_keys_for_value(candidate).values())
    for feed in feed_indexes:
        index = feed.get("index", {})
        if not isinstance(index, dict) or not (keys & set(index.keys())):
            labels.append(str(feed.get("source_label") or feed.get("url") or "Фид"))
    return labels

def file_import_result_filename(original_filename: str) -> str:
    stem = safe_filename(Path(original_filename).stem or "file")
    return f"Новинки_{stem}_{datetime.now(MSK_TZ).strftime('%d-%m-%Y_%H-%M-%S')}.csv"

def compare_file_import_with_feeds() -> Dict[str, object]:
    row = get_file_import_row()
    path = current_file_import_path()
    if not path:
        raise ValueError("Файл не загружен")
    model_field = clean_text(str(row.model_field or ""))
    if not model_field:
        raise ValueError("Укажите название столбца модели")

    exclusions = normalize_file_import_exclusions(row.exclusions)
    replace_rules = normalize_file_import_rules_text(row.replace_rules)
    source_rows = read_file_import_rows(path, model_field)
    feed_indexes = build_file_import_feed_indexes()

    result_rows: List[Dict[str, object]] = []
    processed = excluded = found = missing = empty_model = 0
    for item in source_rows:
        name = str(item.get("name") or "")
        brand = str(item.get("brand") or "")
        if file_import_exclusion_matches(name, brand, exclusions):
            excluded += 1
            continue
        processed += 1
        prepared_model = prepare_file_import_model(name, replace_rules)
        candidates = generate_model_candidates(prepared_model, brand)
        if not candidates:
            empty_model += 1
            result_rows.append(
                {
                    "row": item.get("row_number"),
                    "name": name,
                    "brand": brand,
                    "model_candidates": "",
                    "selected_model": "",
                    "missing_on": "",
                }
            )
            continue
        match = match_candidates_against_feed_indexes(candidates, feed_indexes)
        if match:
            found += 1
            continue
        missing += 1
        selected_model = candidates[0]
        result_rows.append(
            {
                "row": item.get("row_number"),
                "name": name,
                "brand": brand,
                "model_candidates": " | ".join(candidates),
                "selected_model": selected_model,
                "missing_on": ", ".join(missing_feed_labels(candidates, feed_indexes)),
            }
        )

    file_meta = dict(row.file) if isinstance(row.file, dict) else {}
    original_filename = str(file_meta.get("original_filename") or file_meta.get("filename") or path.name)
    remove_file_import_export(row)
    result_path = EXPORT_DIR / file_import_result_filename(original_filename)
    with result_path.open("w", encoding="utf-8-sig", newline="") as csv_file:
        fieldnames = [
            "row",
            "name",
            "brand",
            "model_candidates",
            "selected_model",
            "missing_on",
        ]
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames, delimiter=";")
        writer.writeheader()
        writer.writerows(result_rows)

    file_meta["result_filename"] = result_path.name
    file_meta["result_created_at"] = datetime.now(MSK_TZ).isoformat(timespec="seconds")
    row.export_path = result_path.name
    row.file = file_meta
    return {
        "total_rows": len(source_rows),
        "processed_rows": processed,
        "excluded_rows": excluded,
        "model_not_found_rows": empty_model,
        "found_rows": found,
        "missing_rows": missing,
        "result_filename": result_path.name,
        "result_url": url_for("api_download_file_import_result"),
    }
